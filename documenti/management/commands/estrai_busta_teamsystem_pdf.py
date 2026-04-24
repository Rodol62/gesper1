"""
Estrae campi e valori da un PDF cedolino TeamSystem (pdfplumber + pypdf testo).

Esempio:
  python manage.py estrai_busta_teamsystem_pdf --json /path/busta.pdf

Excel strutturato (sezioni + competenze/trattenute + grezzo):
  python manage.py estrai_busta_teamsystem_pdf --excel busta_estratta.xlsx /path/busta.pdf

Solo due colonne grezze (Campo | Dato):
  python manage.py estrai_busta_teamsystem_pdf --excel busta.xlsx --solo-grezzo /path/busta.pdf
"""

import json
from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import Workbook

from documenti.estrazione_busta_teamsystem import estrai_busta_teamsystem_pdf


def _scrivi_excel_strutturato(wb: Workbook, data: dict) -> None:
    stru = data.get("struttura") or {}
    righe = stru.get("righe_sezioni") or []

    ws1 = wb.active
    ws1.title = "Riepilogo sezioni"
    ws1.append(["Sezione", "Campo", "Valore", "Fonte"])
    for row in righe:
        ws1.append(
            [
                row.get("sezione") or "",
                row.get("campo") or "",
                row.get("valore") or "",
                row.get("fonte") or "",
            ]
        )
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 44
    ws1.column_dimensions["C"].width = 58
    ws1.column_dimensions["D"].width = 18

    ws2 = wb.create_sheet("Competenze")
    ws2.append(["Codice", "Descrizione", "Ore/Giorni", "Base unitaria", "Importo"])
    for v in stru.get("voci_retributive") or []:
        ws2.append(
            [
                v.get("codice"),
                v.get("descrizione"),
                v.get("ore_giorni"),
                v.get("base_unitaria"),
                v.get("importo"),
            ]
        )
    for col, w in ("A", 8), ("B", 40), ("C", 14), ("D", 14), ("E", 14):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Trattenute codice")
    ws3.append(["Codice", "Descrizione", "Ore/Giorni", "Base/Altro", "Importo"])
    for v in stru.get("trattenute_righe") or []:
        ws3.append(
            [
                v.get("codice"),
                v.get("descrizione"),
                v.get("ore_giorni"),
                v.get("base_unitaria"),
                v.get("importo"),
            ]
        )
    for col, w in ("A", 8), ("B", 40), ("C", 14), ("D", 14), ("E", 14):
        ws3.column_dimensions[col].width = w

    ws4 = wb.create_sheet("Grezzo tabelle")
    ws4.append(["Campo", "Dato", "Fonte"])
    for r in data.get("campi_valori") or []:
        ws4.append([r.get("campo"), r.get("valore"), r.get("fonte")])
    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 72
    ws4.column_dimensions["C"].width = 22


class Command(BaseCommand):
    help = "Estrae campi/valori da PDF busta TeamSystem (nessun motore libro paga)."

    def add_arguments(self, parser):
        parser.add_argument(
            "pdf_path",
            nargs="?",
            default="/Users/rosario/Downloads/busta_03_2026_dip_19_p2.pdf",
            help="Percorso del PDF (default: busta esempio 03/2026 dip 19)",
        )
        parser.add_argument(
            "--json",
            action="store_true",
            help="Stampa JSON su stdout",
        )
        parser.add_argument(
            "--max-righe",
            type=int,
            default=0,
            help="Limita righe stampate in modalità testo (0 = tutte)",
        )
        parser.add_argument(
            "--excel",
            metavar="FILE.xlsx",
            default="",
            help="Esporta Excel (default: fogli Riepilogo + Competenze + Trattenute + Grezzo)",
        )
        parser.add_argument(
            "--solo-grezzo",
            action="store_true",
            help="Con --excel: un solo foglio «Campo | Dato» (estrazione tabellare grezza)",
        )

    def handle(self, *args, **options):
        path = options["pdf_path"]
        data = estrai_busta_teamsystem_pdf(path)
        rows = list(data.get("campi_valori") or [])

        excel_path = (options.get("excel") or "").strip()
        if excel_path:
            xlsx = Path(excel_path).expanduser().resolve()
            xlsx.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            if options.get("solo_grezzo"):
                ws = wb.active
                ws.title = "Grezzo"
                ws.append(["Campo", "Dato"])
                for r in rows:
                    ws.append([(r.get("campo") or "").strip(), (r.get("valore") or "").strip()])
                ws.column_dimensions["A"].width = 44
                ws.column_dimensions["B"].width = 72
            else:
                _scrivi_excel_strutturato(wb, data)
            wb.save(str(xlsx))
            n_flat = len((data.get("struttura") or {}).get("righe_sezioni") or [])
            self.stdout.write(
                self.style.SUCCESS(
                    f"Excel creato: {xlsx} — righe riepilogo sezioni: {n_flat}; "
                    f"record grezzi tabella: {len(rows)}."
                )
            )

        if options["json"]:
            self.stdout.write(json.dumps(data, ensure_ascii=False, indent=2, default=str))
            return

        if data.get("errore"):
            self.stderr.write(self.style.ERROR(data["errore"]))
        self.stdout.write(
            self.style.NOTICE(
                f"Pagine: {data.get('num_pagine')} | Campi grezzi: {len(rows)} | ok={data.get('ok')}"
            )
        )
        if excel_path and not options["max_righe"]:
            return

        lim = options["max_righe"] or None
        display = rows[:lim] if lim else rows
        for i, r in enumerate(display, 1):
            c = r.get("campo", "")
            v = r.get("valore", "")
            f = r.get("fonte", "")
            pg = r.get("pagina", "")
            self.stdout.write(f"{i:4d} [p{pg} {f}] {c}  =>  {v}")
