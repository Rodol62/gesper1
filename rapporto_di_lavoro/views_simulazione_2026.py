"""
Simulazione annua (organico e costo personale) — calcolo su scenario ruoli×quantità.
Usa il motore canonico mensile ripetuto sul periodo (es. anno solare 2026).
"""
import logging
import json
from decimal import Decimal
from datetime import date
from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect
from django.contrib import messages
from accounts.formatting import euro_it_str
from anagrafiche.models import Azienda
from .models import (
    CCNL,
    ParametroCCNLTurismo,
    ParametroMaggiorazione,
    ParametroScattiAnnuali,
    ParametroRatei,
    TipoContratto,
    SimulazioneOrganico,
    RuoloOrganico2026,
    PropostaAssunzione,
    RapportoDiLavoro,
)
from presenze.models import Presenza, RiepilogoMensilePresenze
from .utils_calcoli import (
    calcola_netto_dipendente,
    calcola_addizionale_regionale_sicilia,
    calcola_addizionale_comunale_stima,
)
from .services_simulazione import invoca_calcola_busta_paga_mese
from .utils_motore_paga import risolvi_parametro_contributi_ccnl
from .utils_calendario import get_giorni_lavorativi_mese as _get_gg_lav_mese
from .risoluzione_contratto_motore import calcola_scatto_totale_maturato as _calcola_scatto_totale

logger = logging.getLogger(__name__)

# Import costo_lavoro con flag di disponibilità
COSTO_LAVORO_ENABLED = False
CostoLavoroAzienda = None
DatiContrattuali = None
RuleEngine = None
CostiEventuali = None
Decontribuzioni = None

try:
    from costo_lavoro import CostoLavoroAzienda, DatiContrattuali, RuleEngine
    from costo_lavoro.models.costi_eventuali import CostiEventuali
    from costo_lavoro.models.decontribuzioni import Decontribuzioni
    COSTO_LAVORO_ENABLED = True
except Exception as _costo_lavoro_exc:
    logger.warning(
        "Modulo costo_lavoro non disponibile — simulazione userà calcoli legacy. "
        "Errore: %s",
        _costo_lavoro_exc,
    )


def _is_admin_only(user):
    """True se admin o superuser. Protegge da errori DB durante ``.exists()`` sui ruoli (evita 500 sul decorator)."""
    if not user.is_authenticated:
        return False
    if getattr(user, 'is_superuser', False):
        return True
    try:
        return user.has_ruolo('admin')
    except Exception:
        logger.exception('_is_admin_only: errore verifica ruolo (DB o sessione utente)')
        return False


def _get_sim_params(request):
    """Restituisce il QueryDict dei parametri simulazione.

    Priorità: POST (invio form) → GET (link legacy) → session (Excel/CreaProposte).
    Salva sempre nella session quando i parametri vengono trovati in POST o GET,
    così le view secondarie (Excel, CreaProposte) li trovano anche senza querystring.
    """
    from django.http import QueryDict
    if request.method == 'POST':
        post = request.POST
        # Marker nascosto sul form config: garantisce che il POST della simulazione annua
        # non venga scartato (es. parsing ambiguo, scenario senza righe dopo «Rimuovi tutti»).
        if post.get('sim2026_form') or any(str(k).startswith('ruolo_') for k in post.keys()):
            request.session['sim2026_querystring'] = post.urlencode()
            return post
    if any(k.startswith('ruolo_') for k in request.GET):
        request.session['sim2026_querystring'] = request.GET.urlencode()
        return request.GET
    qs = request.session.get('sim2026_querystring', '')
    if qs:
        try:
            if not isinstance(qs, str):
                qs = str(qs)
            return QueryDict(qs)
        except Exception:
            logger.exception('sim2026: session sim2026_querystring non leggibile (QueryDict)')
            try:
                del request.session['sim2026_querystring']
            except Exception:
                pass
            return request.GET
    return request.GET


def _get_azienda_operativa_per_utente(user, session):
    """Azienda da sessione (admin) o da profilo utente; ``has_ruolo`` in try per evitare 500 su DB."""
    is_admin_like = bool(getattr(user, 'is_superuser', False))
    if not is_admin_like:
        try:
            is_admin_like = user.has_ruolo('admin')
        except Exception:
            logger.exception('_get_azienda_operativa_per_utente: errore verifica ruolo admin')
            is_admin_like = False
    if is_admin_like:
        azienda_id = session.get('azienda_operativa_id')
        if azienda_id:
            return Azienda.objects.filter(id=azienda_id).first()
    return user.azienda if hasattr(user, 'azienda') else None


def _get_azienda_con_fallback(user, session):
    """Come _get_azienda_operativa_per_utente ma con fallback robusto.
    Usato dalle view Simulazione annua per garantire il salvataggio anche quando
    la sessione punta a un'azienda non più valida."""
    azienda = _get_azienda_operativa_per_utente(user, session)
    if azienda:
        return azienda
    if hasattr(user, 'azienda') and user.azienda_id:
        try:
            return user.azienda
        except Exception:
            pass
    return Azienda.objects.order_by('id').first()


def _prepara_ruoli_config_per_template(ruoli_config):
    """Aggiunge source_json_html per value= degli hidden (JSON valido + escape HTML).

    Nome senza underscore iniziale: i template Django non consentono ruolo._campo.
    """
    for r in ruoli_config or []:
        try:
            r['source_json_html'] = json.dumps(
                r.get('soggetti_riferimento') or [],
                ensure_ascii=False,
                separators=(',', ':'),
            )
        except (TypeError, ValueError):
            r['source_json_html'] = '[]'


def _somma_testate_ruoli(ruoli_config) -> int:
    """Somma le quantità (testate) dei ruoli nello scenario simulato."""
    tot = 0
    for r in ruoli_config or []:
        try:
            tot += int(r.get('quantita') or 0)
        except (TypeError, ValueError):
            pass
    return tot


def _tipo_rapporto_sim_prefill(tipo_contratto, data_fine_rapporto):
    """
    Valore tipo_rapporto atteso dal motore simulazione / costo_lavoro.
    TD/stagionale/apprendistato da TipoContratto; data fine esplicita (TD o proposta) → determinato.
    """
    if tipo_contratto:
        t = (getattr(tipo_contratto, 'tipo', None) or '').lower()
        if t == 'apprendistato':
            return 'apprendistato'
        if t.startswith('det_') or t.startswith('stag_'):
            return 'determinato'
    if data_fine_rapporto:
        return 'determinato'
    return 'indeterminato'


def _ruoli_precaricati_da_profili(azienda, anno=2026):
    """Genera ruoli_config precompilati per singolo soggetto (attivo/candidato)."""
    from anagrafiche.models import Dipendente

    if not azienda:
        return []

    dipendenti = list(
        Dipendente.objects.filter(azienda=azienda, stato__in=('attivo', 'candidato'))
        .order_by('cognome', 'nome', 'id')
    )
    if not dipendenti:
        return []

    dip_ids = [d.id for d in dipendenti]
    contratti_map = {}
    for c in (
        RapportoDiLavoro.objects.filter(azienda=azienda, dipendente_id__in=dip_ids)
        .select_related('tipo_contratto')
        .order_by('dipendente_id', '-data_modifica', '-id')
    ):
        contratti_map.setdefault(c.dipendente_id, c)

    proposte_map = {}
    for p in (
        PropostaAssunzione.objects.filter(azienda=azienda, dipendente_id__in=dip_ids)
        .select_related('tipo_contratto')
        .order_by('dipendente_id', '-data_modifica', '-id')
    ):
        proposte_map.setdefault(p.dipendente_id, p)

    riepiloghi = RiepilogoMensilePresenze.objects.filter(azienda=azienda, anno=anno, dipendente_id__in=dip_ids)
    riepiloghi_by_dip = {}
    for r in riepiloghi:
        riepiloghi_by_dip.setdefault(r.dipendente_id, {})[r.mese] = r

    livelli_ordered = []
    _seen_lv = set()
    for _lv in ParametroCCNLTurismo.objects.filter(attivo=True).order_by('livello', 'qualifica').values_list('livello', flat=True):
        if _lv not in _seen_lv:
            _seen_lv.add(_lv)
            livelli_ordered.append(_lv)
    livelli_validi = set(livelli_ordered)
    livello_ccnl_fallback = livelli_ordered[0] if livelli_ordered else ''
    livelli_ci = {str(x).lower(): x for x in livelli_ordered}

    def _eta_da_data_nascita(dn):
        if not dn:
            return None
        today = date.today()
        return today.year - dn.year - ((today.month, today.day) < (dn.month, dn.day))

    def _anni_da_data(d):
        if not d:
            return 0
        today = date.today()
        y = today.year - d.year - ((today.month, today.day) < (d.month, d.day))
        return max(0, y)

    ruoli = []
    for idx, dip in enumerate(dipendenti, start=1):
        c = contratti_map.get(dip.id)
        p = proposte_map.get(dip.id)
        src = c or p

        livello = ''
        tipo_contratto_id = ''
        data_inizio_obj = date(anno, 1, 1)
        data_fine_obj = date(anno, 12, 31)
        superminimo = Decimal('0')
        ind_turno = Decimal('0')
        ind_extra = Decimal('0')
        percettore_naspi = None
        tipo_incentivo = None
        tipo_rapporto_sim = 'indeterminato'

        if c:
            livello = (c.livello_ccnl or dip.livello or '').strip()
            tipo_contratto_id = str(c.tipo_contratto_id or '')
            data_inizio_obj = c.data_inizio_rapporto or data_inizio_obj
            data_fine_obj = c.data_fine_rapporto or data_fine_obj
            superminimo = Decimal(str(getattr(c, 'superminimo_mensile', None) or 0)).quantize(Decimal('0.01'))
            ind_extra = Decimal(c.edr_mensile or 0).quantize(Decimal('0.01'))
            tipo_rapporto_sim = _tipo_rapporto_sim_prefill(c.tipo_contratto, c.data_fine_rapporto)
        elif p:
            livello = (p.livello_ccnl or dip.livello or '').strip()
            tipo_contratto_id = str(p.tipo_contratto_id or '')
            data_inizio_obj = p.data_inizio_rapporto or data_inizio_obj
            data_fine_obj = p.data_fine_rapporto or data_fine_obj
            superminimo = Decimal(str(getattr(p, 'superminimo_mensile', None) or 0)).quantize(Decimal('0.01'))
            ind_turno = Decimal(p.indennita_mensile or 0).quantize(Decimal('0.01'))
            ind_extra = Decimal(p.edr_mensile or 0).quantize(Decimal('0.01'))
            tipo_rapporto_sim = _tipo_rapporto_sim_prefill(p.tipo_contratto, p.data_fine_rapporto)
        else:
            livello = (dip.livello or '').strip()
            if dip.data_assunzione:
                data_inizio_obj = dip.data_assunzione
            if dip.data_cessazione:
                data_fine_obj = dip.data_cessazione

        raw_l = (livello or '').strip()
        if raw_l in livelli_validi:
            livello = raw_l
        elif raw_l:
            livello = livelli_ci.get(raw_l.lower()) or ''
        else:
            livello = ''
        if not livello and livello_ccnl_fallback:
            livello = livello_ccnl_fallback

        from .utils_presenze import get_presenze_mese_aggregato

        calendario = {}
        by_mese = riepiloghi_by_dip.get(dip.id, {})
        for m in range(1, 13):
            r = by_mese.get(m)
            if Presenza.objects.filter(dipendente=dip, data__year=anno, data__month=m).exists():
                agg = get_presenze_mese_aggregato(dip, anno, m, azienda)
                calendario[m] = {
                    'ore_straord_diurno': agg['ore_straord_diurno'],
                    'ore_straord_notturno': agg['ore_straord_notturno'],
                    'ore_straord_festivo': agg['ore_straord_festivo'],
                    'ore_straord_domenica': agg['ore_straord_domenica'],
                    'ore_straord_nott_fest': agg['ore_straord_nott_fest'],
                    'ore_ordinarie_retribuite': agg['ore_ordinarie_retribuite'],
                    'ore_domenicali': agg['ore_domenicali'],
                    'giorni_festivi': agg['ore_festivi_lavorati'],
                    'giorni_assenza': agg['giorni_assenza_ingiust'],
                    'trattenute_extra_mese': Decimal('0'),
                    'competenze_extra_non_imponibili': Decimal('0'),
                }
            elif r:
                calendario[m] = {
                    'ore_straord_diurno': Decimal(str(getattr(r, 'ore_straord_diurno', 0) or 0)),
                    'ore_straord_notturno': Decimal(str(getattr(r, 'ore_straord_notturno', 0) or 0)),
                    'ore_straord_festivo': Decimal(str(getattr(r, 'ore_straord_festivo', 0) or 0)),
                    'ore_straord_domenica': Decimal(str(getattr(r, 'ore_straord_domenica', 0) or 0)),
                    'ore_straord_nott_fest': Decimal(str(getattr(r, 'ore_straord_nott_fest', 0) or 0)),
                    'ore_ordinarie_retribuite': Decimal(str(getattr(r, 'ore_ordinarie', 0) or 0)),
                    'ore_domenicali': Decimal(str(getattr(r, 'ore_domenicali', 0) or 0)),
                    'giorni_festivi': Decimal(str(getattr(r, 'ore_festivi', 0) or 0)),
                    'giorni_assenza': Decimal(str(getattr(r, 'giorni_assenza_ingiust', 0) or 0)),
                    'trattenute_extra_mese': Decimal('0'),
                    'competenze_extra_non_imponibili': Decimal('0'),
                }
            else:
                calendario[m] = {
                    'ore_straord_diurno': Decimal('0'),
                    'ore_straord_notturno': Decimal('0'),
                    'ore_straord_festivo': Decimal('0'),
                    'ore_straord_domenica': Decimal('0'),
                    'ore_straord_nott_fest': Decimal('0'),
                    'ore_ordinarie_retribuite': Decimal('0'),
                    'ore_domenicali': Decimal('0'),
                    'giorni_festivi': Decimal('0'),
                    'giorni_assenza': Decimal('0'),
                    'trattenute_extra_mese': Decimal('0'),
                    'competenze_extra_non_imponibili': Decimal('0'),
                }

        cognome_nome = f"{dip.cognome} {dip.nome}".strip()
        # Etichetta mansione: da contratto / proposta (posizione CCNL), altrimenti anagrafica
        if c and (c.posizione or '').strip():
            mansione_label = c.posizione.strip()
        elif p and (p.posizione or '').strip():
            mansione_label = p.posizione.strip()
        elif p and (p.titolo or '').strip():
            mansione_label = (p.titolo or '').strip()[:120]
        else:
            mansione_label = dip.get_mansione_display() if dip.mansione else (dip.ruolo or '')
        anni_anz = _anni_da_data(dip.data_assunzione or data_inizio_obj)

        ruoli.append({
            'id': str(idx),
            'nome': cognome_nome,
            'quantita': 1,
            'livello': livello,
            'tipo_contratto_id': tipo_contratto_id,
            'tipo_rapporto': tipo_rapporto_sim,
            'data_inizio': data_inizio_obj,
            'data_fine': data_fine_obj,
            'regione': 'sicilia',
            'eta': _eta_da_data_nascita(dip.data_nascita),
            'categoria': None,
            'percettore_naspi': percettore_naspi,
            'tipo_incentivo': tipo_incentivo,
            'anni_anzianita': anni_anz,
            'superminimo': superminimo,
            'indennita_turno': ind_turno,
            'indennita_extra': ind_extra,
            'premio_risultato_annuo': Decimal('0'),
            'calendario_mensile': calendario,
            'dipendente_id': dip.id,
            'stato_soggetto': dip.stato,
            'mansione_label': mansione_label,
            'origine_dati': 'auto_profilo',
            'nominativi_riferimento': cognome_nome,
            'soggetti_riferimento': [{
                'dipendente_id': dip.id,
                'stato': dip.stato,
                'cognome_nome': cognome_nome,
                'mansione': mansione_label,
            }],
        })

    return ruoli


def _normalizza_ccnl_key(ccnl_str):
    """Normalizza il nome CCNL a chiave per costo_lavoro."""
    ccnl_lower = (ccnl_str or '').lower()
    if 'turismo' in ccnl_lower or 'fipe' in ccnl_lower or 'ristorazione' in ccnl_lower:
        return 'turismo'
    if 'commercio' in ccnl_lower or 'terziario' in ccnl_lower:
        return 'commercio'
    return 'turismo'


def _stima_dimensione_azienda(azienda):
    """Stima dimensione azienda per aliquote INPS."""
    if not azienda:
        return 'piccola'
    try:
        count = azienda.dipendenti.count()
        if count > 50:
            return 'grande'
        if count > 15:
            return 'media'
    except Exception:
        pass
    return 'piccola'


def _normalizza_livello(livello):
    return str(livello or '').strip().lower().replace(' ', '')


def _carica_regole_lista(rule_engine, category):
    """Carica il JSON completo come lista (senza filtro values)."""
    if not rule_engine:
        return []
    try:
        raw = rule_engine.loader.load(f"{category}.json")
        if isinstance(raw, list):
            return raw
        if isinstance(raw, dict):
            return [raw]
    except Exception:
        return []
    return []


def _risolvi_livello_da_mansione(mansione, rule_engine):
    """Ricava il livello da classificazione_fipe.json in base alla mansione testuale."""
    mansione_norm = str(mansione or '').strip().lower()
    if not mansione_norm:
        return ''

    regole = _carica_regole_lista(rule_engine, 'classificazione_fipe')
    for regola in regole:
        if not isinstance(regola, dict):
            continue
        for m in regola.get('mansioni', []):
            if str(m).strip().lower() == mansione_norm:
                return str(regola.get('livello', '')).strip()
    return ''


def _trova_regola_tabellare(rule_engine, ccnl_key, livello_key, anno, azienda_minore_flag):
    """Trova la riga tabellare in ccnl_fipe_piccola_ristorazione.json."""
    regole_ccnl = _carica_regole_lista(rule_engine, 'ccnl_fipe_piccola_ristorazione')
    livello_norm = _normalizza_livello(livello_key)

    for regola in regole_ccnl:
        if not isinstance(regola, dict):
            continue
        if _normalizza_ccnl_key(regola.get('ccnl', '')) != ccnl_key:
            continue
        if _normalizza_livello(regola.get('livello', '')) != livello_norm:
            continue
        if regola.get('anno') != anno:
            continue
        if regola.get('azienda_minore') != azienda_minore_flag:
            continue
        if isinstance(regola.get('values'), dict):
            return regola
    return None


def _trova_decontribuzione(
    rule_engine,
    *,
    regione=None,
    tipo_contratto=None,
    categoria=None,
    percettore_naspi=None,
    tipo_incentivo=None,
    eta=None,
):
    """Seleziona una decontribuzione coerente da decontribuzioni.json."""
    regole = _carica_regole_lista(rule_engine, 'decontribuzioni')
    tipo_incentivo_norm = (tipo_incentivo or '').strip().lower()

    # Per evitare applicazioni inattese:
    # - senza tipo_incentivo => applica solo regola territoriale (es. Sud/Sicilia)
    # - con tipo_incentivo => filtra la famiglia di incentivo richiesta
    def _rule_matches_incentivo(rule):
        name = str(rule.get('name', '')).strip().lower()
        if not tipo_incentivo_norm:
            has_specific_keys = any(k in rule for k in ('eta', 'categoria', 'percettore_naspi', 'tipo_contratto'))
            return ('regione' in rule) and (not has_specific_keys)

        if tipo_incentivo_norm in ('sud', 'decontribuzione_sud'):
            return 'sud' in name
        if tipo_incentivo_norm in ('under36', 'under_36'):
            return 'under 36' in name
        if tipo_incentivo_norm in ('donna_svantaggiata', 'donne_svantaggiate'):
            return 'donn' in name and 'svant' in name
        if tipo_incentivo_norm in ('naspi',):
            return 'naspi' in name
        # se non riconosciuto, non forza filtro su nome
        return True

    contesto = {
        'regione': regione,
        'tipo_contratto': tipo_contratto,
        'categoria': categoria,
        'percettore_naspi': percettore_naspi,
        'tipo_incentivo': tipo_incentivo,
    }

    best = None
    best_priority = -10**9
    for regola in regole:
        if not isinstance(regola, dict):
            continue

        if not _rule_matches_incentivo(regola):
            continue

        ok = True
        for key in ('regione', 'tipo_contratto', 'categoria', 'percettore_naspi', 'tipo_incentivo'):
            if key in regola and contesto.get(key) is not None and regola.get(key) != contesto.get(key):
                ok = False
                break
            if key in regola and contesto.get(key) is None:
                ok = False
                break

        # Gestione regole con condizione età (es. Under 36)
        if ok and 'eta' in regola:
            cond_eta = regola.get('eta')
            if eta is None:
                ok = False
            elif isinstance(cond_eta, dict):
                op = cond_eta.get('op', '==')
                value = cond_eta.get('value')
                if op == '<' and not (eta < value):
                    ok = False
                elif op == '<=' and not (eta <= value):
                    ok = False
                elif op == '>' and not (eta > value):
                    ok = False
                elif op == '>=' and not (eta >= value):
                    ok = False
                elif op == '==' and not (eta == value):
                    ok = False
            else:
                if eta != cond_eta:
                    ok = False

        if not ok:
            continue

        prio = int(regola.get('priority', 0) or 0)
        if prio > best_priority:
            best = regola
            best_priority = prio

    if best and isinstance(best.get('values'), dict):
        out = dict(best['values'])
        out['_rule_name'] = best.get('name', '')
        return out
    return {}


def _calcola_costo_azienda_ruolo_costo_lavoro(
    parametro,
    coeff_ore,
    giorni_lavorativi_mese,
    giorni_attivi,
    anno,
    mese_num,
    azienda_operativa,
    rule_engine,
    regione='sicilia',
    parametro_tipo_contratto='indeterminato',
    categoria=None,
    percettore_naspi=None,
    tipo_incentivo=None,
    eta=None,
):
    """Calcola costo azienda per ruolo usando il modulo costo_lavoro."""
    if not COSTO_LAVORO_ENABLED or not rule_engine or not DatiContrattuali or not CostoLavoroAzienda:
        return None
    
    try:
        ccnl_key = _normalizza_ccnl_key(parametro.ccnl)
        livello_key = parametro.livello.strip()
        
        dimensione = _stima_dimensione_azienda(azienda_operativa)
        azienda_minore_flag = (dimensione == 'piccola')

        regola_tabellare = _trova_regola_tabellare(
            rule_engine=rule_engine,
            ccnl_key=ccnl_key,
            livello_key=livello_key,
            anno=anno,
            azienda_minore_flag=azienda_minore_flag,
        )
        
        if not regola_tabellare or 'values' not in regola_tabellare:
            return None
        
        values = regola_tabellare['values']
        totale_mensile_tabellare = Decimal(str(values.get('totale_mensile', 0)))
        mensilita = int(values.get('mensilita', 12))
        
        # Base mensile tabellare full-time/part-time (NON pro-rata giorni qui)
        # Il pro-rata giorni viene applicato internamente dal calcolatore
        # tramite giorni_lavorati/giorni_lavorativi_mese.
        lordo_base_full = (totale_mensile_tabellare * coeff_ore).quantize(Decimal('0.01'))
        
        # Model contrattuali.py
        dati = DatiContrattuali(
            retribuzione_lorda_mensile=float(lordo_base_full),
            giorni_lavorativi_mese=giorni_lavorativi_mese,
            giorni_lavorati=giorni_attivi,
            mensilita=mensilita,
            ore_settimanali=40,  # Standard
            livello=livello_key,
            ccnl=ccnl_key,
        )

        try:
            n_dip = azienda_operativa.dipendenti.count() if azienda_operativa else 1
        except Exception:
            n_dip = 1

        # inps.json
        regole_inps = rule_engine.get('inps', ccnl=ccnl_key, dimensione=n_dip)
        if not isinstance(regole_inps, dict) or not regole_inps:
            return None

        # inail.json
        regole_inail = rule_engine.get('inail', ccnl=ccnl_key)
        inail_values = regole_inail if isinstance(regole_inail, dict) else {}

        # ratei.json (fallback generale)
        regole_ratei = rule_engine.get('ratei')
        ratei_values = regole_ratei if isinstance(regole_ratei, dict) else {}

        # Ratei da tabella CCNL (prioritari per 13^/14^ e altri ratei)
        ratei_ccnl_values = {
            'rateo_ferie': values.get('rateo_ferie'),
            'rateo_permessi': values.get('rateo_permessi'),
            'rateo_tredicesima': values.get('rateo_tredicesima'),
            'rateo_quattordicesima': values.get('rateo_quattordicesima'),
        }
        # rimuove chiavi nulle/assenti
        ratei_ccnl_values = {k: v for k, v in ratei_ccnl_values.items() if v is not None}

        # decontribuzioni.json
        decontrib_values = _trova_decontribuzione(
            rule_engine,
            regione=regione,
            tipo_contratto=parametro_tipo_contratto,
            categoria=categoria,
            percettore_naspi=percettore_naspi,
            tipo_incentivo=tipo_incentivo,
            eta=eta,
        )

        # Model costi_eventuali.py / decontribuzioni.py
        extra = CostiEventuali() if CostiEventuali else None
        decontrib_rule_name = decontrib_values.pop('_rule_name', '') if isinstance(decontrib_values, dict) else ''
        decontrib = Decontribuzioni(**decontrib_values).__dict__ if (Decontribuzioni and decontrib_values) else {}

        # Model contributivi.py (via CostoLavoroAzienda)
        # Ordine merge: base INPS/INAIL -> ratei globali -> ratei CCNL (priorità massima)
        combined_regole = {**regole_inps, **inail_values, **ratei_values, **ratei_ccnl_values}

        # Calcolo
        calc_azienda = CostoLavoroAzienda(
            contrattuali=dati,
            regole_inps=combined_regole,
            regole_decontrib=decontrib,
            costi_eventuali=extra,
        )
        risultato = calc_azienda.calcola()
        
        if not risultato:
            return None

        # risultato.py -> chiavi effettive restituite dal calcolatore
        inps_coeff_base = Decimal(str(
            combined_regole.get('aliquota_inps_azienda', 0)
            + combined_regole.get('contributo_fis', 0)
            + combined_regole.get('contributo_naspi', 0)
            + combined_regole.get('contributo_td', 0)
        ))
        inps_coeff_eff = Decimal('0.0000')
        lordo_calc = Decimal(str(risultato.get('retribuzione_proporzionata', 0)))
        inps_calc = Decimal(str(risultato.get('contributi_inps', 0)))
        if lordo_calc > 0:
            inps_coeff_eff = (inps_calc / lordo_calc).quantize(Decimal('0.0001'))

        return {
            'lordo_unit': Decimal(str(risultato.get('retribuzione_proporzionata', 0))).quantize(Decimal('0.01')),
            'inps_azienda_unit': Decimal(str(risultato.get('contributi_inps', 0))).quantize(Decimal('0.01')),
            'inail_unit': Decimal(str(risultato.get('premio_inail', 0))).quantize(Decimal('0.01')),
            'tfr_unit': Decimal(str(risultato.get('tfr', 0))).quantize(Decimal('0.01')),
            'rateo_13_unit': Decimal(str(risultato.get('rateo_tredicesima', 0))).quantize(Decimal('0.01')),
            'rateo_14_unit': Decimal(str(risultato.get('rateo_quattordicesima', 0))).quantize(Decimal('0.01')),
            'costo_azienda_unit': Decimal(str(risultato.get('costo_totale', 0))).quantize(Decimal('0.01')),
            'aliquota_tfr_coeff': Decimal(str(combined_regole.get('aliquota_tfr', Decimal('0.074')))),
            'rateo_13_coeff': Decimal(str(combined_regole.get('rateo_tredicesima', 0))),
            'rateo_14_coeff': Decimal(str(combined_regole.get('rateo_quattordicesima', 0))),
            'inps_coeff_base': inps_coeff_base.quantize(Decimal('0.0001')),
            'inps_coeff_eff': inps_coeff_eff,
            'decontrib_rule_name': decontrib_rule_name,
            'decontrib_tipo': str(decontrib.get('tipo', '')),
            'decontrib_valore': Decimal(str(decontrib.get('valore', 0))) if decontrib else Decimal('0'),
            'lordo_base_full': lordo_base_full,
            'regola_tabellare': values,
        }
    except Exception as e:
        logger.error("Errore calcolo costo_lavoro per livello '%s': %s", getattr(parametro, 'livello', '?'), e)
        return None


def _parse_quantita_ruolo_sim(raw) -> int:
    """Quantità testate scenario: intero ≥ 0; stringa vuota o non numerica → 0 (evita 500 su GET/POST)."""
    if raw is None:
        return 0
    if isinstance(raw, bool):
        return int(raw)
    if isinstance(raw, int):
        return max(0, raw)
    try:
        from decimal import Decimal as _Dec

        if isinstance(raw, _Dec):
            return max(0, int(raw))
    except Exception:
        pass
    s = str(raw).strip()
    if not s:
        return 0
    try:
        return max(0, int(s))
    except (TypeError, ValueError):
        return 0


def _safe_int_nonnegative(raw, default=0) -> int:
    """Intero ≥ 0 per anni anzianità / campi simili; valori non numerici → default."""
    if raw is None:
        return default
    if isinstance(raw, bool):
        return default if not raw else 1
    if isinstance(raw, int):
        return max(0, raw)
    try:
        from decimal import Decimal as _Dec

        if isinstance(raw, _Dec):
            return max(0, int(raw))
    except Exception:
        pass
    s = str(raw).strip()
    if not s:
        return default
    try:
        return max(0, int(s))
    except (TypeError, ValueError):
        return default


def _build_ruoli_config(request):
    """Estrae configurazione ruoli da POST, GET o session."""
    params = _get_sim_params(request)
    ruoli = []
    for key, value in params.items():
        if key.startswith('ruolo_'):
            ruolo_id = key.split('_')[1]
            nome = params.get(f'nome_{ruolo_id}', '')
            quantita = _parse_quantita_ruolo_sim(params.get(f'qta_{ruolo_id}', ''))
            livello = (params.get(f'livello_{ruolo_id}', '') or '').strip()
            tipo_contratto_id = params.get(f'tipo_{ruolo_id}', '')
            data_inizio_str = params.get(f'data_inizio_{ruolo_id}', '2026-01-01')
            data_fine_str = params.get(f'data_fine_{ruolo_id}', '2026-12-31')
            regione = (params.get(f'regione_{ruolo_id}', 'sicilia') or 'sicilia').strip().lower()
            categoria = (params.get(f'categoria_{ruolo_id}', '') or '').strip().lower() or None
            tipo_incentivo = (params.get(f'tipo_incentivo_{ruolo_id}', '') or '').strip().lower() or None

            tipo_rapporto = (params.get(f'tipo_rapporto_{ruolo_id}', '') or '').strip().lower() or 'indeterminato'

            percettore_naspi_raw = (params.get(f'percettore_naspi_{ruolo_id}', '') or '').strip().lower()
            percettore_naspi = None
            if percettore_naspi_raw == 'si':
                percettore_naspi = True
            elif percettore_naspi_raw == 'no':
                percettore_naspi = False

            eta_raw = (params.get(f'eta_{ruolo_id}', '') or '').strip()
            eta = None
            if eta_raw:
                try:
                    eta = int(eta_raw)
                except ValueError:
                    eta = None

            # Anzianità (anni interi, default 0)
            anni_anz_raw = (params.get(f'anni_anz_{ruolo_id}', '') or '').strip()
            try:
                anni_anzianita = max(0, int(anni_anz_raw)) if anni_anz_raw else 0
            except ValueError:
                anni_anzianita = 0

            # Superminimo (€/mese, solo alcuni ruoli)
            superminimo_raw = (params.get(f'superminimo_{ruolo_id}', '') or '').strip()
            try:
                superminimo = Decimal(superminimo_raw) if superminimo_raw else Decimal('0')
            except Exception:
                superminimo = Decimal('0')

            # Indennità di turno (€/mese)
            ind_turno_raw = (params.get(f'ind_turno_{ruolo_id}', '') or '').strip()
            try:
                indennita_turno = Decimal(ind_turno_raw) if ind_turno_raw else Decimal('0')
            except Exception:
                indennita_turno = Decimal('0')

            # Indennità extra €/mese (EDR, El.Dis.Bil., o qualsiasi voce aggiuntiva)
            ind_extra_raw = (params.get(f'ind_extra_{ruolo_id}', '') or '').strip()
            try:
                indennita_extra = Decimal(ind_extra_raw) if ind_extra_raw else Decimal('0')
            except Exception:
                indennita_extra = Decimal('0')

            # Premio di risultato annuo (€/anno, placeholder — default 0)
            premio_raw = (params.get(f'premio_{ruolo_id}', '') or '').strip()
            try:
                premio_risultato_annuo = Decimal(premio_raw) if premio_raw else Decimal('0')
            except Exception:
                premio_risultato_annuo = Decimal('0')

            # Dipendente opzionale per importazione presenze reali
            dip_raw = (params.get(f'dip_{ruolo_id}', '') or '').strip()
            dipendente_id = int(dip_raw) if dip_raw.isdigit() else None
            origine_dati = (params.get(f'origine_{ruolo_id}', '') or 'manuale').strip() or 'manuale'
            nominativi_riferimento = (params.get(f'source_nom_{ruolo_id}', '') or '').strip()
            soggetti_raw = (params.get(f'source_json_{ruolo_id}', '') or '').strip()
            soggetti_riferimento = []
            if soggetti_raw:
                try:
                    parsed = json.loads(soggetti_raw)
                    if isinstance(parsed, list):
                        soggetti_riferimento = parsed
                except Exception:
                    soggetti_riferimento = []

            # Calendario mensile straordinari/maggiorazioni
            # Chiavi: m{mese}_sd_{ruolo_id}    → ore straord. diurno
            #         m{mese}_sn_{ruolo_id}    → ore straord. notturno
            #         m{mese}_sf_{ruolo_id}    → ore straord. festivo (no dom.)
            #         m{mese}_sdom_{ruolo_id}  → ore straord. domenica
            #         m{mese}_snf_{ruolo_id}   → ore straord. nott-fest
            #         m{mese}_ord_{ruolo_id}   → ore ordinarie retribuite (logica cedolino)
            #         m{mese}_dom_{ruolo_id}   → ore domenicali
            #         m{mese}_fest_{ruolo_id}  → ore festive lavorate
            #         m{mese}_assenze_{ruolo_id} → giorni assenza ingiust.
            #         m{mese}_tratt_{ruolo_id} → trattenute extra mese (netto)
            #         m{mese}_comp_{ruolo_id}  → competenze extra non imponibili (netto)
            calendario_mensile = {}
            for mese in range(1, 13):
                def _dec(param_suffix, _p=params, _m=mese, _rid=ruolo_id):
                    raw = (_p.get(f'm{_m}_{param_suffix}_{_rid}', '') or '').strip()
                    try:
                        return Decimal(raw) if raw else Decimal('0')
                    except Exception:
                        return Decimal('0')

                calendario_mensile[mese] = {
                    'ore_straord_diurno':    _dec('sd'),
                    'ore_straord_notturno':  _dec('sn'),
                    'ore_straord_festivo':   _dec('sf'),
                    'ore_straord_domenica':  _dec('sdom'),
                    'ore_straord_nott_fest': _dec('snf'),
                    'ore_ordinarie_retribuite': _dec('ord'),
                    'ore_domenicali':        _dec('dom'),
                    'giorni_festivi':        _dec('fest'),
                    'giorni_assenza':        _dec('assenze'),
                    'trattenute_extra_mese': _dec('tratt'),
                    'competenze_extra_non_imponibili': _dec('comp'),
                }

            # Parse date
            try:
                data_inizio = date.fromisoformat(data_inizio_str) if data_inizio_str else date(2026, 1, 1)
            except (ValueError, AttributeError):
                data_inizio = date(2026, 1, 1)

            try:
                data_fine = date.fromisoformat(data_fine_str) if data_fine_str else date(2026, 12, 31)
            except (ValueError, AttributeError):
                data_fine = date(2026, 12, 31)

            # Includi sempre le testate con qta>0: il motore gestisce livello mancante (riga "missing").
            # Escludere qui le righe senza livello lasciava ruoli_config vuoto e la simulazione senza dati.
            if quantita > 0:
                ruoli.append({
                    'id': ruolo_id,
                    'nome': nome,
                    'quantita': quantita,
                    'livello': livello,
                    'tipo_contratto_id': tipo_contratto_id,
                    'tipo_rapporto': tipo_rapporto,
                    'data_inizio': data_inizio,
                    'data_fine': data_fine,
                    'regione': regione,
                    'categoria': categoria,
                    'percettore_naspi': percettore_naspi,
                    'tipo_incentivo': tipo_incentivo,
                    'eta': eta,
                    # Voci variabili
                    'anni_anzianita': anni_anzianita,
                    'superminimo': superminimo,
                    'indennita_turno': indennita_turno,
                    'indennita_extra': indennita_extra,
                    'premio_risultato_annuo': premio_risultato_annuo,
                    'dipendente_id': dipendente_id,
                    'calendario_mensile': calendario_mensile,
                    'origine_dati': origine_dati,
                    'nominativi_riferimento': nominativi_riferimento,
                    'soggetti_riferimento': soggetti_riferimento,
                })
    return ruoli


def _normalizza_ruolo_per_motore(ruolo):
    """Adatta un ruolo (tracciato soggetto) ai parametri attesi dal motore simulazione."""
    def _dec(v, default='0'):
        try:
            return Decimal(str(v if v is not None else default))
        except Exception:
            return Decimal(default)

    def _to_date(v, fallback):
        if isinstance(v, date):
            return v
        if isinstance(v, str) and v:
            try:
                return date.fromisoformat(v)
            except Exception:
                return fallback
        return fallback

    _qty = max(1, _parse_quantita_ruolo_sim(ruolo.get('quantita')))
    base = {
        'id': ruolo.get('id'),
        'nome': (ruolo.get('nome') or '').strip(),
        'quantita': _qty,
        'livello': (ruolo.get('livello') or '').strip(),
        'tipo_contratto_id': str(ruolo.get('tipo_contratto_id') or ''),
        'data_inizio': _to_date(ruolo.get('data_inizio'), date(2026, 1, 1)),
        'data_fine': _to_date(ruolo.get('data_fine'), date(2026, 12, 31)),
        'regione': ((ruolo.get('regione') or 'sicilia').strip().lower()),
        'eta': ruolo.get('eta'),
        'percettore_naspi': ruolo.get('percettore_naspi'),
        'tipo_incentivo': (ruolo.get('tipo_incentivo') or None),
        'anni_anzianita': _safe_int_nonnegative(ruolo.get('anni_anzianita'), 0),
        'superminimo': _dec(ruolo.get('superminimo')),
        'indennita_turno': _dec(ruolo.get('indennita_turno')),
        'indennita_extra': _dec(ruolo.get('indennita_extra')),
        'premio_risultato_annuo': _dec(ruolo.get('premio_risultato_annuo')),
        'tipo_rapporto': (ruolo.get('tipo_rapporto') or 'indeterminato'),
        'categoria': ruolo.get('categoria'),
        'dipendente_id': ruolo.get('dipendente_id'),
        'stato_soggetto': ruolo.get('stato_soggetto'),
        'mansione_label': ruolo.get('mansione_label'),
        'origine_dati': ruolo.get('origine_dati'),
        'nominativi_riferimento': ruolo.get('nominativi_riferimento'),
        'soggetti_riferimento': ruolo.get('soggetti_riferimento') or [],
    }

    cal_raw = ruolo.get('calendario_mensile') or {}
    cal = {}
    for m in range(1, 13):
        md = cal_raw.get(m) or cal_raw.get(str(m)) or {}
        cal[m] = {
            'ore_straord_diurno': _dec(md.get('ore_straord_diurno')),
            'ore_straord_notturno': _dec(md.get('ore_straord_notturno')),
            'ore_straord_festivo': _dec(md.get('ore_straord_festivo')),
            'ore_straord_domenica': _dec(md.get('ore_straord_domenica')),
            'ore_straord_nott_fest': _dec(md.get('ore_straord_nott_fest')),
            'ore_ordinarie_retribuite': _dec(md.get('ore_ordinarie_retribuite')),
            'ore_domenicali': _dec(md.get('ore_domenicali')),
            'giorni_festivi': _dec(md.get('giorni_festivi')),
            'giorni_assenza': _dec(md.get('giorni_assenza')),
            'trattenute_extra_mese': _dec(md.get('trattenute_extra_mese')),
            'competenze_extra_non_imponibili': _dec(md.get('competenze_extra_non_imponibili')),
        }
    base['calendario_mensile'] = cal
    return base


def _calcola_giorni_attivi_nel_mese(anno, mese_num, data_inizio, data_fine):
    """
    Calcola i giorni in cui il dipendente è attivo nel mese specificato.
    Considera l'intersezione tra il periodo di lavoro e il mese.
    """
    from calendar import monthrange
    
    # Primo e ultimo giorno del mese
    primo_giorno_mese = date(anno, mese_num, 1)
    giorni_nel_mese = monthrange(anno, mese_num)[1]
    ultimo_giorno_mese = date(anno, mese_num, giorni_nel_mese)

    if data_inizio is None:
        data_inizio = primo_giorno_mese
    if data_fine is None:
        data_fine = ultimo_giorno_mese

    # Calcola intersezione
    inizio_effettivo = max(primo_giorno_mese, data_inizio)
    fine_effettiva = min(ultimo_giorno_mese, data_fine)
    
    # Se non c'è sovrapposizione, giorni attivi = 0
    if inizio_effettivo > fine_effettiva:
        return 0
    
    # Altrimenti calcola i giorni
    giorni_attivi = (fine_effettiva - inizio_effettivo).days + 1
    return giorni_attivi


def _conta_mesi_ccnl(data_inizio, data_fine, ref_start, ref_end):
    """
    Conta i mesi di servizio per 13ª/14ª secondo regola CCNL FIPE (art. 214-220 CCNL 2023):
      - frazioni di mese >= 15 giorni = 1 mese intero
      - frazioni di mese < 15 giorni  = non contato

    Parametri:
      data_inizio / data_fine  : periodo effettivo di lavoro del dipendente
      ref_start / ref_end      : periodo di riferimento della mensilità
        - 13ª: ref = 01/01/anno – 31/12/anno
        - 14ª: ref = 01/07/(anno-1) – 30/06/anno  (CCNL FIPE)
    """
    from calendar import monthrange as _mr
    if data_inizio is None:
        data_inizio = ref_start
    if data_fine is None:
        data_fine = ref_end
    eff_start = max(data_inizio, ref_start)
    eff_end   = min(data_fine,   ref_end)
    if eff_start > eff_end:
        return 0
    count = 0
    cur_year, cur_month = eff_start.year, eff_start.month
    while True:
        first = date(cur_year, cur_month, 1)
        last  = date(cur_year, cur_month, _mr(cur_year, cur_month)[1])
        seg_start = max(eff_start, first)
        seg_end   = min(eff_end,   last)
        if seg_start <= seg_end:
            days = (seg_end - seg_start).days + 1
            if days >= 15:
                count += 1
        if (cur_year, cur_month) >= (eff_end.year, eff_end.month):
            break
        if cur_month == 12:
            cur_year, cur_month = cur_year + 1, 1
        else:
            cur_month += 1
    return count


def _giorni_convenzionali_su_base_26(giorni_attivi_calendario, giorni_nel_mese):
    """
    Converte i giorni calendario attivi in giorni convenzionali su base 26.
    - mese pieno -> 26
    - mese parziale -> proporzione su 26
    """
    if giorni_attivi_calendario <= 0 or giorni_nel_mese <= 0:
        return 0

    valore = (Decimal(giorni_attivi_calendario) * Decimal('26')) / Decimal(giorni_nel_mese)
    # arrotondamento commerciale al giorno convenzionale più vicino
    giorni_conv = int(valore.quantize(Decimal('1')))
    if giorni_conv < 0:
        return 0
    if giorni_conv > 26:
        return 26
    return giorni_conv


def _calcola_simulazione_2026(request):
    """
    Calcola simulazione annuale 2026 (12 mesi).
    Ritorna lista di 12 dict, uno per mese.
    """
    from calendar import monthrange as _monthrange

    anno = 2026
    azienda_operativa = _get_azienda_operativa_per_utente(request.user, request.session)
    rule_engine = None
    if COSTO_LAVORO_ENABLED and RuleEngine:
        rule_engine = RuleEngine()

    # Divisore: 173,33 (standard), 172, oppure 26 (FIPE/Turismo giornaliero)
    _params = _get_sim_params(request)
    _div_raw = (_params.get('divisore', '173.33') or '173.33').strip()
    if _div_raw == '172':
        divisore_ore = Decimal('172')
    elif _div_raw == '26':
        divisore_ore = Decimal('26')
    else:
        divisore_ore = Decimal('173.33')

    ruoli_config = _build_ruoli_config(request)
    tipi_contratto = TipoContratto.objects.filter(attivo=True).order_by('nome')
    tipi_contratto_by_id = {tc.id: tc for tc in tipi_contratto}

    # Cache parametri CCNL per mese: usa la versione con la decorrenza più recente
    # non superiore al primo giorno del mese (es. gen-mag 2026 → vers.2025-06,
    # giu-dic 2026 → vers.2026-06).
    # Struttura: { mese_num: QuerySet filtrato e ordinato }
    _parametri_ccnl_per_mese: dict[int, object] = {}
    for _m in range(1, 13):
        _data_mese = date(anno, _m, 1)
        _parametri_ccnl_per_mese[_m] = (
            ParametroCCNLTurismo.objects
            .filter(attivo=True, decorrenza_validita_da__lte=_data_mese)
            .order_by('livello', '-decorrenza_validita_da')
        )

    # Manteniamo anche il queryset "generico" per il form config e per il template
    parametri_ccnl = ParametroCCNLTurismo.objects.filter(attivo=True).order_by('livello', 'qualifica')

    # ------------------------------------------------------------------
    # Parametri DB anno di simulazione (con fallback a valori di default)
    # ------------------------------------------------------------------
    # giorni_lavorativi_mese viene ora calcolato mese per mese dal calendario
    # aziendale (_gg_lav_cache), con fallback a 26 convenzionali CCNL.

    try:
        _ccnl = CCNL.objects.get(sigla='FIPE')
    except Exception:
        _ccnl = None

    # Maggiorazioni (percentuali già /100)
    maggiorazioni_db = {
        'straordinario_feriale':  Decimal('0.15'),
        'straordinario_notturno': Decimal('0.30'),
        'straordinario_festivo':  Decimal('0.30'),
        'lavoro_domenicale':      Decimal('0.15'),
        'lavoro_festivo':         Decimal('0.20'),
    }
    if _ccnl:
        for _m in ParametroMaggiorazione.objects.filter(ccnl=_ccnl, anno=anno, attivo=True):
            maggiorazioni_db[_m.tipo_maggiorazione] = _m.percentuale / Decimal('100')

    # Scatti anzianità: {livello_key: [(soglia_anni, importo), ...]} — ordinato per soglia
    scatti_db: dict = {}
    if _ccnl:
        for _s in ParametroScattiAnnuali.objects.filter(ccnl=_ccnl, anno=anno, attivo=True):
            scatti_db.setdefault(_s.livello, []).append((_s.anni_anzianita, _s.importo_scatto))
        for _k in scatti_db:
            scatti_db[_k].sort(key=lambda x: x[0])

    # Aliquote contributi
    # Aliquote validate sul prospetto di liquidazione reale Feb 2026 (PUNTOZERO SRLS)
    # INPS az 29.31%: IVS 23.81% + DS 1.61% + CUAF 0.68% + Mat. 0.24% + CIG/FSBT ~2.97%
    # INAIL 0.74%: aliquota effettiva su lordo (codice rischio 0210 ristorazione;
    #              la tariffa INAIL ~0.95% è applicata su retrib. convenzionale ~77.7% del lordo)
    inps_az_perc  = Decimal('0.2931')
    inps_dip_perc = Decimal('0.0936')  # CCNL FIPE: 9.19% IVS + ~0.17% EBT/FSBT
    inail_perc    = Decimal('0.0074')
    if _ccnl:
        # Stessa risoluzione categoria/validità del motore busta (evita aliquote «a caso»).
        _pc_inps = risolvi_parametro_contributi_ccnl(
            ccnl_obj=_ccnl, anno=anno, tipo_contributo='inps',
            azienda=azienda_operativa, mese=1,
        )
        if _pc_inps:
            inps_az_perc  = _pc_inps.aliquota_azienda   / Decimal('100')
            inps_dip_perc = _pc_inps.aliquota_dipendente / Decimal('100')
        _pc_inail = risolvi_parametro_contributi_ccnl(
            ccnl_obj=_ccnl, anno=anno, tipo_contributo='inail',
            azienda=azienda_operativa, mese=1,
        )
        if _pc_inail:
            inail_perc = _pc_inail.aliquota_azienda / Decimal('100')

    # Coefficienti ratei
    coeff_tfr_db = Decimal('0.0691')            # TFR = lordo × 6.91% (= 1/13.5 - 0.50% INPS fond.)
    coeff_13_db  = Decimal('1') / Decimal('12') # 13ª = lordo / 12
    coeff_14_db  = Decimal('1') / Decimal('12') # 14ª = lordo / 12
    # Per FIPE piccoli esercizi le mensilità aggiuntive sono incluse nella busta
    # mensile ordinaria (rateo 1/12 pagato ogni mese). Non si genera quindi
    # un pagamento separato in dicembre (13ª) né in luglio (14ª).
    ratei_mensili = coeff_13_db > Decimal('0')
    # Ferie (FIPE: 26 gg/anno su base 26 gg/mese = 1/12 ≈ 8.33%) e ROL/permessi
    # coeff_ferie = giorni_ferie / (26 × 12)
    # coeff_permessi = giorni_rol / (26 × 12)
    _giorni_ferie_ccnl   = Decimal('26')  # FIPE: 26 giorni/anno
    _giorni_rol_ccnl     = Decimal('4')   # FIPE: 4 giorni ex-festività soppresse
    if _ccnl:
        if _ccnl.giorni_ferie_base:
            _giorni_ferie_ccnl = Decimal(str(_ccnl.giorni_ferie_base))
        if _ccnl.giorni_rol_base:
            _giorni_rol_ccnl = Decimal(str(_ccnl.giorni_rol_base))
    coeff_ferie_db    = (_giorni_ferie_ccnl / (Decimal('26') * Decimal('12'))).quantize(Decimal('0.000001'))
    coeff_permessi_db = (_giorni_rol_ccnl   / (Decimal('26') * Decimal('12'))).quantize(Decimal('0.000001'))
    # Override da ParametroRatei se presenti nel DB (una query, stessa semantica .first() per tipo)
    if _ccnl:
        _ratei_row_by_tipo = {}
        for _pr in (
            ParametroRatei.objects.filter(
                ccnl=_ccnl,
                anno=anno,
                attivo=True,
                tipo_rateo__in=('tfr', 'tredicesima', 'quattordicesima', 'ferie', 'permessi'),
            ).order_by('tipo_rateo', 'id')
        ):
            if _pr.tipo_rateo not in _ratei_row_by_tipo:
                _ratei_row_by_tipo[_pr.tipo_rateo] = _pr
        _r_tfr = _ratei_row_by_tipo.get('tfr')
        if _r_tfr:
            coeff_tfr_db = _r_tfr.coefficiente / Decimal('100')
        _r_13 = _ratei_row_by_tipo.get('tredicesima')
        if _r_13:
            coeff_13_db = _r_13.coefficiente / Decimal('12')
        _r_14 = _ratei_row_by_tipo.get('quattordicesima')
        if _r_14:
            coeff_14_db = _r_14.coefficiente / Decimal('12')
        _r_ferie = _ratei_row_by_tipo.get('ferie')
        if _r_ferie:
            coeff_ferie_db = (_r_ferie.coefficiente / Decimal('100'))
        _r_perm = _ratei_row_by_tipo.get('permessi')
        if _r_perm:
            coeff_permessi_db = (_r_perm.coefficiente / Decimal('100'))

    # Cache giorni lavorativi dal calendario aziendale (uno per mese)
    _gg_lav_cache = {}
    for _m in range(1, 13):
        try:
            _info = _get_gg_lav_mese(azienda_operativa, anno, _m)
            _gg_lav_cache[_m] = _info
        except Exception:
            _gg_lav_cache[_m] = {'giorni_conv_26': 26, 'giorni_lavorativi': 26, 'dates_festivita': []}

    risultati_mensili = []
    from anagrafiche.models import Dipendente
    from .risoluzione_contratto_motore import (
        rapporto_sottoscritto_attivo_nel_mese,
        risolvi_parametro_ccnl_per_mese,
        superminimo_da_rapporto_o_ruolo,
    )

    # Dipendenti e contratti sottoscritti per mese: evita N×12 query duplicate nel doppio ciclo.
    dip_by_pk = {}
    rap_cache = {}
    if azienda_operativa:
        dip_ids_sim = set()
        for _ru in ruoli_config:
            _dp = _ru.get('dipendente_id')
            if not _dp:
                continue
            try:
                dip_ids_sim.add(int(_dp))
            except (TypeError, ValueError):
                continue
        if dip_ids_sim:
            for _d in Dipendente.objects.filter(pk__in=dip_ids_sim, azienda=azienda_operativa):
                dip_by_pk[_d.pk] = _d
            for _did, _dip_o in dip_by_pk.items():
                for _mn in range(1, 13):
                    rap_cache[(_did, _mn)] = rapporto_sottoscritto_attivo_nel_mese(
                        dipendente=_dip_o,
                        azienda=azienda_operativa,
                        anno=anno,
                        mese=_mn,
                    )

    for mese_num in range(1, 13):
        # Giorni nel mese: usa monthrange per gestire correttamente anni bisestili
        giorni_nel_mese = _monthrange(anno, mese_num)[1]

        # Giorni lavorativi effettivi dal calendario aziendale (default 26)
        _cal_info = _gg_lav_cache.get(mese_num, {})
        giorni_lavorativi_mese = _cal_info.get('giorni_conv_26', 26) or 26
        
        righe_ruoli = []
        totali = {
            'netto_mensile': Decimal('0'),
            'netto_con_ratei': Decimal('0'),
            'lordo_mensile': Decimal('0'),
            'inps_azienda': Decimal('0'),
            'inps_dipendente': Decimal('0'),
            'irpef_dipendenti': Decimal('0'),
            'inail': Decimal('0'),
            'tfr': Decimal('0'),
            'rateo_13': Decimal('0'),
            'rateo_14': Decimal('0'),
            'rateo_ferie': Decimal('0'),
            'rateo_permessi': Decimal('0'),
            'costo_azienda_totale': Decimal('0'),
            'totale_f24_mese': Decimal('0'),
            'trattamento_integrativo': Decimal('0'),
            'bonus_l207': Decimal('0'),
            # Pagamenti effettivi tredicesima (Dicembre) e quattordicesima (Luglio)
            'pagamento_13_lordo': Decimal('0'),
            'pagamento_13_netto': Decimal('0'),
            'pagamento_14_lordo': Decimal('0'),
            'pagamento_14_netto': Decimal('0'),
            # TFR liquidazione per contratti a termine in scadenza questo mese
            'tfr_liquidazione': Decimal('0'),
            'tfr_liquidazione_netto': Decimal('0'),
            # Stima addizionali (informativa F24 — versate anno successivo)
            'addiz_regionale': Decimal('0'),
            'addiz_comunale': Decimal('0'),
            # F24 breakdown mensile
            'f24_inps': Decimal('0'),
            'f24_inail': Decimal('0'),
            'f24_irpef': Decimal('0'),
            'f24_addiz_reg': Decimal('0'),
            'f24_addiz_com': Decimal('0'),
            'f24_bonus_comp': Decimal('0'),
            'f24_erario_netto': Decimal('0'),
            'f24_tfr_tax': Decimal('0'),
            'f24_netto': Decimal('0'),
        }
        
        for ruolo in ruoli_config:
            ruolo = _normalizza_ruolo_per_motore(ruolo)
            if ruolo['quantita'] <= 0:
                continue
            
            # Contratto reale (se ruolo ancorato a dipendente): livello, date e tabella CCNL come scostamento fiscale.
            rap_sim = None
            _dip_pk = ruolo.get('dipendente_id')
            if _dip_pk and azienda_operativa:
                try:
                    _dip_pk_int = int(_dip_pk)
                except (TypeError, ValueError):
                    _dip_pk_int = None
                if _dip_pk_int:
                    dip_o = dip_by_pk.get(_dip_pk_int)
                    if dip_o:
                        rap_sim = rap_cache.get((_dip_pk_int, mese_num))

            livello_da_ruolo = (ruolo.get('livello') or '').strip() or _risolvi_livello_da_mansione(
                ruolo.get('nome'), rule_engine
            )
            _lc_contr = (rap_sim.livello_ccnl or '').strip() if rap_sim else ''
            livello_effettivo = _lc_contr or livello_da_ruolo

            # Contratto sottoscritto: data_fine_rapporto può essere NULL (indeterminato) → non passare None a max/min sulle date.
            if rap_sim:
                data_inizio_motore = rap_sim.data_inizio_rapporto or ruolo.get('data_inizio') or date(anno, 1, 1)
                data_fine_motore = (
                    rap_sim.data_fine_rapporto
                    or ruolo.get('data_fine')
                    or date(anno, 12, 31)
                )
            else:
                data_inizio_motore = ruolo.get('data_inizio') or date(anno, 1, 1)
                data_fine_motore = ruolo.get('data_fine') or date(anno, 12, 31)

            parametro_ccnl_res, _fonte_pc_sim = risolvi_parametro_ccnl_per_mese(
                rapporto=rap_sim,
                data_primo_giorno_mese=date(anno, mese_num, 1),
                livello_fallback=str(livello_da_ruolo or ''),
            )
            parametro = parametro_ccnl_res or _parametri_ccnl_per_mese[mese_num].filter(
                livello=livello_effettivo
            ).first()
            if not parametro:
                righe_ruoli.append({
                    'ruolo_id': ruolo['id'],
                    'nome': ruolo['nome'],
                    'quantita': ruolo['quantita'],
                    'livello': livello_effettivo,
                    'missing': True,
                })
                continue

            # Tipo contratto e coefficiente ore (contratto ha priorità sul ruolo organico)
            tipo_contratto = None
            coeff_ore = Decimal('1.00')
            if rap_sim and rap_sim.tipo_contratto_id:
                tipo_contratto = rap_sim.tipo_contratto
            elif ruolo['tipo_contratto_id']:
                try:
                    tipo_contratto = tipi_contratto_by_id.get(int(ruolo['tipo_contratto_id']))
                except (TypeError, ValueError):
                    tipo_contratto = None
            if tipo_contratto is not None:
                coeff_ore = Decimal(str(tipo_contratto.coefficiente_ore or Decimal('1.00')))

            # Giorni attivi su calendario (date effettive contratto se presenti)
            giorni_attivi_calendario = _calcola_giorni_attivi_nel_mese(
                anno=anno,
                mese_num=mese_num,
                data_inizio=data_inizio_motore,
                data_fine=data_fine_motore,
            )

            # Se non è attivo in questo mese, salta
            if giorni_attivi_calendario <= 0:
                continue

            # Base CCNL fissa: 26 giorni/mese (indipendente dal calendario)
            giorni_attivi = _giorni_convenzionali_su_base_26(
                giorni_attivi_calendario=giorni_attivi_calendario,
                giorni_nel_mese=giorni_nel_mese,
            )

            if giorni_attivi <= 0:
                continue

            # coefficiente periodo su base contrattuale 26
            coeff_periodo = (Decimal(giorni_attivi) / Decimal(giorni_lavorativi_mese)).quantize(Decimal('0.0001'))
            # coefficiente ratei su giorni calendario (partenza da data inizio/fine rapporto)
            coeff_ratei_calendario = (Decimal(giorni_attivi_calendario) / Decimal(giorni_nel_mese)).quantize(Decimal('0.0001'))
            
            # Tentativo calcolo con costo_lavoro (solo per info decontribuzione)
            costo_lavoro_ruolo = _calcola_costo_azienda_ruolo_costo_lavoro(
                parametro=parametro,
                coeff_ore=coeff_ore,
                giorni_lavorativi_mese=giorni_lavorativi_mese,
                giorni_attivi=giorni_attivi,
                anno=anno,
                mese_num=mese_num,
                azienda_operativa=azienda_operativa,
                rule_engine=rule_engine,
                regione=ruolo.get('regione', 'sicilia'),
                parametro_tipo_contratto=ruolo.get('tipo_rapporto', 'indeterminato'),
                categoria=ruolo.get('categoria'),
                percettore_naspi=ruolo.get('percettore_naspi'),
                tipo_incentivo=ruolo.get('tipo_incentivo'),
                eta=ruolo.get('eta'),
            )

            # ── Motore paga condiviso (engine canonico) ───────────────────────
            fonte = 'motore_paga'

            _cal_mensile = ruolo.get('calendario_mensile', {}).get(mese_num, {})
            _scatto = _calcola_scatto_totale(livello_effettivo, ruolo.get('anni_anzianita', 0), scatti_db)

            _ore_ord_cal = _cal_mensile.get('ore_ordinarie_retribuite', Decimal('0'))
            try:
                _ore_ord_cal = Decimal(str(_ore_ord_cal or 0))
            except Exception:
                _ore_ord_cal = Decimal('0')
            use_modalita_ore = _ore_ord_cal > 0

            _ore_ord_mese = _ore_ord_cal
            if _ore_ord_mese <= 0:
                _ore_mens_param = Decimal(str(parametro.ore_mensili or 0))
                if _ore_mens_param <= 0:
                    _ore_mens_param = (Decimal(str(parametro.ore_settimanali or 40)) * Decimal('4.333333')).quantize(Decimal('0.01'))
                _ore_ord_mese = (_ore_mens_param * coeff_ore * coeff_ratei_calendario).quantize(Decimal('0.01'))

            # Fallback ore festive: solo se non si usano ore effettive da calendario dipendente
            # (se non valorizzate dal calendario mensile, stima da festività del mese — come simulatore_paga)
            _ore_fest_mese = _cal_mensile.get('giorni_festivi', Decimal('0'))
            if _ore_fest_mese <= 0 and not use_modalita_ore:
                from .utils_calendario import get_festivita_mese as _get_fest
                _ore_gg = (Decimal(str(parametro.ore_giornaliere or 0)) * coeff_ore).quantize(Decimal('0.01'))
                if _ore_gg > 0:
                    _n_fest = sum(
                        1 for f in _get_fest(anno, mese_num, azienda_operativa)
                        if f['data'].weekday() != 6
                    )
                    _ore_fest_mese = (Decimal(str(_n_fest)) * _ore_gg).quantize(Decimal('0.01'))

            premio_contr = Decimal('0')
            if rap_sim is not None:
                try:
                    premio_contr = Decimal(str(rap_sim.premio_obiettivi or 0))
                except Exception:
                    premio_contr = Decimal('0')
            ind_extra_tot = (ruolo.get('indennita_extra', Decimal('0')) + premio_contr).quantize(Decimal('0.01'))
            superminimo_eff = superminimo_da_rapporto_o_ruolo(
                rapporto=rap_sim, ruolo_superminimo=ruolo.get('superminimo')
            )

            _reg_res = (ruolo.get('regione') or 'sicilia').strip() or 'sicilia'

            _r = invoca_calcola_busta_paga_mese(
                log_prefix='SIMULAZIONE_2026',
                parametro_ccnl=parametro,
                tipo_contratto=tipo_contratto,
                anno=anno,
                mese=mese_num,
                azienda=azienda_operativa,
                data_inizio_rapporto=data_inizio_motore,
                data_fine_rapporto=data_fine_motore,
                divisore_str=str(divisore_ore),
                superminimo=superminimo_eff,
                indennita_turno=ruolo.get('indennita_turno', Decimal('0')),
                scatto_anzianita=_scatto,
                indennita_extra=ind_extra_tot,
                ore_straord_diurno=_cal_mensile.get('ore_straord_diurno', Decimal('0')),
                ore_straord_notturno=_cal_mensile.get('ore_straord_notturno', Decimal('0')),
                ore_straord_festivo=_cal_mensile.get('ore_straord_festivo', Decimal('0')),
                ore_straord_domenica=_cal_mensile.get('ore_straord_domenica', Decimal('0')),
                ore_straord_nott_fest=_cal_mensile.get('ore_straord_nott_fest', Decimal('0')),
                ore_ordinarie_retribuite=_ore_ord_mese,
                ore_domenicali=_cal_mensile.get('ore_domenicali', Decimal('0')),
                ore_festivi=_ore_fest_mese,
                giorni_assenza_ingiust=_cal_mensile.get('giorni_assenza', Decimal('0')),
                trattenute_extra_mese=_cal_mensile.get('trattenute_extra_mese', Decimal('0')),
                competenze_extra_non_imponibili=_cal_mensile.get('competenze_extra_non_imponibili', Decimal('0')),
                modalita_ore_effettive=use_modalita_ore,
                auto_ore_domenicali_da_calendario=not use_modalita_ore,
                ccnl_obj=_ccnl,
                regione_residenza=_reg_res,
                contratto_esclude_tredicesima=bool(rap_sim is not None and rap_sim.tredicesima is False),
                contratto_esclude_quattordicesima=bool(
                    rap_sim is not None and rap_sim.quattordicesima is False
                ),
                rateo_13_mensile_in_imponibile=bool(
                    rap_sim is not None and getattr(rap_sim, 'tredicesima_rateo_mensile_in_imponibile', False)
                ),
                rateo_14_mensile_in_imponibile=bool(
                    rap_sim is not None and getattr(rap_sim, 'quattordicesima_rateo_mensile_in_imponibile', False)
                ),
            )

            lordo_unit    = _r['lordo_mensile']
            netto_base_unit = _r['netto_base']
            netto_unit    = _r['netto_totale']
            inps_dip_unit = _r['inps_dip']
            irpef_unit    = _r['irpef_netta']
            inps_az_unit  = _r['inps_az']
            inail_unit    = _r['inail_az']
            lordo_base_per_ratei = _r['lordo_base']
            trattamento_integrativo_unit = _r['ti']
            bonus_l207_unit = _r['l207']
            imponibile_annuo_ti = Decimal(str(_r.get('imponibile_ann', Decimal('0'))))

            # Ratei dal motore paga (già pro-ratati internamente)
            tfr_unit         = _r['tfr_m']
            rateo_13_unit    = _r['rat13_m']
            rateo_14_unit    = _r['rat14_m']
            rateo_ferie_unit = _r['rat_fer_m']
            rateo_perm_unit  = (lordo_base_per_ratei * coeff_permessi_db * coeff_ratei_calendario).quantize(Decimal('0.01'))

            aliquota_tfr_coeff = _r['c_tfr']
            rateo_13_coeff     = _r['c_13']
            rateo_14_coeff     = _r['c_14']
            inps_coeff_base    = _r['inps_az_perc'] / 100
            inps_coeff_eff     = _r['inps_az_perc'] / 100

            # Decontribuzione: da costo_lavoro se disponibile, altrimenti vuoto
            if costo_lavoro_ruolo:
                decontrib_rule_name = str(costo_lavoro_ruolo.get('decontrib_rule_name', ''))
                decontrib_tipo      = str(costo_lavoro_ruolo.get('decontrib_tipo', ''))
                decontrib_valore    = Decimal(str(costo_lavoro_ruolo.get('decontrib_valore', 0)))
            else:
                decontrib_rule_name = ''
                decontrib_tipo      = ''
                decontrib_valore    = Decimal('0')

            costo_azienda_unit = (
                lordo_unit + inps_az_unit + inail_unit
                + tfr_unit + rateo_13_unit + rateo_14_unit
                + rateo_ferie_unit + rateo_perm_unit
            ).quantize(Decimal('0.01'))

            # Voci variabili (per diagnostici / output)
            paga_oraria_lorda = _r['paga_oraria']
            scatto_unit          = _r['scatto']
            superminimo_unit     = _r['superminimo']
            indennita_turno_unit = _r['indennita_turno']
            indennita_extra_unit = _r['indennita_extra']
            straord_unit         = _r['tot_straord']
            imp_dom_magg_unit    = _r['imp_dom_magg']
            imp_fest_magg_unit   = _r['imp_fest_magg']
            premio_unit = (ruolo.get('premio_risultato_annuo', Decimal('0')) / Decimal('12') * coeff_ratei_calendario).quantize(Decimal('0.01'))
            lordo_variabile_unit = (scatto_unit + superminimo_unit + indennita_turno_unit + indennita_extra_unit + straord_unit + premio_unit).quantize(Decimal('0.01'))

            # ----------------------------------------------------------------
            # Pagamento tredicesima (Dicembre) e quattordicesima (Luglio)
            # Pro-quota in base ai mesi di servizio nell'anno.
            # IMPORTANTE: il rateo mensile rimane in costo_azienda (logica di
            # competenza). Il pagamento effettivo qui è puramente informativo
            # per la pianificazione del cash-flow (non duplica il costo).
            # ----------------------------------------------------------------
            _data_inizio_ruolo = ruolo.get('data_inizio') or date(anno, 1, 1)
            _data_fine_ruolo   = ruolo.get('data_fine')   or date(anno, 12, 31)

            # Mesi attivi 13ª — riferimento CCNL FIPE: 1 gen – 31 dic anno corrente
            # Frazioni >= 15 giorni = 1 mese intero (art. 214 CCNL FIPE 2023)
            _mesi_13_attivi = _conta_mesi_ccnl(
                _data_inizio_ruolo, _data_fine_ruolo,
                date(anno, 1, 1), date(anno, 12, 31),
            )

            # Mesi attivi 14ª — riferimento CCNL FIPE: 1 lug anno prec – 30 giu anno corrente
            # Frazioni >= 15 giorni = 1 mese intero (art. 219 CCNL FIPE 2023)
            _mesi_14_attivi = _conta_mesi_ccnl(
                _data_inizio_ruolo, _data_fine_ruolo,
                date(anno - 1, 7, 1), date(anno, 6, 30),
            )

            # Mesi attivi TFR — per la liquidazione: tutti i mesi di servizio nell'anno
            # TFR non applica la regola dei 15 giorni: accrual giornaliero (art. 2120 c.c.)
            # Per la liquidazione usiamo lo stesso contatore del 13ª come mesi interi effettivi
            _mesi_attivi = _mesi_13_attivi   # usato solo per display / compatibilità

            pagamento_13_unit     = Decimal('0')
            pagamento_13_inps_dip = Decimal('0')
            pagamento_13_inps_az  = Decimal('0')
            pagamento_13_irpef    = Decimal('0')
            pagamento_13_netto    = Decimal('0')
            pagamento_14_unit     = Decimal('0')
            pagamento_14_inps_dip = Decimal('0')
            pagamento_14_inps_az  = Decimal('0')
            pagamento_14_irpef    = Decimal('0')
            pagamento_14_netto    = Decimal('0')
            tfr_liquidazione_unit = Decimal('0')

            if mese_num == 12 and not ratei_mensili:  # Tredicesima pagata in Dicembre (solo se non mensile)
                # Riferimento: 1 gen – 31 dic; frazioni >= 15 gg = 1 mese intero
                pagamento_13_unit = (
                    lordo_base_per_ratei * coeff_13_db * Decimal(str(_mesi_13_attivi))
                ).quantize(Decimal('0.01'))
                if pagamento_13_unit > 0:
                    pagamento_13_inps_dip = (pagamento_13_unit * inps_dip_perc).quantize(Decimal('0.01'))
                    pagamento_13_inps_az  = (pagamento_13_unit * inps_az_perc).quantize(Decimal('0.01'))
                    # IRPEF incrementale 13ª: stesso anno fiscale e stessa aliquota INPS dip. del motore
                    # (prima: anno di sistema + 9,36% fisso → scostamenti vs cedolino / motore).
                    _irpef_con_13 = Decimal(str(
                        calcola_netto_dipendente(
                            lordo_unit + pagamento_13_unit,
                            anno=anno,
                            aliquota_inps_dip=float(inps_dip_perc),
                        )['irpef_netta']
                    ))
                    pagamento_13_irpef = max(Decimal('0'), _irpef_con_13 - irpef_unit).quantize(Decimal('0.01'))
                    pagamento_13_netto = (
                        pagamento_13_unit - pagamento_13_inps_dip - pagamento_13_irpef
                    ).quantize(Decimal('0.01'))

            if mese_num == 7 and not ratei_mensili:  # Quattordicesima pagata in Luglio (solo se non mensile)
                # Riferimento: 1 lug anno prec – 30 giu anno corrente; frazioni >= 15 gg = 1 mese
                pagamento_14_unit = (
                    lordo_base_per_ratei * coeff_14_db * Decimal(str(_mesi_14_attivi))
                ).quantize(Decimal('0.01'))
                if pagamento_14_unit > 0:
                    pagamento_14_inps_dip = (pagamento_14_unit * inps_dip_perc).quantize(Decimal('0.01'))
                    pagamento_14_inps_az  = (pagamento_14_unit * inps_az_perc).quantize(Decimal('0.01'))
                    _irpef_con_14 = Decimal(str(
                        calcola_netto_dipendente(
                            lordo_unit + pagamento_14_unit,
                            anno=anno,
                            aliquota_inps_dip=float(inps_dip_perc),
                        )['irpef_netta']
                    ))
                    pagamento_14_irpef = max(Decimal('0'), _irpef_con_14 - irpef_unit).quantize(Decimal('0.01'))
                    pagamento_14_netto = (
                        pagamento_14_unit - pagamento_14_inps_dip - pagamento_14_irpef
                    ).quantize(Decimal('0.01'))

            # TFR liquidazione per contratti a tempo determinato (pagata alla scadenza)
            _tipo_rapporto_str = ruolo.get('tipo_rapporto', 'indeterminato')
            _is_last_month = (
                _data_fine_ruolo.year  == anno and
                _data_fine_ruolo.month == mese_num
            )
            if _tipo_rapporto_str == 'determinato' and _is_last_month:
                # TFR: accrual giornaliero (art. 2120 c.c.) — usiamo i mesi interi attivi nell'anno
                # come approssimazione; il calcolo esatto è la somma dei ratei mensili già accumulati
                tfr_liquidazione_unit = (
                    lordo_base_per_ratei * coeff_tfr_db * Decimal(str(_mesi_13_attivi))
                ).quantize(Decimal('0.01'))

            # ----------------------------------------------------------------
            # Coefficiente netto effettivo del mese (INPS dip. + aliquota IRPEF media)
            # Usato per nettizzare ratei e TFR
            # ----------------------------------------------------------------
            _irpef_eff = (irpef_unit / lordo_unit).quantize(Decimal('0.0001')) if lordo_unit > 0 else Decimal('0')
            _net_coeff_ratei = max(Decimal('0'), Decimal('1') - inps_dip_perc - _irpef_eff)

            # ----------------------------------------------------------------
            # TFR netto (tassazione separata — approssimata con aliquota IRPEF eff.)
            # La tassazione separata del TFR usa l'aliquota media degli ultimi 5 anni.
            # Per la pianificazione, si usa l'aliquota IRPEF effettiva del mese come stima.
            # Il datore trattiene tfr_liq_tax_unit e lo versa in F24 (cod. tributo 1012).
            # ----------------------------------------------------------------
            tfr_liq_tax_unit = (tfr_liquidazione_unit * _irpef_eff).quantize(Decimal('0.01'))
            tfr_liq_netto_unit = (
                tfr_liquidazione_unit - tfr_liq_tax_unit
            ).quantize(Decimal('0.01'))

            # ----------------------------------------------------------------
            # Netto busta paga effettivo del mese (cash flow dipendente):
            #   = netto_base + 13ª netta (Dic) + 14ª netta (Lug) + TFR netto (scadenza)
            # In luglio e dicembre e nel mese di scadenza del contratto,
            # il dipendente riceve importi aggiuntivi NETTI (già tassati nelle righe sopra).
            # ----------------------------------------------------------------
            netto_busta_unit = (
                netto_unit + pagamento_13_netto + pagamento_14_netto + tfr_liq_netto_unit
            ).quantize(Decimal('0.01'))

            # ----------------------------------------------------------------
            # Netto "economico mensile" con ratei (vista distribuzione uniforme):
            #   = netto BASE (non busta, per non duplicare i pagamenti straordinari)
            #   + valore netto mensile degli accantonamenti 13ª e 14ª
            # TFR/Ferie/Permessi esclusi: TFR pagato a fine rapporto; ferie/permessi
            # già inclusi nel lordo mensile corrente.
            # ----------------------------------------------------------------
            netto_ratei_unit = ((rateo_13_unit + rateo_14_unit) * _net_coeff_ratei).quantize(Decimal('0.01'))
            netto_con_ratei_unit = (netto_unit + netto_ratei_unit).quantize(Decimal('0.01'))
            # In modalità ratei mensili la busta di ogni mese include già i ratei 13ª/14ª
            if ratei_mensili:
                netto_busta_unit = (netto_unit + netto_ratei_unit + tfr_liq_netto_unit).quantize(Decimal('0.01'))

            # ----------------------------------------------------------------
            # Stima addizionali IRPEF (regionale + comunale) — informativa F24
            # Base: imponibile annuo dal motore canonico
            # Le addizionali non vengono detratte dal netto mensile corrente
            # perché sono versate nell'anno successivo (saldo). Vengono mostrate
            # nel breakdown F24 come voce separata con nota "stima anno succ.".
            # ----------------------------------------------------------------
            addiz_reg_annuo = calcola_addizionale_regionale_sicilia(imponibile_annuo_ti, anno=anno)
            addiz_com_annuo = calcola_addizionale_comunale_stima(imponibile_annuo_ti, anno=anno)
            addiz_reg_unit = (addiz_reg_annuo / Decimal('12') * coeff_ratei_calendario).quantize(Decimal('0.01'))
            addiz_com_unit = (addiz_com_annuo / Decimal('12') * coeff_ratei_calendario).quantize(Decimal('0.01'))

            # Totali per quantità
            qta = Decimal(str(ruolo['quantita']))
            # netto_tot = busta paga effettiva (include 13ª/14ª/TFR netti nei mesi di pagamento)
            netto_tot = (netto_busta_unit * qta).quantize(Decimal('0.01'))
            netto_con_ratei_tot = (netto_con_ratei_unit * qta).quantize(Decimal('0.01'))
            tfr_liq_netto_tot = (tfr_liq_netto_unit * qta).quantize(Decimal('0.01'))
            lordo_tot = (lordo_unit * qta).quantize(Decimal('0.01'))
            inps_az_tot = (inps_az_unit * qta).quantize(Decimal('0.01'))
            inps_dip_tot = (inps_dip_unit * qta).quantize(Decimal('0.01'))
            irpef_tot = (irpef_unit * qta).quantize(Decimal('0.01'))
            tfr_tot = (tfr_unit * qta).quantize(Decimal('0.01'))
            rateo_13_tot = (rateo_13_unit * qta).quantize(Decimal('0.01'))
            rateo_14_tot = (rateo_14_unit * qta).quantize(Decimal('0.01'))
            lordo_con_1314_unit = (lordo_unit + rateo_13_unit + rateo_14_unit).quantize(Decimal('0.01'))
            lordo_con_1314_tot = (lordo_con_1314_unit * qta).quantize(Decimal('0.01'))
            rateo_ferie_tot = (rateo_ferie_unit * qta).quantize(Decimal('0.01'))
            rateo_perm_tot  = (rateo_perm_unit  * qta).quantize(Decimal('0.01'))
            inail_tot = (inail_unit * qta).quantize(Decimal('0.01'))
            costo_tot = (costo_azienda_unit * qta).quantize(Decimal('0.01'))
            trattamento_integrativo_tot = (trattamento_integrativo_unit * qta).quantize(Decimal('0.01'))
            bonus_l207_tot = (bonus_l207_unit * qta).quantize(Decimal('0.01'))

            # Stima addizionali (per quantità)
            addiz_reg_tot = (addiz_reg_unit * qta).quantize(Decimal('0.01'))
            addiz_com_tot = (addiz_com_unit * qta).quantize(Decimal('0.01'))

            # Totali pagamenti effettivi 13ª/14ª e TFR liquidazione (per quantità)
            pagamento_13_tot = (pagamento_13_unit  * qta).quantize(Decimal('0.01'))
            pagamento_13_net_tot = (pagamento_13_netto * qta).quantize(Decimal('0.01'))
            pagamento_13_inps_dip_tot = (pagamento_13_inps_dip * qta).quantize(Decimal('0.01'))
            pagamento_13_inps_az_tot  = (pagamento_13_inps_az  * qta).quantize(Decimal('0.01'))
            pagamento_13_irpef_tot    = (pagamento_13_irpef    * qta).quantize(Decimal('0.01'))
            pagamento_14_tot = (pagamento_14_unit  * qta).quantize(Decimal('0.01'))
            pagamento_14_net_tot = (pagamento_14_netto * qta).quantize(Decimal('0.01'))
            pagamento_14_inps_dip_tot = (pagamento_14_inps_dip * qta).quantize(Decimal('0.01'))
            pagamento_14_inps_az_tot  = (pagamento_14_inps_az  * qta).quantize(Decimal('0.01'))
            pagamento_14_irpef_tot    = (pagamento_14_irpef    * qta).quantize(Decimal('0.01'))
            tfr_liq_tot = (tfr_liquidazione_unit * qta).quantize(Decimal('0.01'))
            tfr_liq_tax_tot = (tfr_liq_tax_unit * qta).quantize(Decimal('0.01'))
            # tfr_liq_netto_tot è già calcolato sopra

            # Aggiungi a totali mese
            totali['netto_mensile'] += netto_tot
            totali['netto_con_ratei'] += netto_con_ratei_tot
            totali['lordo_mensile'] += lordo_con_1314_tot
            # INPS base (lordo mensile ordinario)
            totali['inps_azienda'] += inps_az_tot
            totali['inps_dipendente'] += inps_dip_tot
            totali['irpef_dipendenti'] += irpef_tot
            # INPS e IRPEF su 13ª (Dicembre) e 14ª (Luglio) — F24 mese effettivo
            totali['inps_azienda']     += pagamento_13_inps_az_tot  + pagamento_14_inps_az_tot
            totali['inps_dipendente']  += pagamento_13_inps_dip_tot + pagamento_14_inps_dip_tot
            totali['irpef_dipendenti'] += pagamento_13_irpef_tot    + pagamento_14_irpef_tot
            totali['inail'] += inail_tot
            totali['tfr'] += tfr_tot
            totali['rateo_13'] += rateo_13_tot
            totali['rateo_14'] += rateo_14_tot
            totali['rateo_ferie'] += rateo_ferie_tot
            totali['rateo_permessi'] += rateo_perm_tot
            totali['costo_azienda_totale'] += costo_tot
            totali['trattamento_integrativo'] += trattamento_integrativo_tot
            totali['bonus_l207'] += bonus_l207_tot
            totali['pagamento_13_lordo'] += pagamento_13_tot
            totali['pagamento_13_netto'] += pagamento_13_net_tot
            totali['pagamento_14_lordo'] += pagamento_14_tot
            totali['pagamento_14_netto'] += pagamento_14_net_tot
            totali['tfr_liquidazione'] += tfr_liq_tot
            totali['tfr_liquidazione_netto'] += tfr_liq_netto_tot
            totali['addiz_regionale'] += addiz_reg_tot
            totali['addiz_comunale'] += addiz_com_tot
            # Tassazione separata TFR (versata in F24 nel mese di liquidazione — cod. 1012)
            totali['f24_tfr_tax'] += tfr_liq_tax_tot
            # F24 breakdown: ricalcolato ad ogni ruolo aggiunto (finale = dopo tutti i ruoli)
            totali['f24_inps']       = (totali['inps_azienda'] + totali['inps_dipendente']).quantize(Decimal('0.01'))
            totali['f24_inail']      = totali['inail']  # accrual informativo — versato con autoliquidazione annuale (febbraio)
            totali['f24_irpef']      = totali['irpef_dipendenti']
            totali['f24_addiz_reg']  = totali['addiz_regionale']
            totali['f24_addiz_com']  = totali['addiz_comunale']
            totali['f24_bonus_comp'] = (totali['trattamento_integrativo'] + totali['bonus_l207']).quantize(Decimal('0.01'))
            totali['f24_erario_netto'] = max(Decimal('0'), totali['f24_irpef'] - totali['f24_bonus_comp']).quantize(Decimal('0.01'))
            # F24 mensile = INPS DM10 + Erario netto + Tassazione TFR (INAIL escluso: autoliquidazione annuale)
            totali['f24_netto']      = (totali['f24_inps'] + totali['f24_erario_netto'] + totali['f24_tfr_tax']).quantize(Decimal('0.01'))
            totali['totale_f24_mese'] = totali['f24_netto']
            
            # Costruisci la riga con SPREAD di tutti i dati del motore paga canonico (**_r)
            # seguito dagli override specifici della simulazione annua
            riga_result = {
                **_r,  # Spread TUTTI i risultati del motore: paga_base, contingenza, edr, indennita, 
                       # ore_domenicali, ore_festivi, imp_dom_magg, imp_fest_magg, etc.
                # Meta ruolo
                'ruolo_id': ruolo['id'],
                'nome': ruolo['nome'],
                'quantita': ruolo['quantita'],
                'livello': parametro.livello,
                'qualifica': parametro.qualifica,
                'tipo_contratto': tipo_contratto,
                'coefficiente_ore': coeff_ore,
                'giorni_attivi_calendario': giorni_attivi_calendario,
                'giorni_attivi': giorni_attivi,
                'coeff_ratei_calendario': coeff_ratei_calendario,
                # Lordo/Netto unità e totale (override da _r se presente)
                'lordo_unit': lordo_unit,
                'lordo_con_1314_unit': lordo_con_1314_unit,
                'lordo_con_1314_tot': lordo_con_1314_tot,
                'netto_base_unit': netto_base_unit,   # netto base mensile (senza 13ª/14ª/TFR)
                'netto_busta_unit': netto_busta_unit, # netto effettivo busta paga del mese
                'netto_unit': netto_busta_unit,       # alias per compatibilità template esistente
                'netto_con_ratei_unit': netto_con_ratei_unit,
                'netto_tot': netto_tot,
                'netto_con_ratei_tot': netto_con_ratei_tot,
                'tfr_liq_netto': tfr_liq_netto_unit,  # TFR netto stima (solo mese scadenza)
                'lordo_tot': lordo_tot,
                'inps_azienda_tot': inps_az_tot,
                'inps_dipendente_tot': inps_dip_tot,
                'irpef_tot': irpef_tot,
                'inail_tot': inail_tot,
                'tfr_tot': tfr_tot,
                'rateo_13_tot': rateo_13_tot,
                'rateo_14_tot': rateo_14_tot,
                'rateo_ferie_tot': rateo_ferie_tot,
                'rateo_perm_tot': rateo_perm_tot,
                'aliquota_tfr_coeff': aliquota_tfr_coeff,
                'rateo_13_coeff': rateo_13_coeff,
                'rateo_14_coeff': rateo_14_coeff,
                'inps_coeff_base': inps_coeff_base,
                'inps_coeff_eff': inps_coeff_eff,
                'decontrib_rule_name': decontrib_rule_name,
                'decontrib_tipo': decontrib_tipo,
                'decontrib_valore': decontrib_valore,
                'costo_azienda_tot': costo_tot,
                'fonte_calcolo': fonte,
                'trattamento_integrativo_unit': trattamento_integrativo_unit,
                'trattamento_integrativo_tot': trattamento_integrativo_tot,
                'bonus_l207_unit': bonus_l207_unit,
                'bonus_l207_tot': bonus_l207_tot,
                # Voci variabili (per trasparenza nel risultato)
                'versione_ccnl': getattr(parametro, 'versione', ''),
                'paga_oraria_lorda': paga_oraria_lorda,
                'scatto_unit': scatto_unit,
                'superminimo_unit': superminimo_unit,
                'indennita_turno_unit': indennita_turno_unit,
                'indennita_extra_unit': indennita_extra_unit,
                'straord_unit': straord_unit,
                'imp_dom_magg_unit': imp_dom_magg_unit,
                'imp_fest_magg_unit': imp_fest_magg_unit,
                'premio_unit': premio_unit,
                'lordo_variabile_unit': lordo_variabile_unit,
                # Pagamenti effettivi 13ª (Dicembre) e 14ª (Luglio)
                'pagamento_13_unit': pagamento_13_unit,
                'pagamento_13_inps_dip': pagamento_13_inps_dip,
                'pagamento_13_inps_az': pagamento_13_inps_az,
                'pagamento_13_irpef': pagamento_13_irpef,
                'pagamento_13_netto': pagamento_13_netto,
                'pagamento_13_tot': pagamento_13_tot,
                'pagamento_13_net_tot': pagamento_13_net_tot,
                'pagamento_14_unit': pagamento_14_unit,
                'pagamento_14_inps_dip': pagamento_14_inps_dip,
                'pagamento_14_inps_az': pagamento_14_inps_az,
                'pagamento_14_irpef': pagamento_14_irpef,
                'pagamento_14_netto': pagamento_14_netto,
                'pagamento_14_tot': pagamento_14_tot,
                'pagamento_14_net_tot': pagamento_14_net_tot,
                # TFR liquidazione (contratti determinato in scadenza)
                'tfr_liquidazione_unit': tfr_liquidazione_unit,
                'tfr_liq_tot': tfr_liq_tot,
                'tfr_liq_tax_unit': tfr_liq_tax_unit,
                'mesi_attivi': _mesi_attivi,         # alias generico (= _mesi_13_attivi)
                'mesi_13_attivi': _mesi_13_attivi,   # mesi per 13ª (1 gen – 31 dic)
                'mesi_14_attivi': _mesi_14_attivi,   # mesi per 14ª (1 lug prec – 30 giu)
                # Stima addizionali (F24 informativa)
                'addiz_reg_unit': addiz_reg_unit,
                'addiz_com_unit': addiz_com_unit,
                'addiz_reg_tot': addiz_reg_tot,
                'addiz_com_tot': addiz_com_tot,
            }
            righe_ruoli.append(riga_result)
        
        risultati_mensili.append({
            'anno': anno,
            'mese': mese_num,
            'giorni_nel_mese': giorni_nel_mese,
            'righe': righe_ruoli,
            'totali': totali,
        })

    # Riepilogo annuo: somma per colonna dei subtotali mensili
    totali_annui = {
        'lordo_mensile': Decimal('0.00'),
        'netto_mensile': Decimal('0.00'),
        'netto_con_ratei': Decimal('0.00'),
        'inps_azienda': Decimal('0.00'),
        'inps_dipendente': Decimal('0.00'),
        'irpef_dipendenti': Decimal('0.00'),
        'inail': Decimal('0.00'),
        'tfr': Decimal('0.00'),
        'rateo_13': Decimal('0.00'),
        'rateo_14': Decimal('0.00'),
        'rateo_ferie': Decimal('0.00'),
        'rateo_permessi': Decimal('0.00'),
        'costo_azienda_totale': Decimal('0.00'),
        'totale_f24_mese': Decimal('0.00'),
        'trattamento_integrativo': Decimal('0.00'),
        'bonus_l207': Decimal('0.00'),
        'pagamento_13_lordo': Decimal('0.00'),
        'pagamento_13_netto': Decimal('0.00'),
        'pagamento_14_lordo': Decimal('0.00'),
        'pagamento_14_netto': Decimal('0.00'),
        'tfr_liquidazione': Decimal('0.00'),
        'tfr_liquidazione_netto': Decimal('0.00'),
        'addiz_regionale': Decimal('0.00'),
        'addiz_comunale': Decimal('0.00'),
        'f24_inps': Decimal('0.00'),
        'f24_inail': Decimal('0.00'),
        'f24_irpef': Decimal('0.00'),
        'f24_addiz_reg': Decimal('0.00'),
        'f24_addiz_com': Decimal('0.00'),
        'f24_bonus_comp': Decimal('0.00'),
        'f24_erario_netto': Decimal('0.00'),
        'f24_tfr_tax': Decimal('0.00'),
        'f24_netto': Decimal('0.00'),
    }

    for mese_data in risultati_mensili:
        subtot = mese_data.get('totali', {})
        for k in totali_annui.keys():
            totali_annui[k] += Decimal(str(subtot.get(k, 0)))

    costo_medio_mensile = Decimal('0.00')
    if risultati_mensili:
        costo_medio_mensile = (
            totali_annui['costo_azienda_totale'] / Decimal(len(risultati_mensili))
        ).quantize(Decimal('0.01'))

    # F24 annuo = somma mensili (INPS DM10 + Erario netto + TFR tax; INAIL escluso = autoliquidazione annuale)
    totale_f24_annuo = (
        totali_annui['f24_inps']
        + totali_annui['f24_erario_netto']
        + totali_annui['f24_tfr_tax']
    ).quantize(Decimal('0.01'))
    # Totale INAIL annuo per autoliquidazione (mostrato separatamente)
    totale_inail_annuo = totali_annui['f24_inail']
    
    # Data effettiva: min inizio e max fine tra i ruoli configurati
    date_inizio_ruoli = [r['data_inizio'] for r in ruoli_config if r.get('data_inizio')]
    date_fine_ruoli   = [r['data_fine']   for r in ruoli_config if r.get('data_fine')]
    data_inizio_effettiva = min(date_inizio_ruoli) if date_inizio_ruoli else date(anno, 1, 1)
    data_fine_effettiva   = max(date_fine_ruoli)   if date_fine_ruoli   else date(anno, 12, 31)

    # Segnala se la Decontribuzione Sud è attiva su almeno un ruolo
    # (richiede DURC in regola — L. 296/2006 art. 1 c. 1175 + art. 4 DL 5/2021)
    _has_decontrib_sud = any(
        'sud' in str(r.get('decontrib_rule_name', '')).lower() or
        'sud' in str(r.get('decontrib_tipo', '')).lower()
        for mese_data in risultati_mensili
        for r in mese_data['righe']
        if not r.get('missing')
    )

    return {
        'anno': anno,
        'data_inizio': data_inizio_effettiva,
        'data_fine': data_fine_effettiva,
        'azienda_operativa': azienda_operativa,
        'usa_costo_lavoro': COSTO_LAVORO_ENABLED,
        'divisore_ore': divisore_ore,
        'ruoli_config': ruoli_config,
        'risultati_mensili': risultati_mensili,
        'totali_annui': totali_annui,
        'costo_medio_mensile': costo_medio_mensile,
        'totale_f24_annuo': totale_f24_annuo,
        'totale_inail_annuo': totale_inail_annuo,
        'has_decontrib_sud': _has_decontrib_sud,
    }


def _load_ruoli_da_db(azienda):
    """Carica i ruoli da RuoloOrganico2026 e li restituisce nel formato ruoli_config
    (stesso output di _build_ruoli_config), pronti per il template."""
    from datetime import date as date_type
    role_field_names = {f.name for f in RuoloOrganico2026._meta.fields}
    has_indennita_extra = 'indennita_extra' in role_field_names
    rows = RuoloOrganico2026.objects.filter(azienda=azienda).order_by('ordinamento', 'id')
    ruoli = []
    for idx, row in enumerate(rows, start=1):
        # Il calendario_mensile è salvato in JSON con chiavi stringa; lo convertiamo
        # con chiavi int e valori Decimal per coerenza con _build_ruoli_config.
        cal_raw = row.calendario_mensile or {}
        calendario = {}
        for m in range(1, 13):
            mese_data = cal_raw.get(str(m)) or cal_raw.get(m) or {}
            calendario[m] = {
                'ore_straord_diurno':   Decimal(str(mese_data.get('ore_straord_diurno', 0))),
                'ore_straord_notturno': Decimal(str(mese_data.get('ore_straord_notturno', 0))),
                'ore_straord_festivo':  Decimal(str(mese_data.get('ore_straord_festivo', 0))),
                'ore_straord_domenica': Decimal(str(mese_data.get('ore_straord_domenica', 0))),
                'ore_straord_nott_fest': Decimal(str(mese_data.get('ore_straord_nott_fest', 0))),
                'ore_ordinarie_retribuite': Decimal(str(mese_data.get('ore_ordinarie_retribuite', 0))),
                'ore_domenicali':       Decimal(str(mese_data.get('ore_domenicali', 0))),
                'giorni_festivi':       Decimal(str(mese_data.get('giorni_festivi', 0))),
                'giorni_assenza':       Decimal(str(mese_data.get('giorni_assenza', 0))),
                'trattenute_extra_mese': Decimal(str(mese_data.get('trattenute_extra_mese', 0))),
                'competenze_extra_non_imponibili': Decimal(str(mese_data.get('competenze_extra_non_imponibili', 0))),
            }
        ruoli.append({
            'id': str(idx),
            'nome': row.nome,
            'quantita': row.quantita,
            'livello': row.livello,
            'tipo_contratto_id': row.tipo_contratto_id,
            'tipo_rapporto': row.tipo_rapporto,
            'data_inizio': row.data_inizio,
            'data_fine': row.data_fine,
            'regione': row.regione,
            'eta': row.eta,
            'categoria': row.categoria,
            'percettore_naspi': row.percettore_naspi,
            'tipo_incentivo': row.tipo_incentivo,
            'anni_anzianita': row.anni_anzianita,
            'superminimo': row.superminimo,
            'indennita_turno': row.indennita_turno,
            'indennita_extra': getattr(row, 'indennita_extra', Decimal('0')) if has_indennita_extra else Decimal('0'),
            'premio_risultato_annuo': row.premio_risultato_annuo,
            'calendario_mensile': calendario,
            'origine_dati': getattr(row, 'origine_dati', 'manuale'),
            'nominativi_riferimento': getattr(row, 'nominativi_riferimento', '') or '',
            'soggetti_riferimento': getattr(row, 'soggetti_riferimento', []) or [],
            'dipendente_id': getattr(row, 'dipendente_id', None),
            'stato_soggetto': getattr(row, 'stato_soggetto', '') or '',
            'mansione_label': (
                getattr(row, 'mansione_label', '') or
                (((getattr(row, 'soggetti_riferimento', []) or [{}])[0].get('mansione'))
                 if (getattr(row, 'soggetti_riferimento', []) or []) else '')
            ),
        })
    return ruoli


def _salva_ruoli_nel_db(azienda, user, ruoli_config):
    """Sostituisce i ruoli in RuoloOrganico2026 per l'azienda con quelli calcolati."""
    role_field_names = {f.name for f in RuoloOrganico2026._meta.fields}
    has_indennita_extra = 'indennita_extra' in role_field_names
    RuoloOrganico2026.objects.filter(azienda=azienda).delete()
    for idx, r in enumerate(ruoli_config):
        cal_raw = r.get('calendario_mensile') or {}
        cal_json = {
            str(m): {k: float(v) for k, v in mese_data.items()}
            for m, mese_data in cal_raw.items()
        }
        _q_save = _parse_quantita_ruolo_sim(r.get('quantita'))
        if _q_save < 1:
            _q_save = 1
        create_kwargs = dict(
            azienda=azienda,
            ordinamento=idx,
            dipendente_id=r.get('dipendente_id'),
            stato_soggetto=(r.get('stato_soggetto') or ''),
            mansione_label=(r.get('mansione_label') or ''),
            nome=r.get('nome') or '',
            quantita=_q_save,
            livello=str(r.get('livello') or ''),
            tipo_contratto_id=str(r.get('tipo_contratto_id') or ''),
            tipo_rapporto=r.get('tipo_rapporto') or 'indeterminato',
            data_inizio=r.get('data_inizio') or date(2026, 1, 1),
            data_fine=r.get('data_fine') or date(2026, 12, 31),
            regione=r.get('regione') or 'sicilia',
            eta=r.get('eta'),
            categoria=r.get('categoria') or None,
            percettore_naspi=r.get('percettore_naspi'),
            tipo_incentivo=r.get('tipo_incentivo') or None,
            anni_anzianita=_safe_int_nonnegative(r.get('anni_anzianita'), 0),
            superminimo=r.get('superminimo') or Decimal('0'),
            indennita_turno=r.get('indennita_turno') or Decimal('0'),
            premio_risultato_annuo=r.get('premio_risultato_annuo') or Decimal('0'),
            calendario_mensile=cal_json,
            origine_dati=r.get('origine_dati') or 'manuale',
            nominativi_riferimento=r.get('nominativi_riferimento') or '',
            soggetti_riferimento=r.get('soggetti_riferimento') or [],
            modificato_da=user,
        )
        if has_indennita_extra:
            create_kwargs['indennita_extra'] = r.get('indennita_extra') or Decimal('0')
        RuoloOrganico2026.objects.create(**create_kwargs)


@login_required
@user_passes_test(_is_admin_only)
def simulazione_2026_config(request):
    """Pagina configurazione ruoli per la simulazione annua.

    Flusso:
    - Primo accesso o ?reset=1 → form vuoto
    - Accesso normale senza parametri ruolo_X → carica ruoli da RuoloOrganico2026
    - Accesso con parametri ruolo_X nel GET → usa quei parametri (modifica in corso)
    """
    # Usa fallback robusto: session → user.azienda → prima azienda disponibile
    azienda = _get_azienda_con_fallback(request.user, request.session)

    ha_parametri_ruoli = any(k.startswith('ruolo_') for k in request.GET)

    if not ha_parametri_ruoli:
        # Nuova regola: sorgente canonica = anagrafica/contratti/proposte correnti.
        # Rigenera sempre i box per singolo soggetto e riallinea la tabella RuoloOrganico2026.
        ruoli_config = _ruoli_precaricati_da_profili(azienda, anno=2026) if azienda else []
        if azienda:
            try:
                _salva_ruoli_nel_db(azienda, request.user, ruoli_config)
            except Exception:
                logger.exception('Errore riallineamento automatico RuoloOrganico2026 da anagrafica')
    else:
        ruoli_config = _build_ruoli_config(request)

    parametri_ccnl = ParametroCCNLTurismo.objects.filter(attivo=True).order_by('livello', 'qualifica')
    tipi_contratto = TipoContratto.objects.filter(attivo=True).order_by('nome')

    # Data ultimo salvataggio per il badge nel template
    ultima_modifica = None
    if azienda:
        ultimo_ruolo = RuoloOrganico2026.objects.filter(azienda=azienda).order_by('-data_modifica').first()
        if ultimo_ruolo:
            ultima_modifica = ultimo_ruolo.data_modifica

    mesi_abbr = [
        (1, 'Gen'), (2, 'Feb'), (3, 'Mar'), (4, 'Apr'),
        (5, 'Mag'), (6, 'Giu'), (7, 'Lug'), (8, 'Ago'),
        (9, 'Set'), (10, 'Ott'), (11, 'Nov'), (12, 'Dic'),
    ]
    from anagrafiche.models import Dipendente

    dipendenti_attivi_count = None
    dipendenti_qs = Dipendente.objects.none()
    if azienda:
        dipendenti_attivi_count = Dipendente.objects.filter(
            azienda=azienda, stato='attivo'
        ).count()
        dipendenti_qs = (
            Dipendente.objects.filter(azienda=azienda)
            .exclude(stato='cessato')
            .order_by('cognome', 'nome')
        )
    testate_scenario = _somma_testate_ruoli(ruoli_config)
    delta_organico = (
        (testate_scenario - dipendenti_attivi_count)
        if dipendenti_attivi_count is not None
        else None
    )

    _prepara_ruoli_config_per_template(ruoli_config)

    return render(
        request,
        'rapporto_di_lavoro/simulazione_2026_config.html',
        {
            'ruoli_config': ruoli_config,
            'parametri_ccnl': parametri_ccnl,
            'tipi_contratto': tipi_contratto,
            'azienda_operativa': azienda,
            'mesi_abbr': mesi_abbr,
            'ultima_modifica': ultima_modifica,
            'anno': 2026,
            'dipendenti': dipendenti_qs,
            'dipendenti_attivi_count': dipendenti_attivi_count,
            'testate_scenario': testate_scenario,
            'delta_organico': delta_organico,
        }
    )


def _simulazione_2026_risultato_response(request, _sim_params):
    """Corpo vista risultato (calcolo + render). Usata da ``simulazione_2026_risultato`` con try/except."""
    # Azienda con fallback robusto — non blocca il salvataggio se la sessione è invalida
    azienda = _get_azienda_con_fallback(request.user, request.session)

    # ── 1. Salva subito i ruoli in tabella ──────────────────────────────────
    # Operazione indipendente dal calcolo: avviene sempre quando ci sono ruoli.
    ruoli_dal_form = _build_ruoli_config(request)
    if ruoli_dal_form and azienda:
        try:
            _salva_ruoli_nel_db(azienda, request.user, ruoli_dal_form)
        except Exception:
            logger.exception('Errore salvataggio ruoli in RuoloOrganico2026')

    # ── 2. Calcola la simulazione ────────────────────────────────────────────
    context = _calcola_simulazione_2026(request)
    context['querystring'] = _sim_params.urlencode()
    context['show_debug_aliquote'] = bool(
        getattr(request.user, 'is_superuser', False) or getattr(request.user, 'ruolo', '') == 'admin'
    )
    context['mesi_nomi'] = [
        '', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
        'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre',
    ]

    az_ctx = context.get('azienda_operativa') or azienda
    if az_ctx:
        from anagrafiche.models import Dipendente

        context['dipendenti_attivi_count'] = Dipendente.objects.filter(
            azienda=az_ctx, stato='attivo'
        ).count()
    else:
        context['dipendenti_attivi_count'] = None
    context['testate_scenario'] = _somma_testate_ruoli(context.get('ruoli_config'))
    dac = context.get('dipendenti_attivi_count')
    context['delta_organico'] = (
        (context['testate_scenario'] - dac) if dac is not None else None
    )

    # ── 3. Archivia il risultato in SimulazioneOrganico (storico) ────────────
    az_per_archivio = context.get('azienda_operativa') or azienda
    if az_per_archivio and ruoli_dal_form:
        try:
            def _decimal_to_float(obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                if isinstance(obj, date):
                    return obj.isoformat()
                if isinstance(obj, dict):
                    return {k: _decimal_to_float(v) for k, v in obj.items()}
                if isinstance(obj, (list, tuple)):
                    return [_decimal_to_float(i) for i in obj]
                if obj is None or isinstance(obj, (str, int, float, bool)):
                    return obj
                # UUID, set, oggetti ORM residui in JSON da sessione/ruoli → stringa (evita TypeError su JSONField)
                return str(obj)

            risultato_serializzabile = _decimal_to_float({
                'totali_annui': context.get('totali_annui', {}),
                'costo_medio_mensile': context.get('costo_medio_mensile', 0),
                'totale_f24_annuo': context.get('totale_f24_annuo', 0),
                'riepilogo_mensile': [
                    {'mese': m['mese'], 'anno': m['anno'], 'totali': m['totali']}
                    for m in context.get('risultati_mensili', [])
                ],
            })
            simulazione = SimulazioneOrganico.objects.create(
                azienda=az_per_archivio,
                utente=request.user,
                mese_riferimento='2026-annuale',
                parametri_json=_decimal_to_float({'ruoli': context.get('ruoli_config', [])}),
                risultato_json=risultato_serializzabile,
                querystring=context['querystring'],
            )
            context['simulazione_salvata_id'] = simulazione.pk
            _safe_message(
                request,
                'success',
                f'Ruoli salvati — simulazione archiviata (ID: {simulazione.pk})',
            )
        except Exception as e:
            logger.exception('Errore archiviazione SimulazioneOrganico')
            _safe_message(
                request,
                'warning',
                f'Ruoli salvati, errore archivio: {e}',
            )

    return render(request, 'rapporto_di_lavoro/simulazione_2026_risultato.html', context)


def _safe_message(request, level, text):
    """Evita 500 se la sessione / framework messaggi non accetta il messaggio (es. sessione corrotta)."""
    try:
        if level == 'warning':
            messages.warning(request, text)
        elif level == 'error':
            messages.error(request, text)
        elif level == 'success':
            messages.success(request, text)
        else:
            messages.info(request, text)
    except Exception:
        logger.warning('simulazione_2026: impossibile registrare messaggio utente', exc_info=True)


@login_required
@user_passes_test(_is_admin_only)
def simulazione_2026_risultato(request):
    """Pagina risultato simulazione annua (12 tabelle mensili)."""
    try:
        _sim_params = _get_sim_params(request)
        if not _sim_params:
            _safe_message(
                request,
                'warning',
                'Configura prima ruoli e quantità nello scenario organico.',
            )
            return redirect('simulazione_2026_config')
        return _simulazione_2026_risultato_response(request, _sim_params)
    except Exception as e:
        logger.exception('simulazione_2026_risultato: errore non gestito')
        msg = (
            'Errore nel calcolo o nella visualizzazione della simulazione annua. '
            'Torna allo scenario, verifica ruoli e quantità, poi invia di nuovo il modulo.'
        )
        if getattr(request.user, 'is_superuser', False) or settings.DEBUG:
            det = f'{type(e).__name__}: {e}'
            if len(det) > 420:
                det = det[:417] + '...'
            msg = f'{msg} ({det})'
        _safe_message(request, 'error', msg)
        return redirect('simulazione_2026_config')


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL — simulazione annua
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(_is_admin_only)
def simulazione_2026_excel(request):
    """Esporta il piano organico e costo personale (scenario annuo) in .xlsx.

    Fogli:
    - Riepilogo Mensile: titolo, riga KPI (attivi vs testate scenario vs delta), totali mensili
    - Dettaglio Ruoli:   stesso header contestuale, dettaglio ruolo×mese
    - F24:               stesso header contestuale, breakdown F24

    Etichetta prodotto in UI: «Simulazione annua»; export riporta lo stesso naming.
    """
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    if not _get_sim_params(request):
        return redirect('simulazione_2026_config')

    ctx = _calcola_simulazione_2026(request)
    azienda = ctx.get('azienda_operativa')
    anno = ctx.get('anno', 2026)
    mesi_nomi = [
        '', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
        'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre',
    ]

    from anagrafiche.models import Dipendente

    dipendenti_attivi_count = None
    delta_organico = None
    testate_scenario = _somma_testate_ruoli(ctx.get('ruoli_config'))
    if azienda:
        dipendenti_attivi_count = Dipendente.objects.filter(
            azienda=azienda, stato='attivo'
        ).count()
        delta_organico = testate_scenario - dipendenti_attivi_count

    def _kpi_organico_excel_text():
        parts = [
            'Simulazione annua (organico e costo personale)',
            f'Anno di riferimento: {anno}',
        ]
        if azienda:
            parts.append(f'Azienda: {azienda.nome}')
        if dipendenti_attivi_count is not None:
            parts.append(f'Dipendenti in carico (stato attivo): {dipendenti_attivi_count}')
        parts.append(f'Testate scenario (Σ quantità ruoli): {testate_scenario}')
        if delta_organico is not None:
            if delta_organico > 0:
                parts.append(f'Delta vs organico attuale: +{delta_organico} testate (prospettiva aumento)')
            elif delta_organico < 0:
                parts.append(
                    f'Delta vs organico attuale: {delta_organico} testate (prospettiva riduzione)'
                )
            else:
                parts.append('Delta vs organico attuale: 0 (stesso ordine di grandezza)')
        return '  |  '.join(parts)

    kpi_line = _kpi_organico_excel_text()

    wb = openpyxl.Workbook()

    # ── Stili comuni ──────────────────────────────────────────────────────────
    _euro = '#,##0.00 €'
    _perc = '0.00%'
    _thin = Side(style='thin', color='BBBBBB')
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _hdr(ws, row, col, val, bg='1B3A5F', fg='FFFFFF', bold=True, wrap=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=8)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
        c.border = _border
        return c

    def _cell(ws, row, col, val, fmt=None, bold=False, align='right', color=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=8, color=color or '000000')
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border = _border
        if fmt:
            c.number_format = fmt
        return c

    def _euro_val(d):
        """Converte Decimal/str in float per Excel."""
        try:
            return float(d)
        except Exception:
            return 0.0

    # ══════════════════════════════════════════════════════════════════════════
    # FOGLIO 1: Riepilogo Mensile
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Riepilogo Mensile'
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = 'B4'

    # Titolo
    ws1.merge_cells('A1:Q1')
    t = ws1.cell(
        row=1,
        column=1,
        value=(
            f'SIMULAZIONE ANNUA — {anno} — '
            f'{(azienda.nome if azienda else "Azienda non indicata")} — Riepilogo mensile'
        ),
    )
    t.font = Font(bold=True, size=11, color='1B3A5F')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 20

    # Confronto organico (stesso significato della pagina web / risultato)
    ws1.merge_cells('A2:Q2')
    k = ws1.cell(row=2, column=1, value=kpi_line)
    k.font = Font(size=8, color='2D4E7A')
    k.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    k.fill = PatternFill('solid', fgColor='E8EEF5')
    for col in range(1, 18):
        ws1.cell(row=2, column=col).border = _border
    ws1.row_dimensions[2].height = 36

    hdrs1 = [
        'Mese', 'Lordo', 'TI + Bonus', 'Netto dip.',
        'INPS dip.', 'IRPEF', 'INPS az.', 'INAIL',
        'TFR rateo', '13ª rateo', '14ª rateo', 'Ferie', 'Permessi',
        'Costo az.', 'F24 mensile', '13ª cash', '14ª cash',
    ]
    for ci, h in enumerate(hdrs1, 1):
        _hdr(ws1, 3, ci, h)
    ws1.row_dimensions[3].height = 28

    for mese_data in ctx['risultati_mensili']:
        mn = mese_data['mese']
        t = mese_data['totali']
        r = ws1.max_row + 1
        _cell(ws1, r, 1, mesi_nomi[mn], fmt=None, bold=True, align='left')
        _cell(ws1, r, 2, _euro_val(t.get('lordo_mensile')), fmt=_euro)
        ti_bonus = _euro_val(t.get('trattamento_integrativo')) + _euro_val(t.get('bonus_l207'))
        _cell(ws1, r, 3, ti_bonus, fmt=_euro, color='0A6640')
        _cell(ws1, r, 4, _euro_val(t.get('netto_mensile')), fmt=_euro)
        _cell(ws1, r, 5, _euro_val(t.get('inps_dipendente')), fmt=_euro)
        _cell(ws1, r, 6, _euro_val(t.get('irpef_dipendenti')), fmt=_euro)
        _cell(ws1, r, 7, _euro_val(t.get('inps_azienda')), fmt=_euro)
        _cell(ws1, r, 8, _euro_val(t.get('inail')), fmt=_euro)
        _cell(ws1, r, 9, _euro_val(t.get('tfr')), fmt=_euro)
        _cell(ws1, r, 10, _euro_val(t.get('rateo_13')), fmt=_euro)
        _cell(ws1, r, 11, _euro_val(t.get('rateo_14')), fmt=_euro)
        _cell(ws1, r, 12, _euro_val(t.get('rateo_ferie')), fmt=_euro)
        _cell(ws1, r, 13, _euro_val(t.get('rateo_permessi')), fmt=_euro)
        _cell(ws1, r, 14, _euro_val(t.get('costo_azienda_totale')), fmt=_euro, bold=True, color='8B1C1C')
        _cell(ws1, r, 15, _euro_val(t.get('f24_netto')), fmt=_euro, bold=True, color='7A4E00')
        pg13 = _euro_val(t.get('pagamento_13_lordo'))
        pg14 = _euro_val(t.get('pagamento_14_lordo'))
        _cell(ws1, r, 16, pg13 if pg13 else '', fmt=_euro if pg13 else None)
        _cell(ws1, r, 17, pg14 if pg14 else '', fmt=_euro if pg14 else None)

    # Riga TOTALI ANNUI
    ta = ctx['totali_annui']
    rr = ws1.max_row + 1
    _cell(ws1, rr, 1, 'TOTALE ANNUO', bold=True, align='left')
    cols_ta = [
        _euro_val(ta.get('lordo_mensile')),
        _euro_val(ta.get('trattamento_integrativo')) + _euro_val(ta.get('bonus_l207')),
        _euro_val(ta.get('netto_mensile')),
        _euro_val(ta.get('inps_dipendente')),
        _euro_val(ta.get('irpef_dipendenti')),
        _euro_val(ta.get('inps_azienda')),
        _euro_val(ta.get('inail')),
        _euro_val(ta.get('tfr')),
        _euro_val(ta.get('rateo_13')),
        _euro_val(ta.get('rateo_14')),
        _euro_val(ta.get('rateo_ferie')),
        _euro_val(ta.get('rateo_permessi')),
        _euro_val(ta.get('costo_azienda_totale')),
        _euro_val(ctx.get('totale_f24_annuo')),
        _euro_val(ta.get('pagamento_13_lordo')),
        _euro_val(ta.get('pagamento_14_lordo')),
    ]
    for ci, v in enumerate(cols_ta, 2):
        c = ws1.cell(row=rr, column=ci, value=v)
        c.font = Font(bold=True, size=8, color='1B3A5F')
        c.fill = PatternFill('solid', fgColor='D8E8F5')
        c.number_format = _euro
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = _border

    # Larghezze colonne
    col_widths1 = [12, 13, 11, 13, 11, 10, 11, 9, 9, 9, 9, 9, 9, 13, 13, 11, 11]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # FOGLIO 2: Dettaglio Ruoli
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Dettaglio Ruoli')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'C4'

    ws2.merge_cells('A1:S1')
    t2 = ws2.cell(
        row=1,
        column=1,
        value=(
            f'SIMULAZIONE ANNUA — {anno} — '
            f'{(azienda.nome if azienda else "Azienda non indicata")} — Dettaglio per ruolo'
        ),
    )
    t2.font = Font(bold=True, size=11, color='1B3A5F')
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 20

    ws2.merge_cells('A2:S2')
    k2 = ws2.cell(row=2, column=1, value=kpi_line)
    k2.font = Font(size=8, color='2D4E7A')
    k2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    k2.fill = PatternFill('solid', fgColor='E8EEF5')
    for col in range(1, 20):
        ws2.cell(row=2, column=col).border = _border
    ws2.row_dimensions[2].height = 36

    hdrs2 = [
        'Mese', 'Ruolo', 'Qtà', 'Livello', 'Tipo contr.', 'PT coeff.',
        'Lordo unit.', 'Netto unit.', 'INPS dip. tot.', 'IRPEF tot.',
        'INPS az. tot.', 'INAIL tot.', 'TFR tot.',
        '13ª rateo', '14ª rateo', 'Ferie', 'Permessi',
        'Costo az. tot.', 'Variabile unit.',
    ]
    for ci, h in enumerate(hdrs2, 1):
        _hdr(ws2, 3, ci, h, bg='2D4E7A')
    ws2.row_dimensions[3].height = 28

    for mese_data in ctx['risultati_mensili']:
        mn = mese_data['mese']
        mese_nome = mesi_nomi[mn]
        for riga in mese_data['righe']:
            if riga.get('missing'):
                continue
            r = ws2.max_row + 1
            _cell(ws2, r, 1, mese_nome, bold=True, align='left')
            _cell(ws2, r, 2, riga.get('nome', ''), align='left')
            _cell(ws2, r, 3, riga.get('quantita', 1), align='center')
            _cell(ws2, r, 4, str(riga.get('livello', '')), align='center')
            tc = riga.get('tipo_contratto')
            _cell(ws2, r, 5, str(tc) if tc else 'full-time', align='center')
            _cell(ws2, r, 6, _euro_val(riga.get('coefficiente_ore', 1)), fmt='0.00', align='center')
            _cell(ws2, r, 7, _euro_val(riga.get('lordo_unit')), fmt=_euro)
            _cell(ws2, r, 8, _euro_val(riga.get('netto_unit')), fmt=_euro)
            _cell(ws2, r, 9, _euro_val(riga.get('inps_dipendente_tot')), fmt=_euro)
            _cell(ws2, r, 10, _euro_val(riga.get('irpef_tot')), fmt=_euro)
            _cell(ws2, r, 11, _euro_val(riga.get('inps_azienda_tot')), fmt=_euro)
            _cell(ws2, r, 12, _euro_val(riga.get('inail_tot')), fmt=_euro)
            _cell(ws2, r, 13, _euro_val(riga.get('tfr_tot')), fmt=_euro)
            _cell(ws2, r, 14, _euro_val(riga.get('rateo_13_tot')), fmt=_euro)
            _cell(ws2, r, 15, _euro_val(riga.get('rateo_14_tot')), fmt=_euro)
            _cell(ws2, r, 16, _euro_val(riga.get('rateo_ferie_tot')), fmt=_euro)
            _cell(ws2, r, 17, _euro_val(riga.get('rateo_perm_tot')), fmt=_euro)
            _cell(ws2, r, 18, _euro_val(riga.get('costo_azienda_tot')), fmt=_euro, bold=True, color='8B1C1C')
            _cell(ws2, r, 19, _euro_val(riga.get('lordo_variabile_unit')), fmt=_euro)

    col_widths2 = [11, 18, 5, 7, 12, 8, 11, 11, 11, 10, 11, 9, 9, 9, 9, 9, 9, 13, 11]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # FOGLIO 3: F24
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('F24')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'B4'

    ws3.merge_cells('A1:I1')
    t3 = ws3.cell(
        row=1,
        column=1,
        value=(
            f'SIMULAZIONE ANNUA — {anno} — '
            f'{(azienda.nome if azienda else "Azienda non indicata")} — F24 mensile (da scenario)'
        ),
    )
    t3.font = Font(bold=True, size=11, color='7A5000')
    t3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 20

    ws3.merge_cells('A2:I2')
    k3 = ws3.cell(row=2, column=1, value=kpi_line)
    k3.font = Font(size=8, color='5C4A00')
    k3.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    k3.fill = PatternFill('solid', fgColor='FFF6E0')
    for col in range(1, 10):
        ws3.cell(row=2, column=col).border = _border
    ws3.row_dimensions[2].height = 36

    hdrs3 = [
        'Mese', 'INPS DM10 (az.+dip.)', 'INPS az.', 'INPS dip.',
        'IRPEF (cod.1001)', 'TI cod.1701 (credito)', 'Bonus L207 cod.1704 (credito)',
        'Erario netto', 'F24 NETTO DA VERSARE',
    ]
    for ci, h in enumerate(hdrs3, 1):
        _hdr(ws3, 3, ci, h, bg='7A5000')
    ws3.row_dimensions[3].height = 28

    for mese_data in ctx['risultati_mensili']:
        mn = mese_data['mese']
        t = mese_data['totali']
        r = ws3.max_row + 1
        _cell(ws3, r, 1, mesi_nomi[mn], bold=True, align='left')
        _cell(ws3, r, 2, _euro_val(t.get('f24_inps')), fmt=_euro, bold=True)
        _cell(ws3, r, 3, _euro_val(t.get('inps_azienda')), fmt=_euro)
        _cell(ws3, r, 4, _euro_val(t.get('inps_dipendente')), fmt=_euro)
        _cell(ws3, r, 5, _euro_val(t.get('irpef_dipendenti')), fmt=_euro)
        _cell(ws3, r, 6, -_euro_val(t.get('trattamento_integrativo')), fmt=_euro, color='0A6640')
        _cell(ws3, r, 7, -_euro_val(t.get('bonus_l207')), fmt=_euro, color='0A6640')
        _cell(ws3, r, 8, _euro_val(t.get('f24_erario_netto')), fmt=_euro, bold=True)
        _cell(ws3, r, 9, _euro_val(t.get('f24_netto')), fmt=_euro, bold=True, color='7A5000')

    # Totale annuo F24
    ra = ws3.max_row + 1
    _cell(ws3, ra, 1, 'TOTALE ANNUO', bold=True, align='left')
    f24_tots = [
        _euro_val(ta.get('f24_inps')),
        _euro_val(ta.get('inps_azienda')),
        _euro_val(ta.get('inps_dipendente')),
        _euro_val(ta.get('irpef_dipendenti')),
        -_euro_val(ta.get('trattamento_integrativo')),
        -_euro_val(ta.get('bonus_l207')),
        _euro_val(ta.get('f24_erario_netto')),
        _euro_val(ctx.get('totale_f24_annuo')),
    ]
    for ci, v in enumerate(f24_tots, 2):
        c = ws3.cell(row=ra, column=ci, value=v)
        c.font = Font(bold=True, size=8, color='7A5000')
        c.fill = PatternFill('solid', fgColor='FFF6E0')
        c.number_format = _euro
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = _border

    # Nota INAIL sotto la tabella
    rn = ws3.max_row + 2
    ws3.cell(row=rn, column=1,
             value=(
                 'INAIL annuo (autoliquidazione febbraio anno succ.): '
                 f'€ {euro_it_str(ctx.get("totale_inail_annuo", 0))}'
             )
             ).font = Font(italic=True, size=8, color='666666')
    rn2 = rn + 1
    ws3.cell(row=rn2, column=1,
             value='Addizionali regionali/comunali: saldo anno precedente — versate in F24 anno successivo (mese rif. = anno prec.)'
             ).font = Font(italic=True, size=8, color='666666')

    col_widths3 = [12, 18, 12, 12, 16, 20, 22, 14, 20]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Imposta orientamento pagina stampa per tutti i fogli ──────────────────
    for ws in [ws1, ws2, ws3]:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

    # ── Genera response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    azienda_slug = (azienda.nome if azienda else 'scenario').replace(' ', '_')[:20]
    filename = f'Simulazione_annua_{anno}_{azienda_slug}.xlsx'

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# CREA PROPOSTE DA SIMULAZIONE — simulazione_2026_crea_proposte
# ─────────────────────────────────────────────────────────────────────────────

def _risolvi_tipo_contratto_proposta(tipo_rapporto, coeff_ore, data_inizio=None, data_fine=None):
    """
    Mappa i due parametri della simulazione (tipo_rapporto + coefficiente ore del
    TipoContratto PT/FT) al TipoContratto legale corretto per PropostaAssunzione.

    Logica:
    - indeterminato + FT (coeff=1)  → TC "Indeterminato"
    - indeterminato + PT (coeff<1)  → TC "Part-Time Indeterminato"
    - determinato                   → TC con durata_giorni più vicina alla durata effettiva
    - apprendistato                 → TC "Apprendistato"
    """
    tipo_rapporto = (tipo_rapporto or '').lower()
    coeff = Decimal(str(coeff_ore or 1))

    if tipo_rapporto == 'apprendistato':
        tc = TipoContratto.objects.filter(attivo=True, nome__icontains='apprendistato').first()
        return tc or TipoContratto.objects.filter(attivo=True).order_by('id').first()

    if tipo_rapporto == 'determinato':
        det_qs = (
            TipoContratto.objects.filter(attivo=True, nome__icontains='determinato')
            .exclude(nome__icontains='indeterminato')
        )
        # Trova il più vicino alla durata effettiva del ruolo
        if data_inizio and data_fine:
            durata_gg = (data_fine - data_inizio).days
            best = None
            best_diff = 10 ** 9
            for tc in det_qs:
                if tc.durata_giorni is not None:
                    diff = abs(tc.durata_giorni - durata_gg)
                    if diff < best_diff:
                        best_diff = diff
                        best = tc
            if best:
                return best
        tc = det_qs.order_by('durata_giorni').first()
        return tc or TipoContratto.objects.filter(attivo=True).order_by('id').first()

    # indeterminato (default)
    if coeff >= Decimal('1'):
        tc = TipoContratto.objects.filter(attivo=True, nome__iexact='Indeterminato').first()
    else:
        tc = TipoContratto.objects.filter(attivo=True, nome__icontains='Part-Time Indeterminato').first()
        if not tc:
            tc = TipoContratto.objects.filter(attivo=True, nome__iexact='Indeterminato').first()
    return tc or TipoContratto.objects.filter(attivo=True).order_by('id').first()


def _risolvi_coeff_ore_ruolo(ruolo):
    """Restituisce il coefficiente ore del TipoContratto scelto in simulazione (per PT)."""
    tipo_contratto_id = ruolo.get('tipo_contratto_id')
    if tipo_contratto_id:
        try:
            tc = TipoContratto.objects.get(pk=int(tipo_contratto_id))
            return tc.coefficiente_ore, tc
        except (TipoContratto.DoesNotExist, ValueError, TypeError):
            pass
    return Decimal('1'), None


@login_required
@user_passes_test(_is_admin_only)
def simulazione_2026_crea_proposte(request):
    """
    Converte i ruoli simulati in PropostaAssunzione.

    GET  — mostra form di selezione ruoli con dati CCNL e tipo contratto risolto
    POST — crea un Dipendente(stato='candidato') + PropostaAssunzione per ogni slot
           di ogni ruolo selezionato
    """
    from django.utils import timezone
    from anagrafiche.models import Dipendente
    from .models import ModuloContrattuale, PropostaAssunzione, ParametroCCNLTurismo

    azienda = _get_azienda_con_fallback(request.user, request.session)
    if not azienda:
        messages.error(request, 'Nessuna azienda operativa. Selezionane una prima.')
        return redirect('simulazione_2026_config')

    ruoli_config = _build_ruoli_config(request)
    if not ruoli_config:
        messages.warning(request, 'Nessun ruolo configurato nella simulazione.')
        return redirect('simulazione_2026_config')

    scatti_db: dict = {}
    try:
        _ccnl = CCNL.objects.filter(sigla__icontains='FIPE').first()
        if _ccnl:
            for _s in ParametroScattiAnnuali.objects.filter(ccnl=_ccnl, anno=2026, attivo=True):
                scatti_db.setdefault(_s.livello, []).append((_s.anni_anzianita, _s.importo_scatto))
            for _k in scatti_db:
                scatti_db[_k].sort(key=lambda x: x[0])
    except Exception:
        scatti_db = {}

    # Modulo contrattuale di default (proposta_assunzione) — crea se non esiste
    modulo, _ = ModuloContrattuale.objects.get_or_create(
        nome='Proposta Assunzione Standard',
        defaults={
            'categoria': 'proposta_assunzione',
            'descrizione': 'Modulo standard generato dalla simulazione annua.',
            'compilabile_da_dipendente': True,
            'attivo': True,
        },
    )

    if request.method == 'POST':
        ruoli_selezionati = request.POST.getlist('ruoli_selezionati')
        proposte_create = 0
        errori = []

        for ruolo in ruoli_config:
            rid = str(ruolo['id'])
            if rid not in ruoli_selezionati:
                continue

            # Quantità da creare (può essere ridotta dal form)
            try:
                qta = int(request.POST.get(f'qta_{rid}', ruolo['quantita']))
                qta = max(1, min(qta, ruolo['quantita']))
            except (ValueError, TypeError):
                qta = ruolo['quantita']

            # Coefficiente ore (da TipoContratto PT scelto in simulazione)
            coeff_ore, _ = _risolvi_coeff_ore_ruolo(ruolo)

            # Tipo contratto LEGALE per la proposta (Indeterminato / Determinato / Apprendistato)
            tipo_contr = _risolvi_tipo_contratto_proposta(
                ruolo.get('tipo_rapporto'), coeff_ore,
                data_inizio=ruolo.get('data_inizio'), data_fine=ruolo.get('data_fine'),
            )

            # Parametro CCNL per il livello simulato
            data_rif = ruolo.get('data_inizio') or date(2026, 1, 1)
            parametro_ccnl = ParametroCCNLTurismo.objects.filter(
                livello=ruolo['livello'],
                attivo=True,
                decorrenza_validita_da__lte=data_rif,
            ).order_by('-decorrenza_validita_da').first()

            payload = {}
            if parametro_ccnl:
                from .utils_motore_paga import ricava_parametri_proposta_contrattuale
                scatto_totale = _calcola_scatto_totale(
                    ruolo.get('livello'),
                    ruolo.get('anni_anzianita', 0),
                    scatti_db,
                )
                payload = ricava_parametri_proposta_contrattuale(
                    parametro_ccnl=parametro_ccnl,
                    tipo_contratto=tipo_contr,
                    anno=data_rif.year,
                    mese=data_rif.month,
                    azienda=azienda,
                    data_inizio_rapporto=data_rif,
                    data_fine_rapporto=ruolo.get('data_fine') if ruolo.get('tipo_rapporto') == 'determinato' else None,
                    superminimo=ruolo.get('superminimo') or Decimal('0'),
                    indennita_turno=ruolo.get('indennita_turno') or Decimal('0'),
                    scatto_anzianita=scatto_totale,
                    indennita_extra=ruolo.get('indennita_extra') or Decimal('0'),
                )

            lordo_mensile = payload.get('stipendio_lordo_mensile', Decimal('0'))
            paga_base = payload.get('paga_base_mensile', Decimal('0'))
            contingenza = payload.get('contingenza_mensile', Decimal('0'))
            edr = payload.get('edr_mensile', Decimal('0'))
            ore_sett = payload.get('ore_settimanali', (Decimal('40') * coeff_ore).quantize(Decimal('0.01')))
            ore_mens = payload.get('ore_mensili', (Decimal('173') * coeff_ore).quantize(Decimal('0.01')))
            ore_giorn = payload.get('ore_giornaliere', Decimal('8'))
            sc_per = payload.get('scatto_periodicita_mesi', 24)
            sc_imp = payload.get('scatto_importo', Decimal('0'))
            sc_max = payload.get('numero_scatti_massimi', 10)
            qualifica = payload.get('qualifica', ruolo['nome'])
            tredicesima = payload.get('tredicesima', True)
            quattordicesima = payload.get('quattordicesima', False)
            giorni_ferie_annuali = payload.get('giorni_ferie_annuali', 26)
            giorni_permesso_annuali = payload.get('giorni_permesso_annuali', 3)

            for slot in range(1, qta + 1):
                try:
                    cognome_temp = f"{ruolo['nome']} #{slot}" if qta > 1 else ruolo['nome']
                    dip = Dipendente.objects.create(
                        azienda=azienda,
                        nome='Candidato',
                        cognome=cognome_temp[:100],
                        codice_fiscale=None,
                        data_nascita=None,
                        data_assunzione=None,
                        indirizzo='',
                        email='',
                        ruolo=ruolo['nome'],
                        livello=ruolo['livello'],
                        stato='candidato',
                    )

                    numero = f"SIM2026-{rid}-{slot}-{dip.pk}"

                    PropostaAssunzione.objects.create(
                        azienda=azienda,
                        dipendente=dip,
                        modulo=modulo,
                        parametro_ccnl=parametro_ccnl,
                        tipo_contratto=tipo_contr,
                        numero_proposta=numero,
                        titolo=f"Proposta assunzione — {ruolo['nome']} L.{ruolo['livello']}",
                        posizione=ruolo['nome'],
                        livello_ccnl=ruolo['livello'],
                        qualifica=qualifica,
                        stipendio_lordo_mensile=lordo_mensile,
                        paga_base_mensile=paga_base,
                        contingenza_mensile=contingenza,
                        edr_mensile=edr,
                        tredicesima=tredicesima,
                        quattordicesima=quattordicesima,
                        giorni_ferie_annuali=giorni_ferie_annuali,
                        giorni_permesso_annuali=giorni_permesso_annuali,
                        ore_settimanali=ore_sett,
                        ore_mensili=ore_mens,
                        ore_giornaliere=ore_giorn,
                        scatto_periodicita_mesi=sc_per,
                        scatto_importo=sc_imp,
                        numero_scatti_massimi=sc_max,
                        data_inizio_rapporto=ruolo.get('data_inizio') or date(2026, 1, 1),
                        data_fine_rapporto=ruolo.get('data_fine') if ruolo.get('tipo_rapporto') == 'determinato' else None,
                        stato='bozza',
                        creato_da=request.user,
                    )
                    proposte_create += 1
                except Exception as exc:
                    logger.exception("Errore creazione proposta per ruolo '%s' slot %d", ruolo['nome'], slot)
                    errori.append(f"{ruolo['nome']} #{slot}: {exc}")

        if proposte_create:
            messages.success(request, f'{proposte_create} proposta/e create con successo.')
        if errori:
            for e in errori:
                messages.error(request, e)

        return redirect('lista_proposte_assunzione')

    # GET — form selezione ruoli con dati risolti
    ruoli_con_dati = []
    for ruolo in ruoli_config:
        data_rif = ruolo.get('data_inizio') or date(2026, 1, 1)
        parametro_ccnl = ParametroCCNLTurismo.objects.filter(
            livello=ruolo['livello'], attivo=True,
            decorrenza_validita_da__lte=data_rif,
        ).order_by('-decorrenza_validita_da').first()

        coeff_ore, tc_pt = _risolvi_coeff_ore_ruolo(ruolo)
        tipo_contr_proposta = _risolvi_tipo_contratto_proposta(
            ruolo.get('tipo_rapporto'), coeff_ore,
            data_inizio=ruolo.get('data_inizio'), data_fine=ruolo.get('data_fine'),
        )

        lordo_ft = parametro_ccnl.importo_lordo_mensile if parametro_ccnl else None
        lordo_eff = (Decimal(str(lordo_ft)) * coeff_ore).quantize(Decimal('0.01')) if lordo_ft else None
        if parametro_ccnl:
            try:
                from .utils_motore_paga import ricava_parametri_proposta_contrattuale
                scatto_totale = _calcola_scatto_totale(
                    ruolo.get('livello'),
                    ruolo.get('anni_anzianita', 0),
                    scatti_db,
                )
                payload = ricava_parametri_proposta_contrattuale(
                    parametro_ccnl=parametro_ccnl,
                    tipo_contratto=tipo_contr_proposta,
                    anno=data_rif.year,
                    mese=data_rif.month,
                    azienda=azienda,
                    data_inizio_rapporto=data_rif,
                    data_fine_rapporto=ruolo.get('data_fine') if ruolo.get('tipo_rapporto') == 'determinato' else None,
                    superminimo=ruolo.get('superminimo') or Decimal('0'),
                    indennita_turno=ruolo.get('indennita_turno') or Decimal('0'),
                    scatto_anzianita=scatto_totale,
                    indennita_extra=ruolo.get('indennita_extra') or Decimal('0'),
                )
                lordo_eff = payload.get('stipendio_lordo_mensile', lordo_eff)
            except Exception:
                logger.exception("Errore calcolo payload proposta per ruolo '%s'", ruolo.get('nome'))

        ruoli_con_dati.append({
            **ruolo,
            'parametro_ccnl': parametro_ccnl,
            'lordo_ft': lordo_ft,
            'lordo_mensile': lordo_eff,
            'coeff_ore': coeff_ore,
            'tc_pt_nome': tc_pt.nome if tc_pt else 'Full-time',
            'tipo_contratto_proposta': tipo_contr_proposta,
        })

    context = {
        'azienda': azienda,
        'ruoli': ruoli_con_dati,
        'querystring': _get_sim_params(request).urlencode(),
    }
    return render(request, 'rapporto_di_lavoro/simulazione_2026_crea_proposte.html', context)
