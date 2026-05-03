"""
Import proforma / parcelle studio per area consulente: estrazione testo da PDF,
parsing campi (PROFORMA, numero, data, totale da pagare), righe dare/avere e saldo progressivo.

Estrazione: prima ``pdfplumber``; se il testo è quasi vuoto, tentativo OCR via
``pdftoppm`` + ``tesseract`` (stesso schema di ``scripts/analizza_pdf_paghe.py``).
"""
from __future__ import annotations

import io
import re
import subprocess
import tempfile
import uuid
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

# ── Estrazione testo ─────────────────────────────────────────────────────────


def _pdf_text_pdfplumber(pdf_path: Path) -> str:
    import pdfplumber

    chunks: list[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            try:
                t = page.extract_text() or ""
            except Exception:
                t = ""
            chunks.append(t)
    return "\n".join(chunks)


def _pdf_text_tesseract(pdf_path: Path) -> str:
    """OCR con binari di sistema (pdftoppm + tesseract, lingua italiana)."""
    with tempfile.TemporaryDirectory(prefix="gesper_regstudio_ocr_") as td:
        tmpdir = Path(td)
        ppm_cmd = ["pdftoppm", "-png", "-r", "200", str(pdf_path), str(tmpdir / "page")]
        subprocess.run(ppm_cmd, check=True, capture_output=True, text=True)
        page_imgs = sorted(tmpdir.glob("page-*.png"))
        pages_txt: list[str] = []
        for img in page_imgs:
            out_base = img.with_suffix("")
            tess_cmd = [
                "tesseract",
                str(img),
                str(out_base),
                "-l",
                "ita",
                "--psm",
                "6",
            ]
            subprocess.run(tess_cmd, check=True, capture_output=True, text=True)
            txt_path = Path(f"{out_base}.txt")
            if txt_path.exists():
                pages_txt.append(txt_path.read_text(encoding="utf-8", errors="ignore"))
            else:
                pages_txt.append("")
        return "\n".join(pages_txt)


def estrai_testo_da_pdf(pdf_path: Path) -> tuple[str, str]:
    """
    Restituisce (testo, metodo) con metodo in {'pdfplumber', 'tesseract', 'vuoto'}.
    """
    try:
        txt = _pdf_text_pdfplumber(pdf_path)
    except Exception:
        txt = ""
    compact = re.sub(r"\s+", "", txt)
    if len(compact) >= 40:
        return txt, "pdfplumber"
    try:
        txt2 = _pdf_text_tesseract(pdf_path)
        if re.sub(r"\s+", "", txt2).strip():
            return txt2, "tesseract"
    except (FileNotFoundError, subprocess.CalledProcessError, OSError):
        pass
    return txt or "", "vuoto" if not txt.strip() else "pdfplumber"


# ── Parsing importi / date ───────────────────────────────────────────────────


def _normalizza_testo_pdf(testo: str) -> str:
    """Spazi Unicode e a capo rumorosi: migliora il match delle etichette importo."""
    if not testo:
        return ""
    t = testo.replace("\u00a0", " ").replace("\u202f", " ").replace("\u2009", " ")
    t = re.sub(r"[ \t]+\n", "\n", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t


def _normalizza_etichette_numero_ocr(testo: str) -> str:
    """
    Ripara etichette spezzate da PDF/OCR (es. «n.» + «umero» → «numero»),
    così non finisce in DB il frammento «umero» al posto del vero protocollo.
    """
    if not testo:
        return ""
    t = testo
    t = re.sub(r"(?i)n\.\s*\n\s*umero\b", "numero", t)
    t = re.sub(r"(?i)n\.\s+umero\b", "numero", t)
    t = re.sub(r"(?i)\bnu\s*\n\s*mero\b", "numero", t)
    t = re.sub(r"(?i)\bnum\s*\.\s*\n\s*ero\b", "numero", t)
    return t


def parse_importo_form(raw: str) -> Decimal | None:
    """Parse importo da campo form (formato italiano o semplice)."""
    return _parse_it_decimal(raw or "")


def _parse_it_decimal(raw: str) -> Decimal | None:
    if not raw:
        return None
    s = raw.strip()
    s = re.sub(r"[\s€EUR]", "", s, flags=re.I)
    if not s or s in "-—":
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    if s.startswith("-"):
        neg = True
        s = s[1:]
    # Italiano: migliaia con . e decimali con ,
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2 and parts[1].isdigit():
            s = parts[0].replace(".", "") + "." + parts[1]
        else:
            s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        d = Decimal(s)
        return -d if neg else d
    except InvalidOperation:
        return None


def _parse_date(s: str) -> date | None:
    s = s.strip()
    m = re.match(r"^(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})$", s)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000 if y < 70 else 1900
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def _first_date_in_text(text: str) -> date | None:
    for m in re.finditer(r"\b(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})\b", text):
        d = _parse_date(m.group(0))
        if d:
            return d
    return None


# ── Risultato parsing ─────────────────────────────────────────────────────────


@dataclass
class EsitoParsingProforma:
    tipo_documento: str  # proforma | parcella | sconosciuto
    numero_documento: str
    data_documento: date | None
    totale_da_pagare: Decimal | None
    avvisi: list[str] = field(default_factory=list)


# Frammenti errati da match parziale su «numero» / OCR (non sono numeri documento)
_NUMERO_DOCUMENTO_RIFIUTA = frozenset(
    {
        "umero",
        "mero",
        "ero",
        "ro",
        "num",
        "numo",
        "umo",
        "bero",
    }
)


def _is_plausible_numero_documento(s: str) -> bool:
    """Evita di confondere date, importi o frammenti di «numero» con il numero documento."""
    s = (s or "").strip()
    if len(s) < 2 or len(s) > 80:
        return False
    sl = s.lower()
    if sl in _NUMERO_DOCUMENTO_RIFIUTA:
        return False
    if re.match(r"^\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4}$", s):
        return False
    compact = re.sub(r"[\s€]", "", s, flags=re.I)
    if re.match(r"^-?[\d.]+,\d{2}$", compact) or re.match(r"^-?[\d,]+\.\d{2}$", compact):
        return False
    return True


def _estrai_numero_documento_proforma_parcella(testo: str) -> str:
    """
    Numero proforma/parcella: etichette tipiche (anche su riga diversa dal titolo).
    Ordine dal più specifico al fallback legacy.
    """
    patterns = (
        # Protocollo / Prot. / Rif. (mai «PROT» generico: matcherebbe dentro …/PROT-…)
        r"(?i)(?:protocollo|prot\.)\s*[:#.]?\s*([A-Za-z0-9/_\-\.]{2,40})",
        r"(?i)\brif\.?\s*(?:doc\.?|documento|fatt\.?)?\s*[:#.]?\s*([A-Za-z0-9/_\-\.]{2,40})",
        # Fattura proforma / proforma fattura (layout commercialisti)
        r"(?i)(?:fattura\s+proforma|proforma\s+fattura|nota\s+proforma)\s*(?:nr\.?|num\.?|n°|n\.\s*°?)\s*[:.]?\s*"
        r"([A-Za-z0-9/_\-\.]{2,40})",
        # «Numero proforma …» / «N. proforma …» / «Numero documento» (NUMERO intero: evita num\. su «numero»)
        r"(?:NUMERO|N\.)[\s\n:]+(?:DEL\s+|DELLA\s+)?(?:PRO[-\s]?FORMA|PROFORMA|PARCHELLA|PARCELLA|DOCUMENTO)\s*[:.]?\s*"
        r"([A-Za-z0-9/_\-\.]{2,40})",
        r"(?:NUMERO|N\.)[\s\n:]+(?:DOCUMENTO|DOC\.)\s*[:.]?\s*([A-Za-z0-9/_\-\.]{2,40})",
        # «Proforma n. …» sulla stessa riga (NUM\. solo con punto fermo, non prefisso di «numero»)
        r"(?:PRO[-\s]?FORMA|PROFORMA|PARCHELLA|PARCELLA)\s*(?:NR\.?\s*|NUM\.\s*|NUMERO\s+|N\.?\s*°?\s*)[:.]?\s*"
        r"([A-Za-z0-9/_\-\.]{2,40})",
        # Legacy: dopo PROFORMA/PARCELLA — nr. prima di n. (evita che n.? mangi la «n» di «nr.»)
        r"(?:PROFORMA|PARCHELLA|PARCELLA)[^\n]{0,120}?(?:nr\.?\s*|n\.?\s*°?\s*|numero\s*[:.]?\s*)"
        r"([A-Za-z0-9/_\-\.]{1,40})",
        # Proforma/parcella e poi entro poche righe «n. …» (PDF a blocchi / OCR)
        r"(?is)(?:PRO[-\s]?FORMA|PROFORMA|PARCHELLA|PARCELLA)[\s\S]{0,480}?"
        r"\b(?:nr\.?|n\.\s*°?|n\.\s+|n\.(?=\d))\s*[:.]?\s*([A-Za-z0-9/_\-\.]{2,40})\b",
    )
    for pat in patterns:
        m = re.search(pat, testo, re.I | re.DOTALL)
        if m:
            cand = m.group(1).strip()
            if _is_plausible_numero_documento(cand):
                return cand[:80]
    # Titolo su una riga, numero sulla successiva (PDF a blocchi)
    lines = testo.splitlines()
    for i, line in enumerate(lines):
        if not re.search(r"\b(?:PRO[-\s]?FORMA|PROFORMA|PARCHELLA|PARCELLA)\b", line, re.I):
            continue
        for j in range(i + 1, min(i + 10, len(lines))):
            seg = lines[j]
            m = re.search(
                r"^\s*(?:NUMERO|NR\.?|N\.?\s*°?|NUM\.(?!\w)|RIF\.?)\s*[:#.]?\s*([A-Za-z0-9/_\-\.]{2,40})\b",
                seg,
                re.I,
            )
            if m:
                cand = m.group(1).strip()
                if _is_plausible_numero_documento(cand):
                    return cand[:80]
            m2 = re.search(
                r"^\s*(?:PRO[-\s]?FORMA|PROFORMA)\s*(?:N\.?\s*°?\s*|NR\.?\s*)[:.]?\s*([A-Za-z0-9/_\-\.]{2,40})\b",
                seg,
                re.I,
            )
            if m2:
                cand = m2.group(1).strip()
                if _is_plausible_numero_documento(cand):
                    return cand[:80]
            # Riga che è solo (o quasi) «N. 12/2022» dopo il titolo
            m3 = re.match(
                r"^\s*N\.?\s*°?\s*[:.]?\s*([A-Za-z0-9/_\-\.]{2,40})\s*(?:\s|$)",
                seg,
                re.I,
            )
            if m3 and len(seg.strip()) <= 48:
                cand = m3.group(1).strip()
                if _is_plausible_numero_documento(cand):
                    return cand[:80]
    return ""


def classifica_tipo(testo: str, nome_file: str) -> str:
    u = testo.upper()
    n = nome_file.upper()
    if "PROFORMA" in u or "PRO-FORMA" in u:
        return "proforma"
    if "PARCHELLA" in u or "PARCELLA" in u or "PARCELL" in u:
        return "parcella"
    if "PROFORMA" in n:
        return "proforma"
    if "PARCHELLA" in n or "PARCELLA" in n:
        return "parcella"
    return "sconosciuto"


def _totale_etichette_estese_proforma(testo: str) -> Decimal | None:
    """Etichette totali esplicite (affidabili come i pattern principali; nessun avviso euristico)."""
    if not (testo or "").strip():
        return None
    extra_patterns = (
        r"(?:NETTO\s+A\s+PAGARE|TOTALE\s+A\s+PAGARE|IMPORTO\s+A\s+PAGARE)\s*[:\s€]*([\d\.\s]+(?:,\d{1,2})?)",
        r"(?:TOTALE\s+COMPLESSIVO|TOTALE\s+FATTURA|TOTALE\s+DOCUMENTO|TOTALE\s+GENERALE)\s*[:\s€]*([\d\.\s]+(?:,\d{1,2})?)",
        r"(?:IMPORTO\s+TOTALE\s+DOCUMENTO|IMPORTO\s+COMPLESSIVO|QUOTA\s+COMPLESSIVA|COMPENSO\s+TOTALE)\s*[:\s€]*"
        r"([\d\.\s]+(?:,\d{1,2})?)",
        r"(?:Saldo|Importo)\s+documento\s*[:\s€]*([\d\.\s]+(?:,\d{1,2})?)",
        r"(?:TOTALE|T\.)\s+(?:DELLA\s+|DEL\s+)?(?:PARCHELLA|PARCELLA)\s*(?:PROFESSIONALE\s*)?[:\s€]*([\d\.\s]+(?:,\d{1,2})?)",
        r"(?:TOTALE|T\.)\s+(?:DELLA\s+|DEL\s+)?(?:PRO[-\s]?FORMA|PROFORMA)\s*[:\s€]*([\d\.\s]+(?:,\d{1,2})?)",
        r"(?:IMPORTO\s+)?TOTALE\s+(?:IVA\s+)?(?:COMPRESA\s+)?(?:PARCHELLA|PARCELLA|PRO[-\s]?FORMA)\s*[:\s€]*"
        r"([\d\.\s]+(?:,\d{1,2})?)",
    )
    for pat in extra_patterns:
        m = re.search(pat, testo, re.I)
        if m:
            d = _parse_it_decimal(m.group(1))
            if d is not None and d > 0:
                return d
    return None


def _totale_euristica_righe_finali_proforma(testo: str) -> Decimal | None:
    """Ultime righe con «TOTALE» e importo in coda (meno affidabile)."""
    if not (testo or "").strip():
        return None
    for line in reversed(testo.splitlines()[-150:]):
        if not re.search(r"\bTOTALE\b", line, re.I):
            continue
        m = re.search(r"([\d]{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})\s*(?:EUR|€)?\s*$", line.strip(), re.I)
        if not m:
            m = re.search(r"([\d\.\s]+(?:,\d{1,2}))\s*(?:EUR|€)?\s*$", line.strip(), re.I)
        if m:
            d = _parse_it_decimal(m.group(1))
            if d is not None and d > 0:
                return d
    return None


def parse_testo_proforma_parcella(testo: str, nome_file: str) -> EsitoParsingProforma:
    testo = _normalizza_etichette_numero_ocr(_normalizza_testo_pdf(testo))
    tipo = classifica_tipo(testo, nome_file)
    avvisi: list[str] = []

    numero = _estrai_numero_documento_proforma_parcella(testo)
    if not numero:
        for ap in (
            r"(?i)(?:n\.?\s*°?\s*documento|documento\s+n\.?)\s*[:.]?\s*([A-Za-z0-9/_\-\.]{1,32})",
            r"(?i)(?:codice|identificativo)\s+documento\s*[:.]?\s*([A-Za-z0-9/_\-\.]{2,40})",
            r"(?i)(?:numero|n\.?)\s+del\s+(?:documento|fattura)\s*[:.]?\s*([A-Za-z0-9/_\-\.]{2,40})",
            r"(?i)(?:numero|n\.?)\s+fattura\s*[:.]?\s*([A-Za-z0-9/_\-\.]{2,40})",
            r"(?i)(?:rep\.?\s*n\.?|repertorio\s*n\.?)\s*[:.]?\s*([A-Za-z0-9/_\-\.]{2,40})",
        ):
            m_alt = re.search(ap, testo)
            if m_alt:
                cand = m_alt.group(1).strip()
                if _is_plausible_numero_documento(cand):
                    numero = cand
                    break

    data_doc: date | None = None
    m_data = re.search(
        r"(?:DATA\s*(?:DOCUMENTO|FATTURA|PROFORMA)?|DATA\s*EMISSIONE)\s*[:.]?\s*(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})",
        testo,
        re.I,
    )
    if m_data:
        data_doc = _parse_date(m_data.group(1))
    if data_doc is None:
        data_doc = _first_date_in_text(testo[:2500])

    totale: Decimal | None = None
    m_tot = re.search(
        r"TOTALE\s+DA\s+PAGAR[EI]\s*[:.]?\s*([\d\.\s,]+(?:,\d{1,2})?)",
        testo,
        re.I,
    )
    if m_tot:
        totale = _parse_it_decimal(m_tot.group(1))
    if totale is None:
        m_tot2 = re.search(
            r"(?:IMPORTO\s+TOTALE|TOTALE\s+DOCUMENTO)\s*[:.]?\s*([\d\.\s,]+(?:,\d{1,2})?)",
            testo,
            re.I,
        )
        if m_tot2:
            totale = _parse_it_decimal(m_tot2.group(1))

    # Etichette esplicite «totale parcella» / «totale proforma» (spesso su riga unica col valore)
    if totale is None:
        m_tp = re.search(
            r"(?:TOTALE|T\.)\s+(?:DELLA\s+|DEL\s+)?(?:PARCHELLA|PARCELLA)\s*(?:PROFESSIONALE\s*)?[:\s€]*\s*"
            r"([\d]{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}|[\d\.\s]+,\d{2})",
            testo,
            re.I,
        )
        if m_tp:
            totale = _parse_it_decimal(m_tp.group(1))
    if totale is None:
        m_pf = re.search(
            r"(?:TOTALE|T\.)\s+(?:DELLA\s+|DEL\s+)?(?:PRO[-\s]?FORMA|PROFORMA)\s*[:\s€]*\s*"
            r"([\d]{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}|[\d\.\s]+,\d{2})",
            testo,
            re.I,
        )
        if m_pf:
            totale = _parse_it_decimal(m_pf.group(1))

    if totale is None:
        tot_lbl = _totale_etichette_estese_proforma(testo)
        if tot_lbl is not None:
            totale = tot_lbl
    if totale is None:
        tot_heur = _totale_euristica_righe_finali_proforma(testo)
        if tot_heur is not None:
            totale = tot_heur
            avvisi.append("Totale ricavato da pattern alternativo (verificare sul PDF).")
    if totale is None:
        avvisi.append("Totale da pagare non rilevato automaticamente.")
    if data_doc is None:
        avvisi.append("Data documento non rilevata; usare ordinamento per data import.")
    if not numero:
        avvisi.append("Numero documento non rilevato.")

    return EsitoParsingProforma(
        tipo_documento=tipo,
        numero_documento=numero[:80] if numero else "",
        data_documento=data_doc,
        totale_da_pagare=totale,
        avvisi=avvisi,
    )


def elenco_pdf_cartella(
    root: Path,
    *,
    escludi_prefisso_nome: str = "riepilogo",
) -> list[Path]:
    """PDF ricorsivi, esclusi file il cui nome (senza path) inizia con il prefisso (case-insensitive)."""
    pref = escludi_prefisso_nome.lower()
    out: list[Path] = []
    for p in sorted(root.rglob("*")):
        if not p.is_file():
            continue
        if p.suffix.lower() != ".pdf":
            continue
        if p.name.lower().startswith(pref):
            continue
        out.append(p)
    # Ordine cronologico preferito: data modifica file (proxy finché non si parsa)
    out.sort(key=lambda x: (x.stat().st_mtime, x.name.lower()))
    return out


def _numeri_documento_aggancio_coerenti(num_libro: str, num_pdf: str) -> bool:
    """
    True se il numero in libro (es. «182», «PAR-2021-1») e quello estratto dal PDF
    (es. «PARCELLA 182», «182/2021») si riferiscono allo stesso documento.

    Usato per aggancio PDF → righe pregresse e per deduplica upload coerente con l’aggancio.
    """
    a = (num_libro or "").strip()
    b = (num_pdf or "").strip()
    if len(a) < 2 or len(b) < 2:
        return False
    if a.casefold() == b.casefold():
        return True

    def _strip_lab(x: str) -> str:
        t = re.sub(
            r"(?i)\b(parcella|parchell|proforma|pro-forma|fattura|nr\.?|num\.?|n\.?°?|"
            r"protocollo|prot\.?|documento|doc\.?|repertorio|rep\.?)\b",
            " ",
            x,
        )
        return re.sub(r"[\s._\-/]+", "", t)

    sa, sb = _strip_lab(a), _strip_lab(b)
    if sa and sb and sa.casefold() == sb.casefold():
        return True

    toks_a = {t for t in re.findall(r"\d{2,}", a) if len(t) >= 3 and not (len(t) == 4 and 1990 <= int(t) <= 2100)}
    toks_b = {t for t in re.findall(r"\d{2,}", b) if len(t) >= 3 and not (len(t) == 4 and 1990 <= int(t) <= 2100)}
    if toks_a & toks_b:
        return True
    return False


def _movimento_documento_duplicato_upload(azienda, parsed: EsitoParsingProforma):
    """
    Movimento documento già presente che blocca un secondo upload dello stesso PDF «logico».

    Criterio a regime: **data documento + numero** (stesso numero in anni diversi = righe distinte).
    Il confronto sul numero accetta anche varianti testuali (libro «182» vs PDF «PARCELLA 182»).
    Se dal PDF non si estrae la data, si considera duplicato solo un movimento con **stesso numero**
    e **senza data** in libro (evita doppioni ma non blocca numeri ripetuti tra anni se una riga ha la data).
    Il tipo «sconosciuto» è compatibile con proforma/parcella sullo stesso numero+data.
    """
    from django.db.models import Q

    from .models import MovimentoRegistroStudioConsulente

    num = (parsed.numero_documento or "").strip()
    if len(num) < 2:
        return None
    tipo = (parsed.tipo_documento or "sconosciuto").strip() or "sconosciuto"
    pd = parsed.data_documento
    qs_base = MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda, tipo_riga="documento")
    if tipo == "sconosciuto":
        qs_date = qs_base
    else:
        qs_date = qs_base.filter(Q(tipo_documento=tipo) | Q(tipo_documento="sconosciuto"))
    if pd is not None:
        qs_date = qs_date.filter(data_documento=pd)
    else:
        qs_date = qs_date.filter(data_documento__isnull=True)
    for m in qs_date.order_by("importato_il", "id"):
        nd = (m.numero_documento or "").strip()
        if nd.casefold() == num.casefold() or _numeri_documento_aggancio_coerenti(nd, num):
            return m
    return None


def applica_upload_proforma_parcelle_pdf(azienda, user, uploads) -> list[str]:
    """
    Importa PDF proforma/parcella; restituisce messaggi per ``messages``.

    Anti-duplicazione: stesso **nome file** già importato, oppure stessa coppia **data + numero documento**
    (con regole di compatibilità tra varianti testuali del numero) già presente in libro → file ignorato.
    """
    import os
    import tempfile
    from pathlib import Path

    from django.core.files import File

    from .models import MovimentoRegistroStudioConsulente

    msgs: list[str] = []
    n_ok = n_skip = n_err = 0
    for up in uploads:
        nome = (up.name or "documento.pdf")[:280]
        if nome.lower().startswith("riepilogo"):
            msgs.append(f"Ignorato (prefisso riepilogo): {nome}")
            n_skip += 1
            continue
        if MovimentoRegistroStudioConsulente.objects.filter(
            azienda=azienda, nome_file=nome, tipo_riga="documento"
        ).exists():
            msgs.append(f"Già presente: {nome}")
            n_skip += 1
            continue
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                for chunk in up.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            path = Path(tmp_path)
            testo, metodo = estrai_testo_da_pdf(path)
            parsed = parse_testo_proforma_parcella(testo, nome)
            dup = _movimento_documento_duplicato_upload(azienda, parsed)
            if dup is not None:
                dd = dup.data_documento.isoformat() if dup.data_documento else "—"
                msgs.append(
                    f"Già presente in libro {dup.get_tipo_documento_display()} n. «{parsed.numero_documento.strip()}» "
                    f"del {dd} (file: {dup.nome_file}); ignorato: {nome}"
                )
                n_skip += 1
                continue
            tot = parsed.totale_da_pagare or Decimal("0")
            dare = tot if tot > 0 else Decimal("0")
            note = "; ".join(parsed.avvisi) if parsed.avvisi else ""

            obj = MovimentoRegistroStudioConsulente(
                azienda=azienda,
                tipo_riga="documento",
                tipo_documento=parsed.tipo_documento,
                numero_documento=parsed.numero_documento[:80],
                data_documento=parsed.data_documento,
                totale_da_pagare=parsed.totale_da_pagare,
                dare=dare,
                avere=Decimal("0"),
                nome_file=nome,
                testo_estratto=testo[:50000],
                metodo_estrazione=metodo,
                note=note[:500],
                importato_da=user,
            )
            obj.save()
            if hasattr(up, "seek"):
                up.seek(0)
            obj.file.save(nome, File(up), save=True)
            n_ok += 1
        except Exception as exc:
            msgs.append(f"{nome}: {exc}")
            n_err += 1
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    if n_ok or n_err:
        ricalcola_saldi_progressivi(azienda.id)
    if n_ok:
        msgs.append(f"Importati {n_ok} documenti.")
    if n_err:
        msgs.append(f"{n_err} file con errori.")
    return msgs


def applica_inserimento_manuale_proforma_parcella(
    azienda,
    user,
    *,
    tipo_documento: str,
    numero_documento: str,
    data_documento: date,
    importo_contabile: Decimal,
) -> list[str]:
    """
    Crea una riga ``documento`` (proforma/parcella) **senza PDF**, per documenti non importabili da file.

    ``totale_da_pagare`` e ``dare`` coincidono con l'importo contabile indicato.
    La deduplica è la stessa degli upload PDF (**data + numero**). Per allegare il PDF usare
    «Aggancia PDF a movimenti esistenti».
    """
    from .models import MovimentoRegistroStudioConsulente

    tipo = (tipo_documento or "").strip().lower()
    if tipo not in ("proforma", "parcella"):
        return ["Tipo documento non valido: scegliere Proforma o Parcella."]
    num = (numero_documento or "").strip()[:80]
    if not num or len(num) < 2:
        return ["Indicare un numero documento (almeno 2 caratteri)."]
    if not _is_plausible_numero_documento(num):
        return ["Numero documento non valido o troppo generico."]
    if importo_contabile is None or importo_contabile <= 0:
        return ["L'importo da pagare deve essere maggiore di zero."]
    if data_documento is None:
        return ["Indicare la data documento."]

    parsed = EsitoParsingProforma(
        tipo_documento=tipo,
        numero_documento=num,
        data_documento=data_documento,
        totale_da_pagare=importo_contabile,
        avvisi=[],
    )
    dup = _movimento_documento_duplicato_upload(azienda, parsed)
    if dup is not None:
        dd = dup.data_documento.isoformat() if dup.data_documento else "—"
        return [
            f"Già presente in libro {dup.get_tipo_documento_display()} n. «{num}» del {dd} "
            f"(file: {dup.nome_file}); inserimento annullato."
        ]

    nome_mov = f"manuale/portale/{uuid.uuid4().hex}"[:280]
    if MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda, nome_file=nome_mov).exists():
        nome_mov = f"manuale/portale/{uuid.uuid4().hex}"[:280]

    dare = importo_contabile
    note = "Inserimento manuale da portale; allegare il PDF con «Aggancia PDF a movimenti esistenti»."
    obj = MovimentoRegistroStudioConsulente(
        azienda=azienda,
        tipo_riga="documento",
        tipo_documento=tipo,
        numero_documento=num,
        data_documento=data_documento,
        totale_da_pagare=importo_contabile,
        dare=dare,
        avere=Decimal("0"),
        nome_file=nome_mov,
        testo_estratto="",
        metodo_estrazione="manuale_portale",
        note=note[:500],
        importato_da=user,
    )
    obj.save()
    ricalcola_saldi_progressivi(azienda.id)
    return [
        f"Registrato {obj.get_tipo_documento_display()} n. «{num}» del {data_documento.strftime('%d/%m/%Y')} "
        f"per € {importo_contabile:.2f} (senza PDF). Usa «Aggancia PDF» quando disponibile il file."
    ]


def trova_movimento_documento_senza_pdf_per_aggancio(
    azienda_id: int, parsed: EsitoParsingProforma
) -> tuple[object | None, str | None]:
    """
    Movimento ``documento`` già in libro **senza** ``file``, da agganciare al PDF.

    Criterio: **data documento + numero** (come in anagrafica fiscale: stesso numero in anni diversi
    non coincide). Il numero in libro e quello dal PDF possono differire in forma (es. «182» vs «PARCELLA 182»).
    Richiede la **data** estratta dal PDF; se in libro la riga ha data nulla,
    si accetta **un solo** movimento senza data con quel numero (pregresso Excel incompleto).
    """
    from .models import MovimentoRegistroStudioConsulente

    num = (parsed.numero_documento or "").strip()
    if len(num) < 2:
        return None, "numero_non_estratto"
    if parsed.data_documento is None:
        return None, "data_non_estratto"

    pd = parsed.data_documento

    def _ha_file(m) -> bool:
        return bool(getattr(m.file, "name", None))

    def _numero_abbina(nd: str) -> bool:
        nd = (nd or "").strip()
        if not nd:
            return False
        return nd.casefold() == num.casefold() or _numeri_documento_aggancio_coerenti(nd, num)

    qs_data = (
        MovimentoRegistroStudioConsulente.objects.filter(
            azienda_id=azienda_id,
            tipo_riga="documento",
            data_documento=pd,
        )
        .order_by("importato_il", "id")
    )
    senza_file_data = [c for c in qs_data if not _ha_file(c) and _numero_abbina(c.numero_documento)]
    if len(senza_file_data) == 1:
        return senza_file_data[0], None
    if len(senza_file_data) > 1:
        return None, "ambiguo_numero"

    if any(_ha_file(c) and _numero_abbina(c.numero_documento) for c in qs_data):
        return None, "gia_pdf"

    qs_null = MovimentoRegistroStudioConsulente.objects.filter(
        azienda_id=azienda_id,
        tipo_riga="documento",
        data_documento__isnull=True,
    )
    dig_key = "".join(re.findall(r"\d", num))
    if len(dig_key) >= 3:
        qs_null = qs_null.filter(numero_documento__icontains=dig_key[:24])
    qs_null = qs_null.order_by("importato_il", "id")
    senza_file = [c for c in qs_null if not _ha_file(c) and _numero_abbina(c.numero_documento)]
    if len(senza_file) == 1:
        return senza_file[0], None
    if len(senza_file) > 1:
        return None, "ambiguo_senza_data"

    return None, "nessun_documento"


def applica_aggancia_pdf_proforma_parcelle_a_libro(azienda, user, uploads) -> tuple[list[str], list[dict[str, str]]]:
    """
    Carica PDF già noti (pregresso): per ogni file estrae numero/totale dal testo e **allega il PDF**
    al movimento ``documento`` con **stessa data e stesso numero** documento (estratto dal PDF),
    oppure — se in libro la data manca — **un solo** movimento senza data con quel numero.
    Non crea nuove righe.

    Restituisce ``(messaggi, report_righe)`` dove ogni riga del report è un dict con chiavi fisse
    per export CSV (file, esito, movimento_id, numero_pdf, data_pdf, messaggio).
    """
    import os
    import tempfile
    from pathlib import Path

    from django.core.files import File

    msgs: list[str] = []
    report: list[dict[str, str]] = []
    n_ok = n_skip = n_err = 0
    for up in uploads:
        nome = (up.name or "documento.pdf")[:280]
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                for chunk in up.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            path = Path(tmp_path)
            testo, metodo = estrai_testo_da_pdf(path)
            parsed = parse_testo_proforma_parcella(testo, nome)
            num_pdf = (parsed.numero_documento or "").strip()
            data_pdf = parsed.data_documento.strftime("%Y-%m-%d") if parsed.data_documento else ""
            mov, motivo = trova_movimento_documento_senza_pdf_per_aggancio(azienda.id, parsed)
            if mov is None:
                if motivo == "numero_non_estratto":
                    line = f"{nome}: numero documento non estratto dal PDF; saltato."
                    msgs.append(line)
                    report.append(
                        {
                            "file": nome,
                            "esito": "saltato",
                            "movimento_id": "",
                            "numero_pdf": num_pdf,
                            "data_pdf": data_pdf,
                            "messaggio": line,
                        }
                    )
                elif motivo == "data_non_estratto":
                    line = (
                        f"{nome}: data documento non estratta dal PDF; per agganciare serve data+numero. "
                        f"Saltato (o correggere il PDF / «Rileggi totali» dopo aver messo testo)."
                    )
                    msgs.append(line)
                    report.append(
                        {
                            "file": nome,
                            "esito": "saltato",
                            "movimento_id": "",
                            "numero_pdf": num_pdf,
                            "data_pdf": data_pdf,
                            "messaggio": line,
                        }
                    )
                elif motivo == "ambiguo_numero":
                    line = (
                        f"{nome}: più movimenti in libro alla stessa data con numero compatibile con "
                        f"«{parsed.numero_documento.strip()}»; impostare numeri univoci o unificare le righe. Saltato."
                    )
                elif motivo == "ambiguo_senza_data":
                    line = (
                        f"{nome}: più movimenti senza data con numero «{parsed.numero_documento.strip()}»; "
                        f"impostare la data in Admin o unificare le righe. Saltato."
                    )
                    msgs.append(line)
                    report.append(
                        {
                            "file": nome,
                            "esito": "saltato",
                            "movimento_id": "",
                            "numero_pdf": num_pdf,
                            "data_pdf": data_pdf,
                            "messaggio": line,
                        }
                    )
                elif motivo == "gia_pdf":
                    line = (
                        f"{nome}: movimento con numero «{parsed.numero_documento.strip()}» "
                        f"e data {parsed.data_documento:%d/%m/%Y} ha già PDF; saltato."
                    )
                    msgs.append(line)
                    report.append(
                        {
                            "file": nome,
                            "esito": "saltato",
                            "movimento_id": "",
                            "numero_pdf": num_pdf,
                            "data_pdf": data_pdf,
                            "messaggio": line,
                        }
                    )
                else:
                    line = (
                        f"{nome}: nessun movimento documento senza PDF con numero «{parsed.numero_documento.strip()}» "
                        f"e data {parsed.data_documento:%d/%m/%Y}; saltato."
                    )
                    msgs.append(line)
                    report.append(
                        {
                            "file": nome,
                            "esito": "saltato",
                            "movimento_id": "",
                            "numero_pdf": num_pdf,
                            "data_pdf": data_pdf,
                            "messaggio": line,
                        }
                    )
                n_skip += 1
                continue
            mov.testo_estratto = testo[:50000]
            mov.metodo_estrazione = metodo
            if parsed.totale_da_pagare is not None:
                tot = parsed.totale_da_pagare
                if mov.totale_da_pagare is None:
                    mov.totale_da_pagare = tot
                if mov.dare == 0 or mov.dare is None:
                    mov.dare = tot if tot and tot > 0 else Decimal("0")
            if parsed.data_documento and not mov.data_documento:
                mov.data_documento = parsed.data_documento
            if parsed.tipo_documento and parsed.tipo_documento != "sconosciuto":
                if mov.tipo_documento == "sconosciuto" or not mov.tipo_documento:
                    mov.tipo_documento = parsed.tipo_documento
            extra = f"PDF agganciato da file «{nome}»."
            mov.note = (mov.note + " " + extra).strip()[:500] if mov.note else extra[:500]
            if user is not None:
                mov.importato_da = user
            mov.save()
            if hasattr(up, "seek"):
                up.seek(0)
            mov.file.save(nome[:200], File(up), save=True)
            n_ok += 1
            line = (
                f"{nome}: allegato a movimento id={mov.pk} (n. «{mov.numero_documento}», {mov.get_tipo_documento_display()})."
            )
            msgs.append(line)
            report.append(
                {
                    "file": nome,
                    "esito": "ok",
                    "movimento_id": str(mov.pk),
                    "numero_pdf": (mov.numero_documento or num_pdf or "").strip(),
                    "data_pdf": mov.data_documento.strftime("%Y-%m-%d") if mov.data_documento else data_pdf,
                    "messaggio": line,
                }
            )
        except Exception as exc:
            line = f"{nome}: {exc}"
            msgs.append(line)
            n_err += 1
            report.append(
                {
                    "file": nome,
                    "esito": "errore",
                    "movimento_id": "",
                    "numero_pdf": "",
                    "data_pdf": "",
                    "messaggio": line,
                }
            )
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    if n_ok or n_err:
        ricalcola_saldi_progressivi(azienda.id)
    if n_ok:
        msgs.append(f"Agganciati {n_ok} PDF a movimenti documento già in libro.")
    if n_err:
        msgs.append(f"{n_err} file con errori.")
    return msgs, report


def applica_pdf_su_movimento_documento(
    azienda,
    user,
    movimento_id: int,
    upload,
) -> list[str]:
    """
    Allega un singolo PDF al movimento ``documento`` indicato (es. riga inserita a mano senza file).

    Il contenuto del PDF deve essere **coerente con la riga scelta** (numero documento e, se
    presenti in entrambi, la data), senza risolvere ambiguità su altre righe del libro.
    """
    import os
    import tempfile
    from pathlib import Path

    from django.core.files import File

    from .models import MovimentoRegistroStudioConsulente

    mov = MovimentoRegistroStudioConsulente.objects.filter(
        pk=movimento_id, azienda=azienda, tipo_riga="documento"
    ).first()
    if mov is None:
        return ["Movimento non trovato o non è un documento proforma/parcella."]

    nome = (upload.name or "documento.pdf")[:280]
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            for chunk in upload.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        path = Path(tmp_path)
        testo, metodo = estrai_testo_da_pdf(path)
        parsed = parse_testo_proforma_parcella(testo, nome)
        okc, why = _parsed_proforma_compatibile_con_movimento_documento(mov, parsed)
        if not okc:
            if why == "numero":
                libro_num = (mov.numero_documento or "—").strip()
                pdf_num = (parsed.numero_documento or "—").strip()
                return [
                    f"{nome}: numero documento nel PDF («{pdf_num}») non corrisponde alla riga («{libro_num}»); "
                    "allegato annullato."
                ]
            if why == "data":
                pdf_data = parsed.data_documento.strftime("%d/%m/%Y") if parsed.data_documento else "—"
                libro_data = mov.data_documento.strftime("%d/%m/%Y") if mov.data_documento else "—"
                return [
                    f"{nome}: data nel PDF ({pdf_data}) non coincide con la data della riga ({libro_data}); "
                    "allegato annullato."
                ]
            return [f"{nome}: dal PDF non emergono dati sufficienti per abbinare a questa riga; allegato annullato."]

        replacing_pdf = bool(getattr(mov.file, "name", None))
        mov.testo_estratto = testo[:50000]
        mov.metodo_estrazione = metodo
        if parsed.totale_da_pagare is not None:
            tot = parsed.totale_da_pagare
            if mov.totale_da_pagare is None:
                mov.totale_da_pagare = tot
            if mov.dare == 0 or mov.dare is None:
                mov.dare = tot if tot and tot > 0 else Decimal("0")
        if parsed.data_documento and not mov.data_documento:
            mov.data_documento = parsed.data_documento
        if parsed.tipo_documento and parsed.tipo_documento != "sconosciuto":
            if mov.tipo_documento == "sconosciuto" or not mov.tipo_documento:
                mov.tipo_documento = parsed.tipo_documento
        extra = f"PDF allegato da riga singola «{nome}»."
        mov.note = (mov.note + " " + extra).strip()[:500] if mov.note else extra[:500]
        if user is not None:
            mov.importato_da = user
        mov.save()
        if hasattr(upload, "seek"):
            upload.seek(0)
        if replacing_pdf:
            try:
                mov.file.delete(save=False)
            except OSError:
                pass
        mov.file.save(nome[:200], File(upload), save=True)
        ricalcola_saldi_progressivi(azienda.id)
        if replacing_pdf:
            return [
                f"PDF documento sostituito per n. «{mov.numero_documento or '—'}» — id {mov.pk} "
                f"({mov.get_tipo_documento_display()}). Apri il file dalla colonna documento o dal Libro movimenti "
                f"(file: {nome})."
            ]
        return [
            f"PDF allegato con successo al documento n. «{mov.numero_documento or '—'}» — id {mov.pk} "
            f"({mov.get_tipo_documento_display()}). Apri il file dalla colonna documento o dal Libro movimenti "
            f"(file: {nome})."
        ]
    except Exception as exc:
        return [f"{nome}: {exc}"]
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


def trova_bonifico_senza_pdf_per_aggancio(
    azienda_id: int, parsed: EsitoParsingBonifico
) -> tuple[object | None, str | None]:
    """Movimento ``bonifico`` senza ``file``, compatibile con dati estratti dalla distinta PDF."""

    def _nofile(m):
        return m is not None and not getattr(m.file, "name", None)

    if not parsed.importo or parsed.importo <= 0:
        return None, "importo_non_estratto"

    if parsed.riferimento and len(parsed.riferimento.strip()) >= 6:
        m = trova_movimento_bonifico_per_riferimento(azienda_id, parsed.riferimento)
        if _nofile(m):
            return m, None
        if m is not None:
            return None, "gia_pdf"

    m = trova_bonifico_esistente_stesso_excel(
        azienda_id,
        parsed.data_documento,
        parsed.importo,
        parsed.riferimento or "",
        parsed.causale or "",
    )
    if _nofile(m):
        return m, None
    if m is not None:
        return None, "gia_pdf"

    from .models import MovimentoRegistroStudioConsulente

    qs = (
        MovimentoRegistroStudioConsulente.objects.filter(
            azienda_id=azienda_id,
            tipo_riga="bonifico",
            avere=parsed.importo,
        )
        .order_by("-importato_il")
    )
    pd = parsed.data_documento
    for row in qs[:120]:
        if getattr(row.file, "name", None):
            continue
        rd = row.data_documento
        if pd is None or rd is None or rd == pd:
            return row, None
    return None, "nessun_bonifico"


def applica_aggancia_pdf_bonifici_a_libro(azienda, user, uploads) -> tuple[list[str], list[dict[str, str]]]:
    """
    Allega distinte PDF a bonifici **già** in libro (es. importati da Excel senza file).
    Cerca per CRO/TRN, poi stessa logica «stesso Excel» (data + avere + riferimento/causale), poi stesso importo senza PDF.

    Restituisce ``(messaggi, report_righe)`` con dict: file, esito, movimento_id, riferimento_pdf, importo_pdf, data_pdf, messaggio.
    """
    import os
    import tempfile
    from pathlib import Path

    from django.core.files import File

    msgs: list[str] = []
    report: list[dict[str, str]] = []
    n_ok = n_skip = n_err = 0
    for up in uploads:
        nome = (up.name or "bonifico.pdf")[:280]
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                for chunk in up.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            path = Path(tmp_path)
            testo, metodo = estrai_testo_da_pdf(path)
            parsed = parse_testo_bonifico_pdf(testo, nome)
            rif_pdf = (parsed.riferimento or "").strip()[:160]
            imp_pdf = str(parsed.importo) if parsed.importo else ""
            data_pdf = parsed.data_documento.strftime("%Y-%m-%d") if parsed.data_documento else ""
            mov, motivo = trova_bonifico_senza_pdf_per_aggancio(azienda.id, parsed)
            if mov is None:
                if motivo == "importo_non_estratto":
                    line = f"{nome}: importo non estratto dal PDF; saltato."
                    msgs.append(line)
                    report.append(
                        {
                            "file": nome,
                            "esito": "saltato",
                            "movimento_id": "",
                            "riferimento_pdf": rif_pdf,
                            "importo_pdf": imp_pdf,
                            "data_pdf": data_pdf,
                            "messaggio": line,
                        }
                    )
                elif motivo == "gia_pdf":
                    line = f"{nome}: bonifico compatibile ha già un PDF; saltato."
                    msgs.append(line)
                    report.append(
                        {
                            "file": nome,
                            "esito": "saltato",
                            "movimento_id": "",
                            "riferimento_pdf": rif_pdf,
                            "importo_pdf": imp_pdf,
                            "data_pdf": data_pdf,
                            "messaggio": line,
                        }
                    )
                else:
                    line = f"{nome}: nessun bonifico in libro senza PDF compatibile con distinta; saltato."
                    msgs.append(line)
                    report.append(
                        {
                            "file": nome,
                            "esito": "saltato",
                            "movimento_id": "",
                            "riferimento_pdf": rif_pdf,
                            "importo_pdf": imp_pdf,
                            "data_pdf": data_pdf,
                            "messaggio": line,
                        }
                    )
                n_skip += 1
                continue
            mov.testo_estratto = testo[:50000]
            mov.metodo_estrazione = metodo
            if parsed.riferimento and len(parsed.riferimento.strip()) >= 3:
                mov.riferimento_pagamento = parsed.riferimento[:160]
            if parsed.causale:
                mov.causale_pagamento = parsed.causale[:220]
            if parsed.data_documento and not mov.data_documento:
                mov.data_documento = parsed.data_documento
            extra = f"PDF distinta agganciato da «{nome}»."
            mov.note = (mov.note + " " + extra).strip()[:500] if mov.note else extra[:500]
            if user is not None:
                mov.importato_da = user
            mov.save()
            if hasattr(up, "seek"):
                up.seek(0)
            mov.file.save(nome[:200], File(up), save=True)
            n_ok += 1
            line = f"{nome}: PDF allegato a bonifico id={mov.pk} (rif. {mov.riferimento_pagamento or '—'})."
            msgs.append(line)
            report.append(
                {
                    "file": nome,
                    "esito": "ok",
                    "movimento_id": str(mov.pk),
                    "riferimento_pdf": (mov.riferimento_pagamento or rif_pdf or "").strip()[:160],
                    "importo_pdf": str(mov.avere) if mov.avere is not None else imp_pdf,
                    "data_pdf": mov.data_documento.strftime("%Y-%m-%d") if mov.data_documento else data_pdf,
                    "messaggio": line,
                }
            )
        except Exception as exc:
            line = f"{nome}: {exc}"
            msgs.append(line)
            n_err += 1
            report.append(
                {
                    "file": nome,
                    "esito": "errore",
                    "movimento_id": "",
                    "riferimento_pdf": "",
                    "importo_pdf": "",
                    "data_pdf": "",
                    "messaggio": line,
                }
            )
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    if n_ok or n_err:
        ricalcola_saldi_progressivi(azienda.id)
    if n_ok:
        msgs.append(f"Agganciati {n_ok} PDF a bonifici già in libro.")
    if n_err:
        msgs.append(f"{n_err} file con errori.")
    return msgs, report


def _riferimento_bonifico_pdf_compatibile_con_libro(rif_pdf: str, rif_libro: str) -> bool:
    """True se il riferimento estratto dal PDF «aggancia» al riferimento già in libro (stessa logica CRO/TRN)."""
    a = re.sub(r"\s+", "", (rif_pdf or "")).upper()
    b = re.sub(r"\s+", "", (rif_libro or "")).upper()
    if not a or not b:
        return False
    if len(a) >= 6:
        return a in b or b in a or (len(a) >= 12 and a[-12:] in b) or (len(b) >= 12 and b[-12:] in a)
    return a in b or b in a


def _pipe_data_e_importo_da_riferimento_libro(lr: str) -> tuple[date | None, Decimal | None]:
    """
    Da riferimenti tipo pregresso Excel «…|2021-07-09|130.00» estrae data e importo.
    Serve quando il PDF non espone l'importo ma la riga in elenco è già coerente.
    """
    lr_clean = (lr or "").strip().replace(" ", "")
    m = re.search(r"\|(\d{4})-(\d{2})-(\d{2})\|([\d.,]+)\s*$", lr_clean)
    if not m:
        return None, None
    try:
        dt = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        imp = _parse_it_decimal(m.group(4))
        if imp is None or imp <= 0:
            return dt, None
        return dt, imp
    except (ValueError, InvalidOperation):
        return None, None


def _parsed_bonifico_compatibile_con_movimento(mov, parsed: EsitoParsingBonifico) -> tuple[bool, str, str | None]:
    """
    True se la distinta estratta è coerente con il movimento **scelto dall'utente** (riga + PDF).

    Regole principali:
    - Se l'importo estratto dal PDF coincide con l'avere della riga (e le date note non confliggono),
      si accetta: il PDF spesso contiene codici numerici lunghi che non sono sottostringa del
      riferimento sintetico in libro («BONIFICO …|data|importo»).
    - Se il PDF non ha importo ma il riferimento in libro ha suffisso ``|YYYY-MM-DD|importo``,
      si usa quell'importo come confronto con l'avere.
    - Date «in libro»: ``data_documento`` della riga **e**, se presente, la data nel suffisso Excel
      ``|YYYY-MM-DD|importo`` nel riferimento (spesso più vicina alla data effettiva del bonifico).
    - Se la data estratta dal PDF non coincide con nessuna di queste ma **l'importo coincide**
      con l'avere, si accetta comunque: data contabile in elenco / prima data nel testo PDF possono
      divergere dalla data operazione sulla distinta.

    Ritorna ``(ok, codice_errore, hint_ok)`` dove ``hint_ok`` è solo se ok è True, es. ``data_pdf_diversa``
    per segnalare in messaggio che la data nel PDF differisce da quelle in elenco.
    """
    lr = (mov.riferimento_pagamento or "").strip()
    pipe_d, pipe_imp = _pipe_data_e_importo_da_riferimento_libro(lr)
    effective_imp = parsed.importo if parsed.importo is not None else pipe_imp

    if effective_imp is not None and mov.avere is not None and effective_imp != mov.avere:
        return False, "importo", None

    libro_dates: list[date] = []
    if mov.data_documento:
        libro_dates.append(mov.data_documento)
    if pipe_d and pipe_d not in libro_dates:
        libro_dates.append(pipe_d)

    success_hint: str | None = None
    pd = parsed.data_documento
    if pd is not None and libro_dates:
        if pd not in libro_dates:
            if effective_imp is not None and mov.avere is not None and effective_imp == mov.avere:
                success_hint = "data_pdf_diversa"
            else:
                return False, "data", None

    if effective_imp is not None and mov.avere is not None and effective_imp == mov.avere:
        return True, "", success_hint

    pr = (parsed.riferimento or "").strip()
    if pr and lr:
        if not _riferimento_bonifico_pdf_compatibile_con_libro(pr, lr):
            return False, "riferimento", None
    elif pr and not lr:
        return False, "riferimento", None
    elif not pr and parsed.importo is None:
        return False, "dati", None
    return True, "", success_hint


def _testo_pdf_contiene_importo_coerente_con_avere(testo: str, avere: Decimal) -> bool:
    """Cerca l’importo in avere (centesimi) nel testo grezzo della distinta (parser fallito)."""
    if not testo or avere is None or avere <= 0:
        return False
    av = avere.quantize(Decimal("0.01"))
    cents = int(av * 100)
    if cents % 100 == 0:
        whole = cents // 100
        pats = [
            rf"\b{whole}\s*[,.]\s*00\b",
            rf"\b{whole}\b(?!\d)",
        ]
    else:
        whole, frac = divmod(cents, 100)
        pats = [
            rf"\b{whole}\s*[,.]\s*{frac:02d}\b",
            rf"\b{whole}\s*[,.]\s*{frac}\b",
        ]
    for p in pats:
        if re.search(p, testo, re.I):
            return True
    return False


def _riferimento_libro_ha_codice_lungo_in_testo_pdf(mov, testo: str) -> bool:
    """CRO/TRN lunghi presenti sia nel riferimento in libro che nel testo PDF (senza parsing strutturato)."""
    lr = (mov.riferimento_pagamento or "").strip()
    if not lr or not testo:
        return False
    tc = re.sub(r"\s+", "", testo)
    lr_c = re.sub(r"\s+", "", lr).upper()
    for m in re.finditer(r"\d{9,}", lr):
        s = m.group(0)
        if s in tc or s in lr_c:
            return True
    return False


def _distinta_pdf_coerente_bonifico_per_allegato_manuale(testo: str, mov) -> bool:
    """
    Per allegato singolo su riga già scelta: accetta se importo o codice lungo del riferimento in libro
    compare nel testo estratto dalla distinta (evita falsi negativi quando ``parse_testo_bonifico_pdf`` non riempie i campi).
    """
    if not (testo or "").strip():
        return False
    if mov.avere and _testo_pdf_contiene_importo_coerente_con_avere(testo, mov.avere):
        return True
    if _riferimento_libro_ha_codice_lungo_in_testo_pdf(mov, testo):
        return True
    return False


def _parsed_proforma_compatibile_con_movimento_documento(mov, parsed: EsitoParsingProforma) -> tuple[bool, str]:
    """Coerenza numero (+ data se presente in entrambi) tra PDF e riga documento scelta."""
    num_m = (mov.numero_documento or "").strip()
    num_p = (parsed.numero_documento or "").strip()
    if len(num_p) < 2:
        return False, "numero"
    if not (
        num_m.casefold() == num_p.casefold()
        or _numeri_documento_aggancio_coerenti(num_m, num_p)
    ):
        return False, "numero"
    if mov.data_documento and parsed.data_documento and mov.data_documento != parsed.data_documento:
        return False, "data"
    return True, ""


def applica_pdf_su_movimento_bonifico(
    azienda,
    user,
    movimento_id: int,
    upload,
) -> list[str]:
    """
    Allega una distinta PDF al movimento ``bonifico`` indicato (es. import Excel senza file).

    Il PDF deve essere **coerente con la riga scelta** (data, avere, riferimento come in libro),
    senza cercare un altro bonifico «compatibile» sul libro intero.
    """
    import os
    import tempfile
    from pathlib import Path

    from django.core.files import File

    from .models import MovimentoRegistroStudioConsulente

    mov = MovimentoRegistroStudioConsulente.objects.filter(
        pk=movimento_id, azienda=azienda, tipo_riga="bonifico"
    ).first()
    if mov is None:
        return ["Movimento non trovato o non è un bonifico."]
    replacing_pdf = bool(getattr(mov.file, "name", None))

    nome = (upload.name or "bonifico.pdf")[:280]
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            for chunk in upload.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        path = Path(tmp_path)
        testo, metodo = estrai_testo_da_pdf(path)
        parsed = parse_testo_bonifico_pdf(testo, nome)
        okc, why, hint_ok = _parsed_bonifico_compatibile_con_movimento(mov, parsed)
        if not okc and why in ("dati", "riferimento"):
            if _distinta_pdf_coerente_bonifico_per_allegato_manuale(testo, mov):
                okc, why, hint_ok = True, "", "distinta_testo_coerente"
        if not okc:
            if why == "importo":
                return [
                    f"{nome}: importo estratto dal PDF (€ {parsed.importo}) diverso dall'avere della riga (€ {mov.avere}); "
                    "allegato annullato."
                ]
            if why == "data":
                dp = parsed.data_documento.strftime("%d/%m/%Y") if parsed.data_documento else "—"
                dm = mov.data_documento.strftime("%d/%m/%Y") if mov.data_documento else "—"
                return [
                    f"{nome}: data nel PDF ({dp}) non coincide con la data del bonifico in elenco ({dm}); allegato annullato."
                ]
            if why == "riferimento":
                return [
                    f"{nome}: riferimento estratto dal PDF non riconducibile a «{(mov.riferimento_pagamento or '')[:100]}»; "
                    "allegato annullato. Controllare la distinta o usare l'aggancio massivo."
                ]
            return [f"{nome}: dal PDF non emergono dati sufficienti per abbinare a questa riga; allegato annullato."]

        mov.testo_estratto = testo[:50000]
        mov.metodo_estrazione = metodo
        if not (mov.riferimento_pagamento or "").strip():
            if parsed.riferimento and len(parsed.riferimento.strip()) >= 3:
                mov.riferimento_pagamento = parsed.riferimento[:160]
        if not (mov.causale_pagamento or "").strip() and parsed.causale:
            mov.causale_pagamento = parsed.causale[:220]
        if parsed.data_documento and not mov.data_documento:
            mov.data_documento = parsed.data_documento
        extra = f"PDF distinta allegato da riga singola «{nome}»."
        mov.note = (mov.note + " " + extra).strip()[:500] if mov.note else extra[:500]
        if user is not None:
            mov.importato_da = user
        mov.save()
        if hasattr(upload, "seek"):
            upload.seek(0)
        if replacing_pdf:
            try:
                mov.file.delete(save=False)
            except OSError:
                pass
        mov.file.save(nome[:200], File(upload), save=True)
        ricalcola_saldi_progressivi(azienda.id)
        data_m = mov.data_documento.strftime("%d/%m/%Y") if mov.data_documento else "—"
        extra_data = ""
        if hint_ok == "data_pdf_diversa" and parsed.data_documento:
            extra_data = (
                f" Nota: data operazione sulla distinta ({parsed.data_documento.strftime('%d/%m/%Y')}) "
                f"non coincide con la/e data/e in elenco; verificare la riga se serve."
            )
        if replacing_pdf:
            return [
                f"Distinta PDF sostituita per il bonifico del {data_m} (€ {mov.avere}) — id {mov.pk}. "
                f"Apri il file dalla colonna «PDF bonifico» o dal Libro movimenti (file: {nome})."
                f"{extra_data}"
            ]
        return [
            f"Distinta PDF allegata con successo al bonifico del {data_m} (€ {mov.avere}) — id {mov.pk}. "
            f"Apri il file dalla colonna «PDF bonifico» o dal Libro movimenti (file: {nome})."
            f"{extra_data}"
        ]
    except Exception as exc:
        return [f"{nome}: {exc}"]
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


def render_csv_report_aggancia_documenti(rows: list[dict[str, str]]) -> str:
    """CSV con separatore ``;`` (compatibile Excel IT)."""
    import csv
    import io

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(["file", "esito", "movimento_id", "numero_pdf", "data_pdf", "messaggio"])
    for r in rows:
        w.writerow(
            [
                r.get("file", ""),
                r.get("esito", ""),
                r.get("movimento_id", ""),
                r.get("numero_pdf", ""),
                r.get("data_pdf", ""),
                r.get("messaggio", ""),
            ]
        )
    return buf.getvalue()


def render_csv_report_aggancia_bonifici(rows: list[dict[str, str]]) -> str:
    import csv
    import io

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(["file", "esito", "movimento_id", "riferimento_pdf", "importo_pdf", "data_pdf", "messaggio"])
    for r in rows:
        w.writerow(
            [
                r.get("file", ""),
                r.get("esito", ""),
                r.get("movimento_id", ""),
                r.get("riferimento_pdf", ""),
                r.get("importo_pdf", ""),
                r.get("data_pdf", ""),
                r.get("messaggio", ""),
            ]
        )
    return buf.getvalue()


def render_csv_report_import_proforma_cartella(rows: list[dict[str, str]]) -> str:
    """Riepilogo import PDF proforma/parcelle da cartella (management command); separatore ``;``."""
    import csv
    import io

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(["file", "esito", "movimento_id", "numero_pdf", "data_pdf", "messaggio"])
    for r in rows:
        w.writerow(
            [
                r.get("file", ""),
                r.get("esito", ""),
                r.get("movimento_id", ""),
                r.get("numero_pdf", ""),
                r.get("data_pdf", ""),
                r.get("messaggio", ""),
            ]
        )
    return buf.getvalue()


def applica_upload_bonifici_pdf(azienda, user, uploads) -> list[str]:
    """
    Importa distinte bonifico PDF in avere (con allegato); parsing euristico.

    Anti-duplicazione: se esiste già un bonifico con la stessa **data**, lo stesso **avere** e un
    **riferimento/causale** compatibile (come per l'import Excel), oppure lo stesso CRO/TRN con stesso importo,
    il PDF non crea una nuova riga.
    """
    import os
    import tempfile
    from pathlib import Path

    from django.core.files import File

    from .models import MovimentoRegistroStudioConsulente

    msgs: list[str] = []
    n_ok = n_err = n_skip = 0
    for up in uploads:
        nome = (up.name or "bonifico.pdf")[:280]
        nome_base = (Path(nome).name or "bonifico.pdf")[:200]
        nome_sint = f"pdf-distinta/{uuid.uuid4().hex[:12]}_{nome_base}"[:280]
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                for chunk in up.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            path = Path(tmp_path)
            testo, metodo = estrai_testo_da_pdf(path)
            parsed = parse_testo_bonifico_pdf(testo, nome)
            if not parsed.importo or parsed.importo <= 0:
                msgs.append(f"{nome}: importo non rilevato; usare la sezione Pagamenti con inserimento manuale.")
                n_err += 1
                continue
            if not parsed.riferimento or len(parsed.riferimento.strip()) < 3:
                msgs.append(f"{nome}: riferimento assente; usare inserimento manuale.")
                n_err += 1
                continue
            dup_m = _movimento_bonifico_duplicato_da_parsed_pdf(azienda.id, parsed)
            if dup_m is not None:
                msgs.append(
                    f"{nome}: bonifico già presente in libro (mov. id={dup_m.pk}, "
                    f"rif. «{(dup_m.riferimento_pagamento or '')[:72]}»); importazione ignorata."
                )
                n_skip += 1
                continue
            note = "; ".join(parsed.avvisi) if parsed.avvisi else ""
            obj = MovimentoRegistroStudioConsulente(
                azienda=azienda,
                tipo_riga="bonifico",
                tipo_documento="sconosciuto",
                numero_documento=parsed.riferimento[:80],
                data_documento=parsed.data_documento,
                totale_da_pagare=None,
                dare=Decimal("0"),
                avere=parsed.importo,
                nome_file=nome_sint[:280],
                testo_estratto=testo[:50000],
                metodo_estrazione=metodo,
                note=note[:500],
                riferimento_pagamento=parsed.riferimento[:160],
                causale_pagamento=parsed.causale[:220],
                importato_da=user,
            )
            obj.save()
            if hasattr(up, "seek"):
                up.seek(0)
            obj.file.save(nome, File(up), save=True)
            n_ok += 1
        except Exception as exc:
            msgs.append(f"{nome}: {exc}")
            n_err += 1
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    if n_ok or n_err:
        ricalcola_saldi_progressivi(azienda.id)
    if n_ok:
        msgs.append(f"Registrati {n_ok} bonifici da PDF.")
    if n_skip:
        msgs.append(f"{n_skip} PDF ignorati perché equivalenti a bonifici già in libro.")
    if n_err:
        msgs.append(f"{n_err} PDF bonifici non importati automaticamente.")
    return msgs


def ricalcola_totali_documenti_da_testo_estratto(azienda_id: int) -> dict[str, int | str]:
    """
    Rilegge proforma/parcelle già importate: ``parse_testo_proforma_parcella`` su ``testo_estratto``,
    aggiorna totale/dare e campi anagrafici documento, poi ricalcola i saldi progressivi.

    Non modifica righe ``bonifico`` / ``rettifica``. Righe senza testo salvato vengono saltate.
    """
    from django.db import transaction
    from django.db.models import F

    from .models import MovimentoRegistroStudioConsulente

    n_senza_testo = n_invariati = n_aggiornati = 0
    bulk: list[MovimentoRegistroStudioConsulente] = []

    qs = (
        MovimentoRegistroStudioConsulente.objects.filter(azienda_id=azienda_id, tipo_riga="documento")
        .order_by(F("data_documento").asc(nulls_last=True), "importato_il", "id")
    )
    for row in qs.iterator(chunk_size=100):
        testo = (row.testo_estratto or "").strip()
        if not testo:
            n_senza_testo += 1
            continue
        parsed = parse_testo_proforma_parcella(testo, row.nome_file or "documento.pdf")
        tot = parsed.totale_da_pagare or Decimal("0")
        nuovo_dare = tot if tot > 0 else Decimal("0")
        note = ("; ".join(parsed.avvisi))[:500] if parsed.avvisi else ""

        cambia = (
            row.totale_da_pagare != parsed.totale_da_pagare
            or row.dare != nuovo_dare
            or row.tipo_documento != parsed.tipo_documento
            or (row.numero_documento or "") != (parsed.numero_documento or "")[:80]
            or row.data_documento != parsed.data_documento
            or (row.note or "") != note
        )
        if not cambia:
            n_invariati += 1
            continue

        row.totale_da_pagare = parsed.totale_da_pagare
        row.dare = nuovo_dare
        row.tipo_documento = parsed.tipo_documento
        row.numero_documento = (parsed.numero_documento or "")[:80]
        row.data_documento = parsed.data_documento
        row.note = note
        bulk.append(row)
        n_aggiornati += 1

    if bulk:
        with transaction.atomic():
            MovimentoRegistroStudioConsulente.objects.bulk_update(
                bulk,
                [
                    "totale_da_pagare",
                    "dare",
                    "tipo_documento",
                    "numero_documento",
                    "data_documento",
                    "note",
                ],
                batch_size=100,
            )
        ricalcola_saldi_progressivi(azienda_id)

    msg = (
        f"Aggiornate {n_aggiornati} righe documento; {n_invariati} già allineate; "
        f"{n_senza_testo} senza testo estratto (reimportare il PDF)."
    )
    return {
        "n_aggiornati": n_aggiornati,
        "n_invariati": n_invariati,
        "n_senza_testo": n_senza_testo,
        "message": msg,
    }


def ricalcola_saldi_progressivi(azienda_id: int) -> None:
    """Ricalcola ``saldo_progressivo`` per tutte le righe dell'azienda (ordine: data doc, import, id)."""
    from django.db.models import F

    from .models import MovimentoRegistroStudioConsulente

    qs = (
        MovimentoRegistroStudioConsulente.objects.filter(azienda_id=azienda_id)
        .order_by(F("data_documento").asc(nulls_last=True), "importato_il", "id")
        .only("id", "dare", "avere", "saldo_progressivo")
    )
    saldo = Decimal("0")
    bulk: list[MovimentoRegistroStudioConsulente] = []
    for row in qs:
        saldo = saldo + row.dare - row.avere
        if row.saldo_progressivo != saldo:
            row.saldo_progressivo = saldo
            bulk.append(row)
    if bulk:
        MovimentoRegistroStudioConsulente.objects.bulk_update(bulk, ["saldo_progressivo"], batch_size=200)


# ── Parsing bonifico da PDF (euristica) ───────────────────────────────────────


@dataclass
class EsitoParsingBonifico:
    riferimento: str
    data_documento: date | None
    importo: Decimal | None
    causale: str
    avvisi: list[str] = field(default_factory=list)


def parse_testo_bonifico_pdf(testo: str, nome_file: str) -> EsitoParsingBonifico:
    """Estrae riferimento, data e importo da testo estratto da distinta bonifico (best-effort)."""
    avvisi: list[str] = []
    t = testo or ""
    rif = ""
    for pat in (
        r"(?:CRO|Riferimento\s+CRO|Codice\s+CRO)\s*[:\s]+([A-Za-z0-9/\-\s]{6,45})",
        r"(?:TRN|End\s+To\s+End\s+Id|Riferimento\s+operazione)\s*[:\s]+([A-Za-z0-9/\-\s]{6,45})",
        r"(?:Ordinativo|Numero\s+disposizione)\s*[:\s]+([A-Za-z0-9/\-\s]{6,40})",
    ):
        m = re.search(pat, t, re.I)
        if m:
            rif = re.sub(r"\s+", " ", m.group(1).strip())[:160]
            break
    if not rif:
        digits = re.findall(r"\b\d{10,30}\b", t)
        if digits:
            rif = max(digits, key=len)[:160]
    if not rif and nome_file:
        stem = Path(nome_file).stem
        if len(stem) >= 6:
            rif = stem[:160]
            avvisi.append("Riferimento non trovato nel PDF: usato nome file.")

    data_doc = _first_date_in_text(t[:8000])
    imp: Decimal | None = None
    for pat in (
        r"(?:importo\s+ordinato|importo\s+bonifico|importo\s+accreditato)\s*[:\s€]*([\d\.\s,]+(?:,\d{1,2})?)",
        r"(?:EUR|€)\s*([\d\.\s,]+(?:,\d{1,2})?)\s*(?:accredit|bonific)",
        r"(?:accreditato|versato)\s+([\d\.\s,]+(?:,\d{1,2})?)\s*€?",
    ):
        m = re.search(pat, t, re.I)
        if m:
            imp = _parse_it_decimal(m.group(1))
            if imp and imp > 0:
                break
    if imp is None or imp <= 0:
        avvisi.append("Importo non rilevato automaticamente dal PDF.")
        imp = None

    caus = ""
    m_c = re.search(r"(?:causale|oggetto|descrizione)\s*[:\s]+(.{0,200}?)(?:\n|$)", t, re.I | re.MULTILINE)
    if m_c:
        caus = m_c.group(1).strip()[:220]

    if not rif:
        avvisi.append("Riferimento non rilevato: compilare a mano o rinominare il file.")

    return EsitoParsingBonifico(
        riferimento=rif,
        data_documento=data_doc,
        importo=imp,
        causale=caus,
        avvisi=avvisi,
    )


def _norm_match(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").lower().strip())


def trova_movimento_documento_per_descrizione(azienda_id: int, descrizione: str):
    """Aggancia una riga estratto a un movimento proforma/parcella già importato (PDF)."""
    from .models import MovimentoRegistroStudioConsulente

    desc_n = _norm_match(descrizione)
    if len(desc_n) < 4:
        return None
    qs = MovimentoRegistroStudioConsulente.objects.filter(azienda_id=azienda_id, tipo_riga="documento").order_by(
        "-importato_il"
    )
    for m in qs[:400]:
        candidates = [
            _norm_match(m.numero_documento),
            _norm_match(Path(m.nome_file).stem if m.nome_file else ""),
            _norm_match((m.note or "")[:240]),
        ]
        if m.testo_estratto:
            candidates.append(_norm_match(m.testo_estratto[:400]))
        for c in candidates:
            if len(c) < 3:
                continue
            if c in desc_n or desc_n in c:
                return m
            if m.numero_documento and m.numero_documento.strip() in descrizione:
                return m
    return None


def trova_movimento_bonifico_per_riferimento(azienda_id: int, riferimento: str):
    """Aggancia per CRO / TRN / sottostringa su movimenti bonifico già registrati."""
    from .models import MovimentoRegistroStudioConsulente

    rif_clean = re.sub(r"\s+", "", (riferimento or "")).upper()
    if len(rif_clean) < 6:
        return None
    for m in MovimentoRegistroStudioConsulente.objects.filter(azienda_id=azienda_id, tipo_riga="bonifico").order_by(
        "-importato_il"
    )[:300]:
        mr = re.sub(r"\s+", "", (m.riferimento_pagamento or "")).upper()
        if not mr:
            continue
        if rif_clean in mr or mr in rif_clean or rif_clean[-12:] in mr:
            return m
    return None


def _norm_excel_header_cell(raw) -> str:
    """Testo intestazione foglio: minuscolo, NBSP → spazio, spazi compressi."""
    if raw is None:
        return ""
    s = str(raw).replace("\xa0", " ").replace("\u202f", " ").replace("\n", " ").strip().lower()
    return re.sub(r"\s+", " ", s)


def _pick_excel_col(headers: list[str], patterns: tuple[str, ...]) -> int | None:
    for i, raw in enumerate(headers):
        s = _norm_excel_header_cell(raw)
        for p in patterns:
            if p in s:
                return i
    return None


def _riepilogo_pick_ci_data(headers: list[str]) -> int | None:
    for i, raw in enumerate(headers):
        s = _norm_excel_header_cell(raw)
        if s == "data" or s.startswith("data ") or s.startswith("data:") or s.startswith("data."):
            return i
    return _pick_excel_col(
        headers,
        ("data valuta", "data operazione", "data contabile", "data reg", "data registrazione", "dt oper", "giorno"),
    )


def _riepilogo_pick_ci_documento(headers: list[str]) -> int | None:
    for i, raw in enumerate(headers):
        s = _norm_excel_header_cell(raw)
        if s == "documento" or s == "doc" or s == "doc.":
            return i
        if s.startswith("documento") and not s.startswith("documentazione"):
            return i
        if s.startswith("doc ") or s.startswith("doc.") or s.startswith("n. doc") or s.startswith("n doc"):
            return i
    return _pick_excel_col(
        headers,
        (
            "rif doc",
            "rif. doc",
            "riferimento doc",
            "protocollo",
            "n. protocollo",
            "numero doc",
            "num. fatt",
            "n. fatt",
            "fattura",
            "nota",
        ),
    )


def _riepilogo_pick_ci_descrizione(headers: list[str]) -> int | None:
    for i, raw in enumerate(headers):
        s = _norm_excel_header_cell(raw)
        if s == "descrizione" or s.startswith("descrizione"):
            return i
        if s.startswith("descr") and "documentazione" not in s:
            return i
    return _pick_excel_col(
        headers,
        ("dettaglio", "causale", "oggetto", "operaz", "note", "movimento", "annotaz", "testo", "descri"),
    )


def _riepilogo_pick_ci_importo(headers: list[str]) -> int | None:
    """Colonna importo pagamento: preferisce «importo»; accetta dare/avere di prima nota."""
    for i, raw in enumerate(headers):
        s = _norm_excel_header_cell(raw)
        if s in ("importo", "importo €", "importo eur", "imp.", "imp", "import"):
            return i
        if s.startswith("importo") or s.startswith("import "):
            return i
    idx = _pick_excel_col(
        headers,
        (
            "valore",
            "pagamento",
            "pagato",
            "addebit",
            "uscit",
            "amount",
            "eur",
            "€",
            "tot pag",
            "totale pag",
        ),
    )
    if idx is not None:
        return idx
    for i, raw in enumerate(headers):
        s = _norm_excel_header_cell(raw)
        if s == "dare" or s.startswith("dare ") or s == "avere" or s.startswith("avere "):
            return i
    return None


def _estratto_pick_ci_riferimento_bancario(headers: list[str]) -> int | None:
    """CRO / TRN / riferimento disposizione (non la colonna «Documento» gestionale)."""
    return _pick_excel_col(
        headers,
        (
            "cro",
            "trn",
            "rif. banc",
            "riferimento banc",
            "riferimento dispos",
            "ordinativo",
            "sct",
            "end to end",
        ),
    )


def _estratto_riga_strong_bonifico(desc: str, rif_bank: str, doc_raw: str = "") -> bool:
    """
    Riga da trattare come movimento bancario (match / eventuale creazione bonifico).
    Non usa la sola lunghezza della colonna «Documento» gestionale (evita PF-xx confusi con CRO).
    """
    d = f"{desc or ''} {rif_bank or ''}"
    if re.search(
        r"(?i)\b(cro|trn|bonific\w*|disposizion\w*)\b|\bsepa\b|accredito\s+banc|disposiz.{0,28}\bpag",
        d,
    ):
        return True
    rb = re.sub(r"\s+", "", (rif_bank or "")).upper()
    if len(rb) >= 10:
        return True
    return _excel_riga_sembr_bonifico(desc, doc_raw)


def _excel_riga_sembr_bonifico(descrizione: str, documento: str) -> bool:
    """Riga «pagamento / bonifico» nel riepilogo (non emissione fattura)."""
    d = f"{descrizione or ''} {documento or ''}"
    return bool(
        re.search(
            r"(?i)bonific|bonif\.|sepa\s*(ct|credit)|disposiz.*pag|transfer|"
            r"accredito|versamento|pagamento\s+ft|pag\.\s*ft|credit\s+transfer|"
            r"ordine\s+di\s+bonific",
            d,
        )
    )


def _excel_doc_colonna_riferimento_parcella_proforma(documento: str) -> bool:
    """
    True se la colonna «Documento» dell'Excel è chiaramente un codice proforma/parcella
    (tipico dei riepiloghi «PROFORMA» che non sono distinte bonifici SEPA).
    """
    doc = (documento or "").strip()
    if len(doc) < 2:
        return False
    if re.search(r"(?i)\b(parcella|parchell|proforma)\b", doc):
        return True
    du = doc.upper().replace(" ", "")
    if du.startswith("PAR-") and re.search(r"\d", doc):
        return True
    if (du.startswith("PF-") or du.startswith("PF.")) and re.search(r"\d", doc):
        return True
    return False


def _excel_doc_primo_token_fattura_o_protocollo_n_su_anno(documento: str) -> bool:
    """
    True se il primo token (prima parola) è un progressivo fattura/protocollo tipo «59/2026»,
    «182/2021» — tipico di righe di emissione nel riepilogo importate per errore come bonifico
    (chiave sintetica «59/2026 …|2026-02-10|52.00»). Non coincide con date gg/mm/aaaa su un solo token.
    """
    doc = (documento or "").strip()
    if not doc:
        return False
    first = doc.split()[0]
    return bool(re.match(r"^\d{1,6}/(19|20)\d{2}$", first))


def bonifico_excel_con_riferimento_sintetico_parcella_o_proforma(mov) -> bool:
    """
    True se il movimento è un bonifico da import riepilogo Excel il cui ``riferimento_pagamento``
    è la chiave sintetica ``documento|data|importo`` con documento chiaramente proforma/parcella
    (es. «PARCELLA 182|2021-06-16|130.00») oppure progressivo fattura/protocollo «n/AAAA» in testa
    (es. «59/2026 …|2026-02-10|52.00»), da distinguere dai bonifici bancari reali («BONIFICO …|…»).
    """
    if getattr(mov, "tipo_riga", None) != "bonifico":
        return False
    if (getattr(mov, "metodo_estrazione", None) or "") != "excel_riepilogo":
        return False
    rif = (mov.riferimento_pagamento or "").strip()
    if "|" not in rif:
        return False
    head = rif.split("|", 1)[0].strip()
    if not head:
        return False
    if _excel_doc_colonna_riferimento_parcella_proforma(head):
        return True
    return _excel_doc_primo_token_fattura_o_protocollo_n_su_anno(head)


def _excel_riga_solo_fattura_proforma_senza_bonifico(descrizione: str, documento: str = "") -> bool:
    """Riga di contabilizzazione documento (proforma/parcella), senza movimento di bonifico bancario."""
    if _excel_riga_sembr_bonifico(descrizione, documento):
        return False
    if _excel_doc_colonna_riferimento_parcella_proforma(documento):
        return True
    d = (descrizione or "").strip()
    if not d:
        return False
    return bool(
        re.search(
            r"(?i)\b(proforma|parcella|parchell|fattura\s+(n\.|nr|prot)|emissione|"
            r"documento\s+di\s+contabilit|registrazione\s+fatt)",
            d,
        )
    )


def _cell_to_decimal_excel(val) -> Decimal | None:
    if val is None or val == "":
        return None
    if isinstance(val, Decimal):
        return val.quantize(Decimal("0.01"))
    if isinstance(val, (int, float)):
        try:
            return Decimal(str(val)).quantize(Decimal("0.01"))
        except InvalidOperation:
            return None
    return _parse_it_decimal(str(val))


def trova_movimento_documento_per_colonna_documento(azienda_id: int, doc_raw: str):
    """
    Movimento proforma/parcella (PDF libro) collegato al valore colonna «DOCUMENTO» dell'Excel.
    """
    from .models import MovimentoRegistroStudioConsulente

    doc = (doc_raw or "").strip()
    if len(doc) < 2:
        return None
    qs = MovimentoRegistroStudioConsulente.objects.filter(azienda_id=azienda_id, tipo_riga="documento")
    m = qs.filter(numero_documento__iexact=doc).first()
    if m:
        return m
    doc_l = doc.lower()
    for m in qs.order_by("-importato_il")[:400]:
        nd = (m.numero_documento or "").strip().lower()
        if nd and (doc_l in nd or nd in doc_l):
            return m
        stem = Path(m.nome_file or "").stem.lower()
        if stem and doc_l in stem:
            return m
    return None


def _haystack_bonifico_per_aggancio_testuale(bon) -> str:
    """Testo unificato (minuscolo) per cercare il numero documento in causale / riferimento / note / distinta."""
    parts = [bon.causale_pagamento or "", bon.riferimento_pagamento or "", bon.note or ""]
    te = getattr(bon, "testo_estratto", None) or ""
    if te:
        parts.append(te[:4000])
    return " ".join(parts).lower()


def _numero_documento_in_haystack(num: str, hay: str) -> bool:
    num = (num or "").strip()
    if len(num) < 2:
        return False
    hay = hay or ""
    nl = num.lower()
    if nl not in hay:
        return False
    if len(num) >= 4:
        return True
    try:
        return bool(re.search(rf"(?<![0-9A-Za-z]){re.escape(num)}(?![0-9A-Za-z])", hay, re.I))
    except re.error:
        return len(num) >= 3


def _estrai_duo_numeri_proforma_parcella_con_e(hay: str) -> tuple[list[str] | None, int | None]:
    """
    Riconosce due numeri documento nel testo unificato (hay già minuscolo), in due forme:

    - «… proforma 320 e 367 del 2021 …» (una sola etichetta prima del primo numero);
    - «… proforma 320 e proforma 367 del 2021 …» (etichetta ripetuta, tipico causale in PDF distinta).

    Restituisce (['320','367'], 2021) oppure (None, None).
    """
    if not hay:
        return None, None
    m = re.search(
        r"\b(?:pro[-\s]?forma|proforma|parcella|parchella)\s+(\d{1,8})\s+e\s+(\d{1,8})(?:\s+del\s+(20\d{2}|19\d{2}))?\b",
        hay,
        re.I,
    )
    if m:
        y_raw = m.group(3)
        anno = int(y_raw) if y_raw else None
        return [m.group(1), m.group(2)], anno
    m2 = re.search(
        r"\b(?:pro[-\s]?forma|proforma|parcella|parchella)\s+(\d{1,8})\s+e\s+(?:pro[-\s]?forma|proforma|parcella|parchella)\s+(\d{1,8})(?:\s+del\s+(20\d{2}|19\d{2}))?\b",
        hay,
        re.I,
    )
    if not m2:
        return None, None
    y_raw = m2.group(3)
    anno = int(y_raw) if y_raw else None
    return [m2.group(1), m2.group(2)], anno


def _doppia_proforma_soppressa_per_altra_parcella_stesso_avere(
    hay: str,
    avere: Decimal,
    num_a: str,
    num_b: str,
    documenti: list,
) -> bool:
    """
    Se nel testo compare un altro documento (numero non tra i due della coppia) il cui dare
    coincide con l’avere del bonifico, non applicare la ripartita su due righe: tipico incasso
    BPSA che in causale cita «saldo proforma 320 e proforma 367» ma l’importo va su una sola
    parcella (es. 246) citata altrove nel testo/PDF.
    """
    pair = {num_a.casefold(), num_b.casefold()}
    tol = Decimal("0.02")
    for d in documenti:
        nd = (d.numero_documento or "").strip()
        if len(nd) < 2 or nd.casefold() in pair:
            continue
        if not _numero_documento_in_haystack(nd, hay):
            continue
        dare = d.dare or Decimal(0)
        if abs((dare - avere).quantize(Decimal("0.01"))) <= tol:
            return True
    return False


def _risolvi_documenti_per_numeri_e_anno_espliciti(
    documenti: list, nums: list[str], anno: int | None
) -> list | None:
    """Un documento per ogni numero; se ``anno`` è noto, preferisce un solo match per anno documento."""
    ris: list = []
    seen: set[int] = set()
    for num in nums:
        cand = []
        for d in documenti:
            nd = (d.numero_documento or "").strip()
            if not nd:
                continue
            if nd.casefold() == num.casefold() or _numeri_documento_aggancio_coerenti(nd, num):
                cand.append(d)
        if anno is not None:
            cy = [d for d in cand if d.data_documento and d.data_documento.year == anno]
            if len(cy) == 1:
                cand = cy
            elif len(cy) > 1:
                return None
        if len(cand) != 1:
            return None
        d0 = cand[0]
        if d0.pk in seen:
            return None
        ris.append(d0)
        seen.add(d0.pk)
    return ris


def _allocazione_sequenziale_avere_su_due_documenti(
    tot_avere: Decimal, doc_primo: object, doc_secondo: object
) -> tuple[Decimal, Decimal]:
    """
    Imputa l’avere del bonifico in ordine: fino al dare della prima parcella/proforma, poi il
    residuo fino al dare della seconda (come in contabilità: si «saldano» in cascata le due
    fatture citate nel testo del pagamento). Se l’avere supera la somma dei due dare, le due
    quote coincidono con i due dare (eccedenza oltre non imputata a queste righe).
    """
    cap0 = max(doc_primo.dare or Decimal(0), Decimal(0))
    cap1 = max(doc_secondo.dare or Decimal(0), Decimal(0))
    rem = tot_avere
    q0 = min(rem, cap0).quantize(Decimal("0.01"))
    rem = (rem - q0).quantize(Decimal("0.01"))
    q1 = min(rem, cap1).quantize(Decimal("0.01"))
    return q0, q1


def _allocazione_bonifico_doppia_proforma_da_testo(
    bon, hay: str, documenti: list
) -> list[tuple[object, Decimal]] | None:
    """
    Bonifico unico con in causale/riferimento due proforma/parcelle «N e M del AAAA»:
    imputazione **sequenziale** (prima riga fino al suo dare, residuo sulla seconda), nell’ordine
    dei numeri estratti dal testo.
    """
    nums, anno = _estrai_duo_numeri_proforma_parcella_con_e(hay)
    if not nums or len(nums) != 2:
        return None
    tot = bon.avere or Decimal("0")
    if tot > 0 and _doppia_proforma_soppressa_per_altra_parcella_stesso_avere(
        hay, tot, nums[0], nums[1], documenti
    ):
        return None
    docs = _risolvi_documenti_per_numeri_e_anno_espliciti(documenti, nums, anno)
    if not docs or len(docs) != 2:
        return None
    if tot <= 0:
        return None
    q0, q1 = _allocazione_sequenziale_avere_su_due_documenti(tot, docs[0], docs[1])
    return [(docs[0], q0), (docs[1], q1)]


def _correggi_eccedenza_bonifico_su_secondo_documento_citato(
    doc_assign: dict[int, list[dict]], documenti: list, bonifici: list
) -> None:
    """
    Se un bonifico è stato attribuito per intero a un solo documento (es. «dare più vicino» con
    più candidati) ma l’importo supera il dare di quella riga, e nella causale/distinta compaiono
    anche **entrambi** i numeri di un altro documento il cui dare coincide con l’eccedenza
    (importo − dare del documento «sbagliato»), sposta quella parte sul secondo documento.

    Tipico: € 274,14 tutti sulla parcella 320 (dare € 196,14) mentre in testo si citano 320 e 367:
    restano € 78,00 da imputare alla parcella 367 (dare € 78).
    """
    tol = Decimal("0.02")
    for b in bonifici:
        holders: list[tuple[object, dict]] = []
        for d in documenti:
            for e in doc_assign.get(d.pk, []):
                if e["bon"].pk == b.pk:
                    holders.append((d, e))
        if len(holders) != 1:
            continue
        d_cur, e = holders[0]
        tot_b = b.avere or Decimal("0")
        q = e.get("quota")
        attrib = q if q is not None else tot_b
        dare_cur = d_cur.dare or Decimal(0)
        if attrib <= dare_cur + tol:
            continue
        spill = (attrib - dare_cur).quantize(Decimal("0.01"))
        hay = _haystack_bonifico_per_aggancio_testuale(b)
        nd_cur = (d_cur.numero_documento or "").strip()
        if len(nd_cur) < 2 or not _numero_documento_in_haystack(nd_cur, hay):
            continue
        best_d2 = None
        ambiguous = False
        for d2 in documenti:
            if d2.pk == d_cur.pk:
                continue
            nd2 = (d2.numero_documento or "").strip()
            if len(nd2) < 2 or not _numero_documento_in_haystack(nd2, hay):
                continue
            dare2 = d2.dare or Decimal(0)
            if abs(dare2 - spill) > tol:
                continue
            if best_d2 is not None:
                ambiguous = True
                break
            best_d2 = d2
        if ambiguous or best_d2 is None:
            continue
        lst = doc_assign[d_cur.pk]
        lst.remove(e)
        lst.append({"bon": b, "quota": dare_cur})
        doc_assign[best_d2.pk].append({"bon": b, "quota": spill})


def _documenti_candidati_per_bonifico(azienda_id: int, bon, documenti: list) -> list:
    """
    Documenti proforma/parcella plausibilmente collegati a un bonifico.

    Priorità: prefisso ``documento|data|importo`` nel riferimento (stesso schema usato in libro
    per alcune righe da riepilogo legacy), altrimenti numero documento in causale / riferimento /
    note / testo estratto dalla distinta PDF caricata in Pagamenti.
    """
    riferimento = (bon.riferimento_pagamento or "").strip()
    if "|" in riferimento:
        head = riferimento.split("|", 1)[0].strip()
        if head:
            out: list = []
            seen: set[int] = set()
            m0 = trova_movimento_documento_per_colonna_documento(azienda_id, head)
            if m0 is not None:
                out.append(m0)
                seen.add(m0.pk)
            for d in documenti:
                if d.pk in seen:
                    continue
                nd = (d.numero_documento or "").strip()
                if nd and _numeri_documento_aggancio_coerenti(nd, head):
                    out.append(d)
                    seen.add(d.pk)
            if out:
                return out
    hay = _haystack_bonifico_per_aggancio_testuale(bon)
    out = []
    for d in documenti:
        nd = (d.numero_documento or "").strip()
        if len(nd) < 2:
            continue
        if _numero_documento_in_haystack(nd, hay):
            out.append(d)
    return out


def riferimento_pipe_aggancio_bonifico_documento(bon, doc) -> str:
    """
    Riferimento ``numero|data|avere`` (stesso schema usato da ``_documenti_candidati_per_bonifico``)
    per agganciare esplicitamente un bonifico a un movimento documento in libro.
    """
    num = (doc.numero_documento or "").strip()
    if len(num) < 1:
        raise ValueError("numero documento vuoto")
    dstr = doc.data_documento.isoformat() if doc.data_documento else ""
    imp = (bon.avere or Decimal(0)).quantize(Decimal("0.01"))
    return f"{num}|{dstr}|{imp}"[:160]


def riferimento_pipe_aggancio_bonifico_documenti_importi(pairs: list[tuple[object, Decimal]]) -> str:
    """
    Riferimento su più documenti: ``num|YYYY-MM-DD|importo;num2|…|importo2`` (max 160 caratteri).
    Usato da Pagamenti quando si ripartisce un bonifico su più proforma/parcelle.
    """
    if len(pairs) < 2:
        raise ValueError("Per più documenti servono almeno due righe con importo.")
    parts: list[str] = []
    tot_chk = Decimal(0)
    for doc, imp in pairs:
        num = (doc.numero_documento or "").strip()
        if len(num) < 1:
            raise ValueError("numero documento vuoto")
        impq = imp.quantize(Decimal("0.01"))
        if impq <= 0:
            raise ValueError("Ogni importo deve essere maggiore di zero.")
        tot_chk += impq
        dstr = doc.data_documento.isoformat() if doc.data_documento else ""
        parts.append(f"{num}|{dstr}|{impq}")
    s = ";".join(parts)
    if len(s) > 160:
        raise ValueError("Riferimento oltre 160 caratteri: ridurre i documenti o accorciare i numeri.")
    if tot_chk <= 0:
        raise ValueError("Somma importi non valida.")
    return s


def _parse_segmenti_pipe_aggancio_multiplo(riferimento: str) -> list[tuple[str, str, Decimal]] | None:
    """
    Formato esplicito su più documenti: «num|YYYY-MM-DD|importo;num2|…|importo2».
    Restituisce una lista solo se c'è ``;`` e ogni segmento ha esattamente tre campi con importo > 0.
    """
    s = (riferimento or "").strip()
    if ";" not in s:
        return None
    chunks = [c.strip() for c in s.split(";") if c.strip()]
    if len(chunks) < 2:
        return None
    out: list[tuple[str, str, Decimal]] = []
    for ch in chunks:
        bits = ch.split("|")
        if len(bits) != 3:
            return None
        num, d_raw, imp_raw = bits[0].strip(), bits[1].strip(), bits[2].strip()
        imp = _parse_it_decimal(imp_raw.replace(" ", "").replace("\xa0", ""))
        if imp is None or imp <= 0:
            return None
        d_iso = d_raw.strip()
        if d_iso and not re.match(r"^\d{4}-\d{2}-\d{2}$", d_iso):
            return None
        if len(num) < 1:
            return None
        out.append((num, d_iso, imp.quantize(Decimal("0.01"))))
    return out


def lista_triple_pipe_aggancio_da_riferimento(riferimento: str) -> list[tuple[str, str, Decimal]]:
    """
    Estrae le triple ``(prefisso_numero_documento, data_ISO_o_vuota, importo)`` usate per l’aggancio
    esplicito da Pagamenti: più segmenti ``;`` oppure una sola ``num|YYYY-MM-DD|importo``.
    """
    s = (riferimento or "").strip()
    if not s:
        return []
    if ";" in s:
        parsed = _parse_segmenti_pipe_aggancio_multiplo(s)
        return list(parsed) if parsed else []
    bits = s.split("|")
    if len(bits) != 3:
        return []
    num, d_raw, imp_raw = bits[0].strip(), bits[1].strip(), bits[2].strip()
    imp = _parse_it_decimal(imp_raw.replace(" ", "").replace("\xa0", ""))
    if imp is None or imp <= 0:
        return []
    d_iso = d_raw.strip()
    if d_iso and not re.match(r"^\d{4}-\d{2}-\d{2}$", d_iso):
        return []
    if len(num) < 1:
        return []
    return [(num, d_iso, imp.quantize(Decimal("0.01")))]


def _documento_per_aggancio_manuale_numero(azienda_id: int, num: str, data_iso: str, documenti: list):
    """
    Risolve un movimento documento da numero (e data documento ISO se indicata) tra i movimenti in libro.
    """
    num = (num or "").strip()
    if len(num) < 1:
        return None
    want_date = None
    if data_iso and re.match(r"^\d{4}-\d{2}-\d{2}$", data_iso):
        want_date = date.fromisoformat(data_iso)
    cands = []
    for d in documenti:
        nd = (d.numero_documento or "").strip()
        if not nd:
            continue
        if nd.lower() != num.lower():
            continue
        if want_date is None or d.data_documento == want_date:
            cands.append(d)
    if len(cands) == 1:
        return cands[0]
    if len(cands) > 1 and want_date is not None:
        for d in cands:
            if d.data_documento == want_date:
                return d
    if len(cands) >= 1:
        return cands[0]
    return trova_movimento_documento_per_colonna_documento(azienda_id, num)


def _allocazione_da_segmenti_pipe_espliciti(
    azienda_id: int, bon, documenti: list, segs: list[tuple[str, str, Decimal]]
) -> list[tuple[object, Decimal]] | None:
    out: list[tuple[object, Decimal]] = []
    tot = Decimal(0)
    for num, d_iso, imp in segs:
        d = _documento_per_aggancio_manuale_numero(azienda_id, num, d_iso, documenti)
        if d is None:
            return None
        out.append((d, imp))
        tot += imp
    av = (bon.avere or Decimal(0)).quantize(Decimal("0.01"))
    if abs(tot - av) > Decimal("0.02"):
        return None
    return out


def documenti_con_residuo_quadratura_per_select(azienda_id: int, q_quad: dict | None = None) -> list[dict]:
    """
    Documenti proforma/parcella con **residuo da incassare** > 0 (pipe + piano), per l’aggancio in
    Pagamenti: esclude fatture già saldate; mostra parziali con importo residuo.

    Se ``q_quad`` è già il risultato di ``quadratura_proforma_parcelle_bonifici``, viene riusato
    (evita doppio calcolo nella stessa richiesta).
    """
    q = q_quad if q_quad is not None else quadratura_proforma_parcelle_bonifici(azienda_id)
    rows: list[dict] = []
    for row in q["righe"]:
        res = row["residuo"]
        if res <= Decimal("0.01"):
            continue
        d = row["documento"]
        stato = row.get("stato") or ""
        pezzi = [
            d.get_tipo_documento_display(),
            f"n. {(d.numero_documento or '—').strip()}",
            f"residuo € {res.quantize(Decimal('0.01'))}",
            str(stato),
        ]
        if d.data_documento:
            pezzi.append(f"data doc. {d.data_documento.strftime('%d/%m/%Y')}")
        rows.append({"id": d.pk, "label": " · ".join(pezzi)})
    rows.sort(key=lambda x: (x["label"].lower(), x["id"]))
    return rows


def mappa_bonifico_documenti_stato_da_quadratura(q_quad: dict) -> dict[int, dict[int, str]]:
    """
    Per ogni bonifico presente nelle righe documento della quadratura (pipe + piano allocazione),
    mappa ``bon_pk -> {doc_pk: stato}`` dove ``stato`` è quello della riga documento (saldato, parziale, …).
    """
    out: dict[int, dict[int, str]] = {}
    for row in q_quad.get("righe") or []:
        stato = (row.get("stato") or "").strip() or "aperto"
        d = row.get("documento")
        if d is None:
            continue
        dpk = int(d.pk)
        for e in row.get("bonifici") or []:
            b = e.get("bon")
            if b is None:
                continue
            bpk = int(b.pk)
            out.setdefault(bpk, {})[dpk] = stato
    return out


def metadati_evidenza_bonifico_pagamenti(bon, q_quad: dict, doc_stati_per_bon: dict[int, dict[int, str]]) -> dict[str, str]:
    """
    Dati per la tabella Pagamenti consulente: classe CSS, etichetta breve, tooltip.
    """
    segs = lista_triple_pipe_aggancio_da_riferimento(bon.riferimento_pagamento or "")
    pipe_sembr_aggancio = bool(segs)
    by_bon = doc_stati_per_bon.get(int(bon.pk)) or {}
    if by_bon:
        stati = set(by_bon.values())
        numeri = []
        for row in q_quad.get("righe") or []:
            d = row.get("documento")
            if d is None or int(d.pk) not in by_bon:
                continue
            n = (d.numero_documento or "").strip() or "—"
            t = d.get_tipo_documento_display() if hasattr(d, "get_tipo_documento_display") else ""
            numeri.append(f"{t} n.{n}".strip())
        hint_doc = ", ".join(numeri[:4])
        if len(numeri) > 4:
            hint_doc += "…"
        if stati <= {"saldato"}:
            return {
                "classe": "gesper-bon-pag-saldati",
                "badge": "Saldato",
                "title": (
                    "Bonifico imputato in quadratura a documenti con incasso saldato. "
                    f"Documenti: {hint_doc or '—'}."
                ),
            }
        if "eccedenza" in stati:
            return {
                "classe": "gesper-bon-pag-attivi",
                "badge": "Aggancio",
                "title": (
                    "Bonifico collegato in quadratura; almeno un documento risulta in eccedenza. "
                    f"{hint_doc or '—'}."
                ),
            }
        return {
            "classe": "gesper-bon-pag-attivi",
            "badge": "Aggancio",
            "title": (
                "Bonifico imputato a proforma/parcella con residuo ancora aperto o parziale. "
                f"Documenti: {hint_doc or '—'}."
            ),
        }
    if pipe_sembr_aggancio:
        return {
            "classe": "gesper-bon-pag-pipe-rotto",
            "badge": "Rif. non risolto",
            "title": (
                "Il campo riferimento ha il formato di un aggancio manuale (num|data|importo) ma non risulta "
                "coerente con i documenti in libro o con l’avere del bonifico."
            ),
        }
    return {
        "classe": "gesper-bon-pag-libero",
        "badge": "",
        "title": (
            "Bonifico non ancora imputato in quadratura a parcelle/proforma (solo piano allocazione o "
            "«Salva aggancio» creano il collegamento contabile)."
        ),
    }


def annota_movimenti_bonifici_pagamenti_elenco(righe: list, q_quad: dict) -> None:
    """Aggiunge a ogni movimento ``pagamenti_row_meta`` (dict) per template Pagamenti."""
    mappa = mappa_bonifico_documenti_stato_da_quadratura(q_quad)
    for bon in righe:
        meta = metadati_evidenza_bonifico_pagamenti(bon, q_quad, mappa)
        setattr(bon, "pagamenti_row_meta", meta)


def documenti_proforma_parcella_libro_per_select(azienda_id: int) -> list[dict]:
    """
    Tutte le proforma/parcelle già in libro con residuo e stato (per aggancio manuale da Pagamenti),
    inclusi documenti saldati o in eccedenza.
    """
    q = quadratura_proforma_parcelle_bonifici(azienda_id)
    acc: list[tuple[date, int, dict]] = []
    for row in q["righe"]:
        d = row["documento"]
        res = row["residuo"]
        stato = row.get("stato") or ""
        num = (d.numero_documento or "—").strip()
        pezzi = [
            d.get_tipo_documento_display(),
            f"n. {num}",
            f"residuo € {res.quantize(Decimal('0.01'))}",
            str(stato),
        ]
        if d.data_documento:
            pezzi.append(f"data doc. {d.data_documento.strftime('%d/%m/%Y')}")
        sort_key = d.data_documento or date(1900, 1, 1)
        acc.append((sort_key, d.pk, {"id": d.pk, "label": " · ".join(pezzi)}))
    acc.sort(key=lambda t: (t[0], t[1]))
    return [t[2] for t in acc]


def bonifico_ha_riscontro_documentale_pagamento(mov) -> bool:
    """
    True se sul movimento bonifico è archiviato un PDF in portale (ricevuta bonifico,
    distinta, movimento da estratto conto o estratto conto bancario caricato in Pagamenti).

    Senza tale allegato non si ha evidenza documentale del pagamento: in sede di
    dimostrazione verso il consulente si assume operativamente che il bonifico non sia
    comprovato e che la parcella/proforma resti da pagare finché non si allega un PDF.
    """
    f = getattr(mov, "file", None)
    return bool(f and getattr(f, "name", None))


def _carica_righe_piano_allocazione_bonifici_quad(azienda_id: int) -> list[dict]:
    from .models import PianoAllocazioneBonificiQuad

    obj = PianoAllocazioneBonificiQuad.objects.filter(azienda_id=azienda_id).first()
    if not obj:
        return []
    return list(obj.righe or [])


def quadratura_proforma_parcelle_bonifici_anteprima_allocazione(azienda_id: int, pool_bonifico_ids: set[int]) -> dict:
    """
    Come ``quadratura_proforma_parcelle_bonifici`` ma i bonifici indicati non ricevono
    abbinamento automatico (utile al wizard: pool in costruzione + piano già salvato).
    """
    return _quadratura_proforma_parcelle_bonifici_core(azienda_id, extra_skip_bonifico_ids=frozenset(pool_bonifico_ids))


def _quadratura_proforma_parcelle_bonifici_core(
    azienda_id: int,
    extra_skip_bonifico_ids: frozenset[int] | None = None,
) -> dict:
    """
    Incrocia movimenti ``documento`` e ``bonifico`` in libro usando **solo**:

    - riferimenti in formato pipe esplicito da Pagamenti (``num|data|importo`` o più segmenti con ``;``),
      con somma importi uguale all’avere del bonifico;
    - eventuali righe del **piano allocazione bonifici** salvato in JSON (bonifici nel pool sono
      esclusi dall’attribuzione automatica finché il piano non viene applicato qui).

    Non si usa più l’euristica su causale, note o testo estratto dalla distinta: i bonifici senza
    pipe valido restano tra i «non collegati» finché non si imposta l’aggancio in Pagamenti.
    """
    from django.db.models import F

    from .models import MovimentoRegistroStudioConsulente

    extra_skip = extra_skip_bonifico_ids or frozenset()
    piano_righe = _carica_righe_piano_allocazione_bonifici_quad(azienda_id)
    skip_auto: set[int] = set(extra_skip)
    for r in piano_righe:
        bid = r.get("bonifico_id")
        if bid is None:
            continue
        try:
            skip_auto.add(int(bid))
        except (TypeError, ValueError):
            continue

    documenti = list(
        MovimentoRegistroStudioConsulente.objects.filter(azienda_id=azienda_id, tipo_riga="documento").order_by(
            F("data_documento").asc(nulls_last=True), "importato_il", "id"
        )
    )
    bonifici = list(
        MovimentoRegistroStudioConsulente.objects.filter(azienda_id=azienda_id, tipo_riga="bonifico").order_by(
            F("data_documento").asc(nulls_last=True), "importato_il", "id"
        )
    )

    doc_assign: dict[int, list[dict]] = {d.pk: [] for d in documenti}
    orfani: list = []
    pk_to_bon = {b.pk: b for b in bonifici}

    for b in bonifici:
        if b.pk in skip_auto:
            continue
        segs = lista_triple_pipe_aggancio_da_riferimento(b.riferimento_pagamento or "")
        if segs:
            alloc_espl = _allocazione_da_segmenti_pipe_espliciti(azienda_id, b, documenti, segs)
            if alloc_espl:
                for d, q in alloc_espl:
                    doc_assign[d.pk].append({"bon": b, "quota": q})
                continue
        orfani.append(b)

    for r in piano_righe:
        try:
            bid = int(r["bonifico_id"])
            did = int(r["documento_id"])
        except (KeyError, TypeError, ValueError):
            continue
        b = pk_to_bon.get(bid)
        if b is None or did not in doc_assign:
            continue
        try:
            q = Decimal(str(r.get("quota") or "0")).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError, TypeError):
            continue
        if q <= Decimal("0"):
            continue
        doc_assign[did].append({"bon": b, "quota": q})

    for lista in doc_assign.values():
        for e in lista:
            e["ha_riscontro"] = bonifico_ha_riscontro_documentale_pagamento(e["bon"])

    def _importo_attribuito_entry(entry: dict) -> Decimal:
        q = entry.get("quota")
        if q is not None:
            return q
        return entry["bon"].avere or Decimal("0")

    righe: list[dict] = []
    cum = Decimal("0")
    tot_dare = Decimal("0")
    tot_avere_all = sum((x.avere or Decimal(0)) for x in bonifici)
    tot_avere_attribuito = Decimal("0")

    for d in documenti:
        bs = doc_assign.get(d.pk, [])
        sum_av = sum(_importo_attribuito_entry(x) for x in bs)
        tot_avere_attribuito += sum_av
        dare = d.dare or Decimal("0")
        tot_dare += dare
        residuo = (dare - sum_av).quantize(Decimal("0.01"))
        cum = (cum + residuo).quantize(Decimal("0.01"))
        if residuo <= Decimal("0.01") and residuo >= Decimal("-0.01"):
            stato = "saldato"
        elif sum_av == 0:
            stato = "aperto"
        elif residuo > Decimal("0.01"):
            stato = "parziale"
        else:
            stato = "eccedenza"
        attrib_senza_pdf = sum(
            (_importo_attribuito_entry(x) for x in bs if not x.get("ha_riscontro", True)),
            start=Decimal("0"),
        ).quantize(Decimal("0.01"))
        saldato_ma_senza_evidenza = stato == "saldato" and any(
            not x.get("ha_riscontro", True) and _importo_attribuito_entry(x) > Decimal("0.01")
            for x in bs
        )
        righe.append(
            {
                "documento": d,
                "bonifici": bs,
                "tot_bonifici": sum_av,
                "residuo": residuo,
                "saldo_progressivo_residui": cum,
                "stato": stato,
                "importo_dare": dare,
                "attribuito_senza_evidenza_pdf": attrib_senza_pdf,
                "saldato_ma_con_bonifici_senza_evidenza_pdf": saldato_ma_senza_evidenza,
            }
        )

    tot_orfani = sum((b.avere or Decimal(0)) for b in orfani)
    bonifici_con_evidenza_pdf = sum(1 for b in bonifici if bonifico_ha_riscontro_documentale_pagamento(b))
    bonifici_senza_evidenza_pdf = len(bonifici) - bonifici_con_evidenza_pdf
    tot_avere_attribuito_senza_evidenza = sum(
        (
            _importo_attribuito_entry(e)
            for lista in doc_assign.values()
            for e in lista
            if not e.get("ha_riscontro", True)
        ),
        start=Decimal("0"),
    ).quantize(Decimal("0.01"))
    tot_orfani_senza_evidenza_avere = sum(
        ((b.avere or Decimal(0)) for b in orfani if not bonifico_ha_riscontro_documentale_pagamento(b)),
        start=Decimal("0"),
    ).quantize(Decimal("0.01"))

    bonifici_residuo_post_piano_allocazione: list[dict] = []
    bons_in_piano: set[int] = set()
    for r in piano_righe:
        if r.get("bonifico_id") is None:
            continue
        try:
            bons_in_piano.add(int(r["bonifico_id"]))
        except (TypeError, ValueError):
            continue
    for bid in bons_in_piano:
        b = pk_to_bon.get(bid)
        if b is None:
            continue
        used = sum(
            (
                Decimal(str(r.get("quota") or "0"))
                for r in piano_righe
                if r.get("bonifico_id") is not None and int(r["bonifico_id"]) == bid
            ),
            start=Decimal("0"),
        ).quantize(Decimal("0.01"))
        av = (b.avere or Decimal("0")).quantize(Decimal("0.01"))
        if used + Decimal("0.02") < av:
            bonifici_residuo_post_piano_allocazione.append(
                {
                    "bon": b,
                    "attribuito": used,
                    "residuo": (av - used).quantize(Decimal("0.01")),
                }
            )
    tot_residui_post_piano = sum(
        (x["residuo"] for x in bonifici_residuo_post_piano_allocazione),
        start=Decimal("0"),
    ).quantize(Decimal("0.01"))

    by_bon: dict[int, list[tuple[str, Decimal]]] = {}
    for d in documenti:
        for e in doc_assign.get(d.pk, []):
            q = _importo_attribuito_entry(e)
            if q <= 0:
                continue
            nd = (d.numero_documento or "").strip() or "—"
            bpk = e["bon"].pk
            by_bon.setdefault(bpk, []).append((nd, q))

    bonifici_ripartiti_multi_documento: list[dict] = []
    for bon_pk, parti in by_bon.items():
        if len(parti) < 2:
            continue
        b = pk_to_bon.get(bon_pk)
        if b is None:
            continue
        somma = sum((q for _, q in parti), start=Decimal("0")).quantize(Decimal("0.01"))
        avere_b = (b.avere or Decimal(0)).quantize(Decimal("0.01"))
        bonifici_ripartiti_multi_documento.append(
            {
                "bon": b,
                "parti": [{"numero_documento": n, "quota": q} for n, q in parti],
                "somma_quote": somma,
                "avere": avere_b,
                "coerente": abs(somma - avere_b) <= Decimal("0.02"),
            }
        )

    return {
        "righe": righe,
        "bonifici_orfani": orfani,
        "bonifici_residuo_post_piano_allocazione": bonifici_residuo_post_piano_allocazione,
        "bonifici_ripartiti_multi_documento": bonifici_ripartiti_multi_documento,
        "totali": {
            "documenti_n": len(documenti),
            "bonifici_n": len(bonifici),
            "totale_dare": tot_dare,
            "totale_avere_libro": tot_avere_all,
            "totale_avere_attribuito": tot_avere_attribuito,
            "totale_orfani_avere": tot_orfani,
            "bonifici_con_evidenza_pdf": bonifici_con_evidenza_pdf,
            "bonifici_senza_evidenza_pdf": bonifici_senza_evidenza_pdf,
            "totale_avere_attribuito_senza_evidenza_pdf": tot_avere_attribuito_senza_evidenza,
            "totale_orfani_avere_senza_evidenza_pdf": tot_orfani_senza_evidenza_avere,
            "totale_residui_avere_post_piano_allocazione": tot_residui_post_piano,
        },
    }


def quadratura_proforma_parcelle_bonifici(azienda_id: int) -> dict:
    """Incrocio documenti / bonifici: solo riferimenti pipe espliciti e piano allocazione manuale."""
    return _quadratura_proforma_parcelle_bonifici_core(azienda_id)


def bonifico_ids_con_avere_residuo_utilizzabile_in_quadratura(azienda_id: int) -> set[int]:
    """
    id dei bonifici che, con **pipe espliciti + piano**, hanno ancora **avere non imputato**
    alle parcelle/proforma (orfani = tutto l’avere; altrimenti avere − somma delle quote attribuite
    sulle righe documento).

    Il passo 1 del wizard «Piano bonifici» elenca solo questi movimenti.
    """
    q = quadratura_proforma_parcelle_bonifici(azienda_id)

    def _importo_attribuito_entry(entry: dict) -> Decimal:
        qq = entry.get("quota")
        if qq is not None:
            return qq
        return entry["bon"].avere or Decimal("0")

    attribuito_per_bon: dict[int, Decimal] = {}
    bon_per_pk: dict[int, object] = {}
    for row in q["righe"]:
        for e in row["bonifici"]:
            bpk = e["bon"].pk
            bon_per_pk[bpk] = e["bon"]
            attribuito_per_bon[bpk] = attribuito_per_bon.get(bpk, Decimal("0")) + _importo_attribuito_entry(e)
    for b in q["bonifici_orfani"]:
        bon_per_pk[b.pk] = b
    for row in q.get("bonifici_residuo_post_piano_allocazione") or []:
        bon = row.get("bon")
        if bon is not None:
            bon_per_pk[bon.pk] = bon

    out: set[int] = set()
    for bpk, b in bon_per_pk.items():
        av = (b.avere or Decimal("0")).quantize(Decimal("0.01"))
        used = attribuito_per_bon.get(bpk, Decimal("0")).quantize(Decimal("0.01"))
        if av - used > Decimal("0.02"):
            out.add(bpk)
    return out


def mappa_quadratura_per_export_libro_movimenti(azienda_id: int) -> dict:
    """
    Dati allineati a ``quadratura_proforma_parcelle_bonifici`` (pipe espliciti + piano) per colonne
    «Pagato» / «Residuo» nell’export libro (PDF/Excel): per ogni documento incassi attribuiti e residuo;
    per ogni bonifico l’avere non ancora imputato alle parcelle/proforma; saldo cumulativo finale
    dei residui. Include ``riepilogo_coerenza`` per il blocco riepilogo nel PDF libro.
    """
    q = quadratura_proforma_parcelle_bonifici(azienda_id)

    def _importo_attribuito_entry(entry: dict) -> Decimal:
        qq = entry.get("quota")
        if qq is not None:
            return qq
        return entry["bon"].avere or Decimal("0")

    doc_by_pk: dict[int, dict] = {}
    for row in q["righe"]:
        d = row["documento"]
        doc_by_pk[d.pk] = {
            "residuo": (row.get("residuo") or Decimal("0")).quantize(Decimal("0.01")),
            "tot_bonifici": (row.get("tot_bonifici") or Decimal("0")).quantize(Decimal("0.01")),
            "saldo_progressivo_residui": (row.get("saldo_progressivo_residui") or Decimal("0")).quantize(
                Decimal("0.01")
            ),
        }

    attribuito_per_bon: dict[int, Decimal] = {}
    for row in q["righe"]:
        for e in row["bonifici"]:
            bpk = e["bon"].pk
            attribuito_per_bon[bpk] = attribuito_per_bon.get(bpk, Decimal("0")) + _importo_attribuito_entry(e)

    from django.db.models import F

    from .models import MovimentoRegistroStudioConsulente

    bonifici_all = MovimentoRegistroStudioConsulente.objects.filter(
        azienda_id=azienda_id, tipo_riga="bonifico"
    ).order_by(F("data_documento").asc(nulls_last=True), "importato_il", "id")

    bon_residuo_avere: dict[int, Decimal] = {}
    for b in bonifici_all:
        av = (b.avere or Decimal("0")).quantize(Decimal("0.01"))
        used = attribuito_per_bon.get(b.pk, Decimal("0")).quantize(Decimal("0.01"))
        diff = (av - used).quantize(Decimal("0.01"))
        bon_residuo_avere[b.pk] = diff if diff > Decimal("0") else Decimal("0")

    righe_q = q.get("righe") or []
    if righe_q:
        saldo_cumulativo_finale = (righe_q[-1].get("saldo_progressivo_residui") or Decimal("0")).quantize(
            Decimal("0.01")
        )
    else:
        saldo_cumulativo_finale = Decimal("0")

    tot = q.get("totali") or {}
    tot_dare_parcella = Decimal("0")
    tot_dare_proforma = Decimal("0")
    tot_dare_altro_doc = Decimal("0")
    for row in righe_q:
        d = row["documento"]
        dd = (row.get("importo_dare") or getattr(d, "dare", None) or Decimal("0")).quantize(Decimal("0.01"))
        t = (getattr(d, "tipo_documento", None) or "").strip()
        if t == "parcella":
            tot_dare_parcella += dd
        elif t == "proforma":
            tot_dare_proforma += dd
        else:
            tot_dare_altro_doc += dd

    tot_dare_parcella = tot_dare_parcella.quantize(Decimal("0.01"))
    tot_dare_proforma = tot_dare_proforma.quantize(Decimal("0.01"))
    tot_dare_altro_doc = tot_dare_altro_doc.quantize(Decimal("0.01"))
    tot_dare_documenti_quad = (tot.get("totale_dare") or Decimal("0")).quantize(Decimal("0.01"))
    tot_avere_bonifici_libro = (tot.get("totale_avere_libro") or Decimal("0")).quantize(Decimal("0.01"))
    tot_avere_attribuito_documenti = (tot.get("totale_avere_attribuito") or Decimal("0")).quantize(Decimal("0.01"))
    differenza_dare_meno_avere = (tot_dare_documenti_quad - tot_avere_bonifici_libro).quantize(Decimal("0.01"))
    tot_orfani_avere = (tot.get("totale_orfani_avere") or Decimal("0")).quantize(Decimal("0.01"))
    avere_non_imputato_a_fatture = (tot_avere_bonifici_libro - tot_avere_attribuito_documenti).quantize(
        Decimal("0.01")
    )
    # Σ residui finale = differenza + avere non ancora imputato alle fatture (orfani + quote non allocate)
    scarto_coerenza_sigma_diff = (
        saldo_cumulativo_finale - differenza_dare_meno_avere - avere_non_imputato_a_fatture
    ).quantize(Decimal("0.01"))

    from django.db.models import Sum

    agg_rett = MovimentoRegistroStudioConsulente.objects.filter(
        azienda_id=azienda_id, tipo_riga="rettifica"
    ).aggregate(sd=Sum("dare"), sa=Sum("avere"))
    dare_rettifiche = (agg_rett.get("sd") or Decimal("0")).quantize(Decimal("0.01"))
    avere_rettifiche = (agg_rett.get("sa") or Decimal("0")).quantize(Decimal("0.01"))

    riepilogo_coerenza = {
        "tot_dare_parcella": tot_dare_parcella,
        "tot_dare_proforma": tot_dare_proforma,
        "tot_dare_altro_documento": tot_dare_altro_doc,
        "tot_dare_documenti_quadratura": tot_dare_documenti_quad,
        "tot_avere_bonifici_libro": tot_avere_bonifici_libro,
        "tot_avere_attribuito_documenti": tot_avere_attribuito_documenti,
        "avere_non_imputato_a_fatture": avere_non_imputato_a_fatture,
        "differenza_dare_meno_avere": differenza_dare_meno_avere,
        "tot_orfani_avere": tot_orfani_avere,
        "scarto_coerenza_sigma_diff": scarto_coerenza_sigma_diff,
        "dare_rettifiche": dare_rettifiche,
        "avere_rettifiche": avere_rettifiche,
    }

    return {
        "documento": doc_by_pk,
        "bonifico_residuo_avere": bon_residuo_avere,
        "saldo_cumulativo_residui_finale": saldo_cumulativo_finale,
        "riepilogo_coerenza": riepilogo_coerenza,
    }


def documenti_righe_quadratura_con_residuo_da_coprire(righe_quad: list[dict]) -> list[dict]:
    """
    Righe documento (come in ``quadratura_proforma_parcelle_bonifici`` ``righe``) con residuo
    da incassare su proforma/parcella: stato ``aperto`` o ``parziale`` e residuo > 0.
    Esclude ``saldato`` ed ``eccedenza``.
    """
    out: list[dict] = []
    for row in righe_quad:
        res = (row.get("residuo") or Decimal("0")).quantize(Decimal("0.01"))
        if res <= Decimal("0.01"):
            continue
        if row.get("stato") not in ("aperto", "parziale"):
            continue
        out.append(row)
    return out


def _cap_residui_documenti_per_validazione_piano(azienda_id: int, bon_ids_pool: set[int]) -> dict[int, Decimal]:
    """
    Residuo massimo imputabile per documento rispetto al pool: stessa base della **anteprima**
    wizard (``_quadratura…_core`` con ``extra_skip`` = pool), così i bonifici del pool non
    ricevono abbinamento automatico nella stima del cap; poi si sommano le quote già nel piano
    solo per bonifici del pool (da sostituire in questo salvataggio).

    Usare la quadratura «completa» qui avrebbe effetto contrario: un bonifico del pool ancora
    non in piano verrebbe attribuito in automatico e il cap risulterebbe 0 mentre in anteprima
    il residuo resta aperto.
    """
    from .models import PianoAllocazioneBonificiQuad

    q = _quadratura_proforma_parcelle_bonifici_core(azienda_id, extra_skip_bonifico_ids=frozenset(bon_ids_pool))
    cap: dict[int, Decimal] = {
        row["documento"].pk: (row["residuo"] or Decimal("0")).quantize(Decimal("0.01")) for row in q["righe"]
    }
    obj = PianoAllocazioneBonificiQuad.objects.filter(azienda_id=azienda_id).first()
    if not obj:
        return cap
    for r in obj.righe or []:
        try:
            bid = int(r["bonifico_id"])
            did = int(r["documento_id"])
            qv = Decimal(str(r.get("quota") or "0")).quantize(Decimal("0.01"))
        except (KeyError, TypeError, ValueError, InvalidOperation):
            continue
        if bid not in bon_ids_pool or qv <= 0:
            continue
        cap[did] = (cap.get(did, Decimal("0")) + qv).quantize(Decimal("0.01"))
    return cap


def costruisce_righe_piano_allocazione_bonifici_quad(
    azienda_id: int,
    bon_ids_ordinati: list[int],
    allocazioni: list[tuple[int, Decimal]],
) -> list[dict]:
    """
    Partendo da bonifici in ordine (FIFO sul pool), ripartisce le coppie (documento, importo)
    in righe {documento_id, bonifico_id, quota} per JSON piano.
    """
    from .models import MovimentoRegistroStudioConsulente

    if not bon_ids_ordinati:
        raise ValueError("Selezionare almeno un bonifico nel pool.")
    bons_qs = MovimentoRegistroStudioConsulente.objects.filter(
        pk__in=bon_ids_ordinati, azienda_id=azienda_id, tipo_riga="bonifico"
    )
    by_id = {b.pk: b for b in bons_qs}
    for i in bon_ids_ordinati:
        if i not in by_id:
            raise ValueError("Uno o più bonifici non appartengono a questa azienda o non sono in avere.")
    bons = [by_id[i] for i in bon_ids_ordinati]
    remaining = {b.pk: (b.avere or Decimal("0")).quantize(Decimal("0.01")) for b in bons}
    pool = set(bon_ids_ordinati)
    cap = _cap_residui_documenti_per_validazione_piano(azienda_id, pool)

    alloc_by_doc: dict[int, Decimal] = {}
    for doc_id, amt in allocazioni:
        if amt is None or amt <= 0:
            continue
        amt = amt.quantize(Decimal("0.01"))
        alloc_by_doc[doc_id] = (alloc_by_doc.get(doc_id, Decimal("0")) + amt).quantize(Decimal("0.01"))

    for doc_id, tot_doc in alloc_by_doc.items():
        max_r = cap.get(doc_id)
        if max_r is None:
            raise ValueError(f"Il documento id {doc_id} non risulta in partita con residuo utilizzabile.")
        if tot_doc > max_r + Decimal("0.02"):
            raise ValueError(
                f"Sul documento id {doc_id} la somma richiesta ({tot_doc}) supera il residuo disponibile ({max_r})."
            )

    pool_total = sum((b.avere or Decimal("0")) for b in bons).quantize(Decimal("0.01"))
    total_alloc = sum(alloc_by_doc.values(), start=Decimal("0")).quantize(Decimal("0.01"))
    if total_alloc > pool_total + Decimal("0.02"):
        raise ValueError(
            f"La somma delle imputazioni (€ {total_alloc}) supera il pool bonifici selezionato (€ {pool_total})."
        )

    righe_out: list[dict] = []
    for doc_id, amt in allocazioni:
        if amt is None or amt <= 0:
            continue
        rem = amt.quantize(Decimal("0.01"))
        doc = MovimentoRegistroStudioConsulente.objects.filter(
            pk=doc_id, azienda_id=azienda_id, tipo_riga="documento"
        ).first()
        if doc is None:
            raise ValueError(f"Documento id {doc_id} non valido.")
        for b in bons:
            if rem <= Decimal("0"):
                break
            avail = remaining[b.pk]
            if avail <= Decimal("0"):
                continue
            take = min(rem, avail).quantize(Decimal("0.01"))
            if take <= Decimal("0"):
                continue
            righe_out.append(
                {
                    "documento_id": doc.pk,
                    "bonifico_id": b.pk,
                    "quota": str(take),
                }
            )
            remaining[b.pk] = (remaining[b.pk] - take).quantize(Decimal("0.01"))
            rem = (rem - take).quantize(Decimal("0.01"))
        if rem > Decimal("0.02"):
            raise ValueError(
                "Il pool bonifici non copre tutte le righe richieste: aumentare i bonifici o ridurre gli importi."
            )
    return righe_out


def salva_piano_allocazione_bonifici_quadratura(azienda, bon_ids_ordinati: list[int], allocazioni: list[tuple[int, Decimal]], user) -> None:
    from .models import PianoAllocazioneBonificiQuad

    righe_new = costruisce_righe_piano_allocazione_bonifici_quad(azienda.id, bon_ids_ordinati, allocazioni)
    pool = set(bon_ids_ordinati)
    obj, _ = PianoAllocazioneBonificiQuad.objects.get_or_create(azienda=azienda, defaults={"righe": []})
    kept: list[dict] = []
    for r in obj.righe or []:
        try:
            bid = int(r["bonifico_id"])
        except (KeyError, TypeError, ValueError):
            continue
        if bid not in pool:
            kept.append(r)
    obj.righe = kept + righe_new
    obj.aggiornato_da = user
    obj.save()


def rimuovi_righe_piano_allocazione_per_bonifico(azienda_id: int, bonifico_id: int, user) -> int:
    """
    Rimuove dal piano allocazione manuale tutte le righe JSON che citano ``bonifico_id``.
    Usato prima di eliminare il movimento bonifico dal libro, così non serve svuotare il piano a mano.
    Ritorna il numero di righe eliminate.
    """
    from .models import PianoAllocazioneBonificiQuad

    obj = PianoAllocazioneBonificiQuad.objects.filter(azienda_id=azienda_id).first()
    if not obj or not obj.righe:
        return 0
    old = list(obj.righe)
    kept: list[dict] = []
    for r in old:
        try:
            bid = int(r.get("bonifico_id") or 0)
        except (TypeError, ValueError):
            bid = 0
        if bid != bonifico_id:
            kept.append(r)
    removed = len(old) - len(kept)
    if removed <= 0:
        return 0
    obj.righe = kept
    obj.aggiornato_da = user
    obj.save(update_fields=["righe", "aggiornato_da"])
    return removed


def elimina_piano_allocazione_bonifici_quadratura(azienda_id: int) -> None:
    from .models import PianoAllocazioneBonificiQuad

    PianoAllocazioneBonificiQuad.objects.filter(azienda_id=azienda_id).delete()


def _estrai_riferimento_bonifico_excel(descrizione: str, documento: str, data_doc: date | None, imp: Decimal | None) -> str:
    """CRO/TRN da descrizione, altrimenti chiave sintetica stabile."""
    desc = descrizione or ""
    for pat in (
        r"(?i)(?:CRO|Riferimento\s+CRO)\s*[:\s]+([A-Za-z0-9/\-\s]{6,42})",
        r"(?i)(?:TRN|End\s+To\s+End)\s*[:\s]+([A-Za-z0-9/\-\s]{6,42})",
        r"(?i)(?:Ordinativo|N\.\s*disposizione)\s*[:\s]+([A-Za-z0-9/\-\s]{6,40})",
    ):
        m = re.search(pat, desc)
        if m:
            return re.sub(r"\s+", " ", m.group(1).strip())[:160]
    digits = re.findall(r"\b\d{12,30}\b", desc)
    if digits:
        return max(digits, key=len)[:160]
    parts = [
        (documento or "").strip()[:48],
        data_doc.isoformat() if data_doc else "",
        str(imp) if imp is not None else "",
    ]
    return "|".join(p for p in parts if p)[:160] or "import-excel"


def _bonifico_excel_già_presente(azienda_id: int, data_doc: date | None, avere: Decimal, riferimento: str, causale: str) -> bool:
    return trova_bonifico_esistente_stesso_excel(azienda_id, data_doc, avere, riferimento, causale) is not None


def _movimento_bonifico_duplicato_da_parsed_pdf(azienda_id: int, parsed: EsitoParsingBonifico):
    """
    Bonifico già in libro equivalente ai dati estratti da una distinta PDF.

    Evita doppie righe se lo stesso PDF (o un duplicato logico) viene importato più volte da
    «Upload PDF» / «Solo distinta PDF»: stessi criteri dell'import Excel (data, avere, riferimento o causale),
    più controllo su CRO/TRN se il riferimento è sufficientemente specifico.
    """
    if not parsed.importo or parsed.importo <= 0:
        return None
    ex = trova_bonifico_esistente_stesso_excel(
        azienda_id,
        parsed.data_documento,
        parsed.importo,
        (parsed.riferimento or "").strip(),
        (parsed.causale or "").strip(),
    )
    if ex is not None:
        return ex
    rif = (parsed.riferimento or "").strip()
    if len(rif) >= 6:
        m = trova_movimento_bonifico_per_riferimento(azienda_id, rif)
        if m is not None and m.avere == parsed.importo:
            if (
                parsed.data_documento is None
                or m.data_documento is None
                or m.data_documento == parsed.data_documento
            ):
                return m
    return None


def trova_bonifico_esistente_stesso_excel(
    azienda_id: int,
    data_doc: date | None,
    avere: Decimal | None,
    riferimento: str,
    causale: str,
):
    """
    Bonifico già in libro con stessa data/importo e riferimento o causale compatibile
    (stessa logica di _bonifico_excel_già_presente, ma restituisce il movimento).
    """
    from .models import MovimentoRegistroStudioConsulente

    if avere is None or avere <= 0:
        return None
    rif = (riferimento or "").strip()[:160]
    caus = (causale or "").strip()[:120]
    qs = MovimentoRegistroStudioConsulente.objects.filter(
        azienda_id=azienda_id,
        tipo_riga="bonifico",
        data_documento=data_doc,
        avere=avere,
    )
    if rif and len(rif) >= 8:
        m = qs.filter(riferimento_pagamento__iexact=rif).first()
        if m:
            return m
        m = qs.filter(riferimento_pagamento__icontains=rif[-14:]).first()
        if m:
            return m
    if caus and len(caus) >= 12:
        return qs.filter(causale_pagamento__icontains=caus[:80]).first()
    return None


def bonifico_duplicato_elenco_ids(righe) -> set[int]:
    """
    ID dei bonifici che nell'elenco condividono la stessa chiave con almeno un'altra riga:
    (data valuta, avere, riferimento in minuscolo) oppure, se riferimento vuoto,
    (data valuta, avere, causale troncata in minuscolo).

    Serve solo evidenziazione in UI (es. Pagamenti); non sostituisce la deduplica in import.
    """
    from collections import Counter

    rows = list(righe)

    def row_key(r):
        d = r.data_documento
        av = r.avere
        rif = (r.riferimento_pagamento or "").strip().lower()
        if rif:
            return (d, av, "rif", rif)
        caus = (r.causale_pagamento or "").strip().lower()[:200]
        return (d, av, "caus", caus)

    cnt = Counter(row_key(r) for r in rows)
    dup_keys = {k for k, n in cnt.items() if n > 1}
    return {r.id for r in rows if row_key(r) in dup_keys}


def _riepilogo_excel_convenzione_solo_importi_negativi(all_rows: list[tuple], header_idx: int, ci_imp: int) -> bool:
    """
    True se **tutti** gli importi numerici non nulli della colonna sono < 0 (almeno 5 righe):
    tipico estratto in cui gli incassi compaiono come importi negativi — si importano solo quelle righe
    (gli importi positivi nella colonna vengono ignorati).
    """
    vals: list[Decimal] = []
    for row in all_rows[header_idx + 1 :]:
        if not row or ci_imp >= len(row):
            continue
        v = _cell_to_decimal_excel(row[ci_imp])
        if v is not None and v != 0:
            vals.append(v)
    if len(vals) < 5:
        return False
    if any(v > 0 for v in vals):
        return False
    return any(v < 0 for v in vals)


def import_riepilogo_bonifici_da_excel(
    fileobj,
    nome_file: str,
    azienda,
    user,
) -> list[str]:
    """
    Importa da un Excel «riepilogo» le righe di bonifico (Data, DOCUMENTO, DESCRIZIONE, IMPORTO).

    - Importa solo righe classificate come bonifico.
    - L’importo in avere nel libro è sempre il **valore assoluto** della colonna (importi negativi in Excel
      tipici degli estratti conto vengono registrati come incasso positivo in avere).
    - Se **tutti** gli importi numerici non nulli della colonna sono **negativi** (almeno 5 righe, nessun valore
      positivo in colonna), si importano **solo** le righe con importo < 0 (si ignorano eventuali positivi).
    - Salta righe che sono solo contabilizzazione proforma/parcella (descrizione **o** colonna Documento
      con codice tipo «PARCELLA 182», «PF-…», «PAR-…»), senza segnali da bonifico SEPA: il dare resta
      dai PDF / altre fonti; non si crea un falso bonifico in avere dal riepilogo PROFORMA.
    - Salta bonifici già registrati (stessa data, importo, riferimento/causale).
    """
    from django.db import transaction

    from openpyxl import load_workbook

    from .models import MovimentoRegistroStudioConsulente

    raw_bytes = fileobj.read()
    try:
        fileobj.seek(0)
    except (OSError, AttributeError, io.UnsupportedOperation):
        pass

    msgs: list[str] = []
    stem = Path(nome_file or "riepilogo.xlsx").stem[:80]

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        all_rows: list[tuple] = []
        header_idx: int | None = None
        headers: list[str] = []
        ci_data = ci_doc = ci_desc = ci_imp = None

        for ws in wb.worksheets[:8]:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            for i, row in enumerate(rows[:150]):
                if not row:
                    continue
                hdr = [_norm_excel_header_cell(h) for h in row]
                if not any(hdr):
                    continue
                ci_data = _riepilogo_pick_ci_data(hdr)
                ci_doc = _riepilogo_pick_ci_documento(hdr)
                ci_desc = _riepilogo_pick_ci_descrizione(hdr)
                ci_imp = _riepilogo_pick_ci_importo(hdr)
                if ci_data is not None and ci_desc is not None and ci_imp is not None:
                    header_idx = i
                    headers = hdr
                    all_rows = rows
                    break
            if header_idx is not None:
                break
    finally:
        wb.close()

    if not all_rows or header_idx is None or ci_data is None or ci_desc is None or ci_imp is None:
        raise ValueError(
            "Intestazioni non trovate: servono colonne riconducibili a «Data», «Descrizione» e «Importo» "
            "(primi 150 righe del foglio; si cercano fino a 8 fogli). La colonna «Documento» è opzionale. "
            "Sinonimi accettati: Doc., Dettaglio, Causale, Dare/Avere, Data valuta, ecc."
        )

    data_rows = all_rows[header_idx + 1 :]
    solo_importi_negativi = _riepilogo_excel_convenzione_solo_importi_negativi(all_rows, header_idx, ci_imp)
    n_import = n_skip = n_salt_doc = n_salt_bon = 0

    with transaction.atomic():
        for k, row in enumerate(data_rows, start=header_idx + 2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            doc_raw = ""
            if ci_doc is not None and ci_doc < len(row) and row[ci_doc] is not None:
                doc_raw = str(row[ci_doc]).strip()[:200]
            desc = str(row[ci_desc]).strip()[:600] if ci_desc < len(row) and row[ci_desc] is not None else ""
            d_ex = _cell_to_date(row[ci_data]) if ci_data < len(row) else None
            imp = _cell_to_decimal_excel(row[ci_imp]) if ci_imp < len(row) else None

            if imp is None or imp == 0:
                n_skip += 1
                continue
            if solo_importi_negativi and imp > 0:
                n_skip += 1
                continue

            avere_mov = abs(imp)

            if _excel_riga_solo_fattura_proforma_senza_bonifico(desc, doc_raw):
                n_salt_doc += 1
                continue

            layout_classico = (
                ci_doc is not None
                and ci_data is not None
                and ci_desc is not None
                and ci_imp is not None
            )
            if not _excel_riga_sembr_bonifico(desc, doc_raw):
                if not layout_classico:
                    n_skip += 1
                    continue

            rif = _estrai_riferimento_bonifico_excel(desc, doc_raw, d_ex, avere_mov)
            if _bonifico_excel_già_presente(azienda.id, d_ex, avere_mov, rif, desc):
                n_salt_bon += 1
                continue

            mov_collegato = trova_movimento_documento_per_colonna_documento(azienda.id, doc_raw)
            if mov_collegato and getattr(mov_collegato.file, "name", None):
                pdf_note = "PDF proforma/parcella collegato: presente."
            elif mov_collegato:
                pdf_note = "PDF proforma/parcella collegato: assente (documento in libro senza file)."
            else:
                pdf_note = "PDF proforma/parcella collegato: nessun documento in libro con questo «Documento»."

            nome_mov = f"xlsx-bon/{stem}/R{k}"[:280]
            if MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda, nome_file=nome_mov).exists():
                n_salt_bon += 1
                continue

            note_parts = [f"Bonifici da riepilogo «{stem}»; {pdf_note}"]
            if imp < 0:
                note_parts.append("Importo negativo nel foglio → in avere come incasso (valore assoluto).")
            note = " ".join(note_parts)[:500]
            obj = MovimentoRegistroStudioConsulente(
                azienda=azienda,
                tipo_riga="bonifico",
                tipo_documento="sconosciuto",
                numero_documento=rif[:80],
                data_documento=d_ex,
                totale_da_pagare=None,
                dare=Decimal("0"),
                avere=avere_mov,
                nome_file=nome_mov,
                testo_estratto="",
                metodo_estrazione="excel_riepilogo",
                note=note,
                riferimento_pagamento=rif[:160],
                causale_pagamento=(f"{doc_raw} — {desc}"[:220] if doc_raw else desc[:220]),
                importato_da=user,
            )
            obj.save()
            n_import += 1

    if n_import:
        ricalcola_saldi_progressivi(azienda.id)

    msgs.append(
        f"Bonifici importati: {n_import}; saltati: {n_skip}; "
        f"saltate righe solo proforma/parcella (colonna Documento o testo emissione, senza SEPA): {n_salt_doc}; "
        f"saltati duplicati o riga già importata: {n_salt_bon}."
        + (
            " Convenzione file rilevata: tutti gli importi non nulli in colonna sono negativi — "
            "importate solo le righe con importo negativo (avere = valore assoluto); importi positivi in colonna ignorati."
            if solo_importi_negativi
            else ""
        )
    )
    return msgs


def _excel_cell_for_json(val: object) -> object:
    """Valore serializzabile in JSONField (date/Decimal da openpyxl)."""
    if val is None:
        return None
    if isinstance(val, Decimal):
        return str(val)
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (bytes, bytearray, memoryview)):
        return None
    if isinstance(val, (str, int, float, bool)):
        return val
    return str(val)


def _cell_to_date(val) -> date | None:
    if val is None:
        return None
    if hasattr(val, "year") and hasattr(val, "month"):
        try:
            return date(int(val.year), int(val.month), int(val.day))
        except (ValueError, TypeError):
            pass
    if isinstance(val, str):
        for part in re.split(r"[\s;]+", val.strip()):
            d = _parse_date(part)
            if d:
                return d
    return None


def import_estratto_excel(
    fileobj,
    nome_file: str,
    azienda,
    user,
) -> tuple["ImportEstrattoContoStudio", list[str]]:
    """
    Import Excel con tracciamento per riga (ImportEstrattoContoStudio): uso avanzato e test automatici.

    Per import massivi da riga di comando usare il management command dedicato o chiamare
    ``import_riepilogo_bonifici_da_excel`` (stessa logica bonifici nel libro).
    """
    from django.core.files.base import ContentFile
    from django.db import transaction

    from openpyxl import load_workbook

    from .models import ImportEstrattoContoStudio, MovimentoRegistroStudioConsulente, RigaEstrattoContoStudio

    raw_bytes = fileobj.read()
    try:
        fileobj.seek(0)
    except (OSError, AttributeError, io.UnsupportedOperation):
        pass

    msgs: list[str] = []
    all_rows: list[tuple] = []
    header_idx: int | None = None
    headers_norm: list[str] = []
    header_labels: list[str] = []

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets[:8]:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            for i, row in enumerate(rows[:150]):
                if not row:
                    continue
                hdr_n = [_norm_excel_header_cell(h) for h in row]
                if not any(hdr_n):
                    continue
                ci_data = _riepilogo_pick_ci_data(hdr_n)
                ci_desc = _riepilogo_pick_ci_descrizione(hdr_n)
                ci_imp = _riepilogo_pick_ci_importo(hdr_n)
                if ci_data is not None and ci_desc is not None and ci_imp is not None:
                    header_idx = i
                    headers_norm = hdr_n
                    header_labels = [str(h).strip()[:60] if h is not None else "" for h in row]
                    all_rows = rows
                    break
            if header_idx is not None:
                break
    finally:
        wb.close()

    if not all_rows or header_idx is None:
        raise ValueError(
            "Intestazioni non trovate: servono colonne riconducibili a «Data», «Descrizione» e «Importo» "
            "(primi 150 righe del foglio; si cercano fino a 8 fogli). Opzionali: «Documento», CRO/TRN, saldo progressivo."
        )

    ci_data = _riepilogo_pick_ci_data(headers_norm)
    ci_doc = _riepilogo_pick_ci_documento(headers_norm)
    ci_desc = _riepilogo_pick_ci_descrizione(headers_norm)
    ci_imp = _riepilogo_pick_ci_importo(headers_norm)
    ci_rif_bank = _estratto_pick_ci_riferimento_bancario(headers_norm)

    def row_vals(r: tuple) -> dict[str, object]:
        out: dict[str, object] = {}
        for j, h in enumerate(header_labels):
            if not h:
                continue
            if j < len(r):
                out[h] = _excel_cell_for_json(r[j])
        return out

    data_rows = all_rows[header_idx + 1 :]
    stem = Path(nome_file or "estratto.xlsx").stem[:80]
    n_creati_bon = 0

    with transaction.atomic():
        imp_obj = ImportEstrattoContoStudio(
            azienda=azienda,
            nome_file=(nome_file or "estratto.xlsx")[:280],
            importato_da=user,
        )
        imp_obj.save()
        righe_create: list[RigaEstrattoContoStudio] = []
        agg = 0
        letti = 0
        for di, row in enumerate(data_rows):
            n_row = header_idx + 2 + di
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            desc = ""
            if ci_desc is not None and ci_desc < len(row) and row[ci_desc] is not None:
                desc = str(row[ci_desc]).strip()[:600]
            doc_raw = ""
            if ci_doc is not None and ci_doc < len(row) and row[ci_doc] is not None:
                doc_raw = str(row[ci_doc]).strip()[:200]
            rif_bank = ""
            if ci_rif_bank is not None and ci_rif_bank < len(row) and row[ci_rif_bank] is not None:
                rif_bank = str(row[ci_rif_bank]).strip()[:200]
            rif = rif_bank or doc_raw
            rif = rif[:200]
            imp_ex = None
            if ci_imp is not None and ci_imp < len(row) and row[ci_imp] is not None:
                imp_ex = _cell_to_decimal_excel(row[ci_imp])
            d_ex = None
            if ci_data is not None and ci_data < len(row):
                d_ex = _cell_to_date(row[ci_data])

            raw = row_vals(row)
            mov = None
            esito = "non_trovato"

            desc_match = f"{doc_raw} {desc}".strip()[:600] if doc_raw else desc
            strong_bonif = _estratto_riga_strong_bonifico(desc, rif_bank, doc_raw)
            bonif_lookup = (rif_bank or desc_match or desc or "").strip()

            if strong_bonif:
                mov = trova_movimento_bonifico_per_riferimento(azienda.id, bonif_lookup)
                if mov is None:
                    rif_key_lookup = _estrai_riferimento_bonifico_excel(desc, doc_raw, d_ex, imp_ex)
                    mov = trova_bonifico_esistente_stesso_excel(
                        azienda.id, d_ex, imp_ex, rif_key_lookup, desc
                    )
            else:
                if desc_match:
                    mov = trova_movimento_documento_per_descrizione(azienda.id, desc_match)

            if mov:
                esito = "agganciato"
                agg += 1
            elif (
                strong_bonif
                and imp_ex is not None
                and imp_ex > 0
                and not _excel_riga_solo_fattura_proforma_senza_bonifico(desc, doc_raw)
            ):
                rif_key = _estrai_riferimento_bonifico_excel(desc, doc_raw, d_ex, imp_ex)
                nome_mov = f"xlsx-estratto/{stem}/R{n_row}"[:280]
                existing_nome = MovimentoRegistroStudioConsulente.objects.filter(
                    azienda=azienda, nome_file=nome_mov, tipo_riga="bonifico"
                ).first()
                if existing_nome:
                    mov = existing_nome
                    esito = "agganciato"
                    agg += 1
                else:
                    mov_dup = trova_bonifico_esistente_stesso_excel(azienda.id, d_ex, imp_ex, rif_key, desc)
                    if mov_dup:
                        mov = mov_dup
                        esito = "agganciato"
                        agg += 1
                if (
                    mov is None
                    and strong_bonif
                    and imp_ex is not None
                    and imp_ex > 0
                    and not _excel_riga_solo_fattura_proforma_senza_bonifico(desc, doc_raw)
                    and not _bonifico_excel_già_presente(azienda.id, d_ex, imp_ex, rif_key, desc)
                ):
                    mov_collegato = trova_movimento_documento_per_colonna_documento(azienda.id, doc_raw)
                    if mov_collegato and getattr(mov_collegato.file, "name", None):
                        pdf_note = "PDF proforma/parcella collegato: presente."
                    elif mov_collegato:
                        pdf_note = "PDF proforma/parcella collegato: assente (documento in libro senza file)."
                    else:
                        pdf_note = (
                            "Nessun documento in libro con questo «Documento»."
                            if doc_raw
                            else "Colonna Documento vuota o non associata."
                        )
                    note = f"Bonifico da estratto conto «{stem}» (riga {n_row}). {pdf_note}"[:500]
                    mov = MovimentoRegistroStudioConsulente(
                        azienda=azienda,
                        tipo_riga="bonifico",
                        tipo_documento="sconosciuto",
                        numero_documento=rif_key[:80],
                        data_documento=d_ex,
                        totale_da_pagare=None,
                        dare=Decimal("0"),
                        avere=imp_ex,
                        nome_file=nome_mov,
                        testo_estratto="",
                        metodo_estrazione="excel_estratto_conto",
                        note=note,
                        riferimento_pagamento=rif_key[:160],
                        causale_pagamento=(f"{doc_raw} — {desc}"[:220] if doc_raw else desc[:220]),
                        importato_da=user,
                    )
                    mov.save()
                    esito = "agganciato"
                    agg += 1
                    n_creati_bon += 1
            elif not desc and not doc_raw and not rif_bank and imp_ex is None:
                esito = "saltato"
            letti += 1
            righe_create.append(
                RigaEstrattoContoStudio(
                    importazione=imp_obj,
                    indice_riga=n_row,
                    descrizione=desc,
                    importo_excel=imp_ex,
                    data_excel=d_ex,
                    riferimento_excel=rif,
                    celle_raw=raw,
                    movimento=mov,
                    esito_match=esito,
                )
            )

        RigaEstrattoContoStudio.objects.bulk_create(righe_create, batch_size=300)
        imp_obj.righe_lette = letti
        imp_obj.righe_agganciate = agg
        imp_obj.save(update_fields=["righe_lette", "righe_agganciate"])

    if raw_bytes:
        imp_obj.file.save((nome_file or "estratto.xlsx")[:280], ContentFile(raw_bytes), save=True)

    if n_creati_bon:
        ricalcola_saldi_progressivi(azienda.id)

    msg_finale = (
        f"Lette {imp_obj.righe_lette} righe; agganciate {imp_obj.righe_agganciate} a movimenti PDF/bonifici nel libro."
    )
    if n_creati_bon:
        msg_finale += f" Bonifici registrati automaticamente nel libro (nuovi movimenti): {n_creati_bon}."
    msgs.append(msg_finale)
    return imp_obj, msgs
