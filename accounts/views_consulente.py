"""
views_consulente.py — Interfaccia consulente del lavoro.

Funzionalità:
  1. Dashboard con statistiche rapide; elenco contratti su pagina dedicata
  2. Anagrafiche candidati (firmata_candidato / contratto_attivo) + export Excel/CSV
  3. Approva proposta di assunzione (firmata_candidato → contratto_attivo)
  4. Documenti dipendenti (identità, CF, attestati) + download
  5. Presenze dipendenti + export Excel (esistente) + export CSV
  6. Upload buste paga (singola + massiva per dipendente)
  7. Upload CUD (massivo per dipendente)

Accesso: solo utenti con ruolo 'consulente'.
Isolamento azienda: il consulente vede solo i dati di user.azienda.
"""
import calendar
import csv
import io
import os
import tempfile
import re
from datetime import date
from io import BytesIO
from pathlib import Path
from decimal import Decimal
from urllib.parse import quote

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
from django.conf import settings
from django.core.paginator import Paginator
from django.core.management import call_command
from django.http import HttpResponseForbidden, HttpResponse, FileResponse, Http404
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.db.models import Count, Q, Sum
from gesper_next_url import sanitize_internal_next
from django.db.models.functions import Coalesce

from anagrafiche.models import ComunicazioneRecessoProva, Dipendente
from documenti.models import Documento
from presenze.models import Presenza
from rapporto_di_lavoro.models import AddendumContrattuale, PropostaAssunzione, RapportoDiLavoro
from log_attivita.utils import registra_log
from log_attivita.anomalie import registra_evento_anomalia
from .models import MovimentoImportPaghe
import json
from functools import reduce
from operator import or_

from urllib.parse import urlencode

PDF_BUSTE_PASSWORD = 'DOLCEMASCOLO'


# ── Helpers permessi ─────────────────────────────────────────────────────────

def _is_consulente(user):
    return user.is_authenticated and user.has_ruolo('consulente')


def _is_admin_o_consulente_partitario(user):
    """Partitario studio consulente ↔ azienda: solo superuser, ruolo admin o consulente (non HR né altri)."""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    has_ruolo = getattr(user, 'has_ruolo', None)
    if not callable(has_ruolo):
        return False
    return has_ruolo('admin') or has_ruolo('consulente')


def _partitario_libro_link_admin_movimenti(user) -> bool:
    """Link modifica/elimina verso Django Admin sul libro: solo superuser o ruolo admin (non consulente)."""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    has_ruolo = getattr(user, 'has_ruolo', None)
    return callable(has_ruolo) and has_ruolo('admin')


def _get_azienda_partitario(request):
    """Azienda di contesto: sessione operativa per admin/superuser, FK consulente per consulente."""
    u = request.user
    if u.is_superuser or u.has_ruolo('admin'):
        from accounts.tenant import get_azienda_operativa

        return get_azienda_operativa(u, request.session)
    if u.has_ruolo('consulente'):
        return _get_azienda_consulente(u)
    return None


def _get_azienda_consulente(user):
    """Restituisce l'azienda associata al consulente (FK su User)."""
    return getattr(user, 'azienda', None)


def _bulk_pdf_meta_for_rapporti_consulente(rapporti):
    """
    Per ogni rapporto: id documento scansione cartacea, id «Contratto definitivo»,
    flag PDF caricato sul record, dati proposta collegata (PDF atto firmabile / firmato).
    Usato in elenco consulente per non mostrare solo il PDF generato da ReportLab.
    """
    from documenti.models import Documento

    if not rapporti:
        return {}

    q_parts = []
    for r in rapporti:
        did = getattr(r, 'dipendente_id', None)
        if not did:
            continue
        num = r.numero_contratto
        q_parts.append(Q(dipendente_id=did, descrizione=f'Contratto firmato cartaceo {num}'))
        q_parts.append(Q(dipendente_id=did, descrizione=f'Contratto definitivo {num}'))

    mapping = {}
    if q_parts:
        combined = reduce(or_, q_parts)
        for d in (
            Documento.objects.filter(tipo='contratto', visibile_al_dipendente=True)
            .filter(combined)
            .exclude(file='')
            .order_by('-data_caricamento', '-id')
        ):
            if not d.file or not getattr(d.file, 'name', None):
                continue
            k = (d.dipendente_id, (d.descrizione or '').strip())
            if k not in mapping:
                mapping[k] = d.id

    out = {}
    for r in rapporti:
        did = getattr(r, 'dipendente_id', None)
        num = r.numero_contratto
        dc = f'Contratto firmato cartaceo {num}'
        dd = f'Contratto definitivo {num}'
        meta = {
            'ha_file_su_rapporto': bool(
                getattr(r, 'file_contratto_pdf', None) and getattr(r.file_contratto_pdf, 'name', None)
            ),
            'id_doc_cartaceo': mapping.get((did, dc)) if did else None,
            'id_doc_definitivo': mapping.get((did, dd)) if did else None,
            'proposta_id': None,
            'proposta_firmata': False,
        }
        p = getattr(r, 'proposta_origine', None)
        if p is not None:
            meta['proposta_id'] = getattr(p, 'id', None)
            has_firma = bool(getattr(p, 'data_firma_candidato', None))
            _fn = getattr(p, 'is_firmata_da_candidato', None)
            is_firm = callable(_fn) and bool(_fn())
            meta['proposta_firmata'] = has_firma or is_firm
        out[r.pk] = meta
    return out


# Nomi mesi in italiano
MESI_ITA = [
    '', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
    'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre',
]
MESI_CHOICES = [(i, MESI_ITA[i]) for i in range(1, 13)]
GIORNI_ITA = ['Lun', 'Mar', 'Mer', 'Gio', 'Ven', 'Sab', 'Dom']


def _safe_display(obj, method_name, fallback_attr=''):
    """Restituisce il display di una choice Django senza dipendere dai metodi dinamici per il type checker."""
    method = getattr(obj, method_name, None)
    if callable(method):
        return method()
    return getattr(obj, fallback_attr, '')


def _safe_id(obj):
    """Restituisce l'id del model in modo sicuro per il type checker."""
    return getattr(obj, 'id', None)


def _safe_fk_id(obj, attr_name):
    """Restituisce il valore di una FK *_id in modo sicuro per il type checker."""
    return getattr(obj, attr_name, None)


# ── Dashboard ────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_dashboard(request):
    """Overview consulente: statistiche rapide."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        messages.error(request, "Nessuna azienda associata al tuo account. Contatta l'amministratore.")
        return redirect('home')

    oggi = date.today()
    dip_st = Dipendente.objects.filter(azienda=azienda, stato__in=('attivo', 'candidato')).aggregate(
        dipendenti_attivi=Count('id', filter=Q(stato='attivo')),
        candidati_count=Count('id', filter=Q(stato='candidato')),
    )
    dipendenti_attivi = dip_st['dipendenti_attivi'] or 0
    candidati_count = dip_st['candidati_count'] or 0
    candidati_recenti = Dipendente.objects.filter(
        azienda=azienda,
        stato='candidato',
    ).order_by('-id')[:8]

    # Per ogni candidato: posizione e livello dalla proposta più recente
    candidati_ids = [_safe_id(d) for d in candidati_recenti if _safe_id(d)]
    candidati_posizione = {}
    # Per ogni candidato recupera posizione/livello dalla proposta più recente
    for p in PropostaAssunzione.objects.filter(
        dipendente_id__in=candidati_ids,
    ).order_by('-id').values('dipendente_id', 'posizione', 'livello_ccnl'):
        did = p['dipendente_id']
        if did not in candidati_posizione:
            candidati_posizione[did] = {
                'posizione': p['posizione'] or '',
                'livello': p['livello_ccnl'] or '',
            }
    # Aggiungi ruolo dal dipendente se non disponibile dalla proposta
    # ed è un valore significativo (non numerico, non uguale allo stato)
    _stati_dip = {'candidato', 'attivo', 'cessato'}
    for dip in candidati_recenti:
        did = _safe_id(dip)
        if did is None:
            continue
        if did not in candidati_posizione:
            r = (getattr(dip, 'ruolo', '') or '').strip()
            if r and r.lower() not in _stati_dip and not r.isdigit():
                candidati_posizione[did] = {'posizione': r, 'livello': ''}
            else:
                candidati_posizione[did] = {'posizione': '', 'livello': ''}

    _stati_firmata_equiv = PropostaAssunzione.stati_equivalenti('firmata_candidato')
    proposte_da_approvare = PropostaAssunzione.objects.filter(
        azienda=azienda, stato__in=_stati_firmata_equiv
    ).count()

    # Documenti paghe/fisco: totali per azienda (stessi ambiti delle liste buste / F24 / CUD senza filtri),
    # così i numeri riflettono i PDF in archivio e non un sottoinsieme (es. solo mese solare corrente).
    doc_kpi = Documento.objects.filter(azienda=azienda).aggregate(
        buste_paga_count=Count('id', filter=Q(tipo='busta_paga')),
        f24_documenti_count=Count('id', filter=Q(tipo='altro', descrizione__icontains='F24')),
        cud_certificati_count=Count('id', filter=Q(tipo='certificato')),
    )
    buste_paga_count = doc_kpi['buste_paga_count'] or 0
    f24_documenti_count = doc_kpi['f24_documenti_count'] or 0
    cud_certificati_count = doc_kpi['cud_certificati_count'] or 0

    presenze_mese_count = Presenza.objects.filter(
        azienda=azienda,
        data__year=oggi.year,
        data__month=oggi.month,
    ).count()

    # Ultimi 5 documenti caricati dal consulente
    ultimi_docs = Documento.objects.filter(
        azienda=azienda,
        caricato_da=request.user,
    ).select_related('dipendente').order_by('-data_caricamento')[:5]

    from rapporto_di_lavoro.services_contratti import contratti_td_in_scadenza, contratti_td_scaduti_non_chiusi

    td_in_scadenza = len(contratti_td_in_scadenza(azienda))
    td_scaduti_aperti = len(contratti_td_scaduti_non_chiusi(azienda))

    contratti_registrati_count = RapportoDiLavoro.objects.filter(azienda=azienda).count()
    addenda_recenti = (
        AddendumContrattuale.objects.filter(rapporto__azienda=azienda)
        .select_related('rapporto', 'rapporto__dipendente', 'creato_da')
        .order_by('-data_creazione', '-id')[:10]
    )
    addendum_anno_count = AddendumContrattuale.objects.filter(
        rapporto__azienda=azienda,
        data_creazione__year=oggi.year,
    ).count()

    recesso_qs = (
        ComunicazioneRecessoProva.per_azienda(azienda)
        .filter(stato='in_verifica_consulente')
        .select_related('dipendente', 'rapporto')
        .order_by('-data_modifica')
    )
    recesso_prova_in_verifica_count = recesso_qs.count()
    recesso_prova_in_verifica = list(recesso_qs[:12])

    return render(request, 'consulente/dashboard.html', {
        'azienda': azienda,
        'dipendenti_attivi': dipendenti_attivi,
        'candidati_count': candidati_count,
        'candidati_recenti': candidati_recenti,
        'candidati_posizione': candidati_posizione,
        'proposte_da_approvare': proposte_da_approvare,
        'recesso_prova_in_verifica_count': recesso_prova_in_verifica_count,
        'recesso_prova_in_verifica': recesso_prova_in_verifica,
        'buste_paga_count': buste_paga_count,
        'f24_documenti_count': f24_documenti_count,
        'cud_certificati_count': cud_certificati_count,
        'presenze_mese_count': presenze_mese_count,
        'ultimi_docs': ultimi_docs,
        'mese_corrente': MESI_ITA[oggi.month],
        'mese_corrente_num': oggi.month,
        'anno_corrente': oggi.year,
        'td_in_scadenza': td_in_scadenza,
        'td_scaduti_aperti': td_scaduti_aperti,
        'contratti_registrati_count': contratti_registrati_count,
        'addenda_recenti': addenda_recenti,
        'addendum_anno_count': addendum_anno_count,
    })


@login_required
@user_passes_test(_is_consulente)
def consulente_contratti(request):
    """Elenco contratti / rapporti dell'azienda (tutti gli stati: bozza digitale, sottoscritto anche cartaceo, ecc.)."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        messages.error(request, "Nessuna azienda associata al tuo account. Contatta l'amministratore.")
        return redirect('home')

    qs = (
        RapportoDiLavoro.objects.filter(azienda=azienda)
        .select_related('dipendente', 'tipo_contratto', 'proposta_origine')
        .order_by('-data_modifica', '-id')
    )
    qstr = (request.GET.get('q') or '').strip()
    if qstr:
        qs = qs.filter(
            Q(numero_contratto__icontains=qstr)
            | Q(dipendente__nome__icontains=qstr)
            | Q(dipendente__cognome__icontains=qstr)
            | Q(posizione__icontains=qstr)
        )

    paginator = Paginator(qs, 40)
    page_obj = paginator.get_page(request.GET.get('page'))

    meta_map = _bulk_pdf_meta_for_rapporti_consulente(list(page_obj))
    for c in page_obj:
        c.pdf_meta = meta_map.get(c.id, {})

    oggi = date.today()
    return render(
        request,
        'consulente/contratti.html',
        {
            'azienda': azienda,
            'page_obj': page_obj,
            'q': qstr,
            'mese_corrente': MESI_ITA[oggi.month],
            'anno_corrente': oggi.year,
        },
    )


# ── Dettaglio proposta (read-only per consulente) ───────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_proposta_detail(request, proposta_id):
    """Visualizzazione read-only di una proposta di assunzione per il consulente."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        return HttpResponseForbidden("Nessuna azienda associata.")
    proposta = get_object_or_404(PropostaAssunzione, id=proposta_id, azienda=azienda)
    return render(request, 'consulente/proposta_detail.html', {
        'proposta': proposta,
        'azienda': azienda,
    })


@login_required
@user_passes_test(_is_consulente)
def consulente_proposta_pdf(request, proposta_id):
    """Genera e serve il PDF della proposta per il consulente."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        return HttpResponseForbidden("Nessuna azienda associata.")
    proposta = get_object_or_404(PropostaAssunzione, id=proposta_id, azienda=azienda)
    next_raw = (request.GET.get('next') or '').strip()
    if (request.GET.get('ui') == '1' or next_raw) and request.GET.get('embed') != '1':
        next_safe = sanitize_internal_next(request, next_raw)
        embed_src = request.build_absolute_uri(
            reverse('consulente_proposta_pdf', kwargs={'proposta_id': proposta_id}) + '?embed=1'
        )
        return render(
            request,
            'common/file_viewer_frame.html',
            {
                'titolo': f'Proposta {proposta.numero_proposta}',
                'embed_src': embed_src,
                'next_url': next_safe,
            },
        )
    try:
        from rapporto_di_lavoro.views import _genera_proposta_pdf, _proposta_context_extra
        extra = _proposta_context_extra(proposta)
        buffer = _genera_proposta_pdf(proposta, extra)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        nome_file = f"proposta_{proposta.numero_proposta}.pdf"
        response['Content-Disposition'] = f'inline; filename="{nome_file}"'
        return response
    except Exception as exc:
        return HttpResponse(f"Errore nella generazione del PDF: {exc}", status=500)


# ── Anagrafiche candidati ─────────────────────────────────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_candidati(request):
    """Lista candidati con proposta inviata, firmata o contratto attivo + export Excel/CSV."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        return HttpResponseForbidden("Nessuna azienda associata.")

    stato_filter = request.GET.get('stato', '')
    formato = request.GET.get('formato', '')

    _inv = PropostaAssunzione.stati_equivalenti('inviata_candidato')
    _firm = PropostaAssunzione.stati_equivalenti('firmata_candidato')
    _att = PropostaAssunzione.stati_equivalenti('contratto_attivo')
    STATI_VISIBILI = tuple(dict.fromkeys((*_inv, *_firm, *_att)))

    proposte_qs = PropostaAssunzione.objects.filter(
        azienda=azienda,
        stato__in=STATI_VISIBILI,
    ).select_related('dipendente', 'tipo_contratto').order_by(
        'dipendente__cognome', 'dipendente__nome'
    )

    if stato_filter == 'inviata_candidato':
        proposte_qs = proposte_qs.filter(stato__in=_inv)
    elif stato_filter == 'firmata_candidato':
        proposte_qs = proposte_qs.filter(stato__in=_firm)
    elif stato_filter == 'contratto_attivo':
        proposte_qs = proposte_qs.filter(stato__in=_att)

    cnt_inviata = PropostaAssunzione.objects.filter(azienda=azienda, stato__in=_inv).count()
    cnt_firmata = PropostaAssunzione.objects.filter(azienda=azienda, stato__in=_firm).count()
    cnt_attivo = PropostaAssunzione.objects.filter(azienda=azienda, stato__in=_att).count()

    if formato == 'csv':
        return _export_candidati_csv(proposte_qs, azienda)
    if formato == 'excel':
        return _export_candidati_excel(proposte_qs, azienda)

    paginator = Paginator(proposte_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    return render(request, 'consulente/candidati.html', {
        'azienda': azienda,
        'proposte': page_obj,
        'page_obj': page_obj,
        'stato_filter': stato_filter,
        'totale': cnt_inviata + cnt_firmata + cnt_attivo,
        'cnt_inviata': cnt_inviata,
        'cnt_firmata': cnt_firmata,
        'cnt_attivo': cnt_attivo,
    })


def _export_candidati_csv(proposte_qs, azienda):
    """Export CSV anagrafiche candidati con BOM UTF-8 per compatibilità Excel IT."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    nome = azienda.nome.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="candidati_{nome}.csv"'
    response.write('\ufeff')  # BOM UTF-8

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Cognome', 'Nome', 'Codice Fiscale', 'Data Nascita',
        'Posizione', 'Livello CCNL', 'Tipo Contratto',
        'Data Inizio', 'Data Fine', 'Stipendio Lordo Mensile',
        'Ore Settimanali', 'Stato', 'Data Firma Candidato',
    ])
    for p in proposte_qs:
        dip = p.dipendente
        writer.writerow([
            dip.cognome, dip.nome,
            dip.codice_fiscale or '',
            dip.data_nascita.strftime('%d/%m/%Y') if dip.data_nascita else '',
            p.posizione, p.livello_ccnl,
            p.tipo_contratto.nome if p.tipo_contratto else '',
            p.data_inizio_rapporto.strftime('%d/%m/%Y') if p.data_inizio_rapporto else '',
            p.data_fine_rapporto.strftime('%d/%m/%Y') if p.data_fine_rapporto else '',
            str(p.stipendio_lordo_mensile),
            str(p.ore_settimanali),
            p.get_stato_display(),
            p.data_firma_candidato.strftime('%d/%m/%Y %H:%M') if p.data_firma_candidato else '',
        ])
    return response


def _export_candidati_excel(proposte_qs, azienda):
    """Export Excel (.xlsx) anagrafiche candidati con stile."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non disponibile.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        return HttpResponse("Impossibile creare il file Excel.", status=500)
    ws.title = "Candidati"

    thin = Side(border_style='thin', color='CCCCCC')
    bordo = Border(left=thin, right=thin, top=thin, bottom=thin)
    al_c = Alignment(horizontal='center', vertical='center')
    al_l = Alignment(horizontal='left', vertical='center')

    intestazioni = [
        'Cognome', 'Nome', 'Cod. Fiscale', 'Data Nascita',
        'Posizione', 'Livello CCNL', 'Tipo Contratto',
        'Data Inizio', 'Data Fine', 'Stipendio Lordo (€)',
        'Ore/Sett.', 'Stato', 'Firma Candidato',
    ]
    fill_h = PatternFill('solid', fgColor='1b3a5f')
    for col, h in enumerate(intestazioni, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = fill_h
        c.alignment = al_c
        c.border = bordo
    ws.row_dimensions[1].height = 20

    fill_alt = PatternFill('solid', fgColor='f0f4f9')
    for row_n, p in enumerate(proposte_qs, 2):
        dip = p.dipendente
        riga = [
            dip.cognome, dip.nome,
            dip.codice_fiscale or '',
            dip.data_nascita.strftime('%d/%m/%Y') if dip.data_nascita else '',
            p.posizione, p.livello_ccnl,
            p.tipo_contratto.nome if p.tipo_contratto else '',
            p.data_inizio_rapporto.strftime('%d/%m/%Y') if p.data_inizio_rapporto else '',
            p.data_fine_rapporto.strftime('%d/%m/%Y') if p.data_fine_rapporto else '',
            float(p.stipendio_lordo_mensile),
            float(p.ore_settimanali),
            _safe_display(p, 'get_stato_display', 'stato'),
            p.data_firma_candidato.strftime('%d/%m/%Y %H:%M') if p.data_firma_candidato else '',
        ]
        fill_row = fill_alt if row_n % 2 == 0 else None
        for col, val in enumerate(riga, 1):
            c = ws.cell(row=row_n, column=col, value=val)
            c.border = bordo
            c.alignment = al_l
            if fill_row:
                c.fill = fill_row

    larghezze = [15, 15, 18, 14, 20, 14, 22, 12, 12, 18, 11, 22, 20]
    for i, w in enumerate(larghezze, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    nome = azienda.nome.replace(' ', '_')
    response = HttpResponse(
        out.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="candidati_{nome}.xlsx"'
    return response


# ── Approva proposta ─────────────────────────────────────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_approva_proposta(request, proposta_id):
    """Firma definitiva del datore: firmata_candidato → contratto_attivo."""
    azienda = _get_azienda_consulente(request.user)
    proposta = get_object_or_404(PropostaAssunzione, id=proposta_id, azienda=azienda)

    if not proposta.is_firmata_da_candidato():
        messages.error(
            request,
            f'La proposta deve essere in stato "Firmata dal candidato" (o equivalente legacy). '
            f"Stato attuale: {_safe_display(proposta, 'get_stato_display', 'stato')}."
        )
        return redirect('consulente_candidati')

    if request.method == 'POST':
        try:
            contratto = proposta.firma_definitiva_admin(request.user)
            messages.success(
                request,
                f'✅ Contratto {contratto.numero_contratto} emesso. '
                f'{proposta.dipendente} è ora un dipendente attivo.'
            )
            registra_log(
                utente=request.user,
                azienda=azienda,
                operazione='approva_proposta_consulente',
                descrizione=(
                    f'Consulente {request.user.username} ha approvato la proposta '
                    f'{proposta.numero_proposta} — contratto {contratto.numero_contratto}'
                ),
                request=request,
            )
        except Exception as exc:
            messages.error(request, f'Errore durante l\'approvazione: {exc}')
        return redirect('consulente_candidati')

    return render(request, 'consulente/approva_proposta.html', {
        'proposta': proposta,
        'azienda': azienda,
    })


# ── Documenti ────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_documenti(request):
    """Lista dipendenti con accesso ai loro documenti."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        return HttpResponseForbidden("Nessuna azienda associata.")

    dipendenti = Dipendente.objects.filter(
        azienda=azienda, stato__in=['attivo', 'candidato']
    ).order_by('cognome', 'nome')

    # Conta documenti personali per dipendente
    TIPI_PERSONALI = [
        'documento_identita', 'permesso_soggiorno', 'codice_fiscale_doc',
        'attestato', 'abilitazione', 'titolo_studio', 'certificazione', 'curriculum',
        'contratto',
    ]
    docs_count = {}
    for d in Documento.objects.filter(
        azienda=azienda, tipo__in=TIPI_PERSONALI,
        dipendente__in=dipendenti,
    ).values('dipendente_id'):
        docs_count[d['dipendente_id']] = docs_count.get(d['dipendente_id'], 0) + 1

    dipendenti_list = []
    for dip in dipendenti:
        dip_id = _safe_id(dip)
        dipendenti_list.append({
            'dip': dip,
            'num_documenti': docs_count.get(dip_id, 0) if dip_id is not None else 0,
        })

    return render(request, 'consulente/documenti.html', {
        'azienda': azienda,
        'dipendenti_list': dipendenti_list,
    })


@login_required
@user_passes_test(_is_consulente)
def consulente_documenti_dipendente(request, dipendente_id):
    """Documenti personali di un singolo dipendente (identità, CF, attestati)."""
    azienda = _get_azienda_consulente(request.user)
    dip = get_object_or_404(Dipendente, id=dipendente_id, azienda=azienda)

    TIPI_VISIBILI = [
        'documento_identita', 'permesso_soggiorno', 'codice_fiscale_doc',
        'attestato', 'abilitazione', 'titolo_studio', 'certificazione', 'curriculum',
        'contratto',
    ]
    documenti = Documento.objects.filter(
        dipendente=dip,
        tipo__in=TIPI_VISIBILI,
    ).order_by('-data_caricamento')

    gruppi = {}
    for d in documenti:
        gruppi.setdefault(_safe_display(d, 'get_tipo_display', 'tipo'), []).append(d)

    from rapporto_di_lavoro.services_contratti import posizione_contrattuale_per_dipendente

    posizioni_contrattuali = posizione_contrattuale_per_dipendente(dip)

    return render(request, 'consulente/documenti_dipendente.html', {
        'dip': dip,
        'gruppi': gruppi,
        'azienda': azienda,
        'posizioni_contrattuali': posizioni_contrattuali,
        'puo_registrare_addendum': False,
    })


# ── Presenze ─────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_presenze(request):
    """Selezione mese/anno + link export Excel (esistente) e CSV (nuovo)."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        return HttpResponseForbidden("Nessuna azienda associata.")

    oggi = date.today()

    def _parse_int(raw_value, default_value):
        raw = str(raw_value or default_value)
        normalized = raw.replace('.', '').replace(' ', '').replace(',', '')
        try:
            return int(normalized)
        except (TypeError, ValueError):
            return default_value

    anno = _parse_int(request.GET.get('anno', oggi.year), oggi.year)
    mese = _parse_int(request.GET.get('mese', oggi.month), oggi.month)
    if not (1 <= mese <= 12):
        mese = oggi.month

    dipendenti = Dipendente.objects.filter(
        azienda=azienda, stato__in=['attivo', 'candidato']
    ).order_by('cognome', 'nome')

    anni = list(range(oggi.year - 2, oggi.year + 2))

    return render(request, 'consulente/presenze.html', {
        'azienda': azienda,
        'dipendenti': dipendenti,
        'anno': anno,
        'mese': mese,
        'mesi': MESI_CHOICES,
        'anni': anni,
        'mese_nome': MESI_ITA[mese],
    })


@login_required
@user_passes_test(_is_consulente)
def consulente_presenze_export_csv(request):
    """Export CSV presenze mensili con BOM UTF-8 per Excel italiano."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        registra_evento_anomalia(
            utente=request.user,
            azienda=None,
            contesto='export_presenze_csv_consulente',
            anomalia={'codice': 'AZIENDA_MANCANTE', 'messaggio': 'Export CSV bloccato: azienda non trovata'},
            request=request,
        )
        return HttpResponse("Azienda non trovata.", status=400)

    oggi = date.today()

    def _parse_int(raw_value, default_value):
        raw = str(raw_value or default_value)
        normalized = raw.replace('.', '').replace(' ', '').replace(',', '')
        try:
            return int(normalized)
        except (TypeError, ValueError):
            return default_value

    anno = _parse_int(request.GET.get('anno', oggi.year), oggi.year)
    mese = _parse_int(request.GET.get('mese', oggi.month), oggi.month)
    if not (1 <= mese <= 12):
        mese = oggi.month
    dip_id = request.GET.get('dipendente', '')

    presenze_qs = Presenza.objects.filter(
        azienda=azienda, data__year=anno, data__month=mese
    ).select_related('dipendente').order_by('dipendente__cognome', 'dipendente__nome', 'data')

    if dip_id:
        presenze_qs = presenze_qs.filter(dipendente_id=dip_id)

    if not presenze_qs.exists():
        registra_evento_anomalia(
            utente=request.user,
            azienda=azienda,
            contesto='export_presenze_csv_consulente',
            anomalia={'codice': 'EXPORT_VUOTO', 'messaggio': f'Export CSV vuoto {mese:02d}/{anno}'},
            request=request,
        )
    else:
        registra_log(
            utente=request.user,
            azienda=azienda,
            operazione='export_presenze_csv_consulente',
            descrizione=f'Export presenze CSV {mese:02d}/{anno}',
            request=request,
        )

    nome_az = azienda.nome.replace(' ', '_')
    nome_file = f"presenze_{nome_az}_{anno}_{mese:02d}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{nome_file}"'
    response.write('\ufeff')  # BOM UTF-8

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Cognome', 'Nome', 'Data', 'Giorno', 'Causale',
        'Entrata 1', 'Uscita 1', 'Entrata 2', 'Uscita 2', 'Entrata 3', 'Uscita 3',
        'Ore Straordinario', 'Tipo Straordinario',
    ])
    for p in presenze_qs:
        d = p.data
        writer.writerow([
            p.dipendente.cognome, p.dipendente.nome,
            d.strftime('%d/%m/%Y'),
            GIORNI_ITA[d.weekday()],
            _safe_display(p, 'get_causale_display', 'causale') or '',
            p.ora_entrata.strftime('%H:%M') if p.ora_entrata else '',
            p.ora_uscita.strftime('%H:%M') if p.ora_uscita else '',
            p.ora_entrata2.strftime('%H:%M') if p.ora_entrata2 else '',
            p.ora_uscita2.strftime('%H:%M') if p.ora_uscita2 else '',
            p.ora_entrata3.strftime('%H:%M') if p.ora_entrata3 else '',
            p.ora_uscita3.strftime('%H:%M') if p.ora_uscita3 else '',
            str(p.ore_straordinario or ''),
            _safe_display(p, 'get_tipo_straordinario_display', 'tipo_straordinario') if p.tipo_straordinario else '',
        ])
    return response


# ── Hub caricamento documenti paghe (modulo consulente) ───────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_carica_documento(request):
    """Unico punto di ingresso verso upload buste paga e CUD."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        return HttpResponseForbidden('Nessuna azienda associata.')
    return render(
        request,
        'consulente/carica_documento.html',
        {'azienda': azienda},
    )


# ── Upload buste paga ─────────────────────────────────────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_upload_buste_paga(request):
    """Upload buste paga mensili: singola o massiva (un file per dipendente)."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        return HttpResponseForbidden("Nessuna azienda associata.")

    dipendenti = Dipendente.objects.filter(
        azienda=azienda, stato='attivo'
    ).order_by('cognome', 'nome')
    dipendenti_tutti = Dipendente.objects.filter(
        azienda=azienda
    ).order_by('cognome', 'nome')

    oggi = date.today()

    def _parse_int(raw_value, default_value):
        raw = str(raw_value or default_value)
        normalized = raw.replace('.', '').replace(' ', '').replace(',', '')
        try:
            return int(normalized)
        except (TypeError, ValueError):
            return default_value

    if request.method == 'POST':
        anno = _parse_int(request.POST.get('anno', oggi.year), oggi.year)
        mese = _parse_int(request.POST.get('mese', oggi.month), oggi.month)
        dipendente_filter = request.POST.get('dipendente', '').strip()
        if mese < 1 or mese > 12:
            mese = oggi.month
        mese_nome = MESI_ITA[mese]

        def _parse_decimal_text(value):
            if value in (None, ''):
                return None
            txt = str(value).strip().replace(' ', '').replace('.', '').replace(',', '.')
            try:
                return Decimal(txt).quantize(Decimal('0.01'))
            except Exception:
                return None

        def _extract_importi_from_pdf(uploaded_file):
            try:
                from pypdf import PdfReader
            except Exception:
                return None, None

            try:
                pos = uploaded_file.tell() if hasattr(uploaded_file, 'tell') else None
                if hasattr(uploaded_file, 'seek'):
                    uploaded_file.seek(0)
                reader = PdfReader(uploaded_file)
                if getattr(reader, 'is_encrypted', False):
                    reader.decrypt(PDF_BUSTE_PASSWORD)
                pages_text = []
                for p in reader.pages:
                    try:
                        pages_text.append(p.extract_text() or '')
                    except Exception:
                        pages_text.append('')
                text = '\n'.join(pages_text)
                if hasattr(uploaded_file, 'seek') and pos is not None:
                    uploaded_file.seek(pos)
            except Exception:
                return None, None

            up = (text or '').upper()
            amount_re = re.compile(r"([0-9]{1,3}(?:[\s\.]?[0-9]{3})*,\s*[0-9]{2}|[0-9]+,\s*[0-9]{2})")

            def _extract_by_labels(label_patterns):
                lines = [re.sub(r'\s+', ' ', ln).strip() for ln in up.splitlines()]
                for ln in lines:
                    if not ln:
                        continue
                    for pat in label_patterns:
                        m_lbl = re.search(pat, ln, re.IGNORECASE)
                        if not m_lbl:
                            continue
                        tail = ln[m_lbl.end():]
                        m_amount = amount_re.search(tail)
                        if m_amount:
                            return _parse_decimal_text(m_amount.group(1))
                for pat in label_patterns:
                    m_lbl = re.search(pat + r"([\s\S]{0,120})", up, re.IGNORECASE)
                    if not m_lbl:
                        continue
                    m_amount = amount_re.search(m_lbl.group(1))
                    if m_amount:
                        return _parse_decimal_text(m_amount.group(1))
                return None

            netto = _extract_by_labels([r"NETTO\s+BUSTA"])
            lordo = _extract_by_labels([r"TOTALE\s+LORDO", r"TOT\.?\s*LORDO", r"RETRIBUZIONE\s+LORDA"])

            return netto, lordo

        caricati = 0
        for dip in dipendenti:
            dip_id = _safe_id(dip)
            if dip_id is None:
                continue
            file_key = f'busta_paga_{dip_id}'
            file_obj = request.FILES.get(file_key)
            if file_obj:
                doc_busta = Documento.objects.create(
                    azienda=azienda,
                    dipendente=dip,
                    tipo='busta_paga',
                    descrizione=f'Busta paga {mese_nome} {anno}',
                    file=file_obj,
                    caricato_da=request.user,
                    caricato_dal_dipendente=False,
                    visibile_al_dipendente=True,
                )

                netto, lordo = _extract_importi_from_pdf(file_obj)
                MovimentoImportPaghe.objects.update_or_create(
                    azienda=azienda,
                    dipendente=dip,
                    tipo='BUSTA',
                    anno=anno,
                    mese=mese,
                    defaults={
                        'documento': doc_busta,
                        'importo': netto,
                        'importo_netto': netto,
                        'importo_lordo': lordo,
                        'cf_estratto': (dip.codice_fiscale or '')[:16],
                        'nominativo_estratto': f'{dip.cognome} {dip.nome}'.strip()[:160],
                        'periodo_label': f'{mese:02d}/{anno}',
                        'source_pdf': getattr(file_obj, 'name', '') or '',
                        'page_number': None,
                    },
                )
                caricati += 1
        if caricati > 0:
            messages.success(
                request,
                f'✅ Caricate {caricati} buste paga per {mese_nome} {anno}.'
            )
            registra_log(
                utente=request.user,
                azienda=azienda,
                operazione='upload_buste_paga_massivo',
                descrizione=f'Consulente ha caricato {caricati} buste paga ({mese_nome} {anno})',
                request=request,
            )
        else:
            messages.warning(request, 'Nessun file selezionato.')
        suffix = f'&dipendente={dipendente_filter}' if dipendente_filter else ''
        return redirect(f'{request.path}?anno={anno}&mese={mese}{suffix}')

    anno = _parse_int(request.GET.get('anno', oggi.year), oggi.year)
    mese = _parse_int(request.GET.get('mese', oggi.month), oggi.month)
    dipendente_filter = request.GET.get('dipendente', '').strip()
    dipendente_filter_int = None
    if dipendente_filter.isdigit():
        dipendente_filter_int = int(dipendente_filter)
    if mese < 1 or mese > 12:
        mese = oggi.month
    mese_nome = MESI_ITA[mese]

    def _match_period_descr(descrizione: str, month: int, year: int) -> bool:
        d = (descrizione or '').upper()
        month_name = MESI_ITA[month].upper()
        return (
            f'{month_name} {year}' in d
            or f'{month:02d}/{year}' in d
            or f'{month}/{year}' in d
        )

    # Buste già caricate per il mese selezionato
    buste_esistenti = {}
    buste_periodo = []
    for d in Documento.objects.filter(
        azienda=azienda,
        tipo='busta_paga',
        dipendente__isnull=False,
    ).select_related('dipendente'):
        if not _match_period_descr(getattr(d, 'descrizione', ''), mese, anno):
            continue
        dip_id = _safe_fk_id(d, 'dipendente_id')
        if dip_id is not None and (dipendente_filter_int is None or dip_id == dipendente_filter_int):
            buste_periodo.append(d)
        if dip_id is not None:
            buste_esistenti[dip_id] = d

    buste_periodo.sort(
        key=lambda x: (
            (x.dipendente.cognome if x.dipendente else ''),
            (x.dipendente.nome if x.dipendente else ''),
            x.data_caricamento,
        )
    )

    return render(request, 'consulente/upload_buste_paga.html', {
        'azienda': azienda,
        'dipendenti': dipendenti,
        'dipendenti_tutti': dipendenti_tutti,
        'anno': anno,
        'mese': mese,
        'mesi': MESI_CHOICES,
        'anni': list(range(oggi.year - 2, oggi.year + 1)),
        'mese_nome': mese_nome,
        'buste_esistenti': buste_esistenti,
        'buste_periodo': buste_periodo,
        'dipendente_filter': dipendente_filter,
    })


# ── Upload CUD ────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_upload_cud(request):
    """Alias legacy: usa il motore unico CUD di documenti."""
    anno = (request.GET.get('anno') or request.POST.get('anno') or '').strip()
    target = reverse('upload_cud_massivo')
    if anno:
        return redirect(f"{target}?anno={anno}")
    return redirect(target)


@login_required
@user_passes_test(_is_consulente)
def consulente_partitario_paghe(request):
    """Alias legacy: reindirizza il consulente alla stessa dashboard buste di admin/HR."""
    query = {
        'categoria': 'buste',
        'tipo': 'busta_paga',
        'anno': '',
        'dipendente': '',
    }

    anno = (request.GET.get('anno') or '').strip()
    dipendente = (request.GET.get('dipendente') or '').strip()
    if anno:
        query['anno'] = anno
    if dipendente:
        query['dipendente'] = dipendente

    return redirect(f"{reverse('lista_documenti')}?{urlencode(query)}")


@login_required
@user_passes_test(_is_consulente)
def consulente_riepilogo_f24_annuale(request):
    """Alias legacy: reindirizza il consulente alla stessa dashboard F24 di admin/HR."""
    query = {
        'tipo': 'altro',
        'anno': '',
        'dipendente': '',
    }

    anno = (request.GET.get('anno') or '').strip()
    if anno:
        query['anno'] = anno

    return redirect(f"{reverse('lista_documenti')}?{urlencode(query)}")


# ── Import PDF unico (buste + F24) ──────────────────────────────────────────

@login_required
@user_passes_test(_is_consulente)
def consulente_import_pdf_unico(request):
    """Importa uno o più PDF unici mensili: crea dipendenti mancanti e allega buste/F24."""
    azienda = _get_azienda_consulente(request.user)
    if not azienda:
        return HttpResponseForbidden("Nessuna azienda associata.")

    risultati = []

    if request.method == 'POST':
        files = request.FILES.getlist('pdf_files')
        if not files:
            messages.warning(request, 'Seleziona almeno un file PDF.')
            return redirect(request.path)

        snapshots_dir = Path(settings.BASE_DIR) / 'snapshots'
        snapshots_dir.mkdir(parents=True, exist_ok=True)

        for idx, up in enumerate(files, start=1):
            nome = getattr(up, 'name', f'file_{idx}.pdf')
            if not nome.lower().endswith('.pdf'):
                risultati.append({
                    'file': nome,
                    'ok': False,
                    'errore': 'Formato non supportato (solo PDF).',
                })
                continue

            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                    for chunk in up.chunks():
                        tmp.write(chunk)
                    tmp_path = tmp.name

                preview_out = snapshots_dir / f'preview_ui_{timezone.now().strftime("%Y%m%d_%H%M%S")}_{idx}.json'

                buff_prev = io.StringIO()
                call_command(
                    'preview_import_paghe_pdf',
                    tmp_path,
                    azienda_id=azienda.id,
                    out=str(preview_out),
                    stdout=buff_prev,
                )

                buff_imp = io.StringIO()
                call_command(
                    'import_paghe_pdf',
                    str(preview_out),
                    azienda_id=azienda.id,
                    apply=True,
                    attach_docs=True,
                    stdout=buff_imp,
                )

                preview_data = json.loads(preview_out.read_text(encoding='utf-8'))
                s = preview_data.get('summary', {})
                risultati.append({
                    'file': nome,
                    'ok': True,
                    'buste_uniche': s.get('buste_uniche', 0),
                    'matched': s.get('matched', 0),
                    'to_create': s.get('to_create', 0),
                    'already_present': s.get('already_present', 0),
                    'f24_pages': s.get('f24_pages', 0),
                    'report': str(preview_out),
                })
            except Exception as exc:
                risultati.append({
                    'file': nome,
                    'ok': False,
                    'errore': str(exc),
                })
            finally:
                if tmp_path:
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass

        ok_n = sum(1 for r in risultati if r.get('ok'))
        ko_n = len(risultati) - ok_n
        if ok_n:
            messages.success(request, f'✅ Import completato: {ok_n} file elaborati con successo.')
        if ko_n:
            messages.warning(request, f'⚠️ {ko_n} file non elaborati. Controlla il dettaglio sotto.')

    return render(request, 'consulente/import_pdf_unico.html', {
        'azienda': azienda,
        'risultati': risultati,
    })


# ── Posizione contabile consulente (proforma / pagamenti / libro) ─

SESSION_REPORT_AGGANCIA_DOCUMENTI = "report_aggancia_documenti_csv_v1"
SESSION_REPORT_AGGANCIA_BONIFICI = "report_aggancia_bonifici_csv_v1"


def _partitario_azienda_o_redirect(request):
    azienda = _get_azienda_partitario(request)
    if azienda:
        return azienda, None
    if request.user.is_superuser or request.user.has_ruolo('admin'):
        messages.warning(
            request,
            "Seleziona un'azienda operativa (profilo o elenco aziende) per la posizione contabile consulente.",
        )
        return None, redirect('lista_aziende')
    messages.error(request, "Nessuna azienda associata al tuo account consulente.")
    return None, redirect('home')


def _partitario_back(request):
    u = request.user
    if u.is_superuser or u.has_ruolo('admin'):
        return {'url_name': 'dashboard_admin', 'label': 'Dashboard admin'}
    return {'url_name': 'consulente_dashboard', 'label': 'Consulente'}


def _hub_saldi_posizione_contabile(azienda):
    """
    Saldo finale libro (dare − avere) e saldo «alla data» oggi (esclude movimenti con data documento futura).
    Convenzione come nel libro: saldo > 0 = residuo da pagare allo studio; saldo < 0 = credito azienda.
    """
    from datetime import date
    from decimal import Decimal

    from django.db.models import F

    from .models import MovimentoRegistroStudioConsulente

    def _fmt_it(val):
        try:
            n = Decimal(val or 0)
        except Exception:
            n = Decimal('0')
        return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    oggi = date.today()
    qs = list(
        MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda)
        .order_by(F('data_documento').asc(nulls_last=True), 'importato_il', 'id')
        .only('dare', 'avere', 'data_documento', 'saldo_progressivo')
    )
    if not qs:
        return {
            'ha_movimenti': False,
            'saldo_finale': Decimal('0'),
            'saldo_alla_data': Decimal('0'),
            'saldo_finale_fmt': _fmt_it(0),
            'saldo_alla_data_fmt': _fmt_it(0),
            'data_riferimento': oggi,
            'ultima_data_documento': None,
            'saldi_coincidono': True,
            'segno': 'zero',
        }

    running = Decimal('0')
    saldo_alla_data = Decimal('0')
    for row in qs:
        running += (row.dare or Decimal('0')) - (row.avere or Decimal('0'))
        d = row.data_documento
        if d is None or d <= oggi:
            saldo_alla_data = running

    saldo_finale = running
    dates = [r.data_documento for r in qs if r.data_documento]
    ultima = max(dates) if dates else None
    segno = 'positivo' if saldo_finale > 0 else ('negativo' if saldo_finale < 0 else 'zero')
    return {
        'ha_movimenti': True,
        'saldo_finale': saldo_finale,
        'saldo_alla_data': saldo_alla_data,
        'saldo_finale_fmt': _fmt_it(saldo_finale),
        'saldo_alla_data_fmt': _fmt_it(saldo_alla_data),
        'data_riferimento': oggi,
        'ultima_data_documento': ultima,
        'saldi_coincidono': saldo_alla_data == saldo_finale,
        'segno': segno,
    }


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_posizione_contabile(request):
    """Hub: Proforma, Pagamenti, Estratto conto, Libro movimenti."""
    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir

    from .models import MovimentoRegistroStudioConsulente

    n_doc = MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda, tipo_riga='documento').count()
    n_bon = MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda, tipo_riga='bonifico').count()
    hub_saldo = _hub_saldi_posizione_contabile(azienda)
    rep_doc = request.session.get(SESSION_REPORT_AGGANCIA_DOCUMENTI) or {}
    rep_bon = request.session.get(SESSION_REPORT_AGGANCIA_BONIFICI) or {}
    report_aggancia_csv_documenti = bool(rep_doc.get('azienda_id') == azienda.id and rep_doc.get('rows'))
    report_aggancia_csv_bonifici = bool(rep_bon.get('azienda_id') == azienda.id and rep_bon.get('rows'))
    return render(
        request,
        'consulente/posizione_contabile_hub.html',
        {
            'azienda': azienda,
            'partitario_back': _partitario_back(request),
            'n_documenti': n_doc,
            'n_bonifici': n_bon,
            'hub_saldo': hub_saldo,
            'posizione_nav': 'hub',
            'report_aggancia_csv_documenti': report_aggancia_csv_documenti,
            'report_aggancia_csv_bonifici': report_aggancia_csv_bonifici,
        },
    )


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_posizione_quadratura(request):
    """Quadratura euristica: ogni proforma/parcella vs bonifici collegati e saldi residui."""
    from .consulente_registro_studio import quadratura_proforma_parcelle_bonifici

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir
    quad = quadratura_proforma_parcelle_bonifici(azienda.id)
    return render(
        request,
        'consulente/posizione_contabile_quadratura.html',
        {
            'azienda': azienda,
            'partitario_back': _partitario_back(request),
            'posizione_nav': 'quadratura',
            'quad': quad,
        },
    )


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_piano_allocazione_bonifici(request):
    """Wizard: pool di bonifici → importi su documenti con residuo; salva ripartizione in quadratura."""
    from django.db.models import F

    from .consulente_registro_studio import (
        elimina_piano_allocazione_bonifici_quadratura,
        parse_importo_form,
        quadratura_proforma_parcelle_bonifici_anteprima_allocazione,
        salva_piano_allocazione_bonifici_quadratura,
    )
    from .models import MovimentoRegistroStudioConsulente, PianoAllocazioneBonificiQuad

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir

    def pool_total(ids: list[int]) -> Decimal:
        if not ids:
            return Decimal('0')
        rows = MovimentoRegistroStudioConsulente.objects.filter(
            pk__in=ids, azienda=azienda, tipo_riga='bonifico'
        )
        return sum((r.avere or Decimal('0')) for r in rows).quantize(Decimal('0.01'))

    bonifici_all = list(
        MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda, tipo_riga='bonifico').order_by(
            F('data_documento').desc(nulls_last=True), '-importato_il', '-id'
        )
    )
    piano_obj = PianoAllocazioneBonificiQuad.objects.filter(azienda=azienda).first()
    piano_count = len(piano_obj.righe) if piano_obj and piano_obj.righe else 0

    pool_ids: list[int] = []
    quad_anteprima = None
    step = 1

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'elimina_piano':
            elimina_piano_allocazione_bonifici_quadratura(azienda.id)
            messages.success(
                request,
                'Piano allocazione eliminato: i bonifici tornano ad essere abbinati solo dall’euristica testuale.',
            )
            return redirect('consulente_piano_allocazione_bonifici')
        if action == 'anteprima':
            id_set: set[int] = set()
            for v in request.POST.getlist('bon_sel'):
                try:
                    id_set.add(int(v))
                except ValueError:
                    pass
            sel = [b.pk for b in bonifici_all if b.pk in id_set]
            if not sel:
                messages.error(request, 'Selezionare almeno un bonifico.')
            else:
                pool_ids = sel
                step = 2
                quad_anteprima = quadratura_proforma_parcelle_bonifici_anteprima_allocazione(azienda.id, set(sel))
        elif action == 'salva_piano':
            raw = (request.POST.get('bon_ids_ordinati') or '').strip()
            ids_order: list[int] = []
            for part in raw.split(','):
                part = part.strip()
                if part.isdigit():
                    ids_order.append(int(part))
            ids_order = list(dict.fromkeys(ids_order))
            pairs: list[tuple[int, Decimal]] = []
            for k, v in request.POST.items():
                if not k.startswith('imp_doc_'):
                    continue
                tail = k.removeprefix('imp_doc_')
                if not tail.isdigit():
                    continue
                did = int(tail)
                imp = parse_importo_form((v or '').strip())
                if imp is not None and imp > 0:
                    pairs.append((did, imp))
            if not ids_order:
                messages.error(request, 'Sessione non valida: tornare allo step bonifici.')
            elif not pairs:
                messages.error(request, 'Indicare almeno un importo da imputare su una riga documento.')
            else:
                try:
                    salva_piano_allocazione_bonifici_quadratura(azienda, ids_order, pairs, request.user)
                except ValueError as exc:
                    messages.error(request, str(exc))
                    pool_ids = ids_order
                    step = 2
                    quad_anteprima = quadratura_proforma_parcelle_bonifici_anteprima_allocazione(
                        azienda.id, set(ids_order)
                    )
                else:
                    messages.success(
                        request,
                        'Piano salvato: in Quadrature le ripartizioni manuali si sommano all’euristica sui bonifici non inclusi nel piano.',
                    )
                    return redirect('consulente_piano_allocazione_bonifici')

    return render(
        request,
        'consulente/piano_allocazione_bonifici.html',
        {
            'azienda': azienda,
            'partitario_back': _partitario_back(request),
            'posizione_nav': 'piano_bonifici',
            'bonifici_all': bonifici_all,
            'pool_ids': pool_ids,
            'pool_total': pool_total(pool_ids) if pool_ids else Decimal('0'),
            'quad_anteprima': quad_anteprima,
            'step': step,
            'piano_count': piano_count,
        },
    )


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_posizione_proforma(request):
    from django.db.models import F

    from .consulente_registro_studio import (
        applica_aggancia_pdf_proforma_parcelle_a_libro,
        applica_inserimento_manuale_proforma_parcella,
        applica_upload_proforma_parcelle_pdf,
        parse_importo_form,
    )
    from .models import MovimentoRegistroStudioConsulente

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'inserimento_manuale_documento':
            raw_data = (request.POST.get('data_documento') or '').strip()
            try:
                d_doc = date.fromisoformat(raw_data) if raw_data else None
            except ValueError:
                d_doc = None
            tipo_doc = (request.POST.get('tipo_documento') or '').strip().lower()
            numero = (request.POST.get('numero_documento') or '').strip()
            imp = parse_importo_form(request.POST.get('importo_da_pagare') or '')
            if d_doc is None:
                messages.error(request, 'Data documento non valida (usare il selettore data o formato AAAA-MM-GG).')
            elif imp is None or imp <= 0:
                messages.error(request, 'Importo da pagare non valido o non indicato.')
            else:
                for msg in applica_inserimento_manuale_proforma_parcella(
                    azienda,
                    request.user,
                    tipo_documento=tipo_doc,
                    numero_documento=numero,
                    data_documento=d_doc,
                    importo_contabile=imp,
                ):
                    if msg.lower().startswith('già presente'):
                        messages.warning(request, msg)
                    elif msg.lower().startswith('registrato'):
                        messages.success(request, msg)
                    else:
                        messages.error(request, msg)
            return _redirect_posizione_con_querystring(request, 'consulente_posizione_proforma')
        uploads_agg = request.FILES.getlist('pdf_aggancia')
        if action == 'aggancia_pdf_documenti' and uploads_agg:
            msgs, report_rows = applica_aggancia_pdf_proforma_parcelle_a_libro(azienda, request.user, uploads_agg)
            request.session[SESSION_REPORT_AGGANCIA_DOCUMENTI] = {
                'azienda_id': azienda.id,
                'rows': report_rows,
            }
            for msg in msgs:
                low = msg.lower()
                if msg.startswith('Agganciati') or ': allegato a movimento' in msg.lower():
                    messages.success(request, msg)
                elif ' file con errori' in low:
                    messages.warning(request, msg)
                elif (
                    'saltato' in low
                    or 'nessun movimento' in low
                    or 'non estratto' in low
                    or 'data documento non estratta' in low
                    or 'più movimenti senza data' in low
                    or 'più movimenti in libro' in low
                    or 'già un movimento' in low
                ):
                    messages.warning(request, msg)
                elif ': ' in msg and 'allegato a movimento' not in msg.lower():
                    messages.error(request, msg)
                else:
                    messages.info(request, msg)
            return _redirect_posizione_con_querystring(request, 'consulente_posizione_proforma')
        uploads = request.FILES.getlist('pdf')
        if uploads:
            for msg in applica_upload_proforma_parcelle_pdf(azienda, request.user, uploads):
                low = msg.lower()
                if msg.startswith('Importati'):
                    messages.success(request, msg)
                elif ' file con errori' in low:
                    messages.warning(request, msg)
                elif 'ignorato' in low or 'già presente' in low:
                    messages.warning(request, msg)
                elif ':' in msg and not msg.startswith('Importati'):
                    messages.error(request, msg)
                else:
                    messages.info(request, msg)
        return _redirect_posizione_con_querystring(request, 'consulente_posizione_proforma')
    from django.db.models.functions import ExtractYear

    filter_params = _libro_filter_params_from_request(request)
    base_righe = (
        MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda, tipo_riga='documento')
        .select_related('importato_da')
        .order_by(F('data_documento').desc(nulls_last=True), '-importato_il')
    )
    righe = _filter_movimenti_qs_by_data_documento(base_righe, filter_params)
    anni_disponibili = sorted(
        {
            y
            for y in MovimentoRegistroStudioConsulente.objects.filter(
                azienda=azienda, data_documento__isnull=False, tipo_riga='documento'
            )
            .annotate(y=ExtractYear('data_documento'))
            .values_list('y', flat=True)
        },
        reverse=True,
    )
    rep_doc = request.session.get(SESSION_REPORT_AGGANCIA_DOCUMENTI) or {}
    report_aggancia_csv_documenti = bool(rep_doc.get('azienda_id') == azienda.id and rep_doc.get('rows'))
    return render(
        request,
        'consulente/posizione_contabile_proforma.html',
        {
            'azienda': azienda,
            'righe': righe,
            'partitario_back': _partitario_back(request),
            'posizione_nav': 'proforma',
            'report_aggancia_csv_documenti': report_aggancia_csv_documenti,
            'libro_filter': filter_params,
            'anni_disponibili': anni_disponibili,
        },
    )


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
@require_POST
def consulente_proforma_allega_pdf_movimento(request, movimento_id: int):
    from .consulente_registro_studio import applica_pdf_su_movimento_documento

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir
    up = request.FILES.get("pdf")
    if not up:
        messages.error(request, "Seleziona un file PDF da allegare.")
        return _redirect_posizione_con_filtri_tabella_post(request, 'consulente_posizione_proforma')
    for msg in applica_pdf_su_movimento_documento(azienda, request.user, movimento_id, up):
        low = msg.lower()
        if (
            "pdf allegato con successo" in low
            or "pdf allegato al movimento" in low
            or "pdf documento sostituito" in low
        ):
            messages.success(request, msg)
        elif "non corrisponde" in low or "annullato" in low or "ambiguo" in low:
            messages.warning(request, msg)
        else:
            messages.error(request, msg)
    return _redirect_posizione_con_filtri_tabella_post(request, 'consulente_posizione_proforma')


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
@require_POST
def consulente_pagamenti_allega_pdf_movimento(request, movimento_id: int):
    from .consulente_registro_studio import applica_pdf_su_movimento_bonifico

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir
    up = request.FILES.get("pdf")
    if not up:
        messages.error(request, "Seleziona un file PDF da allegare.")
        return _redirect_posizione_con_filtri_tabella_post(request, 'consulente_posizione_pagamenti')
    for msg in applica_pdf_su_movimento_bonifico(azienda, request.user, movimento_id, up):
        low = msg.lower()
        if (
            "pdf allegato con successo" in low
            or "pdf allegato a bonifico" in low
            or "allegata con successo al bonifico" in low
            or "sostituita per il bonifico" in low
        ):
            messages.success(request, msg)
        elif "non corrisponde" in low or "annullato" in low:
            messages.warning(request, msg)
        else:
            messages.error(request, msg)
    return _redirect_posizione_con_filtri_tabella_post(request, 'consulente_posizione_pagamenti')


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
@require_POST
def consulente_pagamenti_rimuovi_pdf_movimento(request, movimento_id: int):
    from .consulente_registro_studio import ricalcola_saldi_progressivi
    from .models import MovimentoRegistroStudioConsulente

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir
    mov = MovimentoRegistroStudioConsulente.objects.filter(
        pk=movimento_id, azienda=azienda, tipo_riga="bonifico"
    ).first()
    if mov is None:
        messages.error(request, "Movimento non trovato.")
    elif not getattr(mov.file, "name", None):
        messages.warning(request, "Questa riga non ha un PDF allegato.")
    else:
        try:
            mov.file.delete(save=False)
        except OSError:
            pass
        mov.file = None
        mov.save(update_fields=["file"])
        ricalcola_saldi_progressivi(azienda.id)
        messages.success(request, "PDF distinta rimosso dal bonifico.")
    return _redirect_posizione_con_filtri_tabella_post(request, 'consulente_posizione_pagamenti')


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
@require_POST
def consulente_proforma_rimuovi_pdf_movimento(request, movimento_id: int):
    from .consulente_registro_studio import ricalcola_saldi_progressivi
    from .models import MovimentoRegistroStudioConsulente

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir
    mov = MovimentoRegistroStudioConsulente.objects.filter(
        pk=movimento_id, azienda=azienda, tipo_riga="documento"
    ).first()
    if mov is None:
        messages.error(request, "Movimento non trovato.")
    elif not getattr(mov.file, "name", None):
        messages.warning(request, "Questa riga non ha un PDF allegato.")
    else:
        try:
            mov.file.delete(save=False)
        except OSError:
            pass
        mov.file = None
        mov.save(update_fields=["file"])
        ricalcola_saldi_progressivi(azienda.id)
        messages.success(request, "PDF rimosso dal documento.")
    return _redirect_posizione_con_filtri_tabella_post(request, 'consulente_posizione_proforma')


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_report_aggancia_csv(request):
    """Scarica CSV dell'ultimo aggancio massivo PDF (documenti o bonifici) per l'azienda in sessione partitario."""
    from .consulente_registro_studio import render_csv_report_aggancia_bonifici, render_csv_report_aggancia_documenti

    tipo = (request.GET.get('tipo') or '').strip().lower()
    if tipo not in ('documenti', 'bonifici'):
        return HttpResponse('Parametro tipo=documenti|bonifici obbligatorio.', status=400, content_type='text/plain; charset=utf-8')

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir

    key = SESSION_REPORT_AGGANCIA_DOCUMENTI if tipo == 'documenti' else SESSION_REPORT_AGGANCIA_BONIFICI
    payload = request.session.get(key) or {}
    rows = payload.get('rows') or []
    if payload.get('azienda_id') != azienda.id or not rows:
        messages.warning(
            request,
            'Nessun report CSV disponibile per questa azienda: eseguire prima un aggancio massivo PDF nella pagina corrispondente.',
        )
        if tipo == 'documenti':
            return redirect('consulente_posizione_proforma')
        return redirect('consulente_posizione_pagamenti')

    if tipo == 'documenti':
        body = render_csv_report_aggancia_documenti(rows)
        fname = f"report_aggancia_documenti_{timezone.now().strftime('%Y%m%d_%H%M')}.csv"
    else:
        body = render_csv_report_aggancia_bonifici(rows)
        fname = f"report_aggancia_bonifici_{timezone.now().strftime('%Y%m%d_%H%M')}.csv"

    resp = HttpResponse("\ufeff" + body, content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_posizione_pagamenti(request):
    import uuid
    from datetime import datetime as dt_mod
    from decimal import Decimal

    from django.db.models import F

    from .consulente_registro_studio import (
        applica_aggancia_pdf_bonifici_a_libro,
        applica_upload_bonifici_pdf,
        bonifico_duplicato_elenco_ids,
        documenti_con_residuo_quadratura_per_select,
        parse_importo_form,
        ricalcola_saldi_progressivi,
        riferimento_pipe_aggancio_bonifico_documento,
    )
    from .models import MovimentoRegistroStudioConsulente

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        uploads_bon_agg = request.FILES.getlist('pdf_aggancia_bonifici')
        if action == 'aggancia_pdf_bonifici' and uploads_bon_agg:
            msgs, report_rows = applica_aggancia_pdf_bonifici_a_libro(azienda, request.user, uploads_bon_agg)
            request.session[SESSION_REPORT_AGGANCIA_BONIFICI] = {
                'azienda_id': azienda.id,
                'rows': report_rows,
            }
            for msg in msgs:
                low = msg.lower()
                if msg.startswith('Agganciati') or ': pdf allegato a bonifico' in msg.lower():
                    messages.success(request, msg)
                elif ' file con errori' in low:
                    messages.warning(request, msg)
                elif 'saltato' in low or 'nessun bonifico' in low or 'non estratto' in low or 'già un pdf' in low or 'già pdf' in low:
                    messages.warning(request, msg)
                elif ': ' in msg and 'pdf allegato a bonifico' not in msg.lower():
                    messages.error(request, msg)
                else:
                    messages.info(request, msg)
            return _redirect_posizione_con_querystring(request, 'consulente_posizione_pagamenti')
        if action == 'aggancia_bonifico_a_documento':
            try:
                bid = int((request.POST.get("bonifico_id") or "0").strip())
                did = int((request.POST.get("documento_id") or "0").strip())
            except ValueError:
                messages.error(request, "Dati non validi.")
                return _redirect_posizione_con_filtri_tabella_post(request, "consulente_posizione_pagamenti")
            if did <= 0:
                messages.error(request, "Selezionare una proforma o parcella con residuo da incassare.")
                return _redirect_posizione_con_filtri_tabella_post(request, "consulente_posizione_pagamenti")
            bon = get_object_or_404(
                MovimentoRegistroStudioConsulente,
                pk=bid,
                azienda=azienda,
                tipo_riga="bonifico",
            )
            doc = get_object_or_404(
                MovimentoRegistroStudioConsulente,
                pk=did,
                azienda=azienda,
                tipo_riga="documento",
            )
            consentiti = {o["id"] for o in documenti_con_residuo_quadratura_per_select(azienda.id)}
            if doc.pk not in consentiti:
                messages.warning(
                    request,
                    "Il documento selezionato non risulta con residuo da incassare in quadratura (già saldato o non in elenco). Aggiorna la pagina e riprova.",
                )
                return _redirect_posizione_con_filtri_tabella_post(request, "consulente_posizione_pagamenti")
            try:
                rif = riferimento_pipe_aggancio_bonifico_documento(bon, doc)
            except ValueError as exc:
                messages.error(request, str(exc))
                return _redirect_posizione_con_filtri_tabella_post(request, "consulente_posizione_pagamenti")
            bon.riferimento_pagamento = rif
            bon.save(update_fields=["riferimento_pagamento"])
            ricalcola_saldi_progressivi(azienda.id)
            messages.success(
                request,
                f"Aggancio salvato: bonifico id {bon.pk} collegato a {doc.get_tipo_documento_display()} "
                f"n. {(doc.numero_documento or '—').strip()}. Controllare in Quadrature.",
            )
            return _redirect_posizione_con_filtri_tabella_post(request, "consulente_posizione_pagamenti")
        if action == 'aggiungi_bonifico':
            data_raw = (request.POST.get('data_valuta') or '').strip()
            rif = (request.POST.get('riferimento_pagamento') or '').strip()
            caus = (request.POST.get('causale_pagamento') or '').strip()
            imp_raw = (request.POST.get('importo_avere') or '').strip()
            data_doc = None
            try:
                data_doc = dt_mod.strptime(data_raw, '%Y-%m-%d').date()
            except ValueError:
                pass
            imp = parse_importo_form(imp_raw)
            pdf = request.FILES.get('pdf_bonifico')
            if not data_doc:
                messages.error(request, 'Indicare una data valuta valida.')
            elif not rif or len(rif) < 3:
                messages.error(request, 'Indicare un riferimento pagamento (es. CRO, TRN, ordinativo).')
            elif imp is None or imp <= 0:
                messages.error(request, 'Indicare un importo in avere maggiore di zero.')
            else:
                from .consulente_registro_studio import trova_bonifico_esistente_stesso_excel

                dup_b = trova_bonifico_esistente_stesso_excel(azienda.id, data_doc, imp, rif, caus)
                if dup_b is not None:
                    messages.warning(
                        request,
                        'Esiste già un bonifico con la stessa data, lo stesso importo in avere e un riferimento o causale '
                        f'compatibile (mov. id {dup_b.pk}); operazione annullata.',
                    )
                    return _redirect_posizione_con_querystring(request, 'consulente_posizione_pagamenti')
                nome_sint = (pdf.name if pdf else None) or f"bonifico-{data_doc.isoformat()}-{uuid.uuid4().hex[:10]}"
                nome_sint = nome_sint[:280]
                obj = MovimentoRegistroStudioConsulente(
                    azienda=azienda,
                    tipo_riga='bonifico',
                    tipo_documento='sconosciuto',
                    numero_documento=rif[:80],
                    data_documento=data_doc,
                    totale_da_pagare=None,
                    dare=Decimal('0'),
                    avere=imp,
                    nome_file=nome_sint,
                    riferimento_pagamento=rif[:160],
                    causale_pagamento=caus[:220],
                    importato_da=request.user,
                )
                obj.save()
                if pdf:
                    from django.core.files import File

                    if hasattr(pdf, 'seek'):
                        pdf.seek(0)
                    obj.file.save(pdf.name[:200], File(pdf), save=True)
                ricalcola_saldi_progressivi(azienda.id)
                messages.success(request, 'Bonifico registrato in avere.')
            return _redirect_posizione_con_querystring(request, 'consulente_posizione_pagamenti')
        if action == 'solo_pdf_bonifico':
            from .consulente_registro_studio import applica_upload_bonifici_pdf

            solo = request.FILES.get('pdf_bonifico_solo')
            if not solo:
                messages.error(request, 'Seleziona un PDF distinta (importo e riferimento leggibili dal testo).')
            else:
                for msg in applica_upload_bonifici_pdf(azienda, request.user, [solo]):
                    low = msg.lower()
                    if 'registrati' in low and 'bonifici da pdf' in low:
                        messages.success(request, msg)
                    elif (
                        'non importati' in low
                        or 'non rilevato' in low
                        or 'assente' in low
                        or 'usare inserimento' in low
                        or 'riferimento assente' in low
                        or 'importo non rilevato' in low
                        or 'ignorati perché equivalenti' in low
                        or 'importazione ignorata' in low
                    ):
                        messages.warning(request, msg)
                    else:
                        messages.info(request, msg)
            return _redirect_posizione_con_querystring(request, 'consulente_posizione_pagamenti')
        uploads = request.FILES.getlist('pdf')
        if uploads:
            for msg in applica_upload_bonifici_pdf(azienda, request.user, uploads):
                if 'non importati' in msg or 'non rilevato' in msg.lower() or 'assente' in msg.lower():
                    messages.warning(request, msg)
                else:
                    messages.success(request, msg)
            return _redirect_posizione_con_querystring(request, 'consulente_posizione_pagamenti')
    from django.db.models.functions import ExtractYear

    filter_params = _libro_filter_params_from_request(request)
    base_righe = (
        MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda, tipo_riga='bonifico')
        .select_related('importato_da')
        .order_by(F('data_documento').desc(nulls_last=True), '-importato_il')
    )
    righe = list(_filter_movimenti_qs_by_data_documento(base_righe, filter_params))
    bonifico_duplicato_ids = bonifico_duplicato_elenco_ids(righe)
    anni_disponibili = sorted(
        {
            y
            for y in MovimentoRegistroStudioConsulente.objects.filter(
                azienda=azienda, data_documento__isnull=False, tipo_riga='bonifico'
            )
            .annotate(y=ExtractYear('data_documento'))
            .values_list('y', flat=True)
        },
        reverse=True,
    )
    rep_bon = request.session.get(SESSION_REPORT_AGGANCIA_BONIFICI) or {}
    report_aggancia_csv_bonifici = bool(rep_bon.get('azienda_id') == azienda.id and rep_bon.get('rows'))
    documenti_residuo_select = documenti_con_residuo_quadratura_per_select(azienda.id)
    return render(
        request,
        'consulente/posizione_contabile_pagamenti.html',
        {
            'azienda': azienda,
            'righe': righe,
            'bonifico_duplicato_ids': bonifico_duplicato_ids,
            'partitario_back': _partitario_back(request),
            'posizione_nav': 'pagamenti',
            'report_aggancia_csv_bonifici': report_aggancia_csv_bonifici,
            'libro_filter': filter_params,
            'anni_disponibili': anni_disponibili,
            'documenti_residuo_select': documenti_residuo_select,
        },
    )


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_posizione_libro(request):
    from .consulente_registro_studio import ricalcola_saldi_progressivi
    from .models import MovimentoRegistroStudioConsulente

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'ricalcola_saldi':
            ricalcola_saldi_progressivi(azienda.id)
            messages.success(request, 'Saldi progressivi ricalcolati.')
        elif action == 'rileggi_totali_pdf':
            from .consulente_registro_studio import ricalcola_totali_documenti_da_testo_estratto

            res = ricalcola_totali_documenti_da_testo_estratto(azienda.id)
            messages.success(request, res['message'])
        base = reverse('consulente_posizione_libro')
        q = request.GET.urlencode()
        return redirect(f'{base}?{q}' if q else base)
    from collections import Counter

    from django.db.models.functions import ExtractYear

    filter_params = _libro_filter_params_from_request(request)
    righe_full = list(_qs_libro_movimenti_azienda(azienda, None))
    doc_pairs = [
        (r.id, (r.numero_documento or "").strip().lower(), r.tipo_documento, r.data_documento)
        for r in righe_full
        if r.tipo_riga == "documento" and (r.numero_documento or "").strip()
    ]
    cnt = Counter((num, tipo, data) for _pk, num, tipo, data in doc_pairs)
    dup_keys = {k for k, n in cnt.items() if n > 1}
    documento_duplicato_ids = {pk for pk, num, tipo, data in doc_pairs if (num, tipo, data) in dup_keys}
    righe = list(_qs_libro_movimenti_azienda(azienda, filter_params))
    anni_disponibili = sorted(
        {
            y
            for y in MovimentoRegistroStudioConsulente.objects.filter(
                azienda=azienda, data_documento__isnull=False
            )
            .annotate(y=ExtractYear('data_documento'))
            .values_list('y', flat=True)
        },
        reverse=True,
    )
    filtro_attivo = bool(
        (filter_params.get('anno') or '').strip()
        or (filter_params.get('data_da') or '').strip()
        or (filter_params.get('data_a') or '').strip()
    )
    for r in righe:
        r.nota_libro_display = _libro_nota_consulente_display(getattr(r, 'note', None) or '')
    return render(
        request,
        'consulente/posizione_contabile_libro.html',
        {
            'azienda': azienda,
            'righe': righe,
            'documento_duplicato_ids': documento_duplicato_ids,
            'partitario_back': _partitario_back(request),
            'posizione_nav': 'libro',
            'libro_filter': filter_params,
            'anni_disponibili': anni_disponibili,
            'filtro_attivo': filtro_attivo,
            'libro_link_admin_movimenti': _partitario_libro_link_admin_movimenti(request.user),
        },
    )


def _libro_filter_params_from_request(request) -> dict[str, str]:
    from .formatting import normalize_anno_calendario

    raw_anno = (request.GET.get('anno') or '').strip()
    anno = normalize_anno_calendario(raw_anno) if raw_anno else ''
    return {
        'anno': anno,
        'data_da': (request.GET.get('data_da') or '').strip(),
        'data_a': (request.GET.get('data_a') or '').strip(),
    }


def _redirect_posizione_con_querystring(request, url_name: str):
    """Reindirizza mantenendo anno/data_da/data_a in querystring (filtri elenco)."""
    base = reverse(url_name)
    q = request.GET.urlencode()
    return redirect(f'{base}?{q}' if q else base)


def _redirect_posizione_con_filtri_tabella_post(request, url_name: str):
    """Dopo POST (es. allega PDF riga): redirect con filtri da campi hidden anno/data_da/data_a."""
    from .formatting import normalize_anno_calendario

    params = {}
    for k in ('anno', 'data_da', 'data_a'):
        v = (request.POST.get(k) or '').strip()
        if not v:
            continue
        if k == 'anno':
            v = normalize_anno_calendario(v)
            if not v:
                continue
        params[k] = v
    base = reverse(url_name)
    q = urlencode(params)
    return redirect(f'{base}?{q}' if q else base)


def _filter_movimenti_qs_by_data_documento(qs, filter_params: dict[str, str] | None):
    if not filter_params:
        return qs
    anno = (filter_params.get('anno') or '').strip()
    data_da = (filter_params.get('data_da') or '').strip()
    data_a = (filter_params.get('data_a') or '').strip()
    if anno.isdigit() and len(anno) == 4:
        qs = qs.filter(data_documento__year=int(anno))
    if data_da:
        try:
            qs = qs.filter(data_documento__gte=date.fromisoformat(data_da))
        except ValueError:
            pass
    if data_a:
        try:
            qs = qs.filter(data_documento__lte=date.fromisoformat(data_a))
        except ValueError:
            pass
    return qs


def _qs_libro_movimenti_azienda(azienda, filter_params: dict[str, str] | None = None):
    from django.db.models import F

    from .models import MovimentoRegistroStudioConsulente

    qs = (
        MovimentoRegistroStudioConsulente.objects.filter(azienda=azienda)
        .select_related('importato_da')
        .order_by(F('data_documento').asc(nulls_last=True), 'importato_il', 'id')
    )
    return _filter_movimenti_qs_by_data_documento(qs, filter_params)


def _libro_nota_consulente_display(note: str) -> str:
    """
    Libro consulente: rimuove prefissi legacy legati agli import massivi Excel nelle note,
    lasciando il resto (es. stato PDF proforma/parcella collegata, distinta allegata).
    """
    import re

    s = (note or "").strip()
    if not s:
        return ""
    if not re.search(
        r"(?i)(import\s+excel\s*«|colonna\s+importo\s+negativa\s+in\s+excel|creato\s+da\s+import\s+estratto\s+conto\s*«)",
        s,
    ):
        return s
    s = re.sub(r"(?i)Import\s+Excel\s*«[^»]*»\s*;\s*", "", s)
    s = re.sub(
        r"(?i)Colonna\s+Importo\s+negativa\s+in\s+Excel\s*→\s*avere\s+positivo\s*\(incasso\)\.?\s*",
        "Importo con segno negativo nel riepilogo registrato come incasso (valore assoluto). ",
        s,
    )
    s = re.sub(r"(?i)Creato\s+da\s+import\s+estratto\s+conto\s*«[^»]*»\s*\(riga\s+\d+\)\.\s*", "", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def _fmt_euro_pdf(val) -> str:
    try:
        num = Decimal(val or 0)
    except Exception:
        num = Decimal("0")
    s = f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"EUR {s}"


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_posizione_libro_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non disponibile.", status=500)

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir

    righe = list(_qs_libro_movimenti_azienda(azienda, _libro_filter_params_from_request(request)))

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        return HttpResponse("Impossibile creare il file Excel.", status=500)
    ws.title = "Movimenti"

    headers = ["Data", "Movimento", "Dettaglio", "Dare", "Avere", "Saldo", "Doc."]
    thin = Side(border_style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill('solid', fgColor='1B3A5F')
    h_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = h_fill
        c.font = h_font
        c.alignment = center
        c.border = border
    ws.freeze_panes = "A2"

    alt_fill = PatternFill('solid', fgColor='F7FAFC')
    for idx, r in enumerate(righe, start=2):
        if r.tipo_riga == 'bonifico':
            dettaglio = (r.riferimento_pagamento or '').strip() or "—"
            if r.causale_pagamento:
                dettaglio = f"{dettaglio} | {r.causale_pagamento}"
        else:
            tipo_doc = _safe_display(r, 'get_tipo_documento_display', 'tipo_documento') or ''
            num_doc = f"n. {r.numero_documento}" if r.numero_documento else ''
            dettaglio = f"{tipo_doc} {num_doc}".strip() or "—"

        row = [
            r.data_documento.strftime('%d/%m/%Y') if r.data_documento else '',
            _safe_display(r, 'get_tipo_riga_display', 'tipo_riga'),
            dettaglio,
            float(r.dare or 0),
            float(r.avere or 0),
            float(r.saldo_progressivo or 0),
            "PDF" if getattr(r.file, "name", "") else "",
        ]
        fill = alt_fill if idx % 2 == 0 else None
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=idx, column=col, value=val)
            c.border = border
            c.alignment = left
            if fill is not None:
                c.fill = fill
            if col in (4, 5, 6):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')

    widths = [12, 18, 64, 13, 13, 13, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"libro_movimenti_{quote(azienda.nome.replace(' ', '_'))}.xlsx"
    response = HttpResponse(
        out.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_posizione_libro_pdf(request):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as pdfcanvas
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse("reportlab non disponibile.", status=500)

    azienda, redir = _partitario_azienda_o_redirect(request)
    if redir:
        return redir

    righe = list(_qs_libro_movimenti_azienda(azienda, _libro_filter_params_from_request(request)))
    buffer = BytesIO()
    timestamp_ref = timezone.localtime()
    data_ref = timestamp_ref.strftime('%d/%m/%Y')
    header_title = (
        f"Posizione contabile {azienda.nome} "
        "Societa a Responsabilita Limitata Semplificata - SRLS verso Studio di Consulenza del Lavoro"
    )

    class NumberedCanvas(pdfcanvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                _draw_page_chrome(self, doc, self._pageNumber, total_pages)
                super().showPage()
            super().save()

    def _draw_page_chrome(canv, _doc, page_num, total_pages):
        canv.saveState()
        canv.setFont("Helvetica-Bold", 8.6)
        canv.setFillColor(colors.HexColor("#1B3A5F"))
        page_h = landscape(A4)[1]
        canv.drawString(_doc.leftMargin, page_h - 11 * mm, header_title[:180])

        footer = f"Posizione aggiornata al {data_ref} - pagina {page_num} di {total_pages}"
        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#4B5563"))
        footer_y = 8.2 * mm
        page_w = landscape(A4)[0]
        text_w = canv.stringWidth(footer, "Helvetica", 8)
        canv.drawString((page_w - text_w) / 2, footer_y, footer)
        canv.restoreState()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=16 * mm,
        bottomMargin=12 * mm,
        title=f"Posizione contabile - {azienda.nome}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "LibroTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=colors.HexColor("#1B3A5F"),
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "LibroMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "LibroCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.4,
        leading=9,
        wordWrap="CJK",
    )
    mov_style = ParagraphStyle(
        "LibroMov",
        parent=cell_style,
        fontName="Helvetica-Bold",
        fontSize=7.3,
        leading=8.8,
    )
    det_style = ParagraphStyle(
        "LibroDet",
        parent=cell_style,
        fontName="Helvetica",
        fontSize=7.2,
        leading=8.8,
    )

    story = [Spacer(1, 2 * mm)]
    fp = _libro_filter_params_from_request(request)
    filtro_frasi: list[str] = []
    if fp.get('anno'):
        filtro_frasi.append(f"anno {fp['anno']}")
    if fp.get('data_da'):
        filtro_frasi.append(f"dal {fp['data_da']}")
    if fp.get('data_a'):
        filtro_frasi.append(f"al {fp['data_a']}")
    if filtro_frasi:
        story.append(Paragraph("Filtro elenco: " + ", ".join(filtro_frasi), meta_style))
        story.append(Spacer(1, 2 * mm))

    data = [[
        "Data",
        "Movimento",
        "Dettaglio",
        "Da pagare",
        "Pagato",
        "Saldo residuo",
        "Doc.",
    ]]
    row_styles: list[tuple] = []
    for r in righe:
        if r.tipo_riga == 'bonifico':
            dettaglio_raw = (r.riferimento_pagamento or "—").strip()
            if r.causale_pagamento:
                dettaglio_raw = f"{dettaglio_raw}<br/><font size='6' color='#6B7280'>{r.causale_pagamento[:90]}</font>"
        else:
            tipo_doc = _safe_display(r, 'get_tipo_documento_display', 'tipo_documento') or ''
            num_doc = f" n. {r.numero_documento}" if r.numero_documento else ''
            dettaglio_raw = f"{tipo_doc}{num_doc}" or "—"

        mov_label = _safe_display(r, 'get_tipo_riga_display', 'tipo_riga') or "—"
        if len(mov_label) > 24:
            mov_label = f"{mov_label[:21]}..."

        row_idx = len(data)
        if r.tipo_riga == "documento":
            row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#FFF7D6")))
            row_styles.append(("FONTNAME", (0, row_idx), (2, row_idx), "Helvetica-Bold"))
            row_styles.append(("TEXTCOLOR", (0, row_idx), (2, row_idx), colors.HexColor("#4A3B00")))
        else:
            row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#F8FAFC")))

        data.append([
            r.data_documento.strftime('%d/%m/%Y') if r.data_documento else "—",
            Paragraph(mov_label, mov_style),
            Paragraph(dettaglio_raw, det_style),
            _fmt_euro_pdf(r.dare),
            _fmt_euro_pdf(r.avere),
            _fmt_euro_pdf(r.saldo_progressivo),
            "PDF" if getattr(r.file, "name", "") else "—",
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[23 * mm, 25 * mm, 127 * mm, 24 * mm, 24 * mm, 24 * mm, 12 * mm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (2, -1), "LEFT"),
        ("ALIGN", (3, 1), (5, -1), "RIGHT"),
        ("ALIGN", (6, 1), (6, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7.2),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        *row_styles,
    ]))
    story.append(table)
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(f"Totale righe: {len(righe)}", meta_style))

    saldo_finale = righe[-1].saldo_progressivo if righe else Decimal("0")
    if saldo_finale > 0:
        saldo_msg = f"Resta da pagare: {_fmt_euro_pdf(saldo_finale)}"
    elif saldo_finale < 0:
        saldo_msg = f"Importo credito da compensare in conto nuove Proforma: {_fmt_euro_pdf(abs(saldo_finale))}"
    else:
        saldo_msg = "Saldo residuo pari a zero."
    story.append(Paragraph(saldo_msg, title_style))

    doc.build(story, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    filename = f"libro_movimenti_{quote(azienda.nome.replace(' ', '_'))}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
@user_passes_test(_is_admin_o_consulente_partitario)
def consulente_registro_studio(request):
    """Compat URL: reindirizza all'hub posizione contabile."""
    return redirect('consulente_posizione_contabile')
