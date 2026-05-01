"""Parsing importi proforma/parcella da testo estratto PDF."""
import io
import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import SimpleTestCase, TestCase

from accounts.consulente_registro_studio import (
    EsitoParsingBonifico,
    EsitoParsingProforma,
    applica_pdf_su_movimento_bonifico,
    bonifico_duplicato_elenco_ids,
    _numeri_documento_aggancio_coerenti,
    applica_inserimento_manuale_proforma_parcella,
    applica_pdf_su_movimento_documento,
    bonifico_excel_con_riferimento_sintetico_parcella_o_proforma,
    applica_aggancia_pdf_bonifici_a_libro,
    applica_aggancia_pdf_proforma_parcelle_a_libro,
    applica_upload_bonifici_pdf,
    applica_upload_proforma_parcelle_pdf,
    import_estratto_excel,
    import_riepilogo_bonifici_da_excel,
    parse_testo_proforma_parcella,
    render_csv_report_aggancia_documenti,
    render_csv_report_import_proforma_cartella,
    ricalcola_saldi_progressivi,
    ricalcola_totali_documenti_da_testo_estratto,
)
from accounts.models import ImportEstrattoContoStudio, MovimentoRegistroStudioConsulente, RigaEstrattoContoStudio
from anagrafiche.models import Azienda


class NumeroAggancioCoerenzaTests(SimpleTestCase):
    def test_parcella_vs_numero_corto(self):
        self.assertTrue(_numeri_documento_aggancio_coerenti("182", "PARCELLA 182"))
        self.assertTrue(_numeri_documento_aggancio_coerenti("PARCELLA 182", "182"))

    def test_prefisso_pf_par(self):
        self.assertTrue(_numeri_documento_aggancio_coerenti("PAR-2021-1", "par-2021-1"))

    def test_non_confonde_anni_solo_anno(self):
        self.assertFalse(_numeri_documento_aggancio_coerenti("PF-1", "PF-2"))


class ProformaUploadDeduplicaNumeroTests(TestCase):
    """Stesso numero documento su due PDF con nome file diverso → una sola riga in libro."""

    _pdf_min = b"%PDF-1.4\n1 0 obj<<>>endobj trailer<<>>\n%%EOF"

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Dedup Numero",
            partita_iva="IT99888777666",
            indirizzo="Via Dedup 1",
            email="dedup@num.it",
        )

    @patch("accounts.consulente_registro_studio.parse_testo_proforma_parcella")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_secondo_pdf_stesso_numero_ignorato(self, mock_estrai, mock_parse):
        mock_estrai.return_value = ("PROFORMA testo", "pdf")
        mock_parse.return_value = EsitoParsingProforma(
            tipo_documento="proforma",
            numero_documento="PF-2026-DEDUP",
            data_documento=date(2026, 1, 10),
            totale_da_pagare=Decimal("99.00"),
            avvisi=[],
        )
        up1 = SimpleUploadedFile("primo.pdf", self._pdf_min, content_type="application/pdf")
        up2 = SimpleUploadedFile("secondo.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_upload_proforma_parcelle_pdf(self.az, None, [up1, up2])
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="documento").count(),
            1,
        )
        self.assertTrue(any("Già presente in libro" in x for x in msgs), msgs)
        self.assertTrue(any("PF-2026-DEDUP" in x for x in msgs), msgs)

    @patch("accounts.consulente_registro_studio.parse_testo_proforma_parcella")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_secondo_pdf_stesso_numero_con_primo_sconosciuto_ignorato(self, mock_estrai, mock_parse):
        """Prima riga tipo sconosciuto stesso numero: nuova proforma non deve duplicare."""
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="documento",
            tipo_documento="sconosciuto",
            numero_documento="PF-SC-1",
            data_documento=date(2026, 2, 1),
            dare=Decimal("40.00"),
            nome_file="primo-sconosciuto.pdf",
            testo_estratto="x",
        )
        mock_estrai.return_value = ("PROFORMA", "pdf")
        mock_parse.return_value = EsitoParsingProforma(
            tipo_documento="proforma",
            numero_documento="PF-SC-1",
            data_documento=date(2026, 2, 1),
            totale_da_pagare=Decimal("40.00"),
            avvisi=[],
        )
        up = SimpleUploadedFile("rinnovo.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_upload_proforma_parcelle_pdf(self.az, None, [up])
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="documento").count(),
            1,
        )
        self.assertTrue(any("Già presente in libro" in x for x in msgs), msgs)

    @patch("accounts.consulente_registro_studio.parse_testo_proforma_parcella")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_due_pdf_stesso_numero_date_diverse_entrambi_importati(self, mock_estrai, mock_parse):
        """Stesso numero documento con date diverse → due righe (non collisione tra anni)."""
        mock_estrai.return_value = ("PROFORMA testo", "pdf")
        mock_parse.side_effect = [
            EsitoParsingProforma(
                tipo_documento="proforma",
                numero_documento="PF-2025-2026",
                data_documento=date(2025, 6, 15),
                totale_da_pagare=Decimal("10.00"),
                avvisi=[],
            ),
            EsitoParsingProforma(
                tipo_documento="proforma",
                numero_documento="PF-2025-2026",
                data_documento=date(2026, 6, 15),
                totale_da_pagare=Decimal("20.00"),
                avvisi=[],
            ),
        ]
        up1 = SimpleUploadedFile("doc-2025.pdf", self._pdf_min, content_type="application/pdf")
        up2 = SimpleUploadedFile("doc-2026.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_upload_proforma_parcelle_pdf(self.az, None, [up1, up2])
        n = MovimentoRegistroStudioConsulente.objects.filter(
            azienda=self.az, tipo_riga="documento", numero_documento="PF-2025-2026"
        ).count()
        self.assertEqual(n, 2, msgs)
        self.assertFalse(any("Già presente in libro" in x for x in msgs), msgs)


class AgganciaPdfALibroTests(TestCase):
    """Aggancio PDF a movimenti documento/bonifico già in libro (pregresso)."""

    _pdf_min = b"%PDF-1.4\n1 0 obj<<>>endobj trailer<<>>\n%%EOF"

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Aggancia",
            partita_iva="IT77665544332",
            indirizzo="Via Agg 1",
            email="agg@libro.it",
        )

    @patch("accounts.consulente_registro_studio.parse_testo_proforma_parcella")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_aggancia_pdf_a_documento_senza_file(self, mock_estrai, mock_parse):
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="documento",
            tipo_documento="proforma",
            numero_documento="PF-AGG-LIBRO-1",
            data_documento=date(2026, 4, 1),
            dare=Decimal("200.00"),
            nome_file="riga-pregresso-excel.xlsx",
            testo_estratto="",
        )
        mock_estrai.return_value = ("PROFORMA", "pdf")
        mock_parse.return_value = EsitoParsingProforma(
            tipo_documento="proforma",
            numero_documento="PF-AGG-LIBRO-1",
            data_documento=date(2026, 4, 1),
            totale_da_pagare=Decimal("200.00"),
            avvisi=[],
        )
        up = SimpleUploadedFile("scansione.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs, report = applica_aggancia_pdf_proforma_parcelle_a_libro(self.az, None, [up])
        m = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, numero_documento="PF-AGG-LIBRO-1")
        self.assertTrue(m.file.name)
        self.assertTrue(any("allegato a movimento" in x.lower() for x in msgs), msgs)
        self.assertEqual(len(report), 1)
        self.assertEqual(report[0]["esito"], "ok")
        self.assertEqual(report[0]["movimento_id"], str(m.pk))

    @patch("accounts.consulente_registro_studio.parse_testo_proforma_parcella")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_aggancia_pdf_senza_data_estratta_saltato(self, mock_estrai, mock_parse):
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="documento",
            tipo_documento="proforma",
            numero_documento="PF-NO-DATA-PDF",
            data_documento=date(2026, 5, 1),
            dare=Decimal("50.00"),
            nome_file="pregresso.xlsx",
            testo_estratto="",
        )
        mock_estrai.return_value = ("PROFORMA", "pdf")
        mock_parse.return_value = EsitoParsingProforma(
            tipo_documento="proforma",
            numero_documento="PF-NO-DATA-PDF",
            data_documento=None,
            totale_da_pagare=Decimal("50.00"),
            avvisi=[],
        )
        up = SimpleUploadedFile("scan.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs, report = applica_aggancia_pdf_proforma_parcelle_a_libro(self.az, None, [up])
        m = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, numero_documento="PF-NO-DATA-PDF")
        self.assertFalse(bool(getattr(m.file, "name", None)))
        self.assertTrue(any("data documento non estratta" in x.lower() for x in msgs), msgs)
        self.assertEqual(len(report), 1)
        self.assertEqual(report[0]["esito"], "saltato")
        self.assertIn("numero_pdf", report[0])

    @patch("accounts.consulente_registro_studio.parse_testo_bonifico_pdf")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_aggancia_pdf_a_bonifico_senza_file(self, mock_estrai, mock_parse_bon):
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2026, 4, 10),
            dare=Decimal("0"),
            avere=Decimal("55.00"),
            nome_file="xlsx-bon/test/R1",
            riferimento_pagamento="CRO98765432100",
            causale_pagamento="Bonifico SEPA",
            metodo_estrazione="excel_riepilogo",
        )
        mock_estrai.return_value = ("distinta", "pdf")
        mock_parse_bon.return_value = EsitoParsingBonifico(
            riferimento="CRO98765432100",
            data_documento=date(2026, 4, 10),
            importo=Decimal("55.00"),
            causale="SEPA",
            avvisi=[],
        )
        up = SimpleUploadedFile("distinta.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs, report = applica_aggancia_pdf_bonifici_a_libro(self.az, None, [up])
        m = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, riferimento_pagamento="CRO98765432100")
        self.assertTrue(m.file.name)
        self.assertTrue(any("pdf allegato a bonifico" in x.lower() for x in msgs), msgs)
        self.assertEqual(len(report), 1)
        self.assertEqual(report[0]["esito"], "ok")

    @patch("accounts.consulente_registro_studio.parse_testo_proforma_parcella")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_aggancia_pdf_numero_forma_diversa_parcella(self, mock_estrai, mock_parse):
        """Libro «182» + PDF che estrae «PARCELLA 182» → stesso movimento se data uguale."""
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="documento",
            tipo_documento="parcella",
            numero_documento="182",
            data_documento=date(2021, 6, 16),
            dare=Decimal("130.00"),
            nome_file="excel-riga.xlsx",
            testo_estratto="",
        )
        mock_estrai.return_value = ("PARCELLA testo", "pdf")
        mock_parse.return_value = EsitoParsingProforma(
            tipo_documento="parcella",
            numero_documento="PARCELLA 182",
            data_documento=date(2021, 6, 16),
            totale_da_pagare=Decimal("130.00"),
            avvisi=[],
        )
        up = SimpleUploadedFile("scan-parcella.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs, report = applica_aggancia_pdf_proforma_parcelle_a_libro(self.az, None, [up])
        m = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, numero_documento="182")
        self.assertTrue(m.file.name)
        self.assertTrue(any("allegato a movimento" in x.lower() for x in msgs), msgs)
        self.assertEqual(len(report), 1)
        self.assertEqual(report[0]["esito"], "ok")

    def test_render_csv_report_aggancia_documenti_header(self):
        csv_out = render_csv_report_aggancia_documenti(
            [{"file": "a.pdf", "esito": "ok", "movimento_id": "1", "numero_pdf": "N1", "data_pdf": "2026-01-01", "messaggio": "ok"}]
        )
        self.assertIn("numero_pdf", csv_out.splitlines()[0])
        self.assertIn("a.pdf", csv_out)

    def test_render_csv_import_proforma_cartella_header(self):
        csv_out = render_csv_report_import_proforma_cartella(
            [{"file": "x/y.pdf", "esito": "ok", "movimento_id": "2", "numero_pdf": "P1", "data_pdf": "2025-06-01", "messaggio": "ok"}]
        )
        self.assertIn("numero_pdf", csv_out.splitlines()[0])
        self.assertIn("x/y.pdf", csv_out)


class BonificoDuplicatoElencoIdsTests(TestCase):
    """Chiave visiva data + avere + riferimento (o causale se rif vuoto) per badge Doppio in Pagamenti."""

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Dup Bon",
            partita_iva="IT11223344556",
            indirizzo="Via Dup 1",
            email="dup@bon.it",
        )
        self.d1 = date(2026, 4, 1)
        self.rif = "CRO-SEPA-987654"

    def _qs_bon(self):
        return MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico")

    def test_stesso_riferimento_case_insensitive(self):
        a = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            tipo_documento="sconosciuto",
            data_documento=self.d1,
            dare=Decimal("0"),
            avere=Decimal("500.00"),
            nome_file="a.xlsx",
            riferimento_pagamento=self.rif,
        )
        b = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            tipo_documento="sconosciuto",
            data_documento=self.d1,
            dare=Decimal("0"),
            avere=Decimal("500.00"),
            nome_file="b.xlsx",
            riferimento_pagamento=self.rif.upper(),
        )
        ids = bonifico_duplicato_elenco_ids(self._qs_bon())
        self.assertEqual(ids, {a.id, b.id})

    def test_riferimenti_diversi_non_segnati(self):
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            tipo_documento="sconosciuto",
            data_documento=self.d1,
            dare=Decimal("0"),
            avere=Decimal("500.00"),
            nome_file="a.xlsx",
            riferimento_pagamento="CRO-A",
        )
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            tipo_documento="sconosciuto",
            data_documento=self.d1,
            dare=Decimal("0"),
            avere=Decimal("500.00"),
            nome_file="b.xlsx",
            riferimento_pagamento="CRO-B",
        )
        self.assertEqual(bonifico_duplicato_elenco_ids(self._qs_bon()), set())

    def test_riferimento_vuoto_stessa_causale_duplicato(self):
        caus = "Bonifico SEPA ordinativo 12"
        a = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            tipo_documento="sconosciuto",
            data_documento=self.d1,
            dare=Decimal("0"),
            avere=Decimal("100.00"),
            nome_file="a.xlsx",
            riferimento_pagamento="",
            causale_pagamento=caus,
        )
        b = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            tipo_documento="sconosciuto",
            data_documento=self.d1,
            dare=Decimal("0"),
            avere=Decimal("100.00"),
            nome_file="b.xlsx",
            riferimento_pagamento="",
            causale_pagamento=caus,
        )
        self.assertEqual(bonifico_duplicato_elenco_ids(self._qs_bon()), {a.id, b.id})


class RegistroStudioPostDeleteSignalTests(TestCase):
    """post_delete su MovimentoRegistroStudioConsulente → ricalcola saldi."""

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Signal",
            partita_iva="IT88776655443",
            indirizzo="Via Sig 1",
            email="sig@nal.it",
        )

    def test_eliminazione_movimento_aggiorna_saldi_rimanenti(self):
        d = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="documento",
            tipo_documento="proforma",
            numero_documento="SIG-1",
            data_documento=date(2026, 3, 1),
            dare=Decimal("100.00"),
            nome_file="sig-doc.pdf",
            testo_estratto="t",
        )
        b = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2026, 3, 5),
            dare=Decimal("0"),
            avere=Decimal("30.00"),
            nome_file="sig-bon.pdf",
            riferimento_pagamento="CRO1",
        )
        ricalcola_saldi_progressivi(self.az.id)
        b.refresh_from_db()
        saldo_prima = b.saldo_progressivo
        d.delete()
        b.refresh_from_db()
        self.assertNotEqual(b.saldo_progressivo, saldo_prima)


class ParseTotaleParcellaProformaTests(SimpleTestCase):
    def test_totale_parcella_etichetta(self):
        txt = """
        PARCELLA PROFESSIONALE
        Totale parcella 1.234,56
        """
        r = parse_testo_proforma_parcella(txt, "fattura.pdf")
        self.assertEqual(r.totale_da_pagare, Decimal("1234.56"))
        self.assertEqual(r.tipo_documento, "parcella")

    def test_totale_proforma_etichetta(self):
        txt = "PROFORMA\nTotale proforma: € 500,00\n"
        r = parse_testo_proforma_parcella(txt, "x.pdf")
        self.assertEqual(r.totale_da_pagare, Decimal("500.00"))

    def test_numero_proforma_stessa_riga(self):
        txt = "PROFORMA N. 2025/P/042\nTotale proforma 100,00\n"
        r = parse_testo_proforma_parcella(txt, "pf.pdf")
        self.assertEqual(r.numero_documento, "2025/P/042")
        self.assertEqual(r.tipo_documento, "proforma")

    def test_numero_proforma_riga_dopo_titolo(self):
        txt = "PROFORMA FATTURA\nNumero: 2024/PROT-99\nTotale proforma 250,00\n"
        r = parse_testo_proforma_parcella(txt, "a.pdf")
        self.assertEqual(r.numero_documento, "2024/PROT-99")

    def test_numero_numero_proforma_etichetta_estesa(self):
        txt = "Numero proforma  PF-2026-0012\nPROFORMA\nTotale proforma 10,00\n"
        r = parse_testo_proforma_parcella(txt, "b.pdf")
        self.assertEqual(r.numero_documento, "PF-2026-0012")

    def test_umero_ocr_non_usato_prende_protocollo(self):
        """OCR «n. umero» non deve finire in numero_documento; si usa es. Prot."""
        txt = (
            "PROFORMA\nn. umero\nProt. 09/2022-PF\nTotale proforma 174,11\n"
        )
        r = parse_testo_proforma_parcella(txt, "pf.pdf")
        self.assertEqual(r.numero_documento, "09/2022-PF")

    def test_normalizza_n_umero_in_numero(self):
        txt = "PROFORMA\nn. umero: 55/REV-1\nTotale proforma 1,00\n"
        r = parse_testo_proforma_parcella(txt, "a.pdf")
        self.assertEqual(r.numero_documento, "55/REV-1")

    def test_totale_parcella_riga_successiva(self):
        txt = "TOTALE PARCELLA\n1.000,50\n"
        r = parse_testo_proforma_parcella(txt, "p.pdf")
        self.assertEqual(r.totale_da_pagare, Decimal("1000.50"))

    def test_nbsp_normalizzato(self):
        txt = f"TOTALE PARCELLA\u00a0:\u00a01.000,00"
        r = parse_testo_proforma_parcella(txt, "p.pdf")
        self.assertEqual(r.totale_da_pagare, Decimal("1000.00"))

    def test_netto_a_pagare_senza_avviso_pattern_alternativo(self):
        txt = "PROFORMA\nNETTO A PAGARE € 174,11\n"
        r = parse_testo_proforma_parcella(txt, "doc.pdf")
        self.assertEqual(r.totale_da_pagare, Decimal("174.11"))
        self.assertFalse(
            any("pattern alternativo" in a.lower() for a in r.avvisi),
            msg=r.avvisi,
        )

    def test_numero_fattura_proforma_e_riga_n_breve(self):
        txt = "FATTURA PROFORMA nr. 2022/09-PF\nNETTO A PAGARE 174,11\n"
        r = parse_testo_proforma_parcella(txt, "x.pdf")
        self.assertEqual(r.numero_documento, "2022/09-PF")

        txt2 = "PROFORMA\n\nN. 2022/09-PF\nTotale proforma 174,11\n"
        r2 = parse_testo_proforma_parcella(txt2, "y.pdf")
        self.assertEqual(r2.numero_documento, "2022/09-PF")

    def test_codice_documento_fallback(self):
        txt = "PARCELLA\nCodice documento: AB-2024-88\nTotale parcella 100,00\n"
        r = parse_testo_proforma_parcella(txt, "p.pdf")
        self.assertEqual(r.numero_documento, "AB-2024-88")


class RileggiTotaliDaTestoEstrattoTests(TestCase):
    def test_aggiorna_totale_e_dare_da_testo_salvato(self):
        az = Azienda.objects.create(
            nome="Az Test Libro",
            partita_iva="IT99988877701",
            indirizzo="Via Test 1",
            email="libro@test.it",
        )
        txt = "PARCELLA\nTotale parcella 250,50\n"
        m = MovimentoRegistroStudioConsulente.objects.create(
            azienda=az,
            tipo_riga="documento",
            tipo_documento="sconosciuto",
            nome_file="parcella_x.pdf",
            testo_estratto=txt,
            totale_da_pagare=None,
            dare=Decimal("0"),
        )
        res = ricalcola_totali_documenti_da_testo_estratto(az.id)
        self.assertEqual(res["n_aggiornati"], 1)
        self.assertEqual(res["n_invariati"], 0)
        m.refresh_from_db()
        self.assertEqual(m.totale_da_pagare, Decimal("250.50"))
        self.assertEqual(m.dare, Decimal("250.50"))
        self.assertEqual(m.tipo_documento, "parcella")

    def test_salta_senza_testo(self):
        az = Azienda.objects.create(
            nome="Az Test 2",
            partita_iva="IT99988877702",
            indirizzo="Via Y 2",
            email="libro2@test.it",
        )
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=az,
            tipo_riga="documento",
            nome_file="vuoto.pdf",
            testo_estratto="",
            totale_da_pagare=None,
            dare=Decimal("0"),
        )
        res = ricalcola_totali_documenti_da_testo_estratto(az.id)
        self.assertEqual(res["n_senza_testo"], 1)
        self.assertEqual(res["n_aggiornati"], 0)


class ImportRiepilogoBonificiExcelTests(TestCase):
    """Import Excel riepilogo → movimenti bonifico + nota PDF documento collegato."""

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Riepilogo Bon",
            partita_iva="IT88877766655",
            indirizzo="Via Bon 1",
            email="bon@riepilogo.it",
        )

    @staticmethod
    def _xlsx_bytes(rows):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for r in rows:
            ws.append(r)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    def test_importa_bonifico_e_nota_pdf_assente(self):
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="documento",
            tipo_documento="proforma",
            numero_documento="PF-2022-01",
            data_documento=date(2022, 2, 9),
            dare=Decimal("174.11"),
            nome_file="pf.pdf",
            testo_estratto="x",
        )
        bio = self._xlsx_bytes(
            [
                ["Data", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"],
                [date(2022, 3, 1), "PF-2022-01", "Bonifico SEPA disposizione di pagamento", 174.11],
            ]
        )
        msgs = import_riepilogo_bonifici_da_excel(bio, "r.xlsx", self.az, None)
        self.assertTrue(any("Bonifici importati: 1" in m for m in msgs), msgs)
        b = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, tipo_riga="bonifico")
        self.assertEqual(b.avere, Decimal("174.11"))
        self.assertIn("assente", b.note.lower())

    def test_importa_bonifico_pdf_documento_presente(self):
        with tempfile.TemporaryDirectory() as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                d = MovimentoRegistroStudioConsulente.objects.create(
                    azienda=self.az,
                    tipo_riga="documento",
                    tipo_documento="proforma",
                    numero_documento="PF-99",
                    data_documento=date(2023, 1, 10),
                    dare=Decimal("100.00"),
                    nome_file="doc99.pdf",
                    testo_estratto="y",
                )
                d.file.save("doc99.pdf", ContentFile(b"%PDF-1.4 test"), save=True)

                bio = self._xlsx_bytes(
                    [
                        ["Data", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"],
                        [date(2023, 2, 1), "PF-99", "Bonifico bonifico accreditato", 100.00],
                    ]
                )
                import_riepilogo_bonifici_da_excel(bio, "r2.xlsx", self.az, None)
                b = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, tipo_riga="bonifico")
                self.assertIn("presente", b.note.lower())

    def test_intestazione_dopo_righe_titolo_e_senza_colonna_documento(self):
        """Riga intestazione oltre la 10ª; colonne Data / Descrizione / Importo senza Documento."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for _ in range(12):
            ws.append([None, None, None, None])
        ws.append(["Riepilogo generale", None, None, None])
        ws.append(["Data", "Descrizione", "Importo"])
        ws.append([date(2024, 1, 15), "Bonifico SEPA disposizione", 99.5])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        msgs = import_riepilogo_bonifici_da_excel(bio, "t.xlsx", self.az, None)
        self.assertTrue(any("Bonifici importati: 1" in m for m in msgs), msgs)
        b = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, tipo_riga="bonifico")
        self.assertEqual(b.avere, Decimal("99.50"))
        self.assertIn("nessun documento", b.note.lower())

    def test_salta_riga_solo_parcella_già_a_libro(self):
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="documento",
            tipo_documento="parcella",
            numero_documento="PAR-2021-1",
            data_documento=date(2021, 6, 1),
            dare=Decimal("500.00"),
            nome_file="p.pdf",
            testo_estratto="z",
        )
        bio = self._xlsx_bytes(
            [
                ["Data", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"],
                [date(2021, 6, 1), "PAR-2021-1", "Emissione parcella professionale", 500],
            ]
        )
        msgs = import_riepilogo_bonifici_da_excel(bio, "r3.xlsx", self.az, None)
        self.assertFalse(any("Bonifici importati: 1" in m for m in msgs), msgs)
        self.assertEqual(MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(), 0)

    def test_salta_riga_parcella_in_colonna_documento_senza_bonifico(self):
        """Riepilogo PROFORMA: colonna Documento «PARCELLA 182» senza testo bonifico → non crea avere (evita doppio con PDF)."""
        bio = self._xlsx_bytes(
            [
                ["Data", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"],
                [date(2021, 6, 16), "PARCELLA 182", "Cliente / saldo", -130.0],
            ]
        )
        msgs = import_riepilogo_bonifici_da_excel(bio, "riepilogo-par.xlsx", self.az, None)
        self.assertEqual(MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(), 0)
        self.assertTrue(any("solo proforma/parcella" in m for m in msgs), msgs)

    def test_parcella_in_documento_ma_bonifico_in_descrizione_si_importa(self):
        bio = self._xlsx_bytes(
            [
                ["Data", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"],
                [date(2021, 7, 9), "PARCELLA 99", "Bonifico SEPA disposizione di pagamento", 130.0],
            ]
        )
        msgs = import_riepilogo_bonifici_da_excel(bio, "bon-con-doc-par.xlsx", self.az, None)
        self.assertTrue(any("Bonifici importati: 1" in m for m in msgs), msgs)
        b = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, tipo_riga="bonifico")
        self.assertEqual(b.avere, Decimal("130.00"))

    def test_layout_classico_importa_senza_keyword_bonifico(self):
        """Excel con Data / Documento / Descrizione / Importo: righe operative anche senza testo «bonifico»."""
        bio = self._xlsx_bytes(
            [
                ["DATA", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"],
                [date(2024, 6, 1), "INV-1", "Incasso saldo fattura cliente Rossi", 250.00],
            ]
        )
        msgs = import_riepilogo_bonifici_da_excel(bio, "r4.xlsx", self.az, None)
        self.assertTrue(any("Bonifici importati: 1" in m for m in msgs), msgs)
        b = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, tipo_riga="bonifico")
        self.assertEqual(b.avere, Decimal("250.00"))

    def test_importo_negativo_excel_registrato_come_avere_positivo(self):
        """File misto (non «solo negativi»): importo negativo su riga bonifico → avere positivo."""
        bio = self._xlsx_bytes(
            [
                ["Data", "DESCRIZIONE", "IMPORTO"],
                [date(2025, 7, 1), "Addebito bollo", 12.0],
                [date(2025, 7, 2), "Bonifico SEPA disposizione di pagamento", -42.5],
            ]
        )
        msgs = import_riepilogo_bonifici_da_excel(bio, "mix-neg.xlsx", self.az, None)
        self.assertTrue(any("Bonifici importati: 1" in m for m in msgs), msgs)
        b = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, tipo_riga="bonifico")
        self.assertEqual(b.avere, Decimal("42.50"))
        self.assertIn("negativa", b.note.lower())

    def test_tutta_colonna_importi_negativi_solo_quelle_righe(self):
        """Tutti gli importi non nulli nel foglio sono negativi: si importano solo righe con importo < 0."""
        rows = [["DATA", "DESCRIZIONE", "IMPORTO"]]
        for i in range(6):
            rows.append([date(2025, 8, i + 1), "Bonifico SEPA accredito", Decimal(-25)])
        bio = self._xlsx_bytes(rows)
        msgs = import_riepilogo_bonifici_da_excel(bio, "solo-neg.xlsx", self.az, None)
        self.assertEqual(MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(), 6)
        self.assertTrue(any("Convenzione file rilevata" in m for m in msgs), msgs)

    def test_colonna_con_almeno_un_positivo_non_attiva_solo_negativi(self):
        """Se in colonna c’è almeno un importo > 0, non si escludono i positivi (avere sempre = valore assoluto)."""
        rows = [["DATA", "DESCRIZIONE", "IMPORTO"]]
        for i in range(5):
            rows.append([date(2025, 8, i + 1), "Bonifico SEPA", Decimal(-10)])
        rows.append([date(2025, 9, 1), "Bonifico incasso", Decimal("55")])
        bio = self._xlsx_bytes(rows)
        import_riepilogo_bonifici_da_excel(bio, "mix-col.xlsx", self.az, None)
        self.assertEqual(MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(), 6)


class ImportEstrattoContoExcelTests(TestCase):
    """Import Excel estratto conto: tutte le righe dati, intestazione non solo in prima riga."""

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Estratto",
            partita_iva="IT77766655544",
            indirizzo="Via Est 1",
            email="est@conto.it",
        )

    def test_importa_tutte_le_righe_con_documento_e_saldo_typo(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Report", None, None, None, None, None])
        ws.append(["DATA", "DOCUMENTO", None, "DESCRIZIONE", "IMPORTO", "SALDO PRGRESSIVO"])
        ws.append([date(2025, 1, 10), "DOC/A/1", None, "Voce contabile", 100.5, 1000])
        ws.append([date(2025, 1, 11), None, None, "Seconda voce", -20, 980])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                imp, msgs = import_estratto_excel(bio, "e.xlsx", self.az, None)
        self.assertEqual(imp.righe_lette, 2, msgs)
        righe = list(RigaEstrattoContoStudio.objects.filter(importazione=imp).order_by("indice_riga"))
        self.assertEqual(righe[0].riferimento_excel, "DOC/A/1")
        self.assertEqual(righe[0].importo_excel, Decimal("100.50"))
        self.assertEqual(righe[1].importo_excel, Decimal("-20.00"))
        self.assertEqual(righe[0].data_excel, date(2025, 1, 10))
        self.assertEqual(righe[1].data_excel, date(2025, 1, 11))
        raw0 = righe[0].celle_raw
        self.assertTrue(any("PRGRESS" in str(k).upper() or "SALDO" in str(k).upper() for k in raw0), raw0)

    def test_bonifico_estratto_crea_movimento_nel_libro_se_assente(self):
        """Riga con testo bonifico e importo a credito: nuovo MovimentoRegistroStudioConsulente bonifico + riga agganciata."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["DATA", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"])
        ws.append([date(2025, 3, 5), "", "Bonifico SEPA accreditato cliente", 150.25])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                imp, msgs = import_estratto_excel(bio, "bon-est.xlsx", self.az, None)
        self.assertEqual(imp.righe_lette, 1)
        self.assertEqual(imp.righe_agganciate, 1)
        self.assertTrue(any("registrati automaticamente" in m for m in msgs), msgs)
        r = imp.righe.get()
        self.assertIsNotNone(r.movimento_id)
        self.assertEqual(r.movimento.tipo_riga, "bonifico")
        self.assertEqual(r.movimento.avere, Decimal("150.25"))
        self.assertEqual(r.movimento.metodo_estrazione, "excel_estratto_conto")

    def test_secondo_import_stesso_excel_aggancia_stesso_bonifico(self):
        """Re-import dello stesso foglio: nessun secondo movimento bonifico, riga agganciata al primo."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["DATA", "DESCRIZIONE", "IMPORTO"])
        ws.append([date(2025, 4, 1), "Bonifico SEPA accredito", 75.00])
        bio = io.BytesIO()
        wb.save(bio)
        data = bio.getvalue()
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                import_estratto_excel(io.BytesIO(data), "dup.xlsx", self.az, None)
                import_estratto_excel(io.BytesIO(data), "dup.xlsx", self.az, None)
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(),
            1,
        )
        imps = list(
            ImportEstrattoContoStudio.objects.filter(azienda=self.az, nome_file="dup.xlsx").order_by(
                "-importato_il"
            )[:2]
        )
        self.assertGreaterEqual(len(imps), 2)
        for imp in imps[:2]:
            r = imp.righe.filter(importo_excel=Decimal("75.00")).first()
            self.assertIsNotNone(r)
            self.assertIsNotNone(r.movimento_id)


class RimuoviBonificiImportExcelStudioCommandTests(TestCase):
    """Management command rimuovi_bonifici_import_excel_studio."""

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Rimuovi Excel",
            partita_iva="IT11122233344",
            indirizzo="Via Cmd 1",
            email="cmd@excel.it",
        )

    @staticmethod
    def _xlsx_rows(rows):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for r in rows:
            ws.append(r)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    def test_anteprima_non_elimina(self):
        bio = self._xlsx_rows(
            [
                ["DATA", "DESCRIZIONE", "IMPORTO"],
                [date(2025, 1, 10), "Bonifico SEPA accredito", Decimal("50")],
            ]
        )
        import_riepilogo_bonifici_da_excel(bio, "ant.xlsx", self.az, None)
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(),
            1,
        )
        out = io.StringIO()
        call_command("rimuovi_bonifici_import_excel_studio", "--azienda-id", self.az.pk, stdout=out)
        self.assertIn("Anteprima sola", out.getvalue())
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(),
            1,
        )

    def test_execute_rimuove_solo_bonifici_traccia_excel(self):
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2024, 1, 1),
            avere=Decimal("10.00"),
            dare=Decimal("0"),
            nome_file="bonifico-manuale.pdf",
            metodo_estrazione="",
        )
        bio = self._xlsx_rows(
            [
                ["DATA", "DESCRIZIONE", "IMPORTO"],
                [date(2025, 2, 1), "Bonifico SEPA accredito", Decimal("33")],
            ]
        )
        import_riepilogo_bonifici_da_excel(bio, "ex.xlsx", self.az, None)
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(),
            2,
        )
        call_command("rimuovi_bonifici_import_excel_studio", "--azienda-id", self.az.pk, "--execute")
        bon = list(MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico"))
        self.assertEqual(len(bon), 1)
        self.assertEqual(bon[0].nome_file, "bonifico-manuale.pdf")

    def test_execute_elimina_import_estratto_e_bonifico_collegato(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["DATA", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"])
        ws.append([date(2025, 3, 5), "", "Bonifico SEPA accreditato cliente", 150.25])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                imp, _msgs = import_estratto_excel(bio, "est-cmd.xlsx", self.az, None)
        self.assertEqual(ImportEstrattoContoStudio.objects.filter(azienda=self.az).count(), 1)
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(
                azienda=self.az, tipo_riga="bonifico", metodo_estrazione="excel_estratto_conto"
            ).count(),
            1,
        )
        imp_id = imp.pk
        call_command("rimuovi_bonifici_import_excel_studio", "--azienda-id", self.az.pk, "--execute")
        self.assertFalse(ImportEstrattoContoStudio.objects.filter(pk=imp_id).exists())
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(),
            0,
        )

    def test_solo_parcella_sintetici_non_tocca_bonifico_banca_ne_import_estratto(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["DATA", "DOCUMENTO", "DESCRIZIONE", "IMPORTO"])
        ws.append([date(2025, 3, 5), "", "Bonifico SEPA accreditato cliente", 150.25])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                imp, _msgs = import_estratto_excel(bio, "est-solo.xlsx", self.az, None)
        imp_id = imp.pk
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2021, 6, 16),
            avere=Decimal("130.00"),
            dare=Decimal("0"),
            nome_file="xlsx-bon/RIEP/R1",
            metodo_estrazione="excel_riepilogo",
            riferimento_pagamento="PARCELLA 182|2021-06-16|130.00",
            causale_pagamento="PARCELLA 182 —",
        )
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2021, 7, 9),
            avere=Decimal("130.00"),
            dare=Decimal("0"),
            nome_file="xlsx-bon/RIEP/R2",
            metodo_estrazione="excel_riepilogo",
            riferimento_pagamento="BONIFICO DOLCEMASCOLO|2021-07-09|130.00",
            causale_pagamento="Studio —",
        )
        self.assertTrue(bonifico_excel_con_riferimento_sintetico_parcella_o_proforma(
            MovimentoRegistroStudioConsulente.objects.get(riferimento_pagamento__startswith="PARCELLA")
        ))
        self.assertFalse(bonifico_excel_con_riferimento_sintetico_parcella_o_proforma(
            MovimentoRegistroStudioConsulente.objects.get(riferimento_pagamento__startswith="BONIFICO DOLCE")
        ))
        call_command(
            "rimuovi_bonifici_import_excel_studio",
            "--azienda-id",
            self.az.pk,
            "--solo-parcella-proforma-sintetici",
            "--execute",
        )
        self.assertTrue(ImportEstrattoContoStudio.objects.filter(pk=imp_id).exists())
        rifs = set(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").values_list(
                "riferimento_pagamento", flat=True
            )
        )
        self.assertIn("BONIFICO DOLCEMASCOLO|2021-07-09|130.00", rifs)
        self.assertNotIn("PARCELLA 182|2021-06-16|130.00", rifs)


class InserimentoManualeProformaTests(TestCase):
    """Riga documento senza PDF: stessa deduplica data+numero degli upload PDF."""

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Manuale",
            partita_iva="IT11223344556",
            indirizzo="Via Man 1",
            email="man@libro.it",
        )

    def test_inserimento_manuale_crea_riga_senza_file(self):
        msgs = applica_inserimento_manuale_proforma_parcella(
            self.az,
            None,
            tipo_documento="parcella",
            numero_documento="PAR-MAN-99",
            data_documento=date(2026, 4, 1),
            importo_contabile=Decimal("150.50"),
        )
        self.assertTrue(any("Registrato" in m for m in msgs), msgs)
        m = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, numero_documento="PAR-MAN-99")
        self.assertEqual(m.tipo_riga, "documento")
        self.assertEqual(m.tipo_documento, "parcella")
        self.assertEqual(m.data_documento, date(2026, 4, 1))
        self.assertEqual(m.totale_da_pagare, Decimal("150.50"))
        self.assertEqual(m.dare, Decimal("150.50"))
        self.assertFalse(bool(getattr(m.file, "name", None)))
        self.assertEqual(m.metodo_estrazione, "manuale_portale")

    def test_inserimento_duplicato_bloccato(self):
        applica_inserimento_manuale_proforma_parcella(
            self.az,
            None,
            tipo_documento="proforma",
            numero_documento="PF-DUP-1",
            data_documento=date(2026, 5, 10),
            importo_contabile=Decimal("10.00"),
        )
        msgs2 = applica_inserimento_manuale_proforma_parcella(
            self.az,
            None,
            tipo_documento="proforma",
            numero_documento="PF-DUP-1",
            data_documento=date(2026, 5, 10),
            importo_contabile=Decimal("99.00"),
        )
        self.assertTrue(any("Già presente" in m for m in msgs2), msgs2)
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(
                azienda=self.az, tipo_riga="documento", numero_documento="PF-DUP-1"
            ).count(),
            1,
        )


class AllegaPdfSuMovimentoSingoloTests(TestCase):
    _pdf_min = b"%PDF-1.4\n1 0 obj<<>>endobj trailer<<>>\n%%EOF"

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Allega",
            partita_iva="IT99887766554",
            indirizzo="Via All 1",
            email="all@libro.it",
        )

    @patch("accounts.consulente_registro_studio.parse_testo_proforma_parcella")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_allega_pdf_su_movimento_manuale_ok(self, mock_estrai, mock_parse):
        msgs_ins = applica_inserimento_manuale_proforma_parcella(
            self.az,
            None,
            tipo_documento="proforma",
            numero_documento="124",
            data_documento=date(2022, 3, 11),
            importo_contabile=Decimal("318.45"),
        )
        self.assertTrue(any("Registrato" in m for m in msgs_ins), msgs_ins)
        mov = MovimentoRegistroStudioConsulente.objects.get(azienda=self.az, numero_documento="124")
        mock_estrai.return_value = ("testo", "pdf")
        mock_parse.return_value = EsitoParsingProforma(
            tipo_documento="proforma",
            numero_documento="124",
            data_documento=date(2022, 3, 11),
            totale_da_pagare=Decimal("318.45"),
            avvisi=[],
        )
        up = SimpleUploadedFile("proforma124.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_pdf_su_movimento_documento(self.az, None, mov.pk, up)
        self.assertTrue(any("pdf allegato" in m.lower() for m in msgs), msgs)
        mov.refresh_from_db()
        self.assertTrue(bool(getattr(mov.file, "name", None)))


class AllegaPdfBonificoSingoloTests(TestCase):
    _pdf_min = b"%PDF-1.4\n1 0 obj<<>>endobj trailer<<>>\n%%EOF"

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Bon Allega",
            partita_iva="IT88776655443",
            indirizzo="Via Bon 1",
            email="bonall@libro.it",
        )

    @patch("accounts.consulente_registro_studio.parse_testo_bonifico_pdf")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_allega_pdf_su_bonifico_excel_ok(self, mock_estrai, mock_parse):
        mov = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2026, 4, 24),
            avere=Decimal("191.22"),
            dare=Decimal("0"),
            nome_file="xlsx-bon/RIEP/R1",
            riferimento_pagamento="BONIFICO SUM UP CO4B2KLJWK|2026-04-24|191.22",
            causale_pagamento="BONIFICO SUM UP — DITTA",
            metodo_estrazione="excel_riepilogo",
        )
        mock_estrai.return_value = ("distinta", "pdf")
        mock_parse.return_value = EsitoParsingBonifico(
            riferimento="CO4B2KLJWK",
            data_documento=date(2026, 4, 24),
            importo=Decimal("191.22"),
            causale="SEPA",
            avvisi=[],
        )
        up = SimpleUploadedFile("distinta.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_pdf_su_movimento_bonifico(self.az, None, mov.pk, up)
        self.assertTrue(any("allegata con successo al bonifico" in m.lower() for m in msgs), msgs)
        mov.refresh_from_db()
        self.assertTrue(bool(getattr(mov.file, "name", None)))

    @patch("accounts.consulente_registro_studio.parse_testo_bonifico_pdf")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_allega_pdf_su_bonifico_riga_esplicita_non_altro_stesso_importo(
        self, mock_estrai, mock_parse
    ):
        """La distinta si allega alla riga scelta, anche se un altro bonifico ha stesso importo/data."""
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2021, 7, 9),
            avere=Decimal("130.00"),
            dare=Decimal("0"),
            nome_file="excel/R1",
            riferimento_pagamento="BONIFICO ALTRO|2021-07-09|130.00",
            causale_pagamento="Altro",
            metodo_estrazione="excel_riepilogo",
        )
        mov_b = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2021, 7, 9),
            avere=Decimal("130.00"),
            dare=Decimal("0"),
            nome_file="excel/R2",
            riferimento_pagamento="BONIFICO DOLCEMASCOLO|2021-07-09|130.00",
            causale_pagamento="Dolcemascolo",
            metodo_estrazione="excel_riepilogo",
        )
        mock_estrai.return_value = ("distinta", "pdf")
        mock_parse.return_value = EsitoParsingBonifico(
            riferimento="DOLCEMASCOLO",
            data_documento=date(2021, 7, 9),
            importo=Decimal("130.00"),
            causale="SEPA",
            avvisi=[],
        )
        up = SimpleUploadedFile("distinta.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_pdf_su_movimento_bonifico(self.az, None, mov_b.pk, up)
        self.assertTrue(any("allegata con successo al bonifico" in m.lower() for m in msgs), msgs)
        mov_b.refresh_from_db()
        self.assertTrue(bool(getattr(mov_b.file, "name", None)))

    @patch("accounts.consulente_registro_studio.parse_testo_bonifico_pdf")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_allega_pdf_ok_cro_pdf_non_in_libro_ma_importo_data_ok(self, mock_estrai, mock_parse):
        """CRO estratto dal PDF spesso non è sottostringa del riferimento sintetico Excel: conta l'importo."""
        ref_libro = "BONIFICO DOLCEMASCOLO|2021-07-09|130.00"
        caus = "BONIFICO DOLCEMASCOLO — Studio Cipriano"
        mov = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2021, 7, 9),
            avere=Decimal("130.00"),
            dare=Decimal("0"),
            nome_file="excel/R-dolce",
            riferimento_pagamento=ref_libro,
            causale_pagamento=caus,
            metodo_estrazione="excel_riepilogo",
        )
        mock_estrai.return_value = ("distinta", "pdf")
        mock_parse.return_value = EsitoParsingBonifico(
            riferimento="00998877665544332211009988776655",
            data_documento=date(2021, 7, 9),
            importo=Decimal("130.00"),
            causale="Testo causale PDF",
            avvisi=[],
        )
        up = SimpleUploadedFile("distinta.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_pdf_su_movimento_bonifico(self.az, None, mov.pk, up)
        self.assertTrue(any("allegata con successo al bonifico" in m.lower() for m in msgs), msgs)
        mov.refresh_from_db()
        self.assertTrue(bool(getattr(mov.file, "name", None)))
        self.assertEqual(mov.riferimento_pagamento, ref_libro)
        self.assertEqual(mov.causale_pagamento, caus)

    @patch("accounts.consulente_registro_studio.parse_testo_bonifico_pdf")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_allega_pdf_ok_data_distinta_diversa_ma_stesso_importo(self, mock_estrai, mock_parse):
        """Data contabile in elenco ≠ data PDF (es. 07/07 vs 19/10): con stesso importo si allega comunque."""
        ref_libro = "BONIFICO DOLCEMASCOLO|2021-07-09|130.00"
        mov = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2021, 7, 7),
            avere=Decimal("130.00"),
            dare=Decimal("0"),
            nome_file="excel/R-dolce",
            riferimento_pagamento=ref_libro,
            causale_pagamento="BONIFICO DOLCEMASCOLO — Studio Cipriano",
            metodo_estrazione="excel_riepilogo",
        )
        mock_estrai.return_value = ("distinta", "pdf")
        mock_parse.return_value = EsitoParsingBonifico(
            riferimento="CROBPSA20211019",
            data_documento=date(2021, 10, 19),
            importo=Decimal("130.00"),
            causale="SALDO PROFORMA",
            avvisi=[],
        )
        up = SimpleUploadedFile(
            "BONIFICO BPSA DEL 19 OTTOBRE 2021.pdf", self._pdf_min, content_type="application/pdf"
        )
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_pdf_su_movimento_bonifico(self.az, None, mov.pk, up)
        self.assertTrue(any("allegata con successo al bonifico" in m.lower() for m in msgs), msgs)
        self.assertTrue(any("data operazione sulla distinta" in m for m in msgs), msgs)
        mov.refresh_from_db()
        self.assertTrue(bool(getattr(mov.file, "name", None)))

    @patch("accounts.consulente_registro_studio.parse_testo_bonifico_pdf")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_allega_pdf_ok_senza_importo_pdf_ma_rif_pipe_excel(self, mock_estrai, mock_parse):
        """PDF senza importo leggibile: si usa |data|importo dal riferimento in libro."""
        ref_libro = "BONIFICO DOLCEMASCOLO|2021-07-09|130.00"
        mov = MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2021, 7, 9),
            avere=Decimal("130.00"),
            dare=Decimal("0"),
            nome_file="excel/R-dolce",
            riferimento_pagamento=ref_libro,
            causale_pagamento="—",
            metodo_estrazione="excel_riepilogo",
        )
        mock_estrai.return_value = ("distinta", "ocr")
        mock_parse.return_value = EsitoParsingBonifico(
            riferimento="",
            data_documento=None,
            importo=None,
            causale="",
            avvisi=["vuoto"],
        )
        up = SimpleUploadedFile("distinta.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_pdf_su_movimento_bonifico(self.az, None, mov.pk, up)
        self.assertTrue(any("allegata con successo al bonifico" in m.lower() for m in msgs), msgs)
        mov.refresh_from_db()
        self.assertTrue(bool(getattr(mov.file, "name", None)))


class UploadBonificoPdfDedupTests(TestCase):
    """Evita doppie righe bonifico se il PDF estratto coincide con un movimento già in libro."""

    _pdf_min = b"%PDF-1.4\n1 0 obj<<>>endobj trailer<<>>\n%%EOF"

    def setUp(self):
        self.az = Azienda.objects.create(
            nome="Az Dedup Bon",
            partita_iva="IT11223344556",
            indirizzo="Via Ded 1",
            email="dedbon@libro.it",
        )

    @patch("accounts.consulente_registro_studio.parse_testo_bonifico_pdf")
    @patch("accounts.consulente_registro_studio.estrai_testo_da_pdf")
    def test_upload_bonifico_pdf_non_duplica_stesso_cro_data_importo(self, mock_estrai, mock_parse):
        MovimentoRegistroStudioConsulente.objects.create(
            azienda=self.az,
            tipo_riga="bonifico",
            data_documento=date(2026, 2, 1),
            avere=Decimal("88.50"),
            dare=Decimal("0"),
            nome_file="pre/esistente",
            riferimento_pagamento="CROZZDEDUP123456789",
            causale_pagamento="Bonifico test dedup",
            metodo_estrazione="excel_riepilogo",
        )
        mock_estrai.return_value = ("distinta", "pdf")
        mock_parse.return_value = EsitoParsingBonifico(
            riferimento="CROZZDEDUP123456789",
            data_documento=date(2026, 2, 1),
            importo=Decimal("88.50"),
            causale="Altra causale",
            avvisi=[],
        )
        up = SimpleUploadedFile("nuova_distinta.pdf", self._pdf_min, content_type="application/pdf")
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory(dir=str(repo_root)) as media_tmp:
            with self.settings(MEDIA_ROOT=media_tmp):
                msgs = applica_upload_bonifici_pdf(self.az, None, [up])
        self.assertTrue(any("importazione ignorata" in m.lower() for m in msgs), msgs)
        self.assertEqual(
            MovimentoRegistroStudioConsulente.objects.filter(azienda=self.az, tipo_riga="bonifico").count(),
            1,
        )
