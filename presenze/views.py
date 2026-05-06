import calendar
from collections import defaultdict
from datetime import date, timedelta, time
from decimal import Decimal
from typing import Optional

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, Value
from django.db.models.functions import Coalesce
from django.forms import modelformset_factory
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from urllib.parse import urlencode
from django.utils import timezone
from django.views.decorators.http import require_POST

def _parse_int(val, default):
    """Convert val to int, rimuovendo separatori migliaia (es. locale it: '2.026' → 2026)."""
    try:
        if val is None or val == '':
            return default
        # Rimuovi separatori migliaia (. e ,) e spazi — evita fallimenti silenziosi sui hidden del form
        clean = str(val).replace(' ', '').replace('.', '').replace(',', '').strip()
        return int(clean)
    except (TypeError, ValueError):
        return default


def _redirect_pianificazione_orari_annuale(anno, mese):
    return redirect(
        f"{reverse('pianificazione_orari_annuale')}?{urlencode({'anno': int(anno), 'mese': int(mese)})}"
    )


def _redirect_riepilogo_mensile_motore(anno, mese):
    return redirect(
        f"{reverse('riepilogo_mensile_motore')}?{urlencode({'anno': int(anno), 'mese': int(mese)})}"
    )


def _parse_ora_hhmm(val):
    """
    Converte stringhe da input time (HTML) in datetime.time.
    Accetta HH:MM, HH:MM:SS, HH:MM:SS.mmm — il vecchio split su due soli ':' falliva e non salvava.
    """
    s = (val or '').strip()
    if not s:
        return None
    try:
        parts = s.split(':')
        if len(parts) < 2:
            return None
        h = int(parts[0])
        m = int(parts[1])
        return time(h, m)
    except (ValueError, TypeError, IndexError):
        return None

from .forms import (
    PresenzaForm,
    GiornoPresenzaForm,
    ConfigurazioneOrarioMensileForm,
    FasciaAperturaMensileForm,
    TurnoLavorativoAziendaleForm,
)
from .models import (
    Presenza,
    ConfigurazioneOrarioAnnuale,
    FasciaAperturaSettimanale,
    ConfigurazioneOrarioMensile,
    FasciaAperturaMensile,
    TurnoLavorativoAziendale,
    AssegnazioneTurnoDipendente,
)
from accounts.formatting import num_it_str
from anagrafiche.models import Dipendente
from anagrafiche.permissions import admin_required, hr_required
from log_attivita.utils import registra_log
from log_attivita.anomalie import registra_evento_anomalia
from accounts.tenant import get_azienda_operativa
from accounts.dipendente_portale import get_dipendente_collegato
from rapporto_di_lavoro.utils_calendario import get_chiusura_settimanale

# ── Giorni settimana (italiano) ─────────────────────────────────────────────
GIORNI_ITA = ['Lun', 'Mar', 'Mer', 'Gio', 'Ven', 'Sab', 'Dom']
MESI_ITA = ['', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
            'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre']


def _ensure_fasce_settimanali(config):
    """Garantisce una fascia per ciascun giorno settimana (0..6)."""
    existing = {f.giorno_settimana: f for f in config.fasce_apertura.all()}
    riposi = set(config.giorni_riposo_settimanale or [])
    for giorno in range(7):
        if giorno in existing:
            continue
        chiuso = giorno in riposi
        FasciaAperturaSettimanale.objects.create(
            configurazione=config,
            giorno_settimana=giorno,
            chiuso=chiuso,
            ora_apertura_mattina=None if chiuso else time(9, 0),
            ora_chiusura_mattina=None if chiuso else time(13, 0),
            ora_apertura_pomeriggio=None if chiuso else time(16, 0),
            ora_chiusura_pomeriggio=None if chiuso else time(20, 0),
        )


def _ensure_fasce_mensili(config_mese):
    """Garantisce una fascia per ciascun giorno settimana (0..6) sul mese selezionato."""
    existing = {f.giorno_settimana: f for f in config_mese.fasce_apertura.all()}
    riposi = set(config_mese.giorni_riposo_settimanale or [])

    annuale = ConfigurazioneOrarioAnnuale.objects.filter(
        azienda=config_mese.azienda,
        anno=config_mese.anno,
    ).first()
    annuale_map = {}
    if annuale:
        _ensure_fasce_settimanali(annuale)
        annuale_map = {f.giorno_settimana: f for f in annuale.fasce_apertura.all()}

    for giorno in range(7):
        if giorno in existing:
            continue

        a = annuale_map.get(giorno)
        chiuso_default = (giorno in riposi) or bool(getattr(a, 'chiuso', False))

        FasciaAperturaMensile.objects.create(
            configurazione=config_mese,
            giorno_settimana=giorno,
            chiuso=chiuso_default,
            ora_apertura_mattina=None if chiuso_default else (getattr(a, 'ora_apertura_mattina', None) or time(9, 0)),
            ora_chiusura_mattina=None if chiuso_default else (getattr(a, 'ora_chiusura_mattina', None) or time(13, 0)),
            ora_apertura_pomeriggio=None if chiuso_default else (getattr(a, 'ora_apertura_pomeriggio', None) or time(14, 0)),
            ora_chiusura_pomeriggio=None if chiuso_default else (getattr(a, 'ora_chiusura_pomeriggio', None) or time(18, 0)),
        )


def _get_latest_config_orario_mensile(azienda, anno):
    """Ultima configurazione mensile modificata nell'anno (fallback operativo)."""
    return (
        ConfigurazioneOrarioMensile.objects
        .filter(azienda=azienda, anno=anno)
        .order_by('-data_modifica', '-mese', '-id')
        .first()
    )


def _snapshot_parametri_orari_mese(config_mese):
    """Snapshot live dal DB dei parametri orari per verifica salvataggio."""
    fresh = ConfigurazioneOrarioMensile.objects.filter(pk=config_mese.pk).first()
    if not fresh:
        return None

    fasce = []
    for f in FasciaAperturaMensile.objects.filter(configurazione=fresh).order_by('giorno_settimana'):
        fasce.append({
            'giorno_num': f.giorno_settimana,
            'giorno_label': f.get_giorno_settimana_display(),
            'chiuso': f.chiuso,
            'ora_apertura_mattina': f.ora_apertura_mattina,
            'ora_chiusura_mattina': f.ora_chiusura_mattina,
            'ora_apertura_pomeriggio': f.ora_apertura_pomeriggio,
            'ora_chiusura_pomeriggio': f.ora_chiusura_pomeriggio,
        })

    return {
        'id': fresh.id,
        'anno': fresh.anno,
        'mese': fresh.mese,
        'giorni_riposo_settimanale': fresh.giorni_riposo_settimanale or [],
        'genera_presenze_teoriche': fresh.genera_presenze_teoriche,
        'data_modifica': fresh.data_modifica,
        'fasce': fasce,
    }


def _get_config_orario_mese(azienda, anno, mese, create=True):
    """Recupera configurazione mensile (fonte primaria), con default da annuale."""
    latest_mese = _get_latest_config_orario_mensile(azienda, anno)
    annuale = ConfigurazioneOrarioAnnuale.objects.filter(azienda=azienda, anno=anno).first()
    defaults = {
        'giorni_riposo_settimanale': (
            (latest_mese.giorni_riposo_settimanale if latest_mese else None)
            or (annuale.giorni_riposo_settimanale if annuale else None)
            or [6]
        ),
        'genera_presenze_teoriche': (
            latest_mese.genera_presenze_teoriche
            if latest_mese is not None
            else (annuale.genera_presenze_teoriche if annuale else True)
        ),
    }

    if create:
        cfg_mese, _ = ConfigurazioneOrarioMensile.objects.get_or_create(
            azienda=azienda,
            anno=anno,
            mese=mese,
            defaults=defaults,
        )
        _ensure_fasce_mensili(cfg_mese)
        fasce_map = {
            f.giorno_settimana: f
            for f in FasciaAperturaMensile.objects.filter(configurazione=cfg_mese)
        }
        return cfg_mese, fasce_map

    cfg_mese = ConfigurazioneOrarioMensile.objects.filter(
        azienda=azienda,
        anno=anno,
        mese=mese,
    ).first()
    if not cfg_mese:
        return None, {}

    _ensure_fasce_mensili(cfg_mese)
    fasce_map = {
        f.giorno_settimana: f
        for f in FasciaAperturaMensile.objects.filter(configurazione=cfg_mese)
    }
    return cfg_mese, fasce_map


def _periodo_rapporto_dipendente_per_mese(dipendente, azienda, anno: int, mese: int):
    """
    Date (inizio, fine) del contratto da usare per il mese ``anno``/``mese``.

    Tra i :class:`~rapporto_di_lavoro.models.RapportoDiLavoro` che **intersecano** quel mese
    (stessa logica dei monti: esclusa solo la «Proposta»; fine nulla = indeterminato),
    si sceglie quello con ``data_inizio_rapporto`` più recente.

    Così il primo giorno «in rapporto» nel mese segue **data_inizio_rapporto** (es. assunzione
    02/04 → il 01/04 è fuori rapporto), non necessariamente ``data_assunzione`` in anagrafica
    né un contratto non vigente nel mese scelto.

    Se nessun rapporto interseca il mese, fallback su ``data_assunzione`` / ``data_cessazione``
    del dipendente (comportamento legacy).
    """
    from rapporto_di_lavoro.models import RapportoDiLavoro

    _, ult = calendar.monthrange(anno, mese)
    d0 = date(anno, mese, 1)
    d1 = date(anno, mese, ult)

    rapporto = (
        RapportoDiLavoro.objects.filter(azienda=azienda, dipendente=dipendente)
        .filter(data_inizio_rapporto__lte=d1)
        .filter(Q(data_fine_rapporto__isnull=True) | Q(data_fine_rapporto__gte=d0))
        .exclude(stato='proposta')
        .order_by('-data_inizio_rapporto', '-id')
        .first()
    )
    if rapporto:
        return rapporto.data_inizio_rapporto, rapporto.data_fine_rapporto

    data_inizio = getattr(dipendente, 'data_assunzione', None)
    data_fine = getattr(dipendente, 'data_cessazione', None)
    return data_inizio, data_fine


def _intervallo_mese_per_rapporto(anno, mese, data_inizio, data_fine):
    """Intersezione tra mese richiesto e periodo di validità rapporto."""
    _, ultimo = calendar.monthrange(anno, mese)
    d_start = date(anno, mese, 1)
    d_end = date(anno, mese, ultimo)

    if data_inizio and data_inizio > d_end:
        return None, None
    if data_fine and data_fine < d_start:
        return None, None

    if data_inizio and data_inizio > d_start:
        d_start = data_inizio
    if data_fine and data_fine < d_end:
        d_end = data_fine

    if d_start > d_end:
        return None, None
    return d_start, d_end


def _giorno_in_periodo_contrattuale(dipendente, azienda, data_giorno):
    """True se ``data_giorno`` cade nell'intersezione mese / periodo rapporto (come riepilogo)."""
    di, df = _periodo_rapporto_dipendente_per_mese(dipendente, azienda, data_giorno.year, data_giorno.month)
    ps, pe = _intervallo_mese_per_rapporto(data_giorno.year, data_giorno.month, di, df)
    return bool(ps and pe and ps <= data_giorno <= pe)


def _calendario_sidebar_ajax_payload(dipendente, anno: int, mese: int | None = None) -> dict:
    """Dati per aggiornare via AJAX il pannello destro (saldi monti + streak A + riepilogo ore mese)."""
    from .utils import riepilogo_ore_mese_sidebar, saldi_monti_calendario, streak_assenza_ingiustificata

    saldi = saldi_monti_calendario(dipendente, dipendente.azienda, anno, mese)
    streak = streak_assenza_ingiustificata(dipendente, anno)

    def _json_val(v):
        if isinstance(v, Decimal):
            return float(v)
        if isinstance(v, date):
            return v.isoformat()
        return v

    out = {
        'saldi': {k: _json_val(v) for k, v in saldi.items()},
        'streak': {k: _json_val(v) for k, v in streak.items()},
    }
    if mese is not None and 1 <= mese <= 12:
        ro = riepilogo_ore_mese_sidebar(dipendente, dipendente.azienda, anno, mese)
        out['riepilogo_ore'] = {
            'teoriche_da_pianificazione': ro['teoriche_da_pianificazione'],
            'ha_note_teorica_auto': ro['ha_note_teorica_auto'],
            'mostra_avviso_teoriche': ro['mostra_avviso_teoriche'],
            'ore_mensili_contratto': float(ro['ore_mensili_contratto']),
            'coefficiente_ore_contratto': float(ro['coefficiente_ore_contratto']),
            'ore_settimanali_contratto': float(ro['ore_settimanali_contratto']),
            'ha_regola_normativa_ccnl': ro['ha_regola_normativa_ccnl'],
            'giorni_lavorativi_teorici': ro['giorni_lavorativi_teorici'],
            'ferie_maturate_mese_gg': float(ro['ferie_maturate_mese_gg']),
            'permessi_maturati_mese_ore': float(ro['permessi_maturati_mese_ore']),
            'ore_teoriche': float(ro['ore_teoriche']),
            'ore_effettive_totali': float(ro['ore_effettive_totali']),
            'ore_entro_contratto_feriali': float(ro['ore_entro_contratto_feriali']),
            'ore_entro_contratto_domeniche': float(ro['ore_entro_contratto_domeniche']),
            'ore_entro_contratto_festivi': float(ro['ore_entro_contratto_festivi']),
            'straord_feriali': float(ro['straord_feriali']),
            'straord_diurno': float(ro['straord_diurno']),
            'straord_notturno': float(ro['straord_notturno']),
            'straord_domeniche': float(ro['straord_domeniche']),
            'straord_festivi': float(ro['straord_festivi']),
            'straord_nott_fest': float(ro['straord_nott_fest']),
        }
    return out


def _get_chiusura_settimanale_presenze(azienda, anno, mese):
    """Chiusure settimanali per presenze: priorità config mensile, poi annuale, poi legacy."""
    cfg_mese = ConfigurazioneOrarioMensile.objects.filter(azienda=azienda, anno=anno, mese=mese).first()
    if cfg_mese:
        fasce = FasciaAperturaMensile.objects.filter(configurazione=cfg_mese)
        if fasce.exists():
            out_fasce = sorted({int(f.giorno_settimana) for f in fasce if f.chiuso})
            if out_fasce:
                return out_fasce
        values = cfg_mese.giorni_riposo_settimanale or []
        try:
            out = sorted({int(v) for v in values if 0 <= int(v) <= 6})
            return out
        except Exception:
            return [6]

    cfg_last = _get_latest_config_orario_mensile(azienda, anno)
    if cfg_last:
        values = cfg_last.giorni_riposo_settimanale or []
        try:
            out = sorted({int(v) for v in values if 0 <= int(v) <= 6})
            return out
        except Exception:
            return [6]

    cfg = ConfigurazioneOrarioAnnuale.objects.filter(azienda=azienda, anno=anno).first()
    if cfg:
        fasce_ann = FasciaAperturaSettimanale.objects.filter(configurazione=cfg)
        if fasce_ann.exists():
            out_fasce = sorted({int(f.giorno_settimana) for f in fasce_ann if f.chiuso})
            if out_fasce:
                return out_fasce
        values = cfg.giorni_riposo_settimanale or []
        try:
            out = sorted({int(v) for v in values if 0 <= int(v) <= 6})
            return out
        except Exception:
            return [6]
    return get_chiusura_settimanale(azienda, anno, mese)


def _time_add_minutes(t, minutes):
    base = t.hour * 60 + t.minute + minutes
    base = max(0, min(base, 23 * 60 + 59))
    return time(base // 60, base % 60)


def _fasce_teoriche_da_config(fascia):
    """Restituisce (in_m, out_m, in_p, out_p) da configurazione, con fallback legacy."""
    if not fascia or fascia.chiuso:
        return None, None, None, None

    in_m = getattr(fascia, 'ora_apertura_mattina', None)
    out_m = getattr(fascia, 'ora_chiusura_mattina', None)
    in_p = getattr(fascia, 'ora_apertura_pomeriggio', None)
    out_p = getattr(fascia, 'ora_chiusura_pomeriggio', None)

    # Fallback compatibilità: vecchia fascia unica (es. 09:00-18:00)
    # → distribuisce in 2 blocchi (mattina/pomeriggio) mantenendo l'orario di fine.
    if in_m and out_m and not in_p and not out_p:
        durata = (out_m.hour * 60 + out_m.minute) - (in_m.hour * 60 + in_m.minute)
        if durata >= 8 * 60:
            out_m_new = _time_add_minutes(in_m, 4 * 60)
            in_p_new = _time_add_minutes(out_m_new, 60)
            if in_p_new < out_m:
                out_m = out_m_new
                in_p = in_p_new
                out_p = getattr(fascia, 'ora_chiusura_mattina', None)

    return in_m, out_m, in_p, out_p


def _genera_presenze_teoriche_mese_azienda(
    azienda, anno, mese, utente=None, solo_dipendente_id=None, solo_dipendenti_ids=None,
):
    """Generazione idempotente presenze teoriche del mese per dipendenti azienda.

    - solo_dipendente_id: un solo dipendente (calendario individuale).
    - solo_dipendenti_ids: elenco PK (es. pagina lista/riepilogo o dipendenti in export).
    - Se entrambi assenti: tutto l'organico (attivo/candidato/cessato) — es. pulsante pianificazione.
    """
    from .utils import presenze_mese_bloccate

    cfg, fasce_map = _get_config_orario_mese(azienda, anno, mese, create=True)

    if not cfg.genera_presenze_teoriche:
        return {
            'enabled': False,
            'created': 0,
            'updated': 0,
            'existing': 0,
            'fuori_rapporto': 0,
            'skipped_mese_chiuso': 0,
        }

    riposi = set(cfg.giorni_riposo_settimanale or [])

    _, ultimo = calendar.monthrange(anno, mese)
    d_inizio = date(anno, mese, 1)
    d_fine = date(anno, mese, ultimo)

    dip_qs = Dipendente.objects.filter(
        azienda=azienda, stato__in=['attivo', 'candidato', 'cessato']
    )
    if solo_dipendenti_ids is not None:
        ids = list(solo_dipendenti_ids)
        if not ids:
            return {
                'enabled': True,
                'created': 0,
                'updated': 0,
                'existing': 0,
                'fuori_rapporto': 0,
                'skipped_mese_chiuso': 0,
            }
        dip_qs = dip_qs.filter(pk__in=ids)
    elif solo_dipendente_id is not None:
        dip_qs = dip_qs.filter(pk=solo_dipendente_id)
    dipendenti = list(dip_qs.order_by('cognome', 'nome'))
    dip_ids = [d.id for d in dipendenti]
    # Una query per tutte le assegnazioni che intersecano il mese (evita N+1 per dipendente)
    assegnazioni_by_dip = defaultdict(list)
    if dip_ids:
        for ass in (
            AssegnazioneTurnoDipendente.objects.filter(
                dipendente_id__in=dip_ids,
                attivo=True,
                data_inizio__lte=d_fine,
            )
            .filter(Q(data_fine__isnull=True) | Q(data_fine__gte=d_inizio))
            .select_related('turno')
            .order_by('dipendente_id', '-data_inizio')
        ):
            assegnazioni_by_dip[ass.dipendente_id].append(ass)

    # Presenze già presenti nel mese (evita get_or_create = SELECT per ogni giorno×dipendente)
    presenza_map = {}
    if dip_ids:
        for pr in Presenza.objects.filter(
            dipendente_id__in=dip_ids,
            azienda=azienda,
            data__year=anno,
            data__month=mese,
        ):
            presenza_map[(pr.dipendente_id, pr.data)] = pr

    _teorica_bulk_update_fields = [
        'causale',
        'ora_entrata', 'ora_uscita',
        'ora_entrata2', 'ora_uscita2',
        'ora_entrata3', 'ora_uscita3',
        'data_modifica',
    ]

    created = 0
    updated = 0
    existing = 0
    fuori_rapporto = 0
    skipped_mese_chiuso = 0

    for dip in dipendenti:
        if presenze_mese_bloccate(dip, azienda, anno, mese):
            skipped_mese_chiuso += 1
            continue

        data_inizio, data_fine = _periodo_rapporto_dipendente_per_mese(dip, azienda, anno, mese)
        assegnazioni_turno = assegnazioni_by_dip.get(dip.id, [])
        to_create = []
        to_update = []

        cur = d_inizio
        while cur <= d_fine:
            if data_inizio and cur < data_inizio:
                fuori_rapporto += 1
                cur += timedelta(days=1)
                continue
            if data_fine and cur > data_fine:
                fuori_rapporto += 1
                cur += timedelta(days=1)
                continue

            wd = cur.weekday()
            fascia = fasce_map.get(wd)
            chiuso = wd in riposi or (fascia.chiuso if fascia else False)
            causale = 'R' if chiuso else 'P'

            in_m, out_m, in_p, out_p = _fasce_teoriche_da_config(fascia)

            if causale == 'R':
                in_m = out_m = in_p = out_p = None

            # ── Override con turni assegnati ─────────────────────────────
            # Raccoglie TUTTI i turni validi per il giorno, ordina per ora_inizio,
            # assegna in sequenza a T1, T2, T3 (Presenza supporta 3 slot).
            in_t3 = out_t3 = None
            if causale != 'R':
                turni_giorno = sorted(
                    (ass.turno for ass in assegnazioni_turno
                     if ass.data_inizio <= cur and (ass.data_fine is None or ass.data_fine >= cur)),
                    key=lambda t: t.ora_inizio,
                )
                if turni_giorno:
                    in_m  = turni_giorno[0].ora_inizio
                    out_m = turni_giorno[0].ora_fine
                    in_p  = turni_giorno[1].ora_inizio if len(turni_giorno) > 1 else None
                    out_p = turni_giorno[1].ora_fine   if len(turni_giorno) > 1 else None
                    in_t3  = turni_giorno[2].ora_inizio if len(turni_giorno) > 2 else None
                    out_t3 = turni_giorno[2].ora_fine   if len(turni_giorno) > 2 else None
            # ─────────────────────────────────────────────────────────────

            key = (dip.id, cur)
            presenza = presenza_map.get(key)
            if presenza is None:
                ts = timezone.now()
                to_create.append(Presenza(
                    dipendente_id=dip.id,
                    data=cur,
                    azienda_id=azienda.pk,
                    causale=causale,
                    ora_entrata=in_m,
                    ora_uscita=out_m,
                    ora_entrata2=in_p,
                    ora_uscita2=out_p,
                    ora_entrata3=in_t3,
                    ora_uscita3=out_t3,
                    note='TEORICA_AUTO',
                    registrata_da_id=utente.pk if utente else None,
                    ore_straordinario=Decimal('0'),
                    data_registrazione=ts,
                    data_modifica=ts,
                ))
            else:
                existing += 1
                note = (presenza.note or '')
                if (
                    'TEORICA_AUTO' in note
                    and 'src=' not in note
                    and presenza.causale in {'P', 'R'}
                ):
                    changed = (
                        presenza.causale != causale
                        or presenza.ora_entrata != in_m
                        or presenza.ora_uscita != out_m
                        or presenza.ora_entrata2 != in_p
                        or presenza.ora_uscita2 != out_p
                        or presenza.ora_entrata3 != in_t3
                        or presenza.ora_uscita3 != out_t3
                    )
                    if changed:
                        presenza.causale = causale
                        presenza.ora_entrata = in_m
                        presenza.ora_uscita = out_m
                        presenza.ora_entrata2 = in_p
                        presenza.ora_uscita2 = out_p
                        presenza.ora_entrata3 = in_t3
                        presenza.ora_uscita3 = out_t3
                        presenza.data_modifica = timezone.now()
                        to_update.append(presenza)

            cur += timedelta(days=1)

        if to_create:
            Presenza.objects.bulk_create(to_create, batch_size=500)
            created += len(to_create)
            for p in to_create:
                presenza_map[(p.dipendente_id, p.data)] = p
        if to_update:
            Presenza.objects.bulk_update(to_update, _teorica_bulk_update_fields)
            updated += len(to_update)

    return {
        'enabled': True,
        'created': created,
        'updated': updated,
        'existing': existing,
        'fuori_rapporto': fuori_rapporto,
        'skipped_mese_chiuso': skipped_mese_chiuso,
    }


@login_required
def pianificazione_orari_annuale(request):
    """Configurazione orari con persistenza mensile e opzione applica a tutti i mesi."""
    if not (request.user.is_superuser or request.user.has_ruolo('admin')):
        return HttpResponseForbidden('Accesso negato')

    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('lista_presenze')

    oggi = date.today()
    anno = _parse_int(request.GET.get('anno') or request.POST.get('anno'), oggi.year)
    mese_raw = request.GET.get('mese') or request.POST.get('mese')
    if mese_raw in (None, ''):
        mese_raw = request.session.get('presenze_pianificazione_mese_sel')
    mese_default = _get_latest_config_orario_mensile(azienda, anno)
    mese_sel = _parse_int(mese_raw, mese_default.mese if mese_default else oggi.month)
    if mese_sel < 1 or mese_sel > 12:
        mese_sel = oggi.month
    request.session['presenze_pianificazione_mese_sel'] = mese_sel

    config, _ = ConfigurazioneOrarioMensile.objects.get_or_create(
        azienda=azienda,
        anno=anno,
        mese=mese_sel,
        defaults={
            'giorni_riposo_settimanale': [6],
            'genera_presenze_teoriche': True,
        },
    )
    _ensure_fasce_mensili(config)

    config_annuale, _ = ConfigurazioneOrarioAnnuale.objects.get_or_create(
        azienda=azienda,
        anno=anno,
        defaults={
            'giorni_riposo_settimanale': [6],
            'genera_presenze_teoriche': True,
        },
    )

    FasceFormSet = modelformset_factory(
        FasciaAperturaMensile,
        form=FasciaAperturaMensileForm,
        extra=0,
        can_delete=False,
    )
    TurniFormSet = modelformset_factory(
        TurnoLavorativoAziendale,
        form=TurnoLavorativoAziendaleForm,
        extra=1,
        can_delete=True,
    )

    fasce_qs = FasciaAperturaMensile.objects.filter(configurazione=config).order_by('giorno_settimana')
    turni_qs = TurnoLavorativoAziendale.objects.filter(configurazione=config_annuale).order_by('ordine', 'ora_inizio')

    if request.method == 'POST':
        config_form = ConfigurazioneOrarioMensileForm(request.POST, instance=config, prefix='cfg')
        fasce_formset = FasceFormSet(request.POST, queryset=fasce_qs, prefix='fasce')
        turni_formset = TurniFormSet(request.POST, queryset=turni_qs, prefix='turni')

        if config_form.is_valid() and fasce_formset.is_valid() and turni_formset.is_valid():
            config = config_form.save()

            riposi = set(config.giorni_riposo_settimanale or [])
            fasce_objs = fasce_formset.save(commit=False)
            for obj in fasce_objs:
                obj.configurazione = config
                if obj.giorno_settimana in riposi:
                    obj.chiuso = True
                obj.save()

            if request.POST.get('applica_tutti_mesi') == '1':
                source_map = {
                    f.giorno_settimana: f
                    for f in FasciaAperturaMensile.objects.filter(configurazione=config)
                }
                for mese_target in range(1, 13):
                    if mese_target == mese_sel:
                        continue
                    cfg_target, _ = ConfigurazioneOrarioMensile.objects.get_or_create(
                        azienda=azienda,
                        anno=anno,
                        mese=mese_target,
                        defaults={
                            'giorni_riposo_settimanale': config.giorni_riposo_settimanale,
                            'genera_presenze_teoriche': config.genera_presenze_teoriche,
                        },
                    )
                    cfg_target.giorni_riposo_settimanale = config.giorni_riposo_settimanale
                    cfg_target.genera_presenze_teoriche = config.genera_presenze_teoriche
                    cfg_target.save(update_fields=['giorni_riposo_settimanale', 'genera_presenze_teoriche', 'data_modifica'])

                    _ensure_fasce_mensili(cfg_target)
                    for wd in range(7):
                        src = source_map.get(wd)
                        dst, _ = FasciaAperturaMensile.objects.get_or_create(
                            configurazione=cfg_target,
                            giorno_settimana=wd,
                            defaults={'chiuso': wd in riposi},
                        )
                        if src:
                            dst.chiuso = src.chiuso
                            dst.ora_apertura_mattina = src.ora_apertura_mattina
                            dst.ora_chiusura_mattina = src.ora_chiusura_mattina
                            dst.ora_apertura_pomeriggio = src.ora_apertura_pomeriggio
                            dst.ora_chiusura_pomeriggio = src.ora_chiusura_pomeriggio
                        else:
                            dst.chiuso = wd in riposi
                        if dst.giorno_settimana in set(cfg_target.giorni_riposo_settimanale or []):
                            dst.chiuso = True
                        dst.save()

            turni_objs = turni_formset.save(commit=False)
            for obj in turni_formset.deleted_objects:
                obj.delete()
            for obj in turni_objs:
                obj.configurazione = config_annuale
                obj.save()

            config = ConfigurazioneOrarioMensile.objects.get(pk=config.pk)

            if request.POST.get('applica_tutti_mesi') == '1':
                messages.success(
                    request,
                    f'Pianificazione mensile salvata (ID={config.id}, aggiornamento={timezone.localtime(config.data_modifica).strftime("%d/%m/%Y %H:%M:%S")}) e applicata a tutti i mesi dell’anno.',
                )
            else:
                messages.success(
                    request,
                    f'Pianificazione mensile salvata (ID={config.id}, aggiornamento={timezone.localtime(config.data_modifica).strftime("%d/%m/%Y %H:%M:%S")}).',
                )
            return _redirect_pianificazione_orari_annuale(anno, mese_sel)
        messages.error(request, 'Verifica i dati inseriti nella pianificazione orari.')
    else:
        config_form = ConfigurazioneOrarioMensileForm(instance=config, prefix='cfg')
        fasce_formset = FasceFormSet(queryset=fasce_qs, prefix='fasce')
        turni_formset = TurniFormSet(queryset=turni_qs, prefix='turni')

    return render(request, 'presenze/pianificazione_orari_annuale.html', {
        'azienda': azienda,
        'anno': anno,
        'anno_prev': anno - 1,
        'anno_next': anno + 1,
        'mese_sel': mese_sel,
        'mese_nome': MESI_ITA[mese_sel],
        'mesi_choices': [(i, MESI_ITA[i]) for i in range(1, 13)],
        'config_form': config_form,
        'fasce_formset': fasce_formset,
        'turni_formset': turni_formset,
        'db_snapshot': _snapshot_parametri_orari_mese(config),
    })


@login_required
def api_parametri_orari_mese(request):
    """Interrogazione dati parametri orari mensili direttamente dal DB (JSON)."""
    if not (request.user.is_superuser or request.user.has_ruolo('admin')):
        return HttpResponseForbidden('Accesso negato')

    azienda = _get_azienda(request)
    if not azienda:
        return JsonResponse({'ok': False, 'error': 'Azienda operativa non selezionata'}, status=400)

    oggi = date.today()
    anno = _parse_int(request.GET.get('anno'), oggi.year)
    mese = _parse_int(request.GET.get('mese'), oggi.month)
    if mese < 1 or mese > 12:
        return JsonResponse({'ok': False, 'error': 'Mese non valido'}, status=400)

    cfg = ConfigurazioneOrarioMensile.objects.filter(azienda=azienda, anno=anno, mese=mese).first()
    if not cfg:
        return JsonResponse({
            'ok': False,
            'error': 'Nessuna configurazione mensile trovata',
            'azienda_id': azienda.id,
            'anno': anno,
            'mese': mese,
        }, status=404)

    fasce = []
    for f in FasciaAperturaMensile.objects.filter(configurazione=cfg).order_by('giorno_settimana'):
        fasce.append({
            'giorno_settimana': f.giorno_settimana,
            'giorno_label': f.get_giorno_settimana_display(),
            'chiuso': f.chiuso,
            'ora_apertura_mattina': f.ora_apertura_mattina.strftime('%H:%M') if f.ora_apertura_mattina else None,
            'ora_chiusura_mattina': f.ora_chiusura_mattina.strftime('%H:%M') if f.ora_chiusura_mattina else None,
            'ora_apertura_pomeriggio': f.ora_apertura_pomeriggio.strftime('%H:%M') if f.ora_apertura_pomeriggio else None,
            'ora_chiusura_pomeriggio': f.ora_chiusura_pomeriggio.strftime('%H:%M') if f.ora_chiusura_pomeriggio else None,
        })

    return JsonResponse({
        'ok': True,
        'azienda_id': azienda.id,
        'azienda_nome': azienda.nome,
        'configurazione_id': cfg.id,
        'anno': cfg.anno,
        'mese': cfg.mese,
        'giorni_riposo_settimanale': cfg.giorni_riposo_settimanale or [],
        'genera_presenze_teoriche': cfg.genera_presenze_teoriche,
        'data_modifica': timezone.localtime(cfg.data_modifica).isoformat(),
        'fasce': fasce,
    })


@login_required
@require_POST
def genera_presenze_teoriche_mese(request):
    """Genera presenze teoriche del mese da pianificazione orari mensile."""
    if not (request.user.is_superuser or request.user.has_ruolo('admin')):
        return HttpResponseForbidden('Accesso negato')

    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('lista_presenze')

    anno = _parse_int(request.POST.get('anno'), date.today().year)
    mese = _parse_int(request.POST.get('mese'), date.today().month)
    if mese < 1 or mese > 12:
        messages.error(request, 'Mese non valido per generazione presenze teoriche.')
        return _redirect_pianificazione_orari_annuale(anno, date.today().month)

    summary = _genera_presenze_teoriche_mese_azienda(azienda, anno, mese, request.user)
    if not summary['enabled']:
        messages.warning(request, 'Generazione disattivata nella configurazione mensile.')
        return _redirect_pianificazione_orari_annuale(anno, mese)

    messages.success(
        request,
        (
            f'Generate presenze teoriche {MESI_ITA[mese]} {anno}: '
            f'create={summary["created"]}, aggiornate={summary["updated"]}, già_esistenti={summary["existing"]}, fuori_rapporto={summary["fuori_rapporto"]}.'
        ),
    )
    if summary.get('skipped_mese_chiuso'):
        messages.info(
            request,
            f'Saltati {summary["skipped_mese_chiuso"]} dipendenti con mese chiuso (riepilogo approvato o elaborato).',
        )
    return _redirect_pianificazione_orari_annuale(anno, mese)


def _ore_tra(t1, t2):
    """Calcola le ore DECIMALI tra due time objects. Restituisce 0.0 se mancano.
    Formato output: decimale (es. 8.5 = 8h30). Mai HH:MM nei calcoli."""
    if not t1 or not t2:
        return 0.0
    mins = (t2.hour * 60 + t2.minute) - (t1.hour * 60 + t1.minute)
    return round(max(0.0, mins / 60.0), 4)   # 4 decimali per precision interna


def _min_tra(t1, t2) -> int:
    """Minuti interi tra due time objects. Restituisce 0 se mancano o negativi."""
    if not t1 or not t2:
        return 0
    m = (t2.hour * 60 + t2.minute) - (t1.hour * 60 + t1.minute)
    return m if m > 0 else 0


# Causali per cui T1–T3 concorrono al «totale ore lavorate» (coerente con la griglia riepilogo / calendario).
_CAUSALI_ORE_LAVORATE = frozenset(('P', 'ST', 'SMART', 'FE'))


def _azzera_turni_secondo_e_terzo_se_causale_giorno_intero(p: Presenza) -> None:
    """
    Per causali che non registro ore da fascia (ferie, malattia, permesso senza orario, assenza,
    riposo, CIG, …): la variazione vale sull’intera giornata — elimina 2° e 3° turno così non restano
    solo le prime fasce aggiornate dal massivo o dal form.
    """
    if p.causale in _CAUSALI_ORE_LAVORATE:
        return
    p.ora_entrata2 = p.ora_uscita2 = None
    p.ora_entrata3 = p.ora_uscita3 = None


def _minuti_turni_presenza(p) -> tuple[int, int, int]:
    """
    Minuti per turno (T1, T2, T3) con deduplica degli intervalli identici.
    Evita il raddoppio quando lo stesso orario è stato copiato su turno 2/3.
    """
    if not p:
        return 0, 0, 0
    raw = [
        _min_tra(p.ora_entrata, p.ora_uscita),
        _min_tra(p.ora_entrata2, p.ora_uscita2),
        _min_tra(p.ora_entrata3, p.ora_uscita3),
    ]
    keys = [
        (p.ora_entrata, p.ora_uscita),
        (p.ora_entrata2, p.ora_uscita2),
        (p.ora_entrata3, p.ora_uscita3),
    ]
    seen = set()
    out = []
    for mins, key in zip(raw, keys):
        if mins <= 0:
            out.append(0)
            continue
        if key in seen:
            out.append(0)
            continue
        seen.add(key)
        out.append(mins)
    return out[0], out[1], out[2]


def _minuti_lavorati_presenza(p) -> int:
    """
    Minuti lavorati T1+T2+T3 da un record Presenza (totale mese / export).
    Riposo (R): 0. Per altre causali (ferie, malattia, …) gli orari eventualmente presenti
    a DB non si sommano: altrimenti copie massive o dati incoerenti gonfiano il totale rispetto
    alle ore mostrate in griglia (solo P/ST/SMART/FE).
    """
    if not p or p.causale == 'R':
        return 0
    if p.causale not in _CAUSALI_ORE_LAVORATE:
        return 0
    m1, m2, m3 = _minuti_turni_presenza(p)
    return m1 + m2 + m3


def _straord_giorno_semplice(ore_giorno: float, ore_std: Decimal, presenza) -> Optional[Decimal]:
    """
    Ore eccedenti rispetto allo std giornaliero (stessa idea della riga «Straordinario»
    nel calendario dipendente: max(0, ore_lavorate − std)).
    """
    if not presenza or presenza.causale == 'R':
        return None
    og = Decimal(str(round(ore_giorno, 4)))
    std = ore_std if isinstance(ore_std, Decimal) else Decimal(str(ore_std))
    ex = og - std
    if ex <= 0:
        return None
    return ex.quantize(Decimal('0.01'))


def _ore_standard_da_fascia(fascia) -> float:
    """Calcola ore standard da una fascia (mensile o annuale)."""
    if not fascia or getattr(fascia, 'chiuso', False):
        return 0.0
    tot_min = 0
    tot_min += _min_tra(getattr(fascia, 'ora_apertura_mattina', None), getattr(fascia, 'ora_chiusura_mattina', None))
    tot_min += _min_tra(getattr(fascia, 'ora_apertura_pomeriggio', None), getattr(fascia, 'ora_chiusura_pomeriggio', None))
    return round(max(0, tot_min) / 60.0, 4)


def _ore_std_giornaliere(azienda, anno, mese, data_rif=None) -> float:
    """
    Calcola le ore contrattuali giornaliere standard dell'azienda per il mese dato.
    Formula: ore_settimanali / giorni_lavorativi_settimana
    Usa chiusura_sett per non penalizzare le aziende che lavorano 6-7 gg/settimana.
    """
    if data_rif is not None:
        try:
            cfg_mese = ConfigurazioneOrarioMensile.objects.filter(
                azienda=azienda,
                anno=data_rif.year,
                mese=data_rif.month,
            ).first()
            if cfg_mese:
                fascia = FasciaAperturaMensile.objects.filter(
                    configurazione=cfg_mese,
                    giorno_settimana=data_rif.weekday(),
                ).first()
                ore = _ore_standard_da_fascia(fascia)
                if ore > 0:
                    return ore

            cfg = ConfigurazioneOrarioAnnuale.objects.filter(
                azienda=azienda,
                anno=data_rif.year,
            ).first()
            if cfg:
                fascia = FasciaAperturaSettimanale.objects.filter(
                    configurazione=cfg,
                    giorno_settimana=data_rif.weekday(),
                ).first()
                ore = _ore_standard_da_fascia(fascia)
                if ore > 0:
                    return ore
        except Exception:
            pass

    chiusura = set(_get_chiusura_settimanale_presenze(azienda, anno, mese))
    giorni_lav = max(1, 7 - len(chiusura))
    return float(azienda.ore_settimanali_standard or 40) / giorni_lav


def _ricalcola_straordinario(presenza, azienda) -> None:
    """
    Ricalcola e salva ore_straordinario sulla presenza.
    Regola: straordinario = max(0, ore_lavorate_totali - ore_contrattuali_giornaliere)
    Deve essere chiamata DOPO aver aggiornato gli orari del turno.
    """
    ore_std = _ore_std_giornaliere(azienda, presenza.data.year, presenza.data.month, presenza.data)
    ore_lav = presenza.ore_lavorate()          # somma T1 + T2 + T3 in ore decimali
    eccedenza = ore_lav - ore_std
    presenza.ore_straordinario = Decimal(str(round(max(0.0, eccedenza), 2)))


def ore_dec_to_hhmm(decimal_ore) -> str:
    """
    Converte ore decimali → stringa HH:MM per la VISUALIZZAZIONE.
    Regola: i calcoli usano sempre decimali; HH:MM è solo per display.
    8.5 → '8:30' | 8.75 → '8:45' | 0.25 → '0:15' | 0 → '0:00'
    Importabile da altri moduli: from presenze.views import ore_dec_to_hhmm
    """
    if not decimal_ore:
        return '0:00'
    try:
        total_min = round(float(decimal_ore) * 60)
        if total_min <= 0:
            return '0:00'
        return f'{total_min // 60}:{total_min % 60:02d}'
    except (ValueError, TypeError):
        return '0:00'


def _get_festivi_mese(anno, mese):
    """Restituisce il set di date festive italiane nel mese dato (calcola Pasqua con algoritmo Butcher)."""
    a = anno % 19; b = anno // 100; c = anno % 100
    d = b // 4; e = b % 4; f = (b + 8) // 25
    gv = (b - f + 1) // 3
    h = (19 * a + b - d - gv + 15) % 30
    i = c // 4; k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    p_mese = (h + l - 7 * m + 114) // 31
    p_giorno = ((h + l - 7 * m + 114) % 31) + 1
    pasqua = date(anno, p_mese, p_giorno)
    tutti = {
        date(anno, 1,  1),              # Capodanno
        date(anno, 1,  6),              # Epifania
        pasqua,                         # Pasqua
        pasqua + timedelta(days=1),     # Lunedì dell'Angelo
        date(anno, 4, 25),              # Liberazione
        date(anno, 5,  1),              # Festa del Lavoro
        date(anno, 6,  2),              # Festa della Repubblica
        date(anno, 8, 15),              # Ferragosto
        date(anno, 11, 1),              # Ognissanti
        date(anno, 12, 8),              # Immacolata Concezione
        date(anno, 12, 25),             # Natale
        date(anno, 12, 26),             # Santo Stefano
    }
    return {d for d in tutti if d.month == mese}


def _get_azienda(request):
    """Restituisce l'azienda operativa in base al ruolo."""
    user = request.user
    if user.is_superuser or user.has_ruolo('admin'):
        return get_azienda_operativa(user, request.session)
    elif user.has_ruolo('hr') or user.has_ruolo('consulente'):
        return getattr(user, 'azienda', None)
    return None


def _is_admin_hr(user):
    """Admin, HR o consulente del lavoro possono accedere alle presenze."""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user.has_ruolo('admin') or user.has_ruolo('hr') or user.has_ruolo('consulente')


def _costruisci_griglia_mese(anno, mese, presenze_dict, chiusura_sett=None, festivi=None):
    """
    Restituisce una lista di settimane (liste da 7 elementi).
    Ogni elemento è un dict giorno oppure None (padding).
    Includes is_festivo, ore_t1/t2/t3 per day.
    """
    chiusura_sett = set(chiusura_sett or [])
    festivi = set(festivi or [])
    _, num_giorni = calendar.monthrange(anno, mese)
    primo_wd = date(anno, mese, 1).weekday()  # 0=Lun

    days = [None] * primo_wd
    for g in range(1, num_giorni + 1):
        d = date(anno, mese, g)
        dw = d.weekday()
        is_chiusura_sett = dw in chiusura_sett
        is_weekend = dw >= 5
        p = presenze_dict.get(d)
        days.append({
            'data': d,
            'dstr': d.isoformat(),
            'num': g,
            'giorno_settimana': GIORNI_ITA[dw],
            'is_weekend': is_weekend,
            'is_chiusura_sett': is_chiusura_sett,
            'is_festivo': d in festivi,
            # Sabato/domenica sono lavorativi se non in chiusura_sett (ristorazione)
            'is_lavorativo': not is_chiusura_sett,
            'presenza': p,
            'ore_t1': _ore_tra(p.ora_entrata,  p.ora_uscita)  if p else 0.0,
            'ore_t2': _ore_tra(p.ora_entrata2, p.ora_uscita2) if p else 0.0,
            'ore_t3': _ore_tra(p.ora_entrata3, p.ora_uscita3) if p else 0.0,
        })

    weeks = []
    for i in range(0, len(days), 7):
        week = days[i:i + 7]
        while len(week) < 7:
            week.append(None)
        weeks.append(week)
    return weeks


def _calendario_applica_periodo_contrattuale(ps, pe, giorni, griglia, panorama_giorni):
    """
    Imposta ``fuori_rapporto`` su ogni giorno e azzera le ore mostrate se il giorno è
    fuori da ``[ps, pe]`` (coerente con riepilogo multi-dipendente).
    """
    def _in(d):
        return bool(ps and pe and ps <= d <= pe)

    for gi in giorni:
        gi['fuori_rapporto'] = not _in(gi['data'])

    for week in griglia:
        for cell in week:
            if cell is None:
                continue
            fuori = not _in(cell['data'])
            cell['fuori_rapporto'] = fuori
            if fuori:
                cell['ore_t1'] = cell['ore_t2'] = cell['ore_t3'] = 0.0

    if panorama_giorni is not None:
        for pg in panorama_giorni:
            fuori = not _in(pg['data'])
            pg['fuori_rapporto'] = fuori
            if fuori:
                pg['ore_t1'] = pg['ore_t2'] = pg['ore_t3'] = 0.0
                pg['ore_giorno'] = 0.0


def _costruisci_mese(anno, mese, presenze_dict, chiusura_sett=None):
    """Costruisce la lista dei giorni del mese con dati presenza."""
    chiusura_sett = set(chiusura_sett or [])
    _, num_giorni = calendar.monthrange(anno, mese)
    giorni = []
    for g in range(1, num_giorni + 1):
        d = date(anno, mese, g)
        dw = d.weekday()  # 0=lun, 6=dom
        is_weekend = dw >= 5
        is_chiusura_sett = dw in chiusura_sett
        presenza = presenze_dict.get(d)
        giorni.append({
            'data': d,
            'num': g,
            'giorno_settimana': GIORNI_ITA[dw],
            'is_weekend': is_weekend,
            'is_chiusura_sett': is_chiusura_sett,
            'is_lavorativo': not is_chiusura_sett,
            'presenza': presenza,
        })
    return giorni


# ── 1. Lista dipendenti (legacy elenco) / redirect griglia mensile ──────────
@login_required
def lista_dipendenti_presenze(request):
    """Admin/HR: per default reindirizza alla griglia mensile (vista calendario tutti i dipendenti)."""
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")

    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('profile')

    if request.GET.get('vista') != 'elenco':
        q = {k: v for k, v in request.GET.items() if k in ('anno', 'mese', 'q', 'page') and v not in (None, '')}
        url = reverse('riepilogo_presenze_mese')
        if q:
            url = f'{url}?{urlencode(q)}'
        return redirect(url)

    oggi = date.today()
    anno = _parse_int(request.GET.get('anno'), oggi.year)
    mese = _parse_int(request.GET.get('mese'), oggi.month)

    from .utils import dipendenti_per_riepilogo_mese

    dipendenti = dipendenti_per_riepilogo_mese(azienda, anno, mese)
    q_filter = (request.GET.get('q') or '').strip()
    if q_filter:
        qobj = (
            Q(cognome__icontains=q_filter)
            | Q(nome__icontains=q_filter)
            | Q(codice_fiscale__icontains=q_filter)
        )
        if q_filter.isdigit():
            qobj |= Q(matricola=int(q_filter))
        dipendenti = dipendenti.filter(qobj)

    paginator = Paginator(dipendenti, 25)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    # Conteggio presenze del mese per ciascun dipendente
    # Usa chiusura_sett dell'azienda e validità del rapporto dipendente.
    chiusura_sett_az = set(_get_chiusura_settimanale_presenze(azienda, anno, mese))
    _, num_giorni = calendar.monthrange(anno, mese)

    dips_list = list(page_obj.object_list)
    dip_ids = [d.id for d in dips_list]
    # Teoriche solo per i dipendenti di questa pagina (lista è paginata)
    _genera_presenze_teoriche_mese_azienda(
        azienda, anno, mese, request.user, solo_dipendenti_ids=dip_ids,
    )

    dip_periods = {}
    for dip in dips_list:
        data_inizio, data_fine = _periodo_rapporto_dipendente_per_mese(dip, azienda, anno, mese)
        periodo_start, periodo_end = _intervallo_mese_per_rapporto(anno, mese, data_inizio, data_fine)
        dip_periods[dip.id] = (periodo_start, periodo_end)

    counts_by_dip = defaultdict(lambda: defaultdict(int))
    if dip_ids:
        for did, ddata, cau in Presenza.objects.filter(
            azienda=azienda,
            dipendente_id__in=dip_ids,
            data__year=anno,
            data__month=mese,
        ).values_list('dipendente_id', 'data', 'causale'):
            ps, pe = dip_periods.get(did, (None, None))
            if not (ps and pe) or not (ps <= ddata <= pe):
                continue
            counts_by_dip[did][cau] += 1

    dip_info = []
    for dip in dips_list:
        periodo_start, periodo_end = dip_periods[dip.id]
        giorni_lavorativi = 0
        cnt = {}
        if periodo_start and periodo_end:
            cur = periodo_start
            while cur <= periodo_end:
                if cur.weekday() not in chiusura_sett_az:
                    giorni_lavorativi += 1
                cur += timedelta(days=1)
            cnt = dict(counts_by_dip[dip.id])

        registrate = sum(cnt.values())
        dip_info.append({
            'dip': dip,
            'registrate': registrate,
            'giorni_lavorativi': giorni_lavorativi,
            'mancanti': max(0, giorni_lavorativi - registrate),
            'conteggi': cnt,
            'periodo_start': periodo_start,
            'periodo_end': periodo_end,
            'fuori_periodo_mese': not (periodo_start and periodo_end),
        })

    # Navigazione mese
    if mese == 1:
        mese_prev, anno_prev = 12, anno - 1
    else:
        mese_prev, anno_prev = mese - 1, anno
    if mese == 12:
        mese_next, anno_next = 1, anno + 1
    else:
        mese_next, anno_next = mese + 1, anno

    return render(request, 'presenze/dipendenti.html', {
        'dip_info': dip_info,
        'page_obj': page_obj,
        'q_filter': q_filter,
        'azienda': azienda,
        'anno': anno,
        'mese': mese,
        'mese_nome': MESI_ITA[mese],
        'anno_prev': anno_prev, 'mese_prev': mese_prev,
        'anno_next': anno_next, 'mese_next': mese_next,
        'causale_choices': Presenza.CAUSALE_CHOICES,
    })


# ── 2. Calendario mensile per un dipendente ─────────────────────────────────
@login_required
def calendario_presenze(request, dipendente_id, anno=None, mese=None):
    """Calendario mensile interattivo per un singolo dipendente."""
    is_admin_hr = _is_admin_hr(request.user)
    dip_self = get_dipendente_collegato(request.user)
    is_self = bool(dip_self and dip_self.id == dipendente_id)
    if not is_admin_hr and not is_self:
        return HttpResponseForbidden("Accesso negato")

    sola_lettura = not is_admin_hr  # dipendente vede in sola lettura

    if is_admin_hr:
        azienda = _get_azienda(request)
        dipendente = get_object_or_404(Dipendente, id=dipendente_id, azienda=azienda)
    else:
        # Stesso dipendente risolto dal portale (anche se il FK utente in anagrafica non è valorizzato)
        dipendente = dip_self

    oggi = date.today()
    anno = _parse_int(anno or request.GET.get('anno'), oggi.year)
    mese = _parse_int(mese or request.GET.get('mese'), oggi.month)

    from .utils import presenze_mese_bloccate

    mese_bloccato_riepilogo = presenze_mese_bloccate(dipendente, dipendente.azienda, anno, mese)
    # Mantiene popolato il mese solo per questo dipendente (idempotente; non tutto l'organico)
    if not mese_bloccato_riepilogo:
        _genera_presenze_teoriche_mese_azienda(
            dipendente.azienda, anno, mese, request.user, solo_dipendente_id=dipendente.id
        )

    calendario_modificabile = is_admin_hr and not mese_bloccato_riepilogo

    # Chiusura settimanale dell'azienda per questo mese
    chiusura_sett = _get_chiusura_settimanale_presenze(dipendente.azienda, anno, mese)

    di, df = _periodo_rapporto_dipendente_per_mese(dipendente, dipendente.azienda, anno, mese)
    ps, pe = _intervallo_mese_per_rapporto(anno, mese, di, df)

    # Presenze del mese
    presenze_qs = Presenza.objects.filter(
        dipendente=dipendente, data__year=anno, data__month=mese
    )
    presenze_dict = {p.data: p for p in presenze_qs}
    festivi = _get_festivi_mese(anno, mese)
    giorni = _costruisci_mese(anno, mese, presenze_dict, chiusura_sett)
    griglia = _costruisci_griglia_mese(anno, mese, presenze_dict, chiusura_sett, festivi)
    festivi_set = festivi
    panorama_giorni = []
    for gi in giorni:
        d = gi['data']
        p = gi['presenza']
        if p and p.causale == 'R':
            m1 = m2 = m3 = 0
        elif p and p.causale not in _CAUSALI_ORE_LAVORATE:
            m1 = m2 = m3 = 0
        else:
            m1, m2, m3 = _minuti_turni_presenza(p)
        min_giorno = m1 + m2 + m3
        panorama_giorni.append({
            'data': d,
            'num': gi['num'],
            'is_weekend': gi['is_weekend'],
            'is_domenica': d.weekday() == 6,
            'is_chiusura_sett': gi['is_chiusura_sett'],
            'is_festivo': d in festivi_set,
            'presenza': p,
            'ore_t1': m1 / 60,
            'ore_t2': m2 / 60,
            'ore_t3': m3 / 60,
            'ore_giorno': min_giorno / 60,
        })
    _calendario_applica_periodo_contrattuale(ps, pe, giorni, griglia, panorama_giorni)

    # Totali mese — solo giorni nell'intervallo contrattuale (come riepilogo multi-dipendente)
    if ps and pe:
        presenze_qs_periodo = presenze_qs.filter(data__gte=ps, data__lte=pe)
    else:
        presenze_qs_periodo = presenze_qs.none()
    totale_ore = sum(_minuti_lavorati_presenza(p) for p in presenze_qs_periodo) / 60
    totale_straord = sum(float(p.ore_straordinario) for p in presenze_qs_periodo)

    # Turni con almeno entrata o uscita nel periodo contratto del mese (stessa regola per tutti i dip./mesi)
    turno1_attivo = any(p.ora_entrata or p.ora_uscita for p in presenze_qs_periodo)
    turno2_attivo = any(p.ora_entrata2 or p.ora_uscita2 for p in presenze_qs_periodo)
    turno3_attivo = any(p.ora_entrata3 or p.ora_uscita3 for p in presenze_qs_periodo)
    conteggi_causale = {}
    for p in presenze_qs_periodo:
        conteggi_causale[p.causale] = conteggi_causale.get(p.causale, 0) + 1

    # Navigazione mese
    if mese == 1:
        mese_prev, anno_prev = 12, anno - 1
    else:
        mese_prev, anno_prev = mese - 1, anno
    if mese == 12:
        mese_next, anno_next = 1, anno + 1
    else:
        mese_next, anno_next = mese + 1, anno

    # Form vuoto per nuovo inserimento
    form_vuoto = GiornoPresenzaForm()

    # Ore contrattuali giornaliere = ore settimanali / giorni lavorativi settimanali
    # Giorni lavorativi = 7 - chiusura settimanale (sabato, domenica e festivi INCLUSI)
    giorni_chiusura = len(set(chiusura_sett or []))
    giorni_lav_sett = max(1, 7 - giorni_chiusura)
    ore_sett = float(dipendente.azienda.ore_settimanali_standard or 40)
    ore_giornaliere_standard = round(ore_sett / giorni_lav_sett, 4)

    from .utils import riepilogo_ore_mese_sidebar, saldi_monti_calendario, streak_assenza_ingiustificata

    saldi_sidebar = saldi_monti_calendario(dipendente, dipendente.azienda, anno, mese)
    streak_assenza_a = streak_assenza_ingiustificata(dipendente, anno)
    riepilogo_ore_mese = riepilogo_ore_mese_sidebar(dipendente, dipendente.azienda, anno, mese)

    return render(request, 'presenze/calendario.html', {
        'dipendente': dipendente,
        'giorni': giorni,
        'anno': anno,
        'mese': mese,
        'mese_nome': MESI_ITA[mese],
        'anno_prev': anno_prev, 'mese_prev': mese_prev,
        'anno_next': anno_next, 'mese_next': mese_next,
        'totale_ore': round(totale_ore, 2),
        'totale_straord': round(totale_straord, 2),
        'conteggi_causale': conteggi_causale,
        'causale_choices': Presenza.CAUSALE_CHOICES,
        'causale_colori': Presenza.CAUSALE_COLORI,
        'form_vuoto': form_vuoto,
        'chiusura_sett': chiusura_sett,
        'festivi': festivi,
        'griglia': griglia,
        'panorama_giorni': panorama_giorni,
        'turno1_attivo': turno1_attivo,
        'turno2_attivo': turno2_attivo,
        'turno3_attivo': turno3_attivo,
        'ore_giornaliere_standard': ore_giornaliere_standard,
        'sola_lettura': sola_lettura,
        'is_admin_hr': is_admin_hr,
        'mese_bloccato_riepilogo': mese_bloccato_riepilogo,
        'calendario_modificabile': calendario_modificabile,
        'saldi_sidebar': saldi_sidebar,
        'streak_assenza_a': streak_assenza_a,
        'riepilogo_ore_mese': riepilogo_ore_mese,
    })


# ── 3. Salva/aggiorna/elimina giorno ────────────────────────────────────────
@login_required
@require_POST
def salva_giorno(request, dipendente_id):
    """Salva o aggiorna la presenza di un giorno specifico (AJAX + form POST)."""
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")

    azienda = _get_azienda(request)
    dipendente = get_object_or_404(Dipendente, id=dipendente_id, azienda=azienda)

    data_str = request.POST.get('data')
    elimina = request.POST.get('elimina') == '1'

    try:
        data_giorno = date.fromisoformat(data_str)
    except (ValueError, TypeError):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': 'Data non valida'})
        messages.error(request, "Data non valida.")
        return redirect('calendario_presenze', dipendente_id=dipendente_id)

    from .utils import presenze_mese_bloccate

    if presenze_mese_bloccate(dipendente, azienda, data_giorno.year, data_giorno.month):
        msg = (
            'Il mese è chiuso (riepilogo approvato o elaborato): non è possibile modificare le presenze. '
            'Per riaprire alle modifiche: Riepilogo mensile motore → «Riapri mese» sulla riga del dipendente '
            '(stato torna in bozza e vengono rimossi i movimenti ferie/ROL legati a quella chiusura).'
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': msg, 'bloccato': True}, status=403)
        messages.error(request, msg)
        return redirect(
            reverse(
                'calendario_presenze_mese',
                kwargs={'dipendente_id': dipendente_id, 'anno': data_giorno.year, 'mese': data_giorno.month},
            )
        )

    if not elimina and not _giorno_in_periodo_contrattuale(dipendente, azienda, data_giorno):
        msg = (
            'Il giorno è fuori dal periodo di contratto effettivo nel mese: '
            'non è possibile inserire o modificare presenze. Eliminare la riga se serve rimuovere dati errati.'
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': msg, 'fuori_contratto': True}, status=403)
        messages.error(request, msg)
        return redirect(
            reverse(
                'calendario_presenze_mese',
                kwargs={'dipendente_id': dipendente_id, 'anno': data_giorno.year, 'mese': data_giorno.month},
            )
        )

    presenza, created = Presenza.objects.get_or_create(
        dipendente=dipendente,
        data=data_giorno,
        defaults={'azienda': azienda, 'registrata_da': request.user},
    )

    if elimina:
        presenza.delete()
        registra_log(request.user, azienda, 'presenza', f"Eliminata presenza {dipendente} {data_giorno}", None)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'ok': True,
                'eliminata': True,
                'calendario_sidebar': _calendario_sidebar_ajax_payload(
                    dipendente, data_giorno.year, data_giorno.month
                ),
            })
        messages.success(request, f"Presenza del {data_giorno.strftime('%d/%m/%Y')} eliminata.")
        return redirect(request.POST.get('next', 'calendario_presenze_default'))

    form = GiornoPresenzaForm(request.POST, instance=presenza)
    if form.is_valid():
        p = form.save(commit=False)
        p.dipendente = dipendente
        p.azienda = azienda
        if not p.registrata_da:
            p.registrata_da = request.user
        _azzera_turni_secondo_e_terzo_se_causale_giorno_intero(p)
        _ricalcola_straordinario(p, azienda)
        p.save()
        registra_log(
            request.user, azienda, 'presenza',
            f"{'Inserita' if created else 'Aggiornata'} presenza {dipendente} {data_giorno} [{p.causale}]",
            p.id,
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'ok': True,
                'causale': p.causale,
                'causale_label': p.get_causale_display(),
                'colore': p.colore_causale(),
                'ore': p.ore_lavorate(),
                'ore_straord': float(p.ore_straordinario),
                'orario': f"{p.ora_entrata.strftime('%H:%M') if p.ora_entrata else ''}{'–' if p.ora_entrata and p.ora_uscita else ''}{p.ora_uscita.strftime('%H:%M') if p.ora_uscita else ''}",
                'note': p.note,
                'calendario_sidebar': _calendario_sidebar_ajax_payload(
                    dipendente, data_giorno.year, data_giorno.month
                ),
            })
        anno, mese_val = data_giorno.year, data_giorno.month
        return redirect(f"{request.POST.get('next', '')}?anno={anno}&mese={mese_val}" or 'lista_dipendenti_presenze')
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'errors': form.errors})
        messages.error(request, "Errore nei dati inseriti.")
        return redirect('lista_dipendenti_presenze')


# ── 3b. Salva massivo — stessa causale su più giorni ────────────────────────
@login_required
@require_POST
def salva_multiplo(request, dipendente_id):
    """Salva la stessa causale su una lista di giorni selezionati."""
    if not _is_admin_hr(request.user):
        return JsonResponse({'ok': False, 'error': 'Accesso negato'}, status=403)

    azienda = _get_azienda(request)
    dipendente = get_object_or_404(Dipendente, id=dipendente_id, azienda=azienda)

    from .utils import presenze_mese_bloccate

    dates_raw = request.POST.getlist('dates')
    causale = request.POST.get('causale', 'P')
    note = request.POST.get('note', '').strip()
    _turno_post = request.POST.getlist('turno')
    turno = (_turno_post[-1].strip() if _turno_post else (request.POST.get('turno') or '1').strip()) or '1'

    ora_entrata = _parse_ora_hhmm(request.POST.get('ora_entrata'))
    ora_uscita  = _parse_ora_hhmm(request.POST.get('ora_uscita'))

    # Per turno 1: azzera orari se la causale non prevede entrata/uscita
    if turno == '1' and causale not in _CAUSALI_ORE_LAVORATE:
        ora_entrata = None
        ora_uscita  = None

    errori = []
    seen_dates = set()
    valid_dates = []
    for dstr in dates_raw:
        try:
            data_giorno = date.fromisoformat(dstr)
        except (ValueError, TypeError):
            errori.append(dstr)
            continue
        if data_giorno in seen_dates:
            continue
        seen_dates.add(data_giorno)
        valid_dates.append(data_giorno)

    valid_dates = [d for d in valid_dates if _giorno_in_periodo_contrattuale(dipendente, azienda, d)]

    if valid_dates and turno in ('2', '3') and ora_entrata is None and ora_uscita is None:
        return JsonResponse(
            {
                'ok': False,
                'error': (
                    'Per il 2° o 3° turno indica almeno entrata o uscita nella barra laterale '
                    '(i campi non possono essere entrambi vuoti).'
                ),
            },
            status=400,
        )

    if seen_dates and not valid_dates:
        return JsonResponse(
            {
                'ok': False,
                'error': (
                    'Nessun giorno selezionato è nel periodo di contratto effettivo nel mese '
                    '(giorni fuori rapporto esclusi).'
                ),
            },
            status=400,
        )

    mesi_toccati = {(d.year, d.month) for d in valid_dates}
    for ay, am in mesi_toccati:
        if presenze_mese_bloccate(dipendente, azienda, ay, am):
            msg = (
                'Uno o più giorni cadono in un mese chiuso (riepilogo approvato o elaborato): '
                'modifica non consentita. Usa «Riapri mese» nel Riepilogo mensile motore per quel mese.'
            )
            return JsonResponse({'ok': False, 'error': msg, 'bloccato': True}, status=403)

    n_created = n_updated = 0
    # Transazione + lock righe esistenti: T2/T3 aggiornano solo quel turno (T1/causale invariati).
    # Con turno 1, causali senza rilevazione oraria azzerano anche 2° e 3° turno (giorno intero).
    with transaction.atomic():
        locked_by_data = {}
        if valid_dates:
            locked_by_data = {
                p.data: p
                for p in Presenza.objects.select_for_update().filter(
                    dipendente=dipendente, data__in=valid_dates
                )
            }

        for data_giorno in valid_dates:
            if turno in ('2', '3'):
                if ora_entrata is None and ora_uscita is None:
                    continue
                presenza = locked_by_data.get(data_giorno)
                created = False
                if presenza is None:
                    presenza = Presenza(
                        dipendente=dipendente,
                        data=data_giorno,
                        azienda=azienda,
                        registrata_da=request.user,
                        causale='P',
                    )
                    created = True
                if turno == '2':
                    presenza.ora_entrata2 = ora_entrata
                    presenza.ora_uscita2 = ora_uscita
                else:
                    presenza.ora_entrata3 = ora_entrata
                    presenza.ora_uscita3 = ora_uscita
            else:
                presenza = locked_by_data.get(data_giorno)
                created = False
                if presenza is None:
                    presenza = Presenza(
                        dipendente=dipendente,
                        data=data_giorno,
                        azienda=azienda,
                        registrata_da=request.user,
                    )
                    created = True
                presenza.causale = causale
                presenza.ora_entrata = ora_entrata
                presenza.ora_uscita = ora_uscita
                presenza.note = note
                _azzera_turni_secondo_e_terzo_se_causale_giorno_intero(presenza)

            _ricalcola_straordinario(presenza, azienda)

            presenza.save()
            if created:
                n_created += 1
            else:
                n_updated += 1

    registra_log(
        request.user, azienda, 'presenza',
        f"Inserimento massivo {dipendente}: {n_created} create, {n_updated} aggiornate [{causale} T{turno}]",
        None,
    )
    anno_sb = _parse_int(request.POST.get('anno_calendario'), 0)
    if anno_sb < 2000:
        anno_sb = valid_dates[0].year if valid_dates else date.today().year
    mese_sb = _parse_int(request.POST.get('mese_calendario'), 0)
    if mese_sb < 1 or mese_sb > 12:
        mese_sb = valid_dates[0].month if valid_dates else None
    return JsonResponse({
        'ok': True,
        'created': n_created,
        'updated': n_updated,
        'errors': errori,
        'calendario_sidebar': _calendario_sidebar_ajax_payload(dipendente, anno_sb, mese_sb),
    })


# ── 3c. Aggiorna orario singolo giorno (AJAX — usato dalla griglia calendario) ─
@login_required
@require_POST
def aggiorna_orario_giorno(request, dipendente_id):
    """Salva/aggiorna l'orario (entrata/uscita) di un giorno dalla griglia inline."""
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")
    azienda = _get_azienda(request)
    dipendente = get_object_or_404(Dipendente, id=dipendente_id, azienda=azienda)

    data_str = request.POST.get('data')
    try:
        data_giorno = date.fromisoformat(data_str)
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Data non valida'})

    from .utils import presenze_mese_bloccate

    if presenze_mese_bloccate(dipendente, azienda, data_giorno.year, data_giorno.month):
        return JsonResponse(
            {
                'ok': False,
                'error': (
                    'Il mese è chiuso (riepilogo approvato o elaborato): modifica non consentita. '
                    'Usa «Riapri mese» nel Riepilogo mensile motore.'
                ),
                'bloccato': True,
            },
            status=403,
        )

    if not _giorno_in_periodo_contrattuale(dipendente, azienda, data_giorno):
        return JsonResponse(
            {
                'ok': False,
                'error': 'Il giorno è fuori dal periodo di contratto effettivo nel mese: modifica non consentita.',
                'fuori_contratto': True,
            },
            status=403,
        )

    turno = request.POST.get('turno', '1')   # '1', '2' o '3'
    ent = _parse_ora_hhmm(request.POST.get('ora_entrata'))
    usc = _parse_ora_hhmm(request.POST.get('ora_uscita'))

    # Non creare una presenza se non ci sono orari e non esiste già
    if ent is None and usc is None:
        try:
            presenza = Presenza.objects.get(dipendente=dipendente, data=data_giorno)
        except Presenza.DoesNotExist:
            return JsonResponse({'ok': True, 'nothing': True})
    else:
        presenza, _ = Presenza.objects.get_or_create(
            dipendente=dipendente,
            data=data_giorno,
            defaults={'azienda': azienda, 'registrata_da': request.user, 'causale': 'P'},
        )

    if turno == '3':
        presenza.ora_entrata3 = ent
        presenza.ora_uscita3  = usc
    elif turno == '2':
        presenza.ora_entrata2 = ent
        presenza.ora_uscita2  = usc
    else:
        presenza.ora_entrata = ent
        presenza.ora_uscita  = usc

    # Straordinario calcolato server-side (non si fida del valore inviato dal frontend):
    # eccedenza = max(0, ore_lavorate_T1+T2+T3 − ore_contrattuali_giornaliere)
    _ricalcola_straordinario(presenza, azienda)

    presenza.save()

    return JsonResponse({
        'ok': True,
        'ore': presenza.ore_lavorate(),
        'ore_straord': float(presenza.ore_straordinario),
        'causale': presenza.causale,
        'colore': presenza.colore_causale(),
        'orario': (
            f"{presenza.ora_entrata.strftime('%H:%M') if presenza.ora_entrata else ''}"
            f"{'–' if presenza.ora_entrata and presenza.ora_uscita else ''}"
            f"{presenza.ora_uscita.strftime('%H:%M') if presenza.ora_uscita else ''}"
        ),
        'calendario_sidebar': _calendario_sidebar_ajax_payload(
            dipendente, data_giorno.year, data_giorno.month
        ),
    })


# ── 3d. Applica schema orario settimanale a tutto il mese ───────────────────
@login_required
@require_POST
def applica_schema_mese(request, dipendente_id):
    """Applica un pattern orario Lun-Dom a tutti i giorni lavorativi del mese."""
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")
    azienda = _get_azienda(request)
    dipendente = get_object_or_404(Dipendente, id=dipendente_id, azienda=azienda)

    anno = _parse_int(request.POST.get('anno'), date.today().year)
    mese = _parse_int(request.POST.get('mese'), date.today().month)
    solo_vuoti = request.POST.get('solo_vuoti', '1') == '1'

    from .utils import presenze_mese_bloccate

    if presenze_mese_bloccate(dipendente, azienda, anno, mese):
        return JsonResponse(
            {
                'ok': False,
                'error': (
                    'Il mese è chiuso (riepilogo approvato o elaborato): applicazione schema non consentita. '
                    'Usa «Riapri mese» nel Riepilogo mensile motore.'
                ),
                'bloccato': True,
            },
            status=403,
        )

    chiusura_sett = set(_get_chiusura_settimanale_presenze(azienda, anno, mese))

    schema = {}
    for wd in range(7):
        schema[wd] = {
            'attivo': request.POST.get(f'wd{wd}_attivo') == '1',
            'ent1': _parse_ora_hhmm(request.POST.get(f'wd{wd}_ent1')),
            'usc1': _parse_ora_hhmm(request.POST.get(f'wd{wd}_usc1')),
            'ent2': _parse_ora_hhmm(request.POST.get(f'wd{wd}_ent2')),
            'usc2': _parse_ora_hhmm(request.POST.get(f'wd{wd}_usc2')),
        }

    _, num_giorni = calendar.monthrange(anno, mese)
    n_creati = n_aggiornati = n_saltati = 0

    date_con_presenza = set()
    if solo_vuoti:
        date_con_presenza = set(
            Presenza.objects.filter(
                dipendente=dipendente,
                data__year=anno,
                data__month=mese,
            ).values_list('data', flat=True)
        )

    for g in range(1, num_giorni + 1):
        d = date(anno, mese, g)
        wd = d.weekday()
        conf = schema.get(wd, {})
        if not conf.get('attivo'):
            continue
        if solo_vuoti and d in date_con_presenza:
            n_saltati += 1
            continue
        if not _giorno_in_periodo_contrattuale(dipendente, azienda, d):
            n_saltati += 1
            continue
        presenza, created = Presenza.objects.get_or_create(
            dipendente=dipendente,
            data=d,
            defaults={'azienda': azienda, 'registrata_da': request.user, 'causale': 'P'},
        )
        presenza.ora_entrata  = conf['ent1']
        presenza.ora_uscita   = conf['usc1']
        presenza.ora_entrata2 = conf['ent2']
        presenza.ora_uscita2  = conf['usc2']
        _ricalcola_straordinario(presenza, azienda)
        presenza.save()
        if created:
            n_creati += 1
        else:
            n_aggiornati += 1

    registra_log(
        request.user, azienda, 'presenza',
        f"Schema mensile {dipendente} {MESI_ITA[mese]} {anno}: "
        f"{n_creati} create, {n_aggiornati} agg., {n_saltati} saltate",
        None,
    )
    return JsonResponse({'ok': True, 'creati': n_creati, 'aggiornati': n_aggiornati, 'saltati': n_saltati})


# ── 4. Riepilogo mese — tutti i dipendenti ──────────────────────────────────
def _straord_feriale_riepilogo(rie):
    """Straordinario su giorni feriali: diurno + notturno (motore)."""
    if not rie:
        return None
    from decimal import Decimal
    s = rie.ore_straord_diurno + rie.ore_straord_notturno
    return s if isinstance(s, Decimal) else Decimal(str(s or 0))


@login_required
def riepilogo_mese(request):
    """Riepilogo mensile di tutti i dipendenti — tabella e link export."""
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")

    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('profile')

    oggi = date.today()

    if request.method == 'POST' and request.POST.get('action') == 'aggrega_riepiloghi_mese':
        from .utils import aggrega_presenze_per_motore, dipendenti_per_riepilogo_mese

        anno_p = _parse_int(request.POST.get('anno'), oggi.year)
        mese_p = _parse_int(request.POST.get('mese'), oggi.month)
        if mese_p < 1 or mese_p > 12:
            mese_p = oggi.month
        ok = skip = err = 0
        for dip in dipendenti_per_riepilogo_mese(azienda, anno_p, mese_p).order_by('id'):
            try:
                aggrega_presenze_per_motore(dip, azienda, anno_p, mese_p, request.user)
                ok += 1
            except ValueError:
                skip += 1
            except Exception:
                err += 1
        if ok or skip or err:
            messages.success(
                request,
                f'Riepiloghi motore salvati per {mese_p:02d}/{anno_p}: {ok} calcolati, '
                f'{skip} non modificati (già approvati/elaborati), {err} errori.',
            )
        return redirect(f"{request.path}?anno={anno_p}&mese={mese_p}")

    anno = _parse_int(request.GET.get('anno'), oggi.year)
    mese = _parse_int(request.GET.get('mese'), oggi.month)
    if mese < 1 or mese > 12:
        mese = oggi.month

    _, num_giorni = calendar.monthrange(anno, mese)
    date_mese = [date(anno, mese, g) for g in range(1, num_giorni + 1)]

    chiusura_sett = set(_get_chiusura_settimanale_presenze(azienda, anno, mese))
    # Stesso insieme festivi del motore cedolini (nazionali + aziendali)
    from rapporto_di_lavoro.utils_calendario import get_festivita_mese
    festivi = {f['data'] for f in get_festivita_mese(anno, mese, azienda)}

    from .utils import dipendenti_per_riepilogo_mese

    dipendenti = dipendenti_per_riepilogo_mese(azienda, anno, mese)
    q_filter = (request.GET.get('q') or '').strip()
    if q_filter:
        qobj = (
            Q(cognome__icontains=q_filter)
            | Q(nome__icontains=q_filter)
            | Q(codice_fiscale__icontains=q_filter)
        )
        if q_filter.isdigit():
            qobj |= Q(matricola=int(q_filter))
        dipendenti = dipendenti.filter(qobj)
    dipendenti = dipendenti.order_by('cognome', 'nome')

    paginator = Paginator(dipendenti, 25)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    dips_list = list(page_obj.object_list)
    dip_ids = [d.id for d in dips_list]
    # Teoriche solo per la pagina corrente; export Excel/CSV allinea tutti gli attivi.
    _genera_presenze_teoriche_mese_azienda(
        azienda, anno, mese, request.user, solo_dipendenti_ids=dip_ids,
    )

    presenze_mese = Presenza.objects.filter(
        azienda=azienda,
        data__year=anno,
        data__month=mese,
        dipendente_id__in=dip_ids,
    ).select_related('dipendente') if dip_ids else Presenza.objects.none()

    periodi = {}
    for dip in dips_list:
        di, df = _periodo_rapporto_dipendente_per_mese(dip, azienda, anno, mese)
        ps, pe = _intervallo_mese_per_rapporto(anno, mese, di, df)
        periodi[dip.id] = (ps, pe)

    # Mappa (dipendente_id, data) → presenza (tutte le righe del mese, come nel calendario dipendente)
    pmap = {}
    for p in presenze_mese:
        pmap[(p.dipendente_id, p.data)] = p

    from .models import RiepilogoMensilePresenze
    rie_map = {}
    if dip_ids:
        for r in RiepilogoMensilePresenze.objects.filter(
            azienda=azienda, anno=anno, mese=mese, dipendente_id__in=dip_ids
        ):
            rie_map[r.dipendente_id] = r

    from .utils import (
        ore_eccesso_vs_contratto,
        ore_std_giornaliere_contratto,
        tipo_eccesso_vs_contratto,
    )

    # Costruisce righe per la tabella (una per dipendente, una cella per giorno).
    righe = []
    for dip in dips_list:
        periodo_start, periodo_end = periodi.get(dip.id, (None, None))
        ore_std_contratto = ore_std_giornaliere_contratto(dip, azienda, anno, mese)
        giorni_riga = []
        for d in date_mese:
            in_periodo = bool(periodo_start and periodo_end and periodo_start <= d <= periodo_end)
            p = pmap.get((dip.id, d))
            # Giorni Riposo: non contribuiscono alle ore (coerente calendario dipendente)
            # Fuori periodo contrattuale: non conteggiare ore (anche se esiste riga Presenza legacy/teorica).
            # Causali diverse da P/ST/SMART/FE: orari in DB non contano nel totale (come celle orario in griglia).
            if p and p.causale == 'R':
                m1 = m2 = m3 = 0
            elif not in_periodo:
                m1 = m2 = m3 = 0
            elif p and p.causale not in _CAUSALI_ORE_LAVORATE:
                m1 = m2 = m3 = 0
            else:
                m1, m2, m3 = _minuti_turni_presenza(p)
            min_giorno = m1 + m2 + m3
            giorni_riga.append({
                'data': d,
                'presenza': p,
                'is_weekend': d.weekday() >= 5,
                'is_domenica': d.weekday() == 6,
                'is_chiusura_sett': d.weekday() in chiusura_sett,
                'is_festivo': d in festivi,
                'fuori_rapporto': not in_periodo,
                'ore_t1': m1 / 60,
                'ore_t2': m2 / 60,
                'ore_t3': m3 / 60,
                'ore_giorno': min_giorno / 60,
            })
        straord_semplice_mese = Decimal('0')
        for g in giorni_riga:
            if g['fuori_rapporto']:
                g['eccesso_tipo'] = None
                g['ore_eccesso'] = Decimal('0')
                g['straord_semplice'] = None
                continue
            g['eccesso_tipo'] = tipo_eccesso_vs_contratto(
                g['ore_giorno'], ore_std_contratto, g,
            )
            g['ore_eccesso'] = ore_eccesso_vs_contratto(
                g['ore_giorno'], ore_std_contratto, g,
            )
            ss = _straord_giorno_semplice(
                g['ore_giorno'], ore_std_contratto, g.get('presenza'),
            )
            g['straord_semplice'] = ss
            if ss is not None:
                straord_semplice_mese += ss
        rie = rie_map.get(dip.id)
        righe.append({
            'dip': dip,
            'giorni': giorni_riga,
            'ore_tot': sum(g['ore_giorno'] for g in giorni_riga),
            'periodo_start': periodo_start,
            'periodo_end': periodo_end,
            'fuori_periodo_mese': not (periodo_start and periodo_end),
            'rie': rie,
            'straord_fer': _straord_feriale_riepilogo(rie),
            'straord_fest': rie.ore_straord_festivo if rie else None,
            'straord_dom': rie.ore_straord_domenica if rie else None,
            'ore_std_contratto': ore_std_contratto,
            'straord_semplice_mese': straord_semplice_mese,
        })

    if mese == 1:
        mese_prev, anno_prev = 12, anno - 1
    else:
        mese_prev, anno_prev = mese - 1, anno
    if mese == 12:
        mese_next, anno_next = 1, anno + 1
    else:
        mese_next, anno_next = mese + 1, anno

    # Colonne tabella: dip + N giorni + Dom + Fest + Str.fer + Str.fest + Str.dom + Tot
    tabella_colspan = 1 + num_giorni + 5 + 1

    # Primo/ultimo mese destinazione per il form «Estendi orari» (stesso mese sorgente → fino a dicembre anno_fine)
    ay, am = (anno, mese + 1) if mese < 12 else (anno + 1, 1)
    estendi_primo_anno, estendi_primo_mese = ay, am
    estendi_ultimo_anno, estendi_ultimo_mese = anno, 12
    estendi_possibile = (estendi_primo_anno, estendi_primo_mese) <= (estendi_ultimo_anno, estendi_ultimo_mese)

    return render(request, 'presenze/riepilogo_mese.html', {
        'azienda': azienda,
        'anno': anno,
        'mese': mese,
        'mese_nome': MESI_ITA[mese],
        'date_mese': date_mese,
        'num_giorni': num_giorni,
        'tabella_colspan': tabella_colspan,
        'giorni_ita': GIORNI_ITA,
        'righe': righe,
        'festivi': festivi,
        'page_obj': page_obj,
        'q_filter': q_filter,
        'anno_prev': anno_prev, 'mese_prev': mese_prev,
        'anno_next': anno_next, 'mese_next': mese_next,
        'estendi_primo_anno': estendi_primo_anno,
        'estendi_primo_mese': estendi_primo_mese,
        'estendi_ultimo_anno': estendi_ultimo_anno,
        'estendi_ultimo_mese': estendi_ultimo_mese,
        'estendi_primo_mese_nome': MESI_ITA[estendi_primo_mese],
        'estendi_ultimo_mese_nome': MESI_ITA[estendi_ultimo_mese],
        'estendi_possibile': estendi_possibile,
    })


def _scostamento_filters_from_request(request):
    today = date.today()
    anno = _parse_int(request.GET.get('anno'), today.year)
    mese_raw = (request.GET.get('mese') or '').strip()
    mese = _parse_int(mese_raw, 0) if mese_raw else 0
    if mese < 0 or mese > 12:
        mese = 0

    q_filter = (request.GET.get('q') or '').strip()
    ordina = (request.GET.get('ordina') or 'scostamento').strip()
    if ordina not in ('scostamento', 'dipendente', 'mese_desc', 'mese_asc'):
        ordina = 'scostamento'
    filtro = (request.GET.get('filtro') or 'tutti').strip()
    if filtro not in ('tutti', 'solo_scostamento', 'senza_cedolino', 'da_conciliare'):
        filtro = 'tutti'
    return anno, mese, q_filter, ordina, filtro


def _build_rows_scostamento_cedolino(azienda, anno: int, mese: int, q_filter: str, ordina: str, filtro: str):
    from .utils import dipendenti_per_riepilogo_mese
    from documenti.models import CedolinoMotoreV4, VoceCedolinoMotoreV4
    from rapporto_di_lavoro.utils_presenze import (
        confronto_tipologie_cal_vs_cedolino_v4,
        get_presenze_mese_aggregato,
    )

    mesi = [mese] if mese else list(range(1, 13))
    dip_qs = dipendenti_per_riepilogo_mese(azienda, anno, mese or 1)
    if q_filter:
        qobj = (
            Q(cognome__icontains=q_filter)
            | Q(nome__icontains=q_filter)
            | Q(codice_fiscale__icontains=q_filter)
        )
        if q_filter.isdigit():
            qobj |= Q(matricola=int(q_filter))
        dip_qs = dip_qs.filter(qobj)
    dipendenti = list(dip_qs.order_by('cognome', 'nome'))

    cod_ord = {'8001'}
    cod_dom = {'8010', '8011'}
    cod_fest = {'8020', '8108', '8109', '109'}
    q2 = Decimal('0.01')

    def _safe_dec(v):
        try:
            return Decimal(str(v or 0))
        except Exception:
            return Decimal('0')

    rows_all = []
    for dip in dipendenti:
        for m in mesi:
            di, df = _periodo_rapporto_dipendente_per_mese(dip, azienda, anno, m)
            ps, pe = _intervallo_mese_per_rapporto(anno, m, di, df)
            qs = Presenza.objects.filter(dipendente=dip, data__year=anno, data__month=m).order_by('data')
            qsp = qs.filter(data__gte=ps, data__lte=pe) if (ps and pe) else qs.none()
            ore_cal = Decimal(str(sum(_minuti_lavorati_presenza(p) for p in qsp) / 60)).quantize(q2)

            if ps and pe:
                agg = get_presenze_mese_aggregato(dip, anno, m, azienda, data_da=ps, data_a=pe)
            else:
                agg = get_presenze_mese_aggregato(dip, anno, m, azienda, data_da=date(2099, 1, 1), data_a=date(2099, 1, 1))

            ced = (
                CedolinoMotoreV4.objects.filter(
                    dipendente_id=dip.id, anno=anno, mese=m, natura_busta='ORDINARIA'
                )
                .order_by('-id')
                .first()
            )
            ord_h = Decimal('0')
            dom_h = Decimal('0')
            fest_h = Decimal('0')
            ced_by: dict[str, Decimal] = defaultdict(lambda: Decimal('0'))
            if ced:
                for v in VoceCedolinoMotoreV4.objects.filter(cedolino=ced):
                    code = str(v.codice or '').strip()
                    if not code:
                        continue
                    h = _safe_dec(v.ore_gg).quantize(q2)
                    ced_by[code] += h
                    if code in cod_ord:
                        ord_h += h
                    elif code in cod_dom:
                        dom_h += h
                    elif code in cod_fest:
                        fest_h += h
            ore_ced = (ord_h + dom_h + fest_h).quantize(q2)
            diff = (ore_ced - ore_cal).quantize(q2)

            tipologie = confronto_tipologie_cal_vs_cedolino_v4(agg, dict(ced_by))
            tip_map = {t['key']: t for t in tipologie}
            cal_ord = tip_map.get('ord', {}).get('cal') or Decimal('0')
            ced_ord = tip_map.get('ord', {}).get('ced') or Decimal('0')
            cal_dom = tip_map.get('dom', {}).get('cal') or Decimal('0')
            ced_dom = tip_map.get('dom', {}).get('ced') or Decimal('0')
            cal_fest = tip_map.get('fest_lav', {}).get('cal') or Decimal('0')
            ced_fest = tip_map.get('fest_lav', {}).get('ced') or Decimal('0')
            cal_str = ((tip_map.get('nott', {}).get('cal') or Decimal('0')) + (tip_map.get('stra_altri', {}).get('cal') or Decimal('0'))).quantize(q2)
            ced_str = ((tip_map.get('nott', {}).get('ced') or Decimal('0')) + (tip_map.get('stra_altri', {}).get('ced') or Decimal('0'))).quantize(q2)

            rows_all.append({
                'id_collapse': f"scd-{dip.id}-{anno}-{m}",
                'dip': dip,
                'mese': m,
                'mese_nome': MESI_ITA[m] if 0 <= m < len(MESI_ITA) else str(m),
                'giorni_mese': calendar.monthrange(anno, m)[1],
                'righe_calendario': qsp.count(),
                'ore_calendario': ore_cal,
                'ore_cedolino': ore_ced,
                'ore_ced_ord': ord_h.quantize(q2),
                'ore_ced_dom': dom_h.quantize(q2),
                'ore_ced_fest': fest_h.quantize(q2),
                'scostamento': diff,
                'has_cedolino': bool(ced),
                'tipologie': tipologie,
                'cal_ord': cal_ord.quantize(q2), 'ced_ord': ced_ord.quantize(q2), 'delta_ord': (cal_ord - ced_ord).quantize(q2),
                'cal_dom': cal_dom.quantize(q2), 'ced_dom': ced_dom.quantize(q2), 'delta_dom': (cal_dom - ced_dom).quantize(q2),
                'cal_fest': cal_fest.quantize(q2), 'ced_fest': ced_fest.quantize(q2), 'delta_fest': (cal_fest - ced_fest).quantize(q2),
                'cal_straord': cal_str, 'ced_straord': ced_str, 'delta_straord': (cal_str - ced_str).quantize(q2),
            })

    stat = {
        'totale': len(rows_all),
        'con_scostamento': sum(1 for r in rows_all if r['scostamento'] != 0),
        'senza_cedolino': sum(1 for r in rows_all if not r['has_cedolino']),
    }
    rows = list(rows_all)
    if filtro == 'solo_scostamento':
        rows = [r for r in rows_all if r['scostamento'] != 0]
    elif filtro == 'senza_cedolino':
        rows = [r for r in rows_all if not r['has_cedolino']]
    elif filtro == 'da_conciliare':
        rows = [r for r in rows_all if (not r['has_cedolino']) or (r['scostamento'] != 0)]

    if ordina == 'dipendente':
        rows.sort(key=lambda r: (r['dip'].cognome.lower(), r['dip'].nome.lower(), -r['mese']))
    elif ordina == 'mese_asc':
        rows.sort(key=lambda r: (r['mese'], r['dip'].cognome.lower(), r['dip'].nome.lower()))
    elif ordina == 'mese_desc':
        rows.sort(key=lambda r: (-r['mese'], r['dip'].cognome.lower(), r['dip'].nome.lower()))
    else:
        rows.sort(key=lambda r: (-abs(r['scostamento']), -r['mese'], r['dip'].cognome.lower(), r['dip'].nome.lower()))

    stat['visualizzate'] = len(rows)
    return rows, stat, mesi


@login_required
def riepilogo_scostamento_cedolino(request):
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")
    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('profile')

    anno, mese, q_filter, ordina, filtro = _scostamento_filters_from_request(request)
    rows, stat, _mesi = _build_rows_scostamento_cedolino(azienda, anno, mese, q_filter, ordina, filtro)
    return render(request, 'presenze/riepilogo_scostamento_cedolino.html', {
        'azienda': azienda,
        'anno': anno,
        'mese': mese,
        'mesi_nomi': MESI_ITA,
        'mesi_options': [(i, MESI_ITA[i]) for i in range(1, 13)],
        'rows': rows,
        'q_filter': q_filter,
        'ordina': ordina,
        'filtro': filtro,
        'stat': stat,
    })


@login_required
def export_scostamento_cedolino_excel(request):
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")
    azienda = _get_azienda(request)
    if not azienda:
        return HttpResponse("Azienda non selezionata.", status=400)
    anno, mese, q_filter, ordina, filtro = _scostamento_filters_from_request(request)
    rows, _stat, mesi = _build_rows_scostamento_cedolino(azienda, anno, mese, q_filter, ordina, filtro)

    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for m in mesi:
        ws = wb.create_sheet(title=f"{MESI_ITA[m][:3]}-{anno}")
        ws.append([
            'Dipendente', 'Mese', 'Ore cal', 'Ore ced', 'Scostamento',
            'Cal Ord', 'Ced Ord', 'Δ Ord', 'Cal Dom', 'Ced Dom', 'Δ Dom',
            'Cal Fest', 'Ced Fest', 'Δ Fest', 'Cal Straord', 'Ced Straord', 'Δ Straord',
        ])
        for r in [x for x in rows if x['mese'] == m]:
            ws.append([
                f"{r['dip'].cognome} {r['dip'].nome}", r['mese_nome'],
                float(r['ore_calendario']), float(r['ore_cedolino']), float(r['scostamento']),
                float(r['cal_ord']), float(r['ced_ord']), float(r['delta_ord']),
                float(r['cal_dom']), float(r['ced_dom']), float(r['delta_dom']),
                float(r['cal_fest']), float(r['ced_fest']), float(r['delta_fest']),
                float(r['cal_straord']), float(r['ced_straord']), float(r['delta_straord']),
            ])
            for t in r.get('tipologie', []):
                ws.append([
                    '', '', '', '', '',
                    t.get('label', ''),
                    float(t['cal']) if t.get('cal') is not None else '',
                    float(t['ced']) if t.get('ced') is not None else '',
                    float(t['delta']) if t.get('delta') is not None else '',
                    '', '', '', '', '', '', '', '',
                ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 14

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    resp = HttpResponse(out.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="scostamento_cedolino_{anno}.xlsx"'
    return resp


@login_required
def export_scostamento_cedolino_pdf(request):
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")
    azienda = _get_azienda(request)
    if not azienda:
        return HttpResponse("Azienda non selezionata.", status=400)
    anno, mese, q_filter, ordina, filtro = _scostamento_filters_from_request(request)
    rows, _stat, _mesi = _build_rows_scostamento_cedolino(azienda, anno, mese, q_filter, ordina, filtro)

    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=16, rightMargin=16, topMargin=16, bottomMargin=16)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"Scostamento calendario vs cedolino — {azienda.nome} — {anno}", styles['Heading3']), Spacer(1, 8)]
    data = [[
        'Dipendente', 'Mese', 'Cal', 'Ced', 'Δ',
        'Cal Ord', 'Ced Ord', 'Δ Ord', 'Cal Dom', 'Ced Dom', 'Δ Dom', 'Cal Fest', 'Ced Fest', 'Δ Fest'
    ]]
    for r in rows:
        data.append([
            f"{r['dip'].cognome} {r['dip'].nome}", r['mese_nome'],
            f"{r['ore_calendario']:.2f}", f"{r['ore_cedolino']:.2f}", f"{r['scostamento']:.2f}",
            f"{r['cal_ord']:.2f}", f"{r['ced_ord']:.2f}", f"{r['delta_ord']:.2f}",
            f"{r['cal_dom']:.2f}", f"{r['ced_dom']:.2f}", f"{r['delta_dom']:.2f}",
            f"{r['cal_fest']:.2f}", f"{r['ced_fest']:.2f}", f"{r['delta_fest']:.2f}",
        ])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1b3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d0d7de')),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(tbl)
    doc.build(story)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="scostamento_cedolino_{anno}.pdf"'
    return resp


def _build_rows_scostamento_fiscale(
    azienda, anno: int, mese: int, q_filter: str, divisore_raw=None, percorso_fiscale=None
):
    from anagrafiche.models import Dipendente
    from documenti.models import CedolinoMotoreV4, VoceCedolinoMotoreV4
    from rapporto_di_lavoro.models import (
        CCNL,
        ParametroScattiAnnuali,
        RuoloOrganico2026,
        TipoContratto,
    )
    from rapporto_di_lavoro.risoluzione_contratto_motore import (
        divisore_str_da_parametro_get,
        kwargs_percorso_fiscale_sim,
        rapporto_sottoscritto_attivo_nel_mese,
        risolvi_parametro_ccnl_per_mese,
        superminimo_da_rapporto_o_ruolo,
    )
    from rapporto_di_lavoro.services_simulazione import invoca_calcola_busta_paga_mese
    from rapporto_di_lavoro.risoluzione_contratto_motore import calcola_scatto_totale_maturato as _calcola_scatto_totale

    q2 = Decimal('0.01')
    mesi = [mese] if mese else list(range(1, 13))
    dip_qs = Dipendente.objects.filter(azienda=azienda, stato='attivo')
    if q_filter:
        qobj = (
            Q(cognome__icontains=q_filter)
            | Q(nome__icontains=q_filter)
            | Q(codice_fiscale__icontains=q_filter)
        )
        if q_filter.isdigit():
            qobj |= Q(matricola=int(q_filter))
        dip_qs = dip_qs.filter(qobj)
    dipendenti = list(dip_qs.order_by('cognome', 'nome'))

    ccnl = CCNL.objects.filter(sigla__icontains='FIPE').first()
    scatti_db: dict = {}
    if ccnl:
        for s in ParametroScattiAnnuali.objects.filter(ccnl=ccnl, anno=anno, attivo=True):
            scatti_db.setdefault(s.livello, []).append((int(s.anni_anzianita or 0), Decimal(str(s.importo_scatto or 0))))
        for k in scatti_db:
            scatti_db[k].sort(key=lambda x: x[0])

    def _d(v):
        try:
            return Decimal(str(v or 0)).quantize(q2)
        except Exception:
            return Decimal('0.00')

    divisore_str = divisore_str_da_parametro_get(divisore_raw)
    rows = []
    for dip in dipendenti:
        ruolo = RuoloOrganico2026.objects.filter(azienda=azienda, dipendente=dip).order_by('-data_modifica').first()
        if not ruolo:
            continue
        for m in mesi:
            data_mese = date(anno, m, 1)
            rapporto = rapporto_sottoscritto_attivo_nel_mese(
                dipendente=dip, azienda=azienda, anno=anno, mese=m
            )
            _lc = (rapporto.livello_ccnl or '').strip() if rapporto else ''
            livello_eff = _lc or str(ruolo.livello or '')
            parametro, fonte_parametro = risolvi_parametro_ccnl_per_mese(
                rapporto=rapporto,
                data_primo_giorno_mese=data_mese,
                livello_fallback=str(ruolo.livello or ''),
            )
            if not parametro:
                continue
            tc = None
            if rapporto and rapporto.tipo_contratto_id:
                tc = rapporto.tipo_contratto
            if tc is None:
                try:
                    tc = TipoContratto.objects.get(pk=int(ruolo.tipo_contratto_id or 0))
                except Exception:
                    tc = TipoContratto.objects.filter(attivo=True).order_by('id').first()
            if not tc:
                continue

            cal_m = (ruolo.calendario_mensile or {}).get(str(m), (ruolo.calendario_mensile or {}).get(m, {})) or {}
            ore_ord = _d(cal_m.get('ore_ordinarie_retribuite', 0))
            scatto = _calcola_scatto_totale(str(livello_eff), int(ruolo.anni_anzianita or 0), scatti_db)
            data_inizio_eff = rapporto.data_inizio_rapporto if rapporto else ruolo.data_inizio
            data_fine_eff = rapporto.data_fine_rapporto if rapporto else ruolo.data_fine
            superminimo_eff = superminimo_da_rapporto_o_ruolo(
                rapporto=rapporto, ruolo_superminimo=ruolo.superminimo
            )
            premio_extra = Decimal('0')
            if rapporto is not None:
                try:
                    premio_extra = Decimal(str(rapporto.premio_obiettivi or 0)).quantize(q2)
                except Exception:
                    premio_extra = Decimal('0.00')
            fiscal_kw = kwargs_percorso_fiscale_sim(percorso_fiscale)

            sim = invoca_calcola_busta_paga_mese(
                log_prefix='SCOSTAMENTO_FISCALE',
                parametro_ccnl=parametro,
                tipo_contratto=tc,
                anno=anno,
                mese=m,
                azienda=azienda,
                data_inizio_rapporto=data_inizio_eff,
                data_fine_rapporto=data_fine_eff,
                divisore_str=divisore_str,
                superminimo=superminimo_eff,
                indennita_turno=Decimal(str(ruolo.indennita_turno or 0)),
                scatto_anzianita=scatto,
                indennita_extra=premio_extra,
                ore_straord_diurno=Decimal(str(cal_m.get('ore_straord_diurno', 0) or 0)),
                ore_straord_notturno=Decimal(str(cal_m.get('ore_straord_notturno', 0) or 0)),
                ore_straord_festivo=Decimal(str(cal_m.get('ore_straord_festivo', 0) or 0)),
                ore_straord_domenica=Decimal(str(cal_m.get('ore_straord_domenica', 0) or 0)),
                ore_straord_nott_fest=Decimal(str(cal_m.get('ore_straord_nott_fest', 0) or 0)),
                ore_ordinarie_retribuite=ore_ord,
                ore_domenicali=Decimal(str(cal_m.get('ore_domenicali', 0) or 0)),
                ore_festivi=Decimal(str(cal_m.get('giorni_festivi', 0) or 0)),
                giorni_assenza_ingiust=Decimal(str(cal_m.get('giorni_assenza', 0) or 0)),
                trattenute_extra_mese=Decimal(str(cal_m.get('trattenute_extra_mese', 0) or 0)),
                competenze_extra_non_imponibili=Decimal(str(cal_m.get('competenze_extra_non_imponibili', 0) or 0)),
                modalita_ore_effettive=ore_ord > 0,
                auto_ore_domenicali_da_calendario=not (ore_ord > 0),
                ccnl_obj=ccnl,
                contratto_esclude_tredicesima=bool(rapporto is not None and rapporto.tredicesima is False),
                contratto_esclude_quattordicesima=bool(
                    rapporto is not None and rapporto.quattordicesima is False
                ),
                rateo_13_mensile_in_imponibile=bool(
                    rapporto is not None and getattr(rapporto, 'tredicesima_rateo_mensile_in_imponibile', False)
                ),
                rateo_14_mensile_in_imponibile=bool(
                    rapporto is not None and getattr(rapporto, 'quattordicesima_rateo_mensile_in_imponibile', False)
                ),
                **fiscal_kw,
            )

            ced = (
                CedolinoMotoreV4.objects
                .filter(dipendente=dip, anno=anno, mese=m, natura_busta='ORDINARIA')
                .order_by('-id')
                .first()
            )
            if not ced:
                continue

            sim_voci = {}
            for v in sim.get('voci_classificate', []):
                code = str(v.get('codice') or '').strip()
                if not code:
                    continue
                sim_voci[code] = _d(v.get('importo'))
            ced_voci = defaultdict(lambda: Decimal('0.00'))
            for v in VoceCedolinoMotoreV4.objects.filter(cedolino=ced):
                code = str(v.codice or '').strip()
                if not code:
                    continue
                ced_voci[code] += _d(v.importo)

            codes = sorted(set(sim_voci.keys()) | set(ced_voci.keys()))
            voci_delta = []
            for code in codes:
                s_val = _d(sim_voci.get(code, 0))
                c_val = _d(ced_voci.get(code, 0))
                delta = (s_val - c_val).quantize(q2)
                if delta != 0:
                    voci_delta.append({
                        'codice': code,
                        'sim_importo': s_val,
                        'ced_importo': c_val,
                        'delta_importo': delta,
                    })
            voci_delta.sort(key=lambda x: abs(x['delta_importo']), reverse=True)

            sim_impon_inps = _d(sim.get('lordo_imponibile_inps_m'))
            ced_impon_inps = _d(ced.imponibile_contrib)
            sim_inps_dip = _d(sim.get('inps_dip'))
            ced_inps_dip = _d(ced.tot_contrib_soc)
            sim_imposte = _d(sim.get('irpef_netta'))
            ced_imposte = _d(ced.tot_trat_irpef)
            sim_netto = _d(sim.get('netto_totale'))
            ced_netto = _d(ced.netto_busta)
            modalita_ore = ore_ord > 0
            parametri_motore = {
                'divisore_str': divisore_str,
                'fonte_parametro_ccnl': fonte_parametro,
                'parametro_ccnl_id': parametro.id,
                'livello_ccnl_eff': livello_eff,
                'contratto': (
                    f"{rapporto.numero_contratto} (dal {data_inizio_eff})"
                    if rapporto
                    else '— (solo ruolo organico 2026)'
                ),
                'tipo_contratto': getattr(tc, 'nome', None) or str(tc.pk),
                'coeff_ore_contratto': _d(getattr(tc, 'coefficiente_ore', None) or 1),
                'superminimo_mese': superminimo_eff,
                'scatto_anzianita': scatto,
                'modalita_ore_effettive': modalita_ore,
                'ore_ordinarie_retribuite': ore_ord,
                'inps_dip_perc_sim': _d(sim.get('inps_dip_perc')),
                'irpef_lorda_sim': _d(sim.get('irpef_lorda')),
                'ced_irpef_erario': _d(ced.irpef_erario),
                'ced_addiz_reg': _d(ced.addiz_regionale),
                'ced_addiz_com': _d(ced.addiz_comunale),
                'nota_inps_cedolino': (
                    'Sul cedolino TS v4 il campo «Tot. contributi sociali» corrisponde alla quota '
                    'INPS dipendente (IVS) in riga contributi, non alla somma azienda+dipendente.'
                ),
                'percorso_fiscale': (percorso_fiscale or 'standard').strip(),
                'premio_obiettivi_mese': premio_extra,
                'contr_escl_13': bool(rapporto is not None and rapporto.tredicesima is False),
                'contr_escl_14': bool(rapporto is not None and rapporto.quattordicesima is False),
                'coeff_rateo_13': _d(sim.get('c_13')),
                'coeff_rateo_14': _d(sim.get('c_14')),
                'rat13_m': _d(sim.get('rat13_m')),
                'rat14_m': _d(sim.get('rat14_m')),
                'fiscale_modalita_cedolino': bool(sim.get('fiscale_modalita_cedolino')),
                'l207_come_detrazione': bool(sim.get('l207_come_detrazione_irpef')),
            }

            rows.append({
                'id_collapse': f"fisc-{dip.id}-{anno}-{m}",
                'dip': dip,
                'mese': m,
                'mese_nome': MESI_ITA[m],
                'sim_impon_inps': sim_impon_inps,
                'ced_impon_inps': ced_impon_inps,
                'delta_impon_inps': (sim_impon_inps - ced_impon_inps).quantize(q2),
                'sim_inps_dip': sim_inps_dip,
                'ced_inps_dip': ced_inps_dip,
                'delta_inps_dip': (sim_inps_dip - ced_inps_dip).quantize(q2),
                'sim_imposte': sim_imposte,
                'ced_imposte': ced_imposte,
                'delta_imposte': (sim_imposte - ced_imposte).quantize(q2),
                'sim_netto': sim_netto,
                'ced_netto': ced_netto,
                'delta_netto': (sim_netto - ced_netto).quantize(q2),
                'sim_lordo': _d(sim.get('lordo_mensile')),
                'ced_lordo': _d(ced.totale_lordo),
                'delta_lordo': (_d(sim.get('lordo_mensile')) - _d(ced.totale_lordo)).quantize(q2),
                'voci_delta': voci_delta[:30],  # top differenze per leggibilità
                'parametri_motore': parametri_motore,
            })
    rows.sort(key=lambda r: (r['dip'].cognome.lower(), r['dip'].nome.lower(), r['mese']))
    return rows


@login_required
def riepilogo_scostamento_fiscale_cedolino(request):
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")
    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('profile')
    today = date.today()
    anno = _parse_int(request.GET.get('anno'), today.year)
    mese_raw = (request.GET.get('mese') or '').strip()
    mese = _parse_int(mese_raw, 0) if mese_raw else 0
    if mese < 0 or mese > 12:
        mese = 0
    q_filter = (request.GET.get('q') or '').strip()
    divisore_raw = (request.GET.get('divisore') or '173.33').strip()
    percorso_fiscale = (request.GET.get('percorso_fiscale') or 'standard').strip()
    rows = _build_rows_scostamento_fiscale(
        azienda, anno, mese, q_filter, divisore_raw=divisore_raw, percorso_fiscale=percorso_fiscale
    )
    return render(request, 'presenze/riepilogo_scostamento_fiscale_cedolino.html', {
        'azienda': azienda,
        'anno': anno,
        'mese': mese,
        'rows': rows,
        'q_filter': q_filter,
        'divisore': divisore_raw,
        'percorso_fiscale': percorso_fiscale,
        'mesi_options': [(i, MESI_ITA[i]) for i in range(1, 13)],
    })


def _presenza_blocca_estensione_orari(p) -> bool:
    """
    In modalità «solo giorni senza presenza», indica se una riga esistente non va sovrascritta.

    Le presenze create dalla generazione teorica (nota TEORICA_AUTO) non bloccano: altrimenti,
    con i mesi futuri già pieni di teoriche, «Estendi orari» non modificherebbe mai nulla.
    """
    note = (p.note or '')
    if 'TEORICA_AUTO' in note:
        return False
    return True


def _schema_presenze_per_weekday_da_mese(dipendente, azienda, anno_src: int, mese_src: int) -> dict:
    """Weekday 0..6 → ultima Presenza del mese con almeno un orario utile (pattern «attuale» in chiusura mese).

    Usiamo l'**ultima** occorrenza per weekday, non la prima: le prime settimane possono essere
    atipiche (assunzioni, festività, turni prova). Solo causali con ore lavorate nel riepilogo
    (stesso insieme usato in griglia: P, ST, SMART, FE), coerente con «Riepilogo presenze».
    """
    schema = {}
    for p in (
        Presenza.objects.filter(
            dipendente=dipendente,
            azienda=azienda,
            data__year=anno_src,
            data__month=mese_src,
        ).order_by('data')
    ):
        if p.causale not in _CAUSALI_ORE_LAVORATE:
            continue
        if not (p.ora_entrata or p.ora_entrata2 or p.ora_entrata3):
            continue
        wd = p.data.weekday()
        schema[wd] = p
    return schema


def _iter_mesi_inclusi(anno_da: int, mese_da: int, anno_a: int, mese_a: int):
    """(anno_da, mese_da) … (anno_a, mese_a) inclusi; richiede (anno_da, mese_da) <= (anno_a, mese_a)."""
    y, m = anno_da, mese_da
    while (y, m) <= (anno_a, mese_a):
        yield y, m
        if m == 12:
            y, m = y + 1, 1
        else:
            m += 1


@login_required
@require_POST
def estendi_orari_riepilogo_mese(request):
    """
    Per ogni dipendente in elenco riepilogo (mese sorgente), copia il pattern settimanale degli orari
    (ultimo giorno del mese per ciascun weekday con orari da causali lavorate) sui mesi destinazione fino alla data fine,
    rispettando il periodo di contratto giorno per giorno e saltando i mesi con riepilogo approvato/elaborato.
    """
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden('Accesso negato')

    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('profile')

    oggi = date.today()
    anno_src = _parse_int(request.POST.get('anno_src'), oggi.year)
    mese_src = _parse_int(request.POST.get('mese_src'), oggi.month)
    if mese_src < 1 or mese_src > 12:
        mese_src = oggi.month

    # Fine periodo: sempre dicembre dell'anno del mese sorgente (coerente con la UI).
    # Non usare anno_fine/mese_fine dal POST: in caso di parsing errato del browser/locale
    # il confronto (y0,m0) > (anno_fine,mese_fine) escludeva tutti i mesi successivi.
    anno_fine = anno_src
    mese_fine = 12

    # Default sovrascrivi: form con hidden=0; se checkbox «solo non manuali», anche value=1 (ultimo vince).
    _sv = request.POST.getlist('solo_vuoti')
    solo_vuoti = (_sv[-1] == '1') if _sv else False

    if mese_src < 12:
        y0, m0 = anno_src, mese_src + 1
    else:
        y0, m0 = anno_src + 1, 1

    if (y0, m0) > (anno_fine, mese_fine):
        messages.error(
            request,
            'Nessun mese destinazione: il mese sorgente è già l’ultimo del periodo richiesto.',
        )
        return redirect(f"{reverse('riepilogo_presenze_mese')}?{urlencode({'anno': anno_src, 'mese': mese_src})}")

    from .utils import dipendenti_per_riepilogo_mese, presenze_mese_bloccate

    # Stesso elenco dell’export / riepilogo (non solo la pagina corrente). Senza presenze nel
    # mese sorgente lo schema settimanale è vuoto → nessuna estensione per quel dipendente.
    dipendenti = list(
        dipendenti_per_riepilogo_mese(azienda, anno_src, mese_src).order_by('id')
    )
    ids_elenco = [d.id for d in dipendenti]

    n_creati = n_aggiornati = n_saltati = 0
    n_senza_schema = 0
    n_saltati_mese_bloccato = 0  # coppie (dipendente, mese) con riepilogo non modificabile

    teo_sum = None
    with transaction.atomic():
        if ids_elenco:
            teo_sum = _genera_presenze_teoriche_mese_azienda(
                azienda, anno_src, mese_src, request.user,
                solo_dipendenti_ids=ids_elenco,
            )

        for dip in dipendenti:
            schema = _schema_presenze_per_weekday_da_mese(dip, azienda, anno_src, mese_src)
            if not schema:
                n_senza_schema += 1
                continue

            for ay, am in _iter_mesi_inclusi(y0, m0, anno_fine, mese_fine):
                if presenze_mese_bloccate(dip, azienda, ay, am):
                    n_saltati_mese_bloccato += 1
                    continue

                date_bloccate = set()
                if solo_vuoti:
                    for pr in Presenza.objects.filter(
                        dipendente=dip,
                        data__year=ay,
                        data__month=am,
                    ):
                        if _presenza_blocca_estensione_orari(pr):
                            date_bloccate.add(pr.data)

                _, num_giorni = calendar.monthrange(ay, am)
                for g in range(1, num_giorni + 1):
                    d = date(ay, am, g)
                    wd = d.weekday()
                    src = schema.get(wd)
                    if not src:
                        n_saltati += 1
                        continue
                    if solo_vuoti and d in date_bloccate:
                        n_saltati += 1
                        continue
                    if not _giorno_in_periodo_contrattuale(dip, azienda, d):
                        n_saltati += 1
                        continue

                    presenza, created = Presenza.objects.get_or_create(
                        dipendente=dip,
                        data=d,
                        defaults={
                            'azienda': azienda,
                            'registrata_da': request.user,
                            'causale': src.causale or 'P',
                        },
                    )
                    presenza.causale = src.causale or presenza.causale or 'P'
                    presenza.ora_entrata = src.ora_entrata
                    presenza.ora_uscita = src.ora_uscita
                    presenza.ora_entrata2 = src.ora_entrata2
                    presenza.ora_uscita2 = src.ora_uscita2
                    presenza.ora_entrata3 = src.ora_entrata3
                    presenza.ora_uscita3 = src.ora_uscita3
                    _ricalcola_straordinario(presenza, azienda)
                    presenza.save()
                    if created:
                        n_creati += 1
                    else:
                        n_aggiornati += 1

    registra_log(
        request.user,
        azienda,
        'presenza',
        f"Estensione orari riepilogo da {MESI_ITA[mese_src]} {anno_src} "
        f"a {MESI_ITA[mese_fine]} {anno_fine}: {n_creati} cre., {n_aggiornati} agg., {n_saltati} salt., "
        f"{n_senza_schema} dip. senza schema, {n_saltati_mese_bloccato} mesi dip. bloccati",
        None,
    )

    msg_ok = (
        f'Orari estesi da {MESI_ITA[mese_src]} {anno_src} fino a {MESI_ITA[mese_fine]} {anno_fine}: '
        f'{n_creati} presenze create, {n_aggiornati} aggiornate, {n_saltati} giorni saltati '
        f'(vuoti/contratto/schema). '
        f'{n_senza_schema} dipendenti senza orari nel mese sorgente; '
        f'{n_saltati_mese_bloccato} combinazioni dipendente/mese saltate (riepilogo approvato o elaborato).'
    )
    messages.success(request, msg_ok)
    if teo_sum is not None and not teo_sum.get('enabled'):
        messages.warning(
            request,
            'La generazione automatica delle presenze teoriche è disattiva per questo mese '
            'in Pianificazione orari: chi non ha già orari inseriti nel mese sorgente non ottiene uno schema da copiare.',
        )
    q = (request.POST.get('q') or '').strip()
    base = reverse('riepilogo_presenze_mese')
    qs = urlencode({k: v for k, v in [('anno', anno_src), ('mese', mese_src), ('q', q)] if v})
    return redirect(f'{base}?{qs}')


# ── 5. Export Excel ──────────────────────────────────────────────────────────
@login_required
def export_excel_presenze(request):
    """Export presenze mese (XLSX/CSV): candidati + attivi con rapporto che interseca il mese."""
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")

    azienda = _get_azienda(request)
    if not azienda:
        registra_evento_anomalia(
            utente=request.user,
            azienda=None,
            contesto='export_presenze_excel',
            anomalia={'codice': 'AZIENDA_MANCANTE', 'messaggio': 'Export bloccato: azienda non selezionata'},
            request=request,
        )
        return HttpResponse("Azienda non selezionata.", status=400)

    oggi = date.today()
    anno = _parse_int(request.GET.get('anno'), oggi.year)
    mese = _parse_int(request.GET.get('mese'), oggi.month)
    formato = (request.GET.get('formato') or 'xlsx').strip().lower()
    if formato not in ('xlsx', 'csv'):
        formato = 'xlsx'

    _, num_giorni = calendar.monthrange(anno, mese)
    date_mese = [date(anno, mese, g) for g in range(1, num_giorni + 1)]

    from .utils import dipendenti_per_riepilogo_mese

    dipendenti = list(dipendenti_per_riepilogo_mese(azienda, anno, mese))
    if dipendenti:
        _genera_presenze_teoriche_mese_azienda(
            azienda,
            anno,
            mese,
            request.user,
            solo_dipendenti_ids=[d.id for d in dipendenti],
        )

    periodi = {}
    for dip in dipendenti:
        di, df = _periodo_rapporto_dipendente_per_mese(dip, azienda, anno, mese)
        ps, pe = _intervallo_mese_per_rapporto(anno, mese, di, df)
        periodi[dip.id] = (ps, pe)

    presenze_mese = Presenza.objects.filter(
        azienda=azienda,
        dipendente__in=dipendenti,
        data__year=anno,
        data__month=mese,
    ).select_related('dipendente')

    pmap = {}
    for p in presenze_mese:
        ps, pe = periodi.get(p.dipendente_id, (None, None))
        if not (ps and pe):
            continue
        if p.data < ps or p.data > pe:
            continue
        pmap[(p.dipendente_id, p.data)] = p

    def _dec2(v):
        return round(float(v or 0), 2)

    def _dec_it(v):
        return num_it_str(_dec2(v), 2)

    mese_nome = MESI_ITA[mese]

    # ── CSV ───────────────────────────────────────────────────────
    if formato == 'csv':
        import csv
        from io import StringIO

        buf = StringIO()
        writer = csv.writer(buf, delimiter=';')
        writer.writerow([
            'Cognome', 'Nome', 'Data', 'Causale',
            'Ore_T1_dec', 'Ore_T2_dec', 'Ore_T3_dec', 'Ore_tot_dec', 'Straord_dec',
            'Entrata1', 'Uscita1', 'Entrata2', 'Uscita2', 'Entrata3', 'Uscita3',
            'Note'
        ])

        for dip in dipendenti:
            ps, pe = periodi.get(dip.id, (None, None))
            if not (ps and pe):
                continue

            cur = ps
            while cur <= pe:
                p = pmap.get((dip.id, cur))
                if p:
                    t1 = _dec2(_ore_tra(p.ora_entrata, p.ora_uscita))
                    t2 = _dec2(_ore_tra(p.ora_entrata2, p.ora_uscita2))
                    t3 = _dec2(_ore_tra(p.ora_entrata3, p.ora_uscita3))
                    tot = _dec2(_minuti_lavorati_presenza(p) / 60.0)
                    st = _dec2(p.ore_straordinario)
                    writer.writerow([
                        dip.cognome,
                        dip.nome,
                        cur.strftime('%d/%m/%Y'),
                        p.causale,
                        _dec_it(t1),
                        _dec_it(t2),
                        _dec_it(t3),
                        _dec_it(tot),
                        _dec_it(st),
                        p.ora_entrata.strftime('%H:%M') if p.ora_entrata else '',
                        p.ora_uscita.strftime('%H:%M') if p.ora_uscita else '',
                        p.ora_entrata2.strftime('%H:%M') if p.ora_entrata2 else '',
                        p.ora_uscita2.strftime('%H:%M') if p.ora_uscita2 else '',
                        p.ora_entrata3.strftime('%H:%M') if p.ora_entrata3 else '',
                        p.ora_uscita3.strftime('%H:%M') if p.ora_uscita3 else '',
                        p.note or '',
                    ])
                cur += timedelta(days=1)

        nome_file = f"presenze_{azienda.nome.replace(' ', '_')}_{mese:02d}_{anno}.csv"
        resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{nome_file}"'

        if not pmap:
            registra_evento_anomalia(
                utente=request.user,
                azienda=azienda,
                contesto='export_presenze_csv',
                anomalia={'codice': 'EXPORT_VUOTO', 'messaggio': f'Export CSV vuoto {mese:02d}/{anno}'},
                request=request,
            )
        else:
            registra_log(
                utente=request.user,
                azienda=azienda,
                operazione='export_presenze_csv',
                descrizione=f'Export presenze CSV {mese:02d}/{anno}',
                request=request,
            )
        return resp

    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Stili ────────────────────────────────────────────────────
    def _fill(hex_col):
        return PatternFill('solid', fgColor=hex_col.lstrip('#'))

    st_header_az = Font(bold=True, color='FFFFFF', size=11)
    fill_header_az = _fill('1b3a5f')
    st_header_col = Font(bold=True, size=9)
    fill_header_col = _fill('dde4ec')
    fill_weekend = _fill('f0f0f0')
    fill_assenza = _fill('fdecea')
    fill_ferie = _fill('e8f0fe')
    fill_totali = _fill('f4f7fa')
    thin = Side(border_style='thin', color='CCCCCC')
    bordo = Border(left=thin, right=thin, top=thin, bottom=thin)
    al_c = Alignment(horizontal='center', vertical='center')
    al_l = Alignment(horizontal='left', vertical='center')

    # ── Foglio RIEPILOGO ─────────────────────────────────────────
    ws_rie = wb.active
    ws_rie.title = f"Riepilogo {mese_nome[:3]}{anno}"
    ws_rie.freeze_panes = 'C3'

    # Riga 1: intestazione azienda + mese
    ws_rie.merge_cells(f'A1:{get_column_letter(3 + num_giorni)}1')
    cell_title = ws_rie['A1']
    cell_title.value = f"PRESENZE {mese_nome.upper()} {anno} — {azienda.nome.upper()}"
    cell_title.font = st_header_az
    cell_title.fill = fill_header_az
    cell_title.alignment = al_c
    ws_rie.row_dimensions[1].height = 22

    # Riga 2: intestazioni colonne
    ws_rie['A2'] = 'Cognome'
    ws_rie['B2'] = 'Nome'
    ws_rie['C2'] = 'Ore Tot.'
    for i, d in enumerate(date_mese):
        col = 4 + i
        c = ws_rie.cell(row=2, column=col)
        c.value = d.day
        c.font = Font(bold=True, size=8)
        c.fill = fill_weekend if d.weekday() >= 5 else fill_header_col
        c.alignment = al_c
        c.border = bordo
        # Sottocella: giorno settimana
    ws_rie.row_dimensions[2].height = 18

    # Riga 3: giorno settimana
    ws_rie['A3'] = ''
    ws_rie['B3'] = ''
    ws_rie['C3'] = ''
    for i, d in enumerate(date_mese):
        col = 4 + i
        c = ws_rie.cell(row=3, column=col)
        c.value = GIORNI_ITA[d.weekday()]
        c.font = Font(size=7, italic=True, color='888888')
        c.fill = fill_weekend if d.weekday() >= 5 else _fill('f4f7fa')
        c.alignment = al_c
        c.border = bordo
    ws_rie.row_dimensions[3].height = 13

    # Righe dati
    for r_idx, dip in enumerate(dipendenti):
        row = 4 + r_idx
        ws_rie.cell(row=row, column=1, value=dip.cognome).font = Font(bold=True, size=9)
        ws_rie.cell(row=row, column=2, value=dip.nome).font = Font(size=9)
        ore_tot = 0
        periodo_start, periodo_end = periodi.get(dip.id, (None, None))
        for i, d in enumerate(date_mese):
            col = 4 + i
            p = pmap.get((dip.id, d))
            c = ws_rie.cell(row=row, column=col)
            c.border = bordo
            c.alignment = al_c
            c.font = Font(size=8)

            in_periodo = bool(periodo_start and periodo_end and periodo_start <= d <= periodo_end)
            if not in_periodo:
                c.value = ''
                c.fill = _fill('f8f9fa')
                continue

            if d.weekday() >= 5:
                c.fill = fill_weekend
            if p:
                ore = _dec2(_minuti_lavorati_presenza(p) / 60.0)
                ore_tot += ore
                cau = p.causale
                if cau in _CAUSALI_ORE_LAVORATE:
                    c.value = ore if ore else cau
                    if ore:
                        c.number_format = '0.00'
                else:
                    c.value = cau
                # Colore sfondo per assenze
                if cau in ('M', 'INF', 'A'):
                    c.fill = fill_assenza
                elif cau in ('F', 'PE', 'MAT', 'CIG'):
                    c.fill = fill_ferie
            else:
                if d.weekday() < 5:
                    c.value = ''
                    c.fill = _fill('fffde7')  # giallo — non registrato

        # Totale ore
        ore_cell = ws_rie.cell(row=row, column=3, value=_dec2(ore_tot))
        ore_cell.font = Font(bold=True, size=9)
        ore_cell.fill = fill_totali
        ore_cell.alignment = al_c
        ore_cell.border = bordo
        ore_cell.number_format = '0.00'
        ws_rie.row_dimensions[row].height = 16

    # Larghezze colonne
    ws_rie.column_dimensions['A'].width = 16
    ws_rie.column_dimensions['B'].width = 13
    ws_rie.column_dimensions['C'].width = 8
    for i in range(num_giorni):
        ws_rie.column_dimensions[get_column_letter(4 + i)].width = 5

    # Legenda causali
    leg_row = 5 + len(dipendenti)
    ws_rie.cell(row=leg_row, column=1, value='LEGENDA CAUSALI').font = Font(bold=True, size=9)
    for j, (cau, label) in enumerate(Presenza.CAUSALE_CHOICES):
        c = ws_rie.cell(row=leg_row + 1 + j // 4, column=1 + (j % 4) * 2)
        c.value = f"{cau} = {label}"
        c.font = Font(size=8)

    # ── Fogli individuali per dipendente ────────────────────────
    for dip in dipendenti:
        ws = wb.create_sheet(title=f"{dip.cognome[:10]}_{dip.nome[:6]}")
        ws.freeze_panes = 'A3'

        # Intestazione
        ws.merge_cells('A1:G1')
        t = ws['A1']
        t.value = f"{dip.cognome} {dip.nome} — {mese_nome} {anno} — {azienda.nome}"
        t.font = st_header_az
        t.fill = fill_header_az
        t.alignment = al_c

        # Intestazioni colonne
        intestazioni = ['Data', 'Giorno', 'Causale', 'Entrata', 'Uscita', 'Ore', 'Straord.', 'Note']
        for ci, h in enumerate(intestazioni):
            c = ws.cell(row=2, column=ci + 1, value=h)
            c.font = Font(bold=True, size=9)
            c.fill = fill_header_col
            c.alignment = al_c
            c.border = bordo
        ws.row_dimensions[2].height = 16

        # Dati
        ore_totali = 0
        periodo_start, periodo_end = periodi.get(dip.id, (None, None))
        for ri, d in enumerate(date_mese):
            row = 3 + ri
            p = pmap.get((dip.id, d))
            is_we = d.weekday() >= 5
            fill_row = fill_weekend if is_we else None

            def _cell(col, val, bold=False, center=False):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(size=9, bold=bold)
                c.border = bordo
                c.alignment = al_c if center else al_l
                if fill_row:
                    c.fill = fill_row
                return c

            _cell(1, d.strftime('%d/%m/%Y'), center=True)
            _cell(2, GIORNI_ITA[d.weekday()], center=True)

            in_periodo = bool(periodo_start and periodo_end and periodo_start <= d <= periodo_end)
            if not in_periodo:
                cfr = ws.cell(row=row, column=3, value='FUORI CONTRATTO')
                cfr.font = Font(size=9, italic=True, color='999999')
                cfr.border = bordo
                cfr.alignment = al_c
                cfr.fill = _fill('f8f9fa')
                for col in range(4, 9):
                    cc = ws.cell(row=row, column=col, value='')
                    cc.border = bordo
                    cc.fill = _fill('f8f9fa')
                ws.row_dimensions[row].height = 15
                continue

            if p:
                ore = _dec2(_minuti_lavorati_presenza(p) / 60.0)
                ore_totali += ore
                cau_c = ws.cell(row=row, column=3, value=p.get_causale_display())
                cau_c.font = Font(size=9, bold=True,
                                   color=Presenza.CAUSALE_COLORI.get(p.causale, '555555').lstrip('#'))
                cau_c.border = bordo
                cau_c.alignment = al_c
                if not is_we:
                    if p.causale in ('M', 'INF', 'A'):
                        cau_c.fill = fill_assenza
                    elif p.causale in ('F', 'PE', 'MAT', 'CIG'):
                        cau_c.fill = fill_ferie
                _cell(4, p.ora_entrata.strftime('%H:%M') if p.ora_entrata else '', center=True)
                _cell(5, p.ora_uscita.strftime('%H:%M') if p.ora_uscita else '', center=True)
                c6 = _cell(6, _dec2(ore) if ore else '', center=True)
                if ore:
                    c6.number_format = '0.00'
                stra = _dec2(p.ore_straordinario)
                c7 = _cell(7, stra if stra else '', center=True)
                if stra:
                    c7.number_format = '0.00'
                _cell(8, p.note or '')
            else:
                _cell(3, '— non registrato —' if not is_we else '')
                for col in range(4, 9):
                    _cell(col, '')
            ws.row_dimensions[row].height = 15

        # Riga totali
        tot_row = 3 + num_giorni
        ws.merge_cells(f'A{tot_row}:E{tot_row}')
        tc = ws[f'A{tot_row}']
        tc.value = 'TOTALE ORE LAVORATE'
        tc.font = Font(bold=True, size=9)
        tc.fill = _fill('1b3a5f')
        tc.font = Font(bold=True, size=9, color='FFFFFF')
        tc.alignment = al_c
        ore_c = ws.cell(row=tot_row, column=6, value=_dec2(ore_totali))
        ore_c.font = Font(bold=True, size=10)
        ore_c.fill = fill_totali
        ore_c.alignment = al_c
        ore_c.border = bordo
        ore_c.number_format = '0.00'

        # Larghezze
        for ci, w in enumerate([12, 8, 20, 8, 8, 7, 8, 25]):
            ws.column_dimensions[get_column_letter(ci + 1)].width = w

    # ── Output ──────────────────────────────────────────────────
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_file = f"presenze_{azienda.nome.replace(' ', '_')}_{mese:02d}_{anno}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{nome_file}"'
    if not pmap:
        registra_evento_anomalia(
            utente=request.user,
            azienda=azienda,
            contesto='export_presenze_excel',
            anomalia={'codice': 'EXPORT_VUOTO', 'messaggio': f'Export Excel vuoto {mese:02d}/{anno}'},
            request=request,
        )
    else:
        registra_log(
            utente=request.user,
            azienda=azienda,
            operazione='export_presenze_excel',
            descrizione=f'Export presenze Excel {mese:02d}/{anno}',
            request=request,
        )
    return resp


# ── Riepilogo mensile motore (HR/Admin) ─────────────────────────────────────

MESI_NOMI = ['', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
             'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre']

def _contratto_attivo(dipendente, azienda, data_rif):
    """Restituisce il RapportoDiLavoro sottoscritto attivo per la data di riferimento."""
    from rapporto_di_lavoro.models import RapportoDiLavoro
    from django.db.models import Q
    return RapportoDiLavoro.objects.filter(
        dipendente=dipendente,
        azienda=azienda,
        stato='sottoscritto',
        data_inizio_rapporto__lte=data_rif,
    ).filter(
        Q(data_fine_rapporto__isnull=True) | Q(data_fine_rapporto__gte=data_rif)
    ).order_by('-data_inizio_rapporto').first()


def _calcola_cedolino_da_riepilogo(riepilogo):
    """
    Chiama il motore paga per un RiepilogoMensilePresenze approvato.
    Restituisce il dict risultato del motore o None se mancano i dati.
    """
    from datetime import date as ddate
    from decimal import Decimal
    from django.db.models import Q
    from rapporto_di_lavoro.models import ParametroCCNLTurismo, CCNL
    from rapporto_di_lavoro.utils_motore_paga import calcola_busta_paga_mese

    data_rif = ddate(riepilogo.anno, riepilogo.mese, 1)
    contratto = _contratto_attivo(riepilogo.dipendente, riepilogo.azienda, data_rif)
    if not contratto:
        return None

    livello = contratto.livello_ccnl
    cp = ParametroCCNLTurismo.objects.filter(
        livello=livello,
        decorrenza_validita_da__lte=data_rif,
    ).order_by('-decorrenza_validita_da').first()
    if not cp:
        return None

    ccnl_obj = CCNL.objects.filter(sigla='FIPE').first()

    kwargs = riepilogo.as_motore_kwargs()
    kwargs.update({
        'parametro_ccnl':    cp,
        'tipo_contratto':    contratto.tipo_contratto,
        'anno':              riepilogo.anno,
        'mese':              riepilogo.mese,
        'azienda':           riepilogo.azienda,
        'data_inizio_rapporto': contratto.data_inizio_rapporto,
        'data_fine_rapporto':   contratto.data_fine_rapporto,
        'divisore_str':      str(round(float(cp.ore_mensili))) if cp.ore_mensili else '173',
        'superminimo':       Decimal('0'),
        'indennita_turno':   Decimal('0'),
        'scatto_anzianita':  Decimal('0'),
        'indennita_extra':   Decimal('0'),
        'ccnl_obj':          ccnl_obj,
        'modalita_ore_effettive': True,
        'contratto_esclude_tredicesima': bool(contratto.tredicesima is False),
        'contratto_esclude_quattordicesima': bool(contratto.quattordicesima is False),
        'rateo_13_mensile_in_imponibile': bool(
            getattr(contratto, 'tredicesima_rateo_mensile_in_imponibile', False)
        ),
        'rateo_14_mensile_in_imponibile': bool(
            getattr(contratto, 'quattordicesima_rateo_mensile_in_imponibile', False)
        ),
    })
    return calcola_busta_paga_mese(**kwargs)


@login_required
def riepilogo_mensile_motore(request):
    """
    Vista admin/HR: gestione riepiloghi mensili presenze.
    Permette di aggregare, cambiare stato e avviare il calcolo cedolino.
    """
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")

    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('profile')

    oggi = date.today()
    anno = _parse_int(request.GET.get('anno'), oggi.year)
    mese = _parse_int(request.GET.get('mese'), oggi.month)

    from .models import RiepilogoMensilePresenze
    from .utils import aggrega_presenze_per_motore

    # ── POST: azioni ─────────────────────────────────────────────────────────
    if request.method == 'POST':
        action   = request.POST.get('action', '')
        dip_id   = _parse_int(request.POST.get('dip_id'), 0)
        anno_p   = _parse_int(request.POST.get('anno'), anno)
        mese_p   = _parse_int(request.POST.get('mese'), mese)
        redirect_url = f"{request.path}?anno={anno_p}&mese={mese_p}"

        if action in ('aggrega', 'aggrega_tutti'):
            if action == 'aggrega' and dip_id:
                dipendenti_da_aggregare = Dipendente.objects.filter(
                    id=dip_id,
                    azienda=azienda,
                    stato='attivo',
                )
            else:
                dipendenti_da_aggregare = Dipendente.objects.filter(
                    azienda=azienda,
                    stato='attivo',
                )
            ok = err = 0
            for dip in dipendenti_da_aggregare:
                try:
                    aggrega_presenze_per_motore(dip, azienda, anno_p, mese_p, utente=request.user)
                    ok += 1
                except ValueError as e:
                    messages.warning(request, str(e))
                    err += 1
            if ok:
                messages.success(request, f"Aggregati {ok} riepilogo/i.")
            return redirect(redirect_url)

        if action == 'cambia_stato' and dip_id:
            nuovo_stato = request.POST.get('nuovo_stato', '')
            rie = RiepilogoMensilePresenze.objects.filter(
                dipendente_id=dip_id, azienda=azienda, anno=anno_p, mese=mese_p
            ).first()
            if rie and nuovo_stato in ('revisione', 'approvata', 'elaborata'):
                stati_consentiti = {
                    'revisione': ['bozza'],
                    'approvata': ['revisione'],
                    'elaborata': ['approvata'],
                }
                if rie.stato in stati_consentiti.get(nuovo_stato, []):
                    if nuovo_stato == 'approvata':
                        rie.approvata_da = request.user
                    rie.stato = nuovo_stato
                    rie.save()
                    if nuovo_stato == 'approvata':
                        from .monte_ledger import applica_movimenti_da_riepilogo

                        applica_movimenti_da_riepilogo(
                            rie,
                            utente=request.user,
                            solo_se_approvato=True,
                        )
                    messages.success(request, f"Stato aggiornato a '{rie.get_stato_display()}'.")
                else:
                    messages.error(request, f"Transizione non consentita da '{rie.get_stato_display()}' a '{nuovo_stato}'.")
            return redirect(redirect_url)

        if action == 'elabora' and dip_id:
            rie = RiepilogoMensilePresenze.objects.filter(
                dipendente_id=dip_id, azienda=azienda, anno=anno_p, mese=mese_p
            ).first()
            if rie and rie.stato == 'approvata':
                return redirect(
                    f"{request.path}anteprima/{dip_id}/{anno_p}/{mese_p}/"
                )
            else:
                messages.error(request, "Il riepilogo deve essere in stato 'Approvata' per avviare l'elaborazione.")
            return redirect(redirect_url)

        if action == 'riapri_mese_presenze' and dip_id:
            rie = RiepilogoMensilePresenze.objects.filter(
                dipendente_id=dip_id, azienda=azienda, anno=anno_p, mese=mese_p
            ).first()
            if rie and rie.stato in ('approvata', 'elaborata'):
                from .monte_ledger import elimina_movimenti_monti_da_riepilogo

                n_mov = elimina_movimenti_monti_da_riepilogo(rie)
                rie.stato = 'bozza'
                rie.approvata_da = None
                ts = timezone.now().strftime('%Y-%m-%d %H:%M')
                un = getattr(request.user, 'username', '') or str(request.user.pk)
                extra = (
                    f"\n[{ts}] Riaperto mese presenze da {un}: stato → bozza; "
                    f"rimossi {n_mov} movimenti monti (ferie/ROL) collegati al riepilogo."
                )
                rie.note = ((rie.note or '') + extra)[:4000]
                rie.save()
                messages.success(
                    request,
                    f"Riepilogo {rie.dipendente.cognome} {rie.dipendente.nome} — {mese_p:02d}/{anno_p} "
                    f"riportato in bozza: puoi modificare le presenze e usare «Aggrega» per ricalcolare. "
                    f"Rimossi {n_mov} movimenti sui monti del mese.",
                )
            elif rie:
                messages.error(
                    request,
                    f"Riapertura disponibile solo da stato Approvata o Elaborata (stato attuale: {rie.get_stato_display()}).",
                )
            else:
                messages.error(request, 'Nessun riepilogo motore per questo dipendente e mese.')
            return redirect(redirect_url)

    # ── GET: costruzione lista ────────────────────────────────────────────────
    dipendenti = Dipendente.objects.filter(
        azienda=azienda,
        stato='attivo',
    ).order_by('cognome', 'nome')

    riepiloghi_map = {
        r.dipendente_id: r
        for r in RiepilogoMensilePresenze.objects.filter(
            azienda=azienda, anno=anno, mese=mese
        ).select_related('dipendente', 'approvata_da')
    }

    righe = []
    for dip in dipendenti:
        rie = riepiloghi_map.get(dip.id)
        righe.append({'dip': dip, 'rie': rie})

    if mese == 1:
        mese_prev, anno_prev = 12, anno - 1
    else:
        mese_prev, anno_prev = mese - 1, anno
    if mese == 12:
        mese_next, anno_next = 1, anno + 1
    else:
        mese_next, anno_next = mese + 1, anno

    return render(request, 'presenze/riepilogo_mensile_motore.html', {
        'anno': anno, 'mese': mese,
        'mese_nome': MESI_NOMI[mese],
        'anno_prev': anno_prev, 'mese_prev': mese_prev,
        'anno_next': anno_next, 'mese_next': mese_next,
        'azienda': azienda,
        'righe': righe,
        'totali': {
            'tot': len(righe),
            'bozza':     sum(1 for r in righe if r['rie'] and r['rie'].stato == 'bozza'),
            'revisione': sum(1 for r in righe if r['rie'] and r['rie'].stato == 'revisione'),
            'approvata': sum(1 for r in righe if r['rie'] and r['rie'].stato == 'approvata'),
            'elaborata': sum(1 for r in righe if r['rie'] and r['rie'].stato == 'elaborata'),
            'assente':   sum(1 for r in righe if not r['rie']),
        },
    })


@login_required
def anteprima_cedolino_riepilogo(request, dip_id, anno, mese):
    """Calcola e mostra il cedolino del motore a partire dal RiepilogoMensilePresenze."""
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden("Accesso negato")

    azienda = _get_azienda(request)
    dip = get_object_or_404(Dipendente, id=dip_id, azienda=azienda)

    from .models import RiepilogoMensilePresenze
    rie = get_object_or_404(RiepilogoMensilePresenze, dipendente=dip, azienda=azienda, anno=anno, mese=mese)

    errore = None
    risultato = None
    if rie.stato not in ('approvata', 'elaborata'):
        errore = "Il riepilogo deve essere in stato 'Approvata' o 'Elaborata' per visualizzare il cedolino."
    else:
        risultato = _calcola_cedolino_da_riepilogo(rie)
        if risultato is None:
            errore = "Nessun contratto attivo trovato per questo dipendente nel mese selezionato, o parametri CCNL mancanti."

    # POST: conferma elaborazione (approvata → elaborata)
    if request.method == 'POST' and request.POST.get('action') == 'conferma_elabora':
        if rie.stato == 'approvata' and risultato:
            rie.stato = 'elaborata'
            rie.save()
            messages.success(request, f"Cedolino confermato ed elaborato per {dip}.")
            return _redirect_riepilogo_mensile_motore(anno, mese)

    return render(request, 'presenze/anteprima_cedolino_riepilogo.html', {
        'dip': dip, 'rie': rie, 'anno': anno, 'mese': mese,
        'mese_nome': MESI_NOMI[mese],
        'risultato': risultato,
        'errore': errore,
        'azienda': azienda,
    })


@login_required
def monti_saldi(request):
    """
    HR/Admin: saldi monti (ferie, ROL, riposi) per anno competenza e impostazione saldo iniziale.
    """
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden('Accesso negato')

    azienda = _get_azienda(request)
    if not azienda:
        messages.warning(request, "Selezionare un'azienda operativa.")
        return redirect('profile')

    from .forms import SaldoMonteInizialeForm
    from .models import SaldoMonteDipendente
    from .monte_ledger import TIPO_MONTE_ORDINE
    from .utils import dipendenti_attivi_con_rapporto_nel_mese

    oggi = date.today()
    anno = _parse_int(request.GET.get('anno'), oggi.year)
    mese_busta = _parse_int(request.GET.get('mese_busta'), oggi.month)
    if mese_busta < 1 or mese_busta > 12:
        mese_busta = oggi.month

    dipendenti = dipendenti_attivi_con_rapporto_nel_mese(azienda, anno, mese_busta)

    form = SaldoMonteInizialeForm()
    form.fields['dipendente'].queryset = dipendenti
    form.fields['anno_competenza'].initial = anno

    if request.method == 'POST' and request.POST.get('action') == 'salva_saldo_iniziale':
        anno = _parse_int(request.POST.get('anno'), anno)
        form = SaldoMonteInizialeForm(request.POST)
        form.fields['dipendente'].queryset = dipendenti
        if form.is_valid():
            cd = form.cleaned_data
            saldo, created = SaldoMonteDipendente.objects.get_or_create(
                dipendente=cd['dipendente'],
                azienda=azienda,
                tipo_monte=cd['tipo_monte'],
                anno_competenza=cd['anno_competenza'],
                defaults={
                    'saldo_iniziale': cd['saldo_iniziale'],
                    'data_saldo_iniziale': cd['data_saldo_iniziale'],
                    'note': cd['note'] or '',
                },
            )
            if not created:
                saldo.saldo_iniziale = cd['saldo_iniziale']
                saldo.data_saldo_iniziale = cd['data_saldo_iniziale']
                saldo.note = cd['note'] or ''
                saldo.save()
            messages.success(request, 'Saldo iniziale salvato.')
            return redirect(f'{request.path}?anno={anno}&mese_busta={mese_busta}')
        messages.error(request, 'Correggere i dati del modulo saldo iniziale.')

    saldi_qs = (
        SaldoMonteDipendente.objects.filter(azienda=azienda, anno_competenza=anno)
        .select_related('dipendente')
        .annotate(mov_sum=Coalesce(Sum('movimenti__quantita'), Value(Decimal('0'))))
    )
    saldi_map = {}
    for s in saldi_qs:
        mov_sum = (s.mov_sum or Decimal('0')).quantize(Decimal('0.01'))
        tot = (s.saldo_iniziale + mov_sum).quantize(Decimal('0.01'))
        saldi_map[(s.dipendente_id, s.tipo_monte)] = {
            'mov_sum': mov_sum,
            'tot': tot,
            'saldo': s,
        }

    tipi_headers = [
        {'label': 'Ferie', 'unita': 'gg'},
        {'label': 'ROL / permessi', 'unita': 'h'},
        {'label': 'Riposi comp.', 'unita': 'gg'},
    ]

    righe = []
    for dip in dipendenti:
        righe.append({
            'dip': dip,
            'cells': [saldi_map.get((dip.id, t)) for t in TIPO_MONTE_ORDINE],
        })

    return render(request, 'presenze/monti_saldi.html', {
        'azienda': azienda,
        'anno': anno,
        'anno_prev': anno - 1,
        'anno_next': anno + 1,
        'mese_busta': mese_busta,
        'mese_busta_nome': MESI_NOMI[mese_busta],
        'mesi_nomi': MESI_NOMI,
        'righe': righe,
        'form': form,
        'tipi_headers': tipi_headers,
    })


@login_required
def export_monti_riconciliazione_csv(request):
    """
    CSV confronto saldo gestionale vs residui su cedolino (CedolinoMotoreV4) per mese busta.
    """
    if not _is_admin_hr(request.user):
        return HttpResponseForbidden('Accesso negato')

    azienda = _get_azienda(request)
    if not azienda:
        return HttpResponse('Azienda non selezionata.', status=400)

    import csv
    from io import StringIO

    from documenti.models import CedolinoMotoreV4

    from .models import SaldoMonteDipendente
    from .monte_ledger import TIPO_MONTE_ORDINE
    from .utils import dipendenti_attivi_con_rapporto_nel_mese

    oggi = date.today()
    anno = _parse_int(request.GET.get('anno'), oggi.year)
    mese = _parse_int(request.GET.get('mese'), oggi.month)
    if mese < 1 or mese > 12:
        mese = oggi.month

    ced_map = {
        c.dipendente_id: c
        for c in CedolinoMotoreV4.objects.filter(
            dipendente__azienda=azienda,
            anno=anno,
            mese=mese,
            natura_busta='ORDINARIA',
        )
    }

    saldi_qs = (
        SaldoMonteDipendente.objects.filter(azienda=azienda, anno_competenza=anno)
        .select_related('dipendente')
        .annotate(mov_sum=Coalesce(Sum('movimenti__quantita'), Value(Decimal('0'))))
    )
    saldi_dict = {}
    for s in saldi_qs:
        mov_sum = (s.mov_sum or Decimal('0')).quantize(Decimal('0.01'))
        tot = (s.saldo_iniziale + mov_sum).quantize(Decimal('0.01'))
        saldi_dict[(s.dipendente_id, s.tipo_monte)] = (s, mov_sum, tot)

    def _residuo_consulente(ced, tipo_monte):
        if not ced:
            return None
        if tipo_monte == 'FERIE_GG':
            return ced.ferie_res
        if tipo_monte == 'ROL_ORE':
            return ced.rol_res
        if tipo_monte == 'RIPOSI_COMP':
            return ced.fest_res
        return None

    def _fmt_it(num):
        return str(num).replace('.', ',')

    buf = StringIO()
    w = csv.writer(buf, delimiter=';')
    w.writerow([
        'Cognome', 'Nome', 'Matricola', 'Tipo monte', 'Unita',
        'Saldo iniziale', 'Somma movimenti', 'Saldo gestionale',
        'Residuo consulente (busta)', 'Delta (gest. - consulente)',
    ])

    dipendenti = dipendenti_attivi_con_rapporto_nel_mese(azienda, anno, mese)
    etichette = dict(SaldoMonteDipendente.TIPO_MONTE_CHOICES)

    for dip in dipendenti:
        ced = ced_map.get(dip.id)
        for tipo in TIPO_MONTE_ORDINE:
            row = saldi_dict.get((dip.id, tipo))
            if row:
                s, mov_sum, tot = row
                ini = s.saldo_iniziale
            else:
                s = None
                mov_sum = Decimal('0')
                tot = Decimal('0')
                ini = Decimal('0')

            rb = _residuo_consulente(ced, tipo)
            if rb is not None:
                rbq = Decimal(str(rb)).quantize(Decimal('0.01'))
                delta = (tot - rbq).quantize(Decimal('0.01'))
            else:
                rbq = None
                delta = None

            unita = 'GG' if tipo == 'FERIE_GG' else ('ORE' if tipo == 'ROL_ORE' else 'GG')

            w.writerow([
                dip.cognome,
                dip.nome,
                dip.matricola if dip.matricola is not None else '',
                etichette.get(tipo, tipo),
                unita,
                _fmt_it(ini),
                _fmt_it(mov_sum),
                _fmt_it(tot),
                _fmt_it(rbq) if rbq is not None else '',
                _fmt_it(delta) if delta is not None else '',
            ])

    nome = f"monti_riconciliazione_{azienda.nome.replace(' ', '_')}_{mese:02d}_{anno}.csv"
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{nome}"'
    return resp


# ── Viste legacy ────────────────────────────────────────────────────────────
@login_required
def lista_presenze(request):
    return redirect('lista_dipendenti_presenze')


@login_required
def seleziona_dipendente_presenza(request):
    azienda = _get_azienda(request)
    if azienda:
        dipendenti = Dipendente.objects.filter(azienda=azienda, stato__in=['attivo', 'candidato']).order_by('cognome', 'nome')
        if dipendenti.count() == 1:
            return redirect('calendario_presenze', dipendente_id=dipendenti.first().id)
    return redirect('lista_dipendenti_presenze')


@login_required
def inserisci_presenza(request, dipendente_id):
    return redirect('calendario_presenze', dipendente_id=dipendente_id)


@login_required
def riepilogo_presenze_mese(request, anno, mese):
    return redirect(f"{request.path.replace(str(anno)+'/'+str(mese), '')}?anno={anno}&mese={mese}")


