from io import StringIO
from collections import defaultdict

from django.contrib import messages
from django.core.management import call_command
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseForbidden, HttpResponse
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.views.decorators.http import require_POST
from .models import EventoStorico, LibroPaga, RegistroUnico, StoricoAccessi, LibroPagaStorico
from anagrafiche.models import Dipendente, Azienda
from anagrafiche.permissions import hr_required, admin_required
from documenti.models import VoceCedolinoMotoreV4
import csv
import openpyxl
from decimal import Decimal


def _libro_sum_attr(lista, attr):
    vals = [getattr(v, attr) for v in lista if getattr(v, attr, None) is not None]
    if not vals:
        return None
    return sum(vals[1:], start=vals[0]) if len(vals) > 1 else vals[0]


def _totali_blocco_libro_paga(lista):
    """Totali per dipendente / anno (tutti i campi numerici del libro)."""
    return {
        'totale_ore_ord': _libro_sum_attr(lista, 'ore_ordinarie'),
        'totale_ore_str': _libro_sum_attr(lista, 'ore_straordinario'),
        'totale_ore_ass': _libro_sum_attr(lista, 'ore_assenza'),
        'totale_retribuzione_base': _libro_sum_attr(lista, 'retribuzione_base'),
        'totale_indennita': _libro_sum_attr(lista, 'indennita_accessorie'),
        'totale_lordo': _libro_sum_attr(lista, 'lordo_mensile'),
        'totale_inps_dip': _libro_sum_attr(lista, 'inps_dipendente'),
        'totale_irpef': _libro_sum_attr(lista, 'irpef'),
        'totale_addizionali': _libro_sum_attr(lista, 'addizionali'),
        'totale_altre_trattenute': _libro_sum_attr(lista, 'altre_trattenute'),
        'totale_ti': _libro_sum_attr(lista, 'trattamento_integrativo'),
        'totale_netto': sum((v.importo for v in lista), start=Decimal('0')),
        'totale_tfr': _libro_sum_attr(lista, 'tfr_mensile'),
        'totale_rateo_13': _libro_sum_attr(lista, 'rateo_13'),
        'totale_rateo_14': _libro_sum_attr(lista, 'rateo_14'),
        'totale_inps_az': _libro_sum_attr(lista, 'inps_azienda'),
        'totale_inail': _libro_sum_attr(lista, 'inail_azienda'),
        'totale_costo': _libro_sum_attr(lista, 'costo_azienda'),
    }


def _libro_paga_storico_excel_response(qs):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Libro Unico'
    headers = [
        'Cognome', 'Nome', 'Codice fiscale', 'Azienda', 'Periodo', 'Data pagamento',
        'Ore ordinarie', 'Ore straord.', 'Ore assenza',
        'Retrib. base €', 'Indenn. accessorie €', 'Lordo €', 'INPS dip. €', 'IRPEF €',
        'Addizionali €', 'Altre trattenute €', 'TI/Bonus €', 'Netto €',
        'TFR acc. €', 'Rateo 13ª €', 'Rateo 14ª €', 'INPS az. €', 'INAIL az. €', 'Costo az. €',
        'Fonte', 'Note', 'Livello CCNL', 'Qualifica', 'Tipo contratto',
    ]
    ws.append(headers)
    for v in qs.select_related('dipendente', 'azienda').order_by(
        'dipendente__cognome', 'dipendente__nome', 'ordinamento', 'data_pagamento'
    ):
        dip = v.dipendente
        ws.append([
            dip.cognome or '',
            dip.nome or '',
            getattr(dip, 'codice_fiscale', None) or '',
            str(v.azienda) if v.azienda_id else '',
            v.periodo_riferimento or '',
            v.data_pagamento.strftime('%d/%m/%Y') if v.data_pagamento else '',
            float(v.ore_ordinarie) if v.ore_ordinarie is not None else None,
            float(v.ore_straordinario) if v.ore_straordinario is not None else None,
            float(v.ore_assenza) if v.ore_assenza is not None else None,
            float(v.retribuzione_base) if v.retribuzione_base is not None else None,
            float(v.indennita_accessorie) if v.indennita_accessorie is not None else None,
            float(v.lordo_mensile) if v.lordo_mensile is not None else None,
            float(v.inps_dipendente) if v.inps_dipendente is not None else None,
            float(v.irpef) if v.irpef is not None else None,
            float(v.addizionali) if v.addizionali is not None else None,
            float(v.altre_trattenute) if v.altre_trattenute is not None else None,
            float(v.trattamento_integrativo) if v.trattamento_integrativo is not None else None,
            float(v.importo),
            float(v.tfr_mensile) if v.tfr_mensile is not None else None,
            float(v.rateo_13) if v.rateo_13 is not None else None,
            float(v.rateo_14) if v.rateo_14 is not None else None,
            float(v.inps_azienda) if v.inps_azienda is not None else None,
            float(v.inail_azienda) if v.inail_azienda is not None else None,
            float(v.costo_azienda) if v.costo_azienda is not None else None,
            v.get_fonte_dati_display() if hasattr(v, 'get_fonte_dati_display') else v.fonte_dati,
            (v.note or '')[:500],
            v.livello_ccnl or '',
            v.qualifica or '',
            v.tipo_contratto or '',
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="libro_unico_lavoro.xlsx"'
    wb.save(response)
    return response


def _serializza_attivita(accessi_qs, eventi_qs):
    attivita = []
    for s in accessi_qs:
        attivita.append({
            'quando': s.data_azione,
            'utente': str(s.utente),
            'categoria': 'Accesso',
            'azione': s.azione,
            'descrizione': s.descrizione,
        })
    for e in eventi_qs:
        attivita.append({
            'quando': e.data_evento,
            'utente': str(getattr(e.dipendente, 'utente', '') or e.dipendente),
            'categoria': f"Evento {e.get_tipo_display()}",
            'azione': e.get_tipo_display(),
            'descrizione': e.descrizione,
        })
    attivita.sort(key=lambda x: x['quando'], reverse=True)
    return attivita


def _render_storico_attivita(
    request,
    accessi_qs,
    eventi_qs,
    azienda=None,
    dipendente=None,
    dipendenti_filtro=None,
    mostra_filtro_dipendente=False,
):
    attivita = _serializza_attivita(accessi_qs, eventi_qs)
    export = request.GET.get('export')
    if export == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="storico_attivita.csv"'
        writer = csv.writer(response)
        writer.writerow(['Utente', 'Data/Ora', 'Categoria', 'Azione', 'Descrizione'])
        for r in attivita:
            writer.writerow([
                r['utente'],
                r['quando'].strftime('%d/%m/%Y %H:%M') if r['quando'] else '',
                r['categoria'],
                r['azione'],
                r['descrizione'],
            ])
        return response
    elif export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(['Utente', 'Data/Ora', 'Categoria', 'Azione', 'Descrizione'])
        for r in attivita:
            ws.append([
                r['utente'],
                r['quando'].strftime('%d/%m/%Y %H:%M') if r['quando'] else '',
                r['categoria'],
                r['azione'],
                r['descrizione'],
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="storico_attivita.xlsx"'
        wb.save(response)
        return response

    return render(request, 'storico/storico_accessi_lista.html', {
        'attivita': attivita,
        'azienda': azienda,
        'dipendente': dipendente,
        'titolo_storico': 'Storico Attività',
        'filtro_dipendente_id': request.GET.get('dipendente_id', ''),
        'dipendenti_filtro': dipendenti_filtro or [],
        'mostra_filtro_dipendente': mostra_filtro_dipendente,
    })

@hr_required
def report_storico_periodo(request, azienda_id, data_inizio, data_fine):
    eventi = EventoStorico.objects.filter(
        azienda_id=azienda_id,
        data_evento__gte=data_inizio,
        data_evento__lte=data_fine
    ).order_by('data_evento')
    return render(request, 'storico/report_periodo.html', {'eventi': eventi, 'data_inizio': data_inizio, 'data_fine': data_fine})

def storico_generale(request):
    storici = EventoStorico.objects.all()
    return render(request, 'storico/lista_generale.html', {'storici': storici})

def storico_dipendente(request, pk=None):
    if pk:
        dipendente = get_object_or_404(Dipendente, pk=pk)
        storici = EventoStorico.objects.filter(dipendente=dipendente)
        return render(request, 'storico/lista_dipendente.html', {'storici': storici, 'dipendente': dipendente})
    else:
        storici = EventoStorico.objects.all()
        return render(request, 'storico/lista_generale.html', {'storici': storici})

# LIBRO PAGA

def libro_paga_lista(request):
    from itertools import groupby
    from anagrafiche.models import Azienda as AziendaModel
    from accounts.tenant import get_azienda_operativa

    # Filtri opzionali
    azienda_id = request.GET.get('azienda_id')
    dipendente_id = request.GET.get('dipendente_id')
    if not azienda_id:
        az_op = get_azienda_operativa(request.user, request.session)
        if az_op is not None:
            azienda_id = str(az_op.id)

    qs = LibroPagaStorico.objects.select_related('dipendente', 'azienda').order_by(
        'dipendente__cognome', 'dipendente__nome',
    )
    if azienda_id:
        qs = qs.filter(azienda_id=azienda_id)
    if dipendente_id:
        qs = qs.filter(dipendente_id=dipendente_id)

    if request.GET.get('export') == 'excel':
        return _libro_paga_storico_excel_response(qs)

    def _parse_periodo_mm_yyyy(periodo: str):
        """Converte periodo MM/YYYY in (mese, anno)."""
        try:
            if not periodo or len(periodo) < 7:
                return None
            mese = int(periodo[:2])
            anno = int(periodo[3:7])
            if 1 <= mese <= 12:
                return mese, anno
        except (TypeError, ValueError):
            return None
        return None

    periodi_target = set()
    for voce in qs:
        p = _parse_periodo_mm_yyyy(voce.periodo_riferimento or '')
        if p:
            periodi_target.add((voce.dipendente_id, p[0], p[1]))

    voci_motore_by_key = defaultdict(list)
    if periodi_target:
        dip_ids = sorted({k[0] for k in periodi_target})
        mesi = sorted({k[1] for k in periodi_target})
        anni = sorted({k[2] for k in periodi_target})
        qs_voci_motore = (
            VoceCedolinoMotoreV4.objects.select_related('cedolino')
            .filter(
                cedolino__dipendente_id__in=dip_ids,
                cedolino__mese__in=mesi,
                cedolino__anno__in=anni,
            )
            .order_by(
                'cedolino__dipendente_id',
                'cedolino__anno',
                'cedolino__mese',
                'cedolino__natura_busta',
                'pk',
            )
        )
        for vm in qs_voci_motore:
            key = (vm.cedolino.dipendente_id, vm.cedolino.mese, vm.cedolino.anno)
            if key in periodi_target:
                voci_motore_by_key[key].append(vm)

    # Raggruppa per dipendente
    dipendenti_gruppi = []
    def _chiave_periodo(v):
        """Ordina MM/YYYY come (anno, mese) intero."""
        p = v.periodo_riferimento or ''
        try:
            return (int(p[3:7]), int(p[0:2]))
        except (ValueError, IndexError):
            return (9999, 99)

    def _anno_da_periodo(voce):
        p = voce.periodo_riferimento or ''
        try:
            return int(p[3:7])
        except (ValueError, IndexError):
            return 9999

    def _fonte_calcolo_libro(voci_motore):
        if not voci_motore:
            return '—'
        codici = {str(getattr(vm, 'codice', '') or '').strip() for vm in voci_motore}
        tipi = {str(getattr(vm, 'tipo', '') or '').strip().upper() for vm in voci_motore}
        parti = []
        if '8001' in codici:
            parti.append('Ore/Ret.base: voce 8001')
        if any(
            str(getattr(vm, 'tipo', '') or '').strip().upper() == 'COMPETENZA'
            and str(getattr(vm, 'codice', '') or '').strip() != '8001'
            for vm in voci_motore
        ):
            parti.append('Indennità: competenze residue')
        if 'BONUS' in tipi:
            parti.append('TI/Bonus: voci BONUS')
        if any(c in codici for c in {'1800', '1802', '1812', '800', '802'}):
            parti.append('Addiz.: codici 1800/1802/1812')
        if any(
            str(getattr(vm, 'tipo', '') or '').strip().upper() == 'TRATTENUTA'
            and str(getattr(vm, 'codice', '') or '').strip() not in {'1800', '1802', '1812', '800', '802'}
            for vm in voci_motore
        ):
            parti.append('Altre tratt.: trattenute residue')
        return ' · '.join(parti) if parti else 'Voci motore v4'

    for dip, voci_iter in groupby(qs, key=lambda v: v.dipendente):
        voci = sorted(voci_iter, key=_chiave_periodo)
        # Inizio/fine rapporto: prendi il minimo/massimo tra tutte le voci del dipendente
        date_inizio = [v.data_inizio_rapporto for v in voci if v.data_inizio_rapporto]
        date_fine = [v.data_fine_rapporto for v in voci if v.data_fine_rapporto]
        # Totali cumulativi per il riepilogo finale
        totali_dip = _totali_blocco_libro_paga(voci)
        # Sottogruppi per anno (periodo MM/YYYY), anni più recenti per primi
        per_anno_raw = []
        for anno, voci_anno_iter in groupby(voci, key=_anno_da_periodo):
            lista_a = list(voci_anno_iter)
            for voce in lista_a:
                p = _parse_periodo_mm_yyyy(voce.periodo_riferimento or '')
                voce.voci_motore = (
                    voci_motore_by_key.get((voce.dipendente_id, p[0], p[1]), []) if p else []
                )
                voce.fonte_calcolo = _fonte_calcolo_libro(voce.voci_motore)
            t = _totali_blocco_libro_paga(lista_a)
            per_anno_raw.append({
                'anno': anno,
                'anno_label': str(anno) if anno < 9000 else 'Altro periodo',
                'voci': lista_a,
                **t,
            })
        per_anno_valid = [b for b in per_anno_raw if b['anno'] < 9000]
        per_anno_valid.sort(key=lambda b: b['anno'], reverse=True)
        per_anno_altro = [b for b in per_anno_raw if b['anno'] >= 9000]
        dipendenti_gruppi.append({
            'dipendente': dip,
            'voci': voci,
            'per_anno': per_anno_valid + per_anno_altro,
            'data_inizio_rapporto': min(date_inizio) if date_inizio else None,
            'data_fine_rapporto': max(date_fine) if date_fine else None,
            **totali_dip,
        })

    aziende = AziendaModel.objects.order_by('nome')
    return render(request, 'storico/libro_paga_lista.html', {
        'dipendenti_gruppi': dipendenti_gruppi,
        'aziende': aziende,
        'azienda_id_sel': azienda_id,
        'dipendente_id_sel': dipendente_id,
        'totale_voci': qs.count(),
    })


def _puo_sincronizzare_libro_paga(user):
    return user.is_authenticated and (
        user.is_superuser
        or user.has_ruolo('admin')
        or user.has_ruolo('hr')
        or user.has_ruolo('consulente')
    )


ANNO_ESTRAZIONE_BUSTE_PAGA = 2026


@login_required
@admin_required
def estrazione_dati_buste_paga(request):
    """
    Solo admin: elenco solo etichette (testi campo) da tutto il PDF busta paga, senza valori/importi.
    Elabora solo documenti il cui periodo (da descrizione / fallback data) è ANNO_ESTRAZIONE_BUSTE_PAGA.
    """
    from storico.management.commands import popolalibropaga as plp
    from documenti.models import Documento
    from documenti.views import estrai_voci_descrittive_busta_paga_pdf

    lim_raw = request.GET.get('limit')
    try:
        lim = int(lim_raw) if lim_raw else None
    except (TypeError, ValueError):
        lim = None
    if lim is not None and lim < 1:
        lim = None
    if lim is not None and lim > 2000:
        lim = 2000

    docs_all = list(
        Documento.objects.filter(tipo='busta_paga')
        .select_related('dipendente', 'azienda')
        .order_by('-data_caricamento', 'pk')
    )
    docs_filtrati = []
    for d in docs_all:
        mese, anno = plp._parse_periodo_busta(d)
        if anno == ANNO_ESTRAZIONE_BUSTE_PAGA:
            docs_filtrati.append((d, mese, anno))

    totale_anno = len(docs_filtrati)
    if lim is not None:
        docs_filtrati = docs_filtrati[:lim]

    stats = {
        'con_file': 0,
        'pdf_ok': 0,
        'errori': 0,
        'senza_dipendente': 0,
        'tot_voci': 0,
        'scartate_non_2026': max(0, len(docs_all) - totale_anno),
    }

    righe = []

    vuoto_estr = {
        'ok': False,
        'errore': 'Senza file',
        'num_pagine': 0,
        'metodo': None,
        'voci': [],
        'n_voci': 0,
        'anteprima_testo': '',
        'solo_etichette': True,
    }

    for doc, mese, anno in docs_filtrati:
        ha_file = bool(getattr(doc, 'file', None) and getattr(doc.file, 'name', None))
        if ha_file:
            stats['con_file'] += 1
        if not doc.dipendente_id:
            stats['senza_dipendente'] += 1

        estr = dict(vuoto_estr)
        if ha_file:
            try:
                estr = estrai_voci_descrittive_busta_paga_pdf(doc, solo_etichette=True)
            except Exception as exc:
                estr = dict(vuoto_estr)
                estr['errore'] = str(exc)
                stats['errori'] += 1
            else:
                if estr.get('ok'):
                    stats['pdf_ok'] += 1
        voci = estr.get('voci') or []
        stats['tot_voci'] += len(voci)

        righe.append({
            'doc': doc,
            'estr': estr,
            'mese': mese,
            'anno': anno,
            'voci_breve': voci[:10],
            'n_voci': len(voci),
        })

    return render(request, 'storico/estrazione_dati_buste_paga.html', {
        'righe': righe,
        'stats': stats,
        'totale_anno': totale_anno,
        'totale_buste_archivio': len(docs_all),
        'limite_applicato': lim,
        'anno_filtro': ANNO_ESTRAZIONE_BUSTE_PAGA,
    })


@login_required
@require_POST
def libro_paga_svuota(request):
    """Elimina le voci LibroPagaStorico (tutto o per azienda). HR/consulente solo la propria azienda."""
    if not _puo_sincronizzare_libro_paga(request.user):
        return HttpResponseForbidden("Permesso negato.")

    scope_all = (request.POST.get("scope") or "").strip().lower() == "tutto"
    aid_raw = (request.POST.get("azienda_id") or "").strip()

    def _redirect_con_filtri():
        q = []
        if aid_raw.isdigit():
            q.append(f"azienda_id={aid_raw}")
        did = (request.POST.get("dipendente_id") or "").strip()
        if did.isdigit():
            q.append(f"dipendente_id={did}")
        if q:
            return redirect(f"{reverse('libro_paga_lista')}?{'&'.join(q)}")
        return redirect("libro_paga_lista")

    qs = LibroPagaStorico.objects.all()

    if scope_all:
        if not (request.user.is_superuser or request.user.has_ruolo("admin")):
            messages.error(
                request,
                "Solo gli amministratori di piattaforma possono svuotare il libro di tutte le aziende.",
            )
            return _redirect_con_filtri()
        n = qs.count()
        qs.delete()
        messages.success(
            request,
            f"Libro Unico svuotato: eliminate {n} voci (tutte le aziende).",
        )
        return redirect("libro_paga_lista")

    if not aid_raw.isdigit():
        from accounts.tenant import get_azienda_operativa
        az_op = get_azienda_operativa(request.user, request.session)
        if az_op is not None:
            aid_raw = str(az_op.id)

    if not aid_raw.isdigit():
        messages.error(
            request,
            "Seleziona un'azienda dal filtro oppure, come amministratore, usa «Svuota tutte le aziende».",
        )
        return _redirect_con_filtri()

    aid = int(aid_raw)
    if not (request.user.is_superuser or request.user.has_ruolo("admin")):
        ua = getattr(request.user, "azienda_id", None)
        if ua != aid:
            messages.error(
                request,
                "Puoi svuotare il libro solo per la tua azienda.",
            )
            return _redirect_con_filtri()

    n = qs.filter(azienda_id=aid).count()
    qs.filter(azienda_id=aid).delete()
    messages.success(
        request,
        f"Libro Unico svuotato per l'azienda selezionata: {n} voci eliminate.",
    )
    return _redirect_con_filtri()


@login_required
@require_POST
def libro_paga_ricarica_da_buste(request):
    """Esegue manage.py popolalibropaga da web: buste=solo Documenti; completo=Documenti+Simulazioni (esclude movimenti orfani)."""
    if not _puo_sincronizzare_libro_paga(request.user):
        return HttpResponseForbidden('Permesso negato.')

    modo = (request.POST.get('modo') or 'buste').strip().lower()
    buf = StringIO()
    try:
        if modo == 'completo':
            call_command('popolalibropaga', stdout=buf, stderr=buf)
        else:
            call_command('popolalibropaga', solo_documenti=True, stdout=buf, stderr=buf)
    except Exception as exc:
        messages.error(request, f'Sincronizzazione non riuscita: {exc}')
    else:
        raw = buf.getvalue().strip()
        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        summary = ' · '.join(lines[-6:]) if lines else 'Operazione completata.'
        if len(summary) > 2000:
            summary = summary[:1997] + '…'
        messages.success(request, summary)

    redir = 'libro_paga_lista'
    q = []
    aid = request.POST.get('azienda_id')
    did = request.POST.get('dipendente_id')
    if aid:
        q.append(f'azienda_id={aid}')
    if did:
        q.append(f'dipendente_id={did}')
    if q:
        return redirect(f"{reverse(redir)}?{'&'.join(q)}")
    return redirect(redir)


def libro_paga_azienda(request, azienda_id):
    azienda = get_object_or_404(Azienda, pk=azienda_id)
    libri = LibroPaga.objects.filter(azienda=azienda)
    export = request.GET.get('export')
    if export == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="libro_paga_azienda.csv"'
        writer = csv.writer(response)
        writer.writerow(['Dipendente', 'Azienda', 'Mese', 'Anno', 'Stipendio', 'Contributi', 'Trattenute', 'Netto', 'Data Creazione'])
        for libro in libri:
            writer.writerow([
                str(libro.dipendente), str(libro.azienda), libro.mese, libro.anno,
                libro.stipendio, libro.contributi, libro.trattenute, libro.netto,
                libro.data_creazione.strftime('%d/%m/%Y')
            ])
        return response
    elif export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(['Dipendente', 'Azienda', 'Mese', 'Anno', 'Stipendio', 'Contributi', 'Trattenute', 'Netto', 'Data Creazione'])
        for libro in libri:
            ws.append([
                str(libro.dipendente), str(libro.azienda), libro.mese, libro.anno,
                libro.stipendio, libro.contributi, libro.trattenute, libro.netto,
                libro.data_creazione.strftime('%d/%m/%Y')
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="libro_paga_azienda.xlsx"'
        wb.save(response)
        return response
    return render(request, 'storico/libro_paga_lista.html', {'libri': libri, 'azienda': azienda})

def libro_paga_dipendente(request, dipendente_id):
    dipendente = get_object_or_404(Dipendente, pk=dipendente_id)
    libri = LibroPaga.objects.filter(dipendente=dipendente)
    export = request.GET.get('export')
    if export == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="libro_paga_dipendente.csv"'
        writer = csv.writer(response)
        writer.writerow(['Dipendente', 'Azienda', 'Mese', 'Anno', 'Stipendio', 'Contributi', 'Trattenute', 'Netto', 'Data Creazione'])
        for libro in libri:
            writer.writerow([
                str(libro.dipendente), str(libro.azienda), libro.mese, libro.anno,
                libro.stipendio, libro.contributi, libro.trattenute, libro.netto,
                libro.data_creazione.strftime('%d/%m/%Y')
            ])
        return response
    elif export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(['Dipendente', 'Azienda', 'Mese', 'Anno', 'Stipendio', 'Contributi', 'Trattenute', 'Netto', 'Data Creazione'])
        for libro in libri:
            ws.append([
                str(libro.dipendente), str(libro.azienda), libro.mese, libro.anno,
                libro.stipendio, libro.contributi, libro.trattenute, libro.netto,
                libro.data_creazione.strftime('%d/%m/%Y')
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="libro_paga_dipendente.xlsx"'
        wb.save(response)
        return response
    return render(request, 'storico/libro_paga_lista.html', {'libri': libri, 'dipendente': dipendente})

# REGISTRO UNICO

def registro_unico_lista(request):
    registri = RegistroUnico.objects.all()
    export = request.GET.get('export')
    if export == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="registro_unico.csv"'
        writer = csv.writer(response)
        writer.writerow(['Utente', 'Azienda', 'Data Accesso', 'Descrizione'])
        for registro in registri:
            writer.writerow([
                str(registro.utente), str(registro.azienda), registro.data_accesso.strftime('%d/%m/%Y %H:%M'), registro.descrizione
            ])
        return response
    elif export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(['Utente', 'Azienda', 'Data Accesso', 'Descrizione'])
        for registro in registri:
            ws.append([
                str(registro.utente), str(registro.azienda), registro.data_accesso.strftime('%d/%m/%Y %H:%M'), registro.descrizione
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="registro_unico.xlsx"'
        wb.save(response)
        return response
    return render(request, 'storico/registro_unico_lista.html', {'registri': registri})

def registro_unico_azienda(request, azienda_id):
    azienda = get_object_or_404(Azienda, pk=azienda_id)
    registri = RegistroUnico.objects.filter(azienda=azienda)
    export = request.GET.get('export')
    if export == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="registro_unico_azienda.csv"'
        writer = csv.writer(response)
        writer.writerow(['Utente', 'Azienda', 'Data Accesso', 'Descrizione'])
        for registro in registri:
            writer.writerow([
                str(registro.utente), str(registro.azienda), registro.data_accesso.strftime('%d/%m/%Y %H:%M'), registro.descrizione
            ])
        return response
    elif export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(['Utente', 'Azienda', 'Data Accesso', 'Descrizione'])
        for registro in registri:
            ws.append([
                str(registro.utente), str(registro.azienda), registro.data_accesso.strftime('%d/%m/%Y %H:%M'), registro.descrizione
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="registro_unico_azienda.xlsx"'
        wb.save(response)
        return response
    return render(request, 'storico/registro_unico_lista.html', {'registri': registri, 'azienda': azienda})

def registro_unico_dipendente(request, dipendente_id):
    dipendente = get_object_or_404(Dipendente, pk=dipendente_id)
    registri = RegistroUnico.objects.filter(utente__dipendente=dipendente)
    export = request.GET.get('export')
    if export == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="registro_unico_dipendente.csv"'
        writer = csv.writer(response)
        writer.writerow(['Utente', 'Azienda', 'Data Accesso', 'Descrizione'])
        for registro in registri:
            writer.writerow([
                str(registro.utente), str(registro.azienda), registro.data_accesso.strftime('%d/%m/%Y %H:%M'), registro.descrizione
            ])
        return response
    elif export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(['Utente', 'Azienda', 'Data Accesso', 'Descrizione'])
        for registro in registri:
            ws.append([
                str(registro.utente), str(registro.azienda), registro.data_accesso.strftime('%d/%m/%Y %H:%M'), registro.descrizione
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="registro_unico_dipendente.xlsx"'
        wb.save(response)
        return response
    return render(request, 'storico/registro_unico_lista.html', {'registri': registri, 'dipendente': dipendente})

# STORICO ACCESSI

@login_required
def storico_accessi_lista(request):
    user = request.user
    ruolo = getattr(user, 'ruolo', None)

    # Dipendente/candidato: solo attività personali
    if ruolo in ('dipendente', 'candidato'):
        dip = Dipendente.objects.filter(utente=user).first()
        accessi = StoricoAccessi.objects.filter(utente=user)
        eventi = EventoStorico.objects.filter(dipendente=dip) if dip else EventoStorico.objects.none()
        return _render_storico_attivita(request, accessi, eventi, dipendente=dip)

    # Datore/admin/hr/superuser: attività azienda (superuser = tutte)
    if user.is_superuser:
        accessi = StoricoAccessi.objects.all()
        eventi = EventoStorico.objects.all()
        dipendenti_filtro = Dipendente.objects.all().order_by('cognome', 'nome')
    else:
        azienda = getattr(user, 'azienda', None)
        if ruolo not in ('admin', 'hr') or not azienda:
            return HttpResponseForbidden("Accesso negato")
        accessi = StoricoAccessi.objects.filter(utente__azienda=azienda)
        eventi = EventoStorico.objects.filter(azienda=azienda)
        dipendenti_filtro = Dipendente.objects.filter(azienda=azienda).order_by('cognome', 'nome')

    dipendente_id = request.GET.get('dipendente_id')
    dipendente_sel = None
    if dipendente_id:
        try:
            dipendente_sel = dipendenti_filtro.get(pk=int(dipendente_id))
            eventi = eventi.filter(dipendente=dipendente_sel)
            if getattr(dipendente_sel, 'utente_id', None):
                accessi = accessi.filter(utente=dipendente_sel.utente)
            else:
                accessi = accessi.none()
        except (ValueError, Dipendente.DoesNotExist):
            dipendente_sel = None

    return _render_storico_attivita(
        request,
        accessi,
        eventi,
        dipendente=dipendente_sel,
        dipendenti_filtro=dipendenti_filtro,
        mostra_filtro_dipendente=True,
    )

def storico_accessi_azienda(request, azienda_id):
    azienda = get_object_or_404(Azienda, pk=azienda_id)
    accessi = StoricoAccessi.objects.filter(utente__azienda=azienda)
    eventi = EventoStorico.objects.filter(azienda=azienda)
    return _render_storico_attivita(request, accessi, eventi, azienda=azienda)

def storico_accessi_dipendente(request, dipendente_id):
    dipendente = get_object_or_404(Dipendente, pk=dipendente_id)
    accessi = StoricoAccessi.objects.filter(utente=dipendente.utente) if getattr(dipendente, 'utente', None) else StoricoAccessi.objects.none()
    eventi = EventoStorico.objects.filter(dipendente=dipendente)
    return _render_storico_attivita(request, accessi, eventi, dipendente=dipendente)
